"""Spreadsheet domain loader."""

import csv
import os
import re
from typing import Optional

from openpyxl import Workbook, load_workbook

from .world import SpreadsheetWorld
from ...log.logger_registry import LoggerRegistry


logger = LoggerRegistry.setup_logger(__name__)


def load_world(paths: list[str], output_path: str, progress_logger=None) -> SpreadsheetWorld:
    """Load spreadsheet files into a world object."""
    workbooks = {}
    for path in paths:
        logger.info(f"Loading spreadsheet file: {path}")
        workbook = ExcelLoader.load_workbook_from_path(path)
        workbooks[path] = workbook

    if progress_logger:
        progress_logger.log(
            f"📊 [Excel] Loaded {len(paths)} file(s)"
        )

    primary_path = paths[0] if paths else ""
    return SpreadsheetWorld(workbooks, output_path, primary_path)


class ExcelLoader:
    """Loads Excel and CSV files into openpyxl workbooks."""

    _XML_ESCAPE_RE = re.compile(r"_x[0-9A-Fa-f]{4}_")

    @staticmethod
    def load_workbook_from_path(excel_path: str):
        _, ext = os.path.splitext(excel_path)
        ext = ext.lower()
        excel_extensions = {".xlsx", ".xlsm", ".xltx", ".xltm"}

        if ext in excel_extensions:
            workbook = load_workbook(excel_path, data_only=True)
            ExcelLoader.normalize_workbook_strings(workbook)
            return workbook

        if ext == ".csv":
            workbook = ExcelLoader.create_workbook_from_csv(excel_path)
            ExcelLoader.normalize_workbook_strings(workbook)
            return workbook

        raise ValueError(
            f"Unsupported file extension '{ext}' for {excel_path}. "
            "Supported formats: .xlsx, .xlsm, .xltx, .xltm, .csv"
        )

    @staticmethod
    def create_workbook_from_csv(csv_path: str) -> Workbook:
        workbook = Workbook()
        sheet = workbook.active

        sheet_name = os.path.splitext(os.path.basename(csv_path))[0][:31] or "Sheet1"
        sheet.title = sheet_name

        with open(csv_path, newline="", encoding="utf-8-sig") as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                processed_row = [ExcelLoader.infer_cell_value(value) for value in row]
                sheet.append(processed_row)

        return workbook

    @staticmethod
    def infer_cell_value(value: Optional[str]):
        if value is None:
            return value

        normalized = ExcelLoader.normalize_text_value(value)
        stripped = normalized.strip()
        if stripped == "":
            return ""

        try:
            if '.' not in stripped:
                return int(stripped)
            return float(stripped)
        except ValueError:
            return stripped

    @staticmethod
    def normalize_text_value(value: str) -> str:
        """Normalize Excel text artifacts before downstream stages."""
        text = str(value)
        # Remove XML-escaped control sequences like `_x000D_`, then trim CR/LF artifacts.
        text = ExcelLoader._XML_ESCAPE_RE.sub("", text)
        text = text.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
        return text.strip()

    @staticmethod
    def normalize_workbook_strings(workbook: Workbook) -> None:
        """Apply text normalization to all string cells in loaded workbook."""
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        normalized = ExcelLoader.normalize_text_value(cell.value)
                        if normalized != cell.value:
                            cell.value = normalized
