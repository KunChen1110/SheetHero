"""Spreadsheet domain helper functions."""

import os
import re
from typing import Any, Dict, List

from openpyxl.utils import range_boundaries

from ..world import SpreadsheetWorld


_XML_ESCAPE_RE = re.compile(r"_x[0-9A-Fa-f]{4}_")
_PERIOD_TEXT_RE = re.compile(
    r"^(?:Q[1-4]\s+\d{4}|\d{4}(?:-\d{2}-\d{2})?|[A-Za-z]{3,9}\s+\d{4})$",
    flags=re.IGNORECASE,
)
_NOTE_MARKERS = {
    "note",
    "notes",
    "comment",
    "comments",
    "remark",
    "remarks",
}


def get_workbook(world: SpreadsheetWorld, file_path: str):
    """Get a workbook by full path or by matching filename."""
    if file_path in world.workbooks:
        return world.workbooks[file_path]
    for path, wb in world.workbooks.items():
        if os.path.basename(path) == os.path.basename(file_path):
            return wb
    raise ValueError(
        f"Workbook not found: {file_path}. Available: {list(world.workbooks.keys())}"
    )


def list_all_workbooks(world: SpreadsheetWorld):
    """List all workbook paths."""
    return list(world.workbooks.keys())


def get_sheet_from_workbook(world: SpreadsheetWorld, file_path: str, sheet_name: str):
    """Get a sheet by workbook path and sheet name."""
    wb = get_workbook(world, file_path)
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    raise ValueError(
        f"Sheet '{sheet_name}' not found in {file_path}. "
        f"Available: {wb.sheetnames}"
    )


def inspector_multi(world: SpreadsheetWorld, file_path: str, range_ref: str,
                    sheet_name: str | None = None):
    """Inspect a cell range from a specific workbook."""
    wb = get_workbook(world, file_path)
    if sheet_name is None:
        sheet = wb.active
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in {os.path.basename(file_path)}. "
                f"Available sheets: {wb.sheetnames}"
            )
        sheet = wb[sheet_name]
    cell_range = sheet[range_ref]
    if hasattr(cell_range, 'value'):
        return [[cell_range.value]]
    return [[cell.value for cell in row] for row in cell_range]


def _normalize_text(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    text = _XML_ESCAPE_RE.sub("", value)
    text = text.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    return text.strip()


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _is_note_sentinel_row(row: List[Any]) -> bool:
    if not row:
        return False
    first = row[0]
    if not isinstance(first, str):
        return False
    marker = first.strip().lower()
    if marker.startswith("note:") or marker.startswith("comment:") or marker.startswith("remark:"):
        return True
    if marker not in _NOTE_MARKERS:
        return False
    return all(_is_empty(v) for v in row[1:])


def _non_empty_pairs(row: List[Any]) -> list[tuple[int, Any]]:
    return [(idx, value) for idx, value in enumerate(row) if not _is_empty(value)]


def _looks_numeric_value(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text or text == "-":
        return False
    text = (
        text.replace(",", "")
        .replace("%", "")
        .replace("£", "")
        .replace("$", "")
    )
    return bool(re.fullmatch(r"[+-]?\d+(?:\.\d+)?", text))


def _looks_period_label(value: Any) -> bool:
    text = _normalize_text(value)
    return isinstance(text, str) and bool(_PERIOD_TEXT_RE.fullmatch(text))


def _find_next_non_empty_row_index(rows: List[List[Any]], start_idx: int) -> int | None:
    for idx in range(start_idx, len(rows)):
        if _non_empty_pairs(rows[idx]):
            return idx
    return None


def _pick_header_row_index(rows: List[List[Any]], first_non_empty_idx: int) -> int:
    first_pairs = _non_empty_pairs(rows[first_non_empty_idx])
    if len(first_pairs) > 1:
        return first_non_empty_idx

    for idx in range(first_non_empty_idx + 1, len(rows)):
        pairs = _non_empty_pairs(rows[idx])
        if len(pairs) < 2:
            continue
        text_like_count = sum(
            1
            for _, value in pairs
            if not _looks_numeric_value(value) and not _looks_period_label(value)
        )
        next_idx = _find_next_non_empty_row_index(rows, idx + 1)
        next_pairs = _non_empty_pairs(rows[next_idx]) if next_idx is not None else []
        next_starts_with_period = bool(next_pairs and _looks_period_label(next_pairs[0][1]))
        if text_like_count >= 2 and (len(pairs) >= 3 or next_starts_with_period):
            return idx

    return first_non_empty_idx


def _is_ignorable_overflow_value(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    marker = value.strip().lower()
    return marker in {"in %", "%"}


def extract_sheet_table(
    worksheet,
    range_ref: str = "A1:Z200",
    *,
    drop_blank_rows: bool = True,
    drop_empty_primary_key: bool = True,
    stop_at_note_row: bool = True,
) -> Dict[str, Any]:
    """
    Extract a normalized tabular view from a worksheet.

    This helper centralizes header detection and column alignment so
    execution/diagnose/preview paths can share the same table semantics.
    """
    cell_range = worksheet[range_ref]
    if hasattr(cell_range, "value"):
        raw_rows = [[cell_range.value]]
    else:
        raw_rows = [[cell.value for cell in row] for row in cell_range]

    normalized_rows = [[_normalize_text(v) for v in row] for row in raw_rows]
    min_col, min_row, _, _ = range_boundaries(range_ref)
    _ = min_col  # explicit ignore; row offset is the useful part here.

    if not normalized_rows:
        return {
            "sheet_name": worksheet.title,
            "header": [],
            "rows": [],
            "excel_rows": [],
            "header_row_index": None,
            "header_excel_row": None,
            "total_raw_rows": 0,
            "kept_rows": 0,
            "dropped_rows": 0,
        }

    first_non_empty_row_index = None
    for idx, row in enumerate(normalized_rows):
        if any(not _is_empty(cell) for cell in row):
            first_non_empty_row_index = idx
            break

    if first_non_empty_row_index is None:
        return {
            "sheet_name": worksheet.title,
            "header": [],
            "rows": [],
            "excel_rows": [],
            "header_row_index": None,
            "header_excel_row": None,
            "total_raw_rows": len(normalized_rows),
            "kept_rows": 0,
            "dropped_rows": len(normalized_rows),
        }

    header_row_index = _pick_header_row_index(normalized_rows, first_non_empty_row_index)
    header_raw = normalized_rows[header_row_index]
    keep_indices = [
        idx for idx, cell in enumerate(header_raw)
        if not _is_empty(cell)
    ]
    header = [str(header_raw[idx]).strip() for idx in keep_indices]
    data_start_index = header_row_index + 1

    preview_data_index = _find_next_non_empty_row_index(normalized_rows, data_start_index)
    if header and keep_indices and preview_data_index is not None:
        preview_pairs = _non_empty_pairs(normalized_rows[preview_data_index])
        if (
            preview_pairs
            and _looks_period_label(preview_pairs[0][1])
            and preview_pairs[0][0] < min(keep_indices)
        ):
            keep_indices = [preview_pairs[0][0]] + keep_indices
            header = ["Time"] + header

    if len(header) <= 1:
        synthetic_data_index = _find_next_non_empty_row_index(normalized_rows, header_row_index + 1)
        while synthetic_data_index is not None:
            data_pairs = _non_empty_pairs(normalized_rows[synthetic_data_index])
            if len(data_pairs) >= 2 and _looks_period_label(data_pairs[0][1]):
                keep_indices = [col_idx for col_idx, _ in data_pairs]
                value_count = len(keep_indices) - 1
                if value_count >= 1:
                    header = ["Time"] + (
                        ["Value"] if value_count == 1
                        else [f"Value {idx}" for idx in range(1, value_count + 1)]
                    )
                    data_start_index = synthetic_data_index
                    header_row_index = synthetic_data_index - 1 if synthetic_data_index > 0 else 0
                    break
            synthetic_data_index = _find_next_non_empty_row_index(normalized_rows, synthetic_data_index + 1)

    if not header:
        return {
            "sheet_name": worksheet.title,
            "header": [],
            "rows": [],
            "excel_rows": [],
            "header_row_index": header_row_index,
            "header_excel_row": min_row + header_row_index,
            "total_raw_rows": len(normalized_rows),
            "kept_rows": 0,
            "dropped_rows": max(0, len(normalized_rows) - 1),
        }

    candidate_rows: List[tuple[List[Any], int, bool]] = []
    dropped_rows = 0
    for offset, raw_row in enumerate(normalized_rows[data_start_index:], start=data_start_index):
        row = [
            raw_row[idx] if idx < len(raw_row) else None
            for idx in keep_indices
        ]
        excel_row = min_row + offset
        overflow = False
        for idx, value in enumerate(raw_row):
            if idx in keep_indices:
                continue
            if not _is_empty(value) and not _is_ignorable_overflow_value(value):
                overflow = True
                break

        if stop_at_note_row and _is_note_sentinel_row(row):
            break

        candidate_rows.append((row, excel_row, overflow))

    while candidate_rows and all(_is_empty(v) for v in candidate_rows[-1][0]):
        candidate_rows.pop()
        dropped_rows += 1

    cleaned_rows: List[List[Any]] = []
    excel_rows: List[int] = []
    overflow_excel_rows: List[int] = []
    for row, excel_row, overflow in candidate_rows:
        if all(_is_empty(v) for v in row):
            if drop_blank_rows:
                dropped_rows += 1
                continue
            cleaned_rows.append(row)
            excel_rows.append(excel_row)
            if overflow:
                overflow_excel_rows.append(excel_row)
            continue

        if drop_empty_primary_key and _is_empty(row[0]):
            dropped_rows += 1
            continue

        cleaned_rows.append(row)
        excel_rows.append(excel_row)
        if overflow:
            overflow_excel_rows.append(excel_row)

    return {
        "sheet_name": worksheet.title,
        "header": header,
        "rows": cleaned_rows,
        "excel_rows": excel_rows,
        "overflow_excel_rows": overflow_excel_rows,
        "header_row_index": header_row_index,
        "header_excel_row": min_row + header_row_index,
        "total_raw_rows": len(normalized_rows),
        "kept_rows": len(cleaned_rows),
        "dropped_rows": dropped_rows,
    }


def read_table_multi(
    world: SpreadsheetWorld,
    file_path: str,
    sheet_name: str | None = None,
    range_ref: str = "A1:Z200",
    require_primary_key: bool = True,
    stop_at_note_row: bool = True,
) -> Dict[str, Any]:
    """
    Read a table from a specific workbook with lightweight normalization.

    Returns:
      {
        "file_path": str,
        "sheet_name": str,
        "header": list[str],
        "rows": list[list[Any]],
        "total_raw_rows": int,
        "kept_rows": int,
        "dropped_rows": int
      }
    """
    wb = get_workbook(world, file_path)
    if sheet_name is None:
        ws = wb.active
    else:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in {os.path.basename(file_path)}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]

    extracted = extract_sheet_table(
        ws,
        range_ref,
        drop_blank_rows=True,
        drop_empty_primary_key=require_primary_key,
        stop_at_note_row=stop_at_note_row,
    )

    return {
        "file_path": file_path,
        "sheet_name": extracted["sheet_name"],
        "header": extracted["header"],
        "rows": extracted["rows"],
        "total_raw_rows": extracted["total_raw_rows"],
        "kept_rows": extracted["kept_rows"],
        "dropped_rows": extracted["dropped_rows"],
    }
