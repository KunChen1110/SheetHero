"""Workbook sheet info helpers."""

from typing import Any, Dict, List, Optional

from openpyxl.utils import get_column_letter


class ExcelSheetInfo:
    """Sheet listing and metadata helpers."""

    def __init__(self, workbook, reader):
        self.workbook = workbook
        self.reader = reader

    def list_sheets(self) -> List[str]:
        """Return list of all sheet names in the workbook."""
        return self.workbook.sheetnames

    def get_sheet_info(self, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """Get information about a specific sheet (name, dimensions)."""
        sheet = self.reader.get_sheet(sheet_name)
        return {
            'name': sheet.title,
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'dimensions': f"{sheet.max_row} rows × {sheet.max_column} columns"
        }

    def get_all_sheets_info(self) -> Dict[str, Dict[str, Any]]:
        """Get information about all sheets in the workbook."""
        result = {}
        for sheet_name in self.workbook.sheetnames:
            result[sheet_name] = self.get_sheet_info(sheet_name)
        return result

    def read_multiple_sheets(self, sheet_names: List[str],
                             range_ref: Optional[str] = None) -> Dict[str, List[List]]:
        """Read data from multiple sheets at once."""
        result = {}
        for sheet_name in sheet_names:
            if sheet_name not in self.workbook.sheetnames:
                result[sheet_name] = None
                continue

            if range_ref:
                result[sheet_name] = self.reader.inspector(range_ref, sheet_name)
            else:
                sheet = self.reader.get_sheet(sheet_name)
                max_range = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
                result[sheet_name] = self.reader.inspector(max_range, sheet_name)

        return result
