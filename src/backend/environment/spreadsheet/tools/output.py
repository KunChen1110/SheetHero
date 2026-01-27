"""Output workbook writer helpers."""

import os
from typing import Any, Dict, List, Optional

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple


class ExcelOutputWriter:
    """Write to a separate output workbook and manage artifacts."""

    def __init__(self, workbook, excel_path: str,
                 output_path: Optional[str], temp_files: list):
        self.workbook = workbook
        self.excel_path = excel_path
        self.output_path = output_path
        self._temp_files = temp_files
        self._output_workbook = None

    def save_workbook(self) -> str:
        """
        Save OUTPUT workbook to new file (input files remain unchanged).
        """
        if self.output_path:
            filename = self.output_path
        else:
            dir_path = os.path.dirname(self.excel_path)
            base_name = os.path.splitext(os.path.basename(self.excel_path))[0]
            filename = os.path.join(dir_path, f"{base_name}_output.xlsx")

        output_wb = self._get_output_workbook()

        if len(output_wb.sheetnames) == 0:
            output_wb.create_sheet("Output")

        output_wb.save(filename)

        for temp_file in self._temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as e:
                print(f"⚠️ Warning: Could not delete temporary file {temp_file}: {e}")

        self._temp_files = []

        print(f"💾 Workbook saved to: {filename}")
        return filename

    def _get_output_workbook(self):
        """Get or create the output workbook (separate from input files)."""
        from openpyxl import Workbook as OpenpyxlWorkbook
        if self._output_workbook is None:
            self._output_workbook = OpenpyxlWorkbook()
            if 'Sheet' in self._output_workbook.sheetnames:
                del self._output_workbook['Sheet']
        return self._output_workbook

    def create_output_sheet(self, sheet_name: str = "Output") -> str:
        """Create a new sheet in the OUTPUT workbook."""
        try:
            output_wb = self._get_output_workbook()

            if sheet_name in output_wb.sheetnames:
                del output_wb[sheet_name]

            output_wb.create_sheet(sheet_name)
            message = f"✅ Created output sheet '{sheet_name}' (in new output file)"
            print(message)
            return message
        except Exception as e:
            error_msg = f"❌ Error creating sheet: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def write_dataframe_to_sheet(self, data: List[List], sheet_name: str,
                                 start_cell: str = "A1",
                                 include_header: bool = True) -> str:
        """
        Write 2D list (DataFrame-like data) to output workbook.
        """
        try:
            output_wb = self._get_output_workbook()

            if sheet_name not in output_wb.sheetnames:
                output_wb.create_sheet(sheet_name)

            sheet = output_wb[sheet_name]

            start_row, start_col = coordinate_to_tuple(start_cell)

            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    current_row = start_row + row_idx
                    current_col = start_col + col_idx
                    sheet.cell(row=current_row, column=current_col, value=value)

            rows_written = len(data)
            cols_written = max(len(row) for row in data) if data else 0
            end_row = start_row + rows_written - 1
            end_col = start_col + cols_written - 1
            end_cell = f"{get_column_letter(end_col)}{end_row}"

            message = f"✅ Wrote {rows_written} rows to {sheet_name}!{start_cell}:{end_cell}"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error writing data to sheet: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def highlight_rows(self, sheet_name: str, row_numbers: List[int],
                       format_dict: Dict[str, Any] = None) -> str:
        """
        Highlight entire rows in output workbook.
        """
        try:
            if format_dict is None:
                format_dict = {"fill_color": "red"}

            output_wb = self._get_output_workbook()
            if sheet_name not in output_wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found in output workbook. Create it first with create_output_sheet()."
                )
            sheet = output_wb[sheet_name]
            max_col = sheet.max_column

            for row_num in row_numbers:
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_num, column=col_idx)

                    if 'fill_color' in format_dict:
                        color = self._parse_color(format_dict['fill_color'])
                        cell.fill = PatternFill(
                            start_color=color, end_color=color, fill_type='solid'
                        )

                    font_kwargs = {}
                    if 'font_color' in format_dict:
                        font_kwargs['color'] = self._parse_color(format_dict['font_color'])
                    if 'bold' in format_dict:
                        font_kwargs['bold'] = format_dict['bold']
                    if font_kwargs:
                        cell.font = Font(**font_kwargs)

            message = f"✅ Highlighted row(s) {row_numbers} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error highlighting rows: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def save_workbook_to(self, output_path: str) -> str:
        """Save the output workbook to a specific path."""
        try:
            output_wb = self._get_output_workbook()

            if len(output_wb.sheetnames) == 0:
                output_wb.create_sheet("Output")

            output_wb.save(output_path)

            for temp_file in self._temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                except Exception:
                    pass
            self._temp_files = []

            message = f"💾 Workbook saved to: {output_path}"
            print(message)
            return output_path

        except Exception as e:
            error_msg = f"❌ Error saving workbook: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def add_summary_row(self, sheet_name: str, row_number: int,
                        summary_data: Dict[str, Any]) -> str:
        """Add a summary row with labeled statistics to the output workbook."""
        try:
            output_wb = self._get_output_workbook()
            if sheet_name not in output_wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in output workbook.")
            sheet = output_wb[sheet_name]
            col = 1

            for label, value in summary_data.items():
                sheet.cell(row=row_number, column=col, value=label)
                sheet.cell(row=row_number, column=col + 1, value=value)
                col += 3

            message = f"✅ Added summary row at row {row_number} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error adding summary row: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def _parse_color(self, color: str) -> str:
        color_names = {
            'red': 'FF0000', 'green': '00FF00', 'blue': '0000FF',
            'yellow': 'FFFF00', 'orange': 'FFA500', 'purple': '800080',
            'pink': 'FFC0CB', 'brown': 'A52A2A', 'black': '000000',
            'white': 'FFFFFF', 'gray': '808080', 'grey': '808080'
        }

        if color.startswith('#'):
            return color[1:]
        if color.lower() in color_names:
            return color_names[color.lower()]
        return color
