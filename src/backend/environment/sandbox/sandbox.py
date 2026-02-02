"""Sandbox facade that builds execution state on init."""

from __future__ import annotations

from types import SimpleNamespace
import os
from typing import Dict, Optional, Tuple

import matplotlib
import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from .runner import SandboxRunner
from ..spreadsheet.loader import load_world
from ..namespace.namespace import SpreadsheetNamespace


class Sandbox:
    """Builds and holds sandbox execution state."""

    def __init__(self,
                 excel_paths: list[str],
                 output_preferences: Dict[str, Optional[str]],
                 output_path: str,
                 enabled_namespaces: Optional[list[str]] = None,
                 progress_logger=None,
                 load_excel: bool = True):
        self.excel_paths = excel_paths
        self.output_preferences = output_preferences
        self.output_path = output_path
        self.enabled_namespaces = enabled_namespaces or ["spreadsheet"]
        self.progress_logger = progress_logger

        self.code_globals = {
            "math": __import__("math"),
            "json": __import__("json"),
            "re": __import__("re"),
            "excel_paths": self.excel_paths,
            "output_preferences": self.output_preferences,
            "output_path": self.output_path,
            "namespaces": SimpleNamespace(),
        }
        self.code_locals = {}
        self.workbooks = {}
        self.runner = SandboxRunner()
        self.world = None

        if not load_excel:
            return

        matplotlib.use("Agg")

        self.world = load_world(self.excel_paths, self.output_path, self.progress_logger)
        self.workbooks = self.world.workbooks

        self.code_globals.update({
            "openpyxl": openpyxl,
            "workbooks": self.workbooks,
            "sheet_names": self.world.primary_workbook.sheetnames,
            "range_boundaries": range_boundaries,
            "get_column_letter": get_column_letter,
            "column_index_from_string": column_index_from_string,
            "pandas": pd,
            "pd": pd,
            "numpy": np,
            "np": np,
        })

        for namespace in self.enabled_namespaces:
            self.load_namespace(namespace)

    def load_namespace(self, name: str):
        """Load a namespace into the sandbox globals."""
        if name == "spreadsheet":
            if not self.world:
                raise RuntimeError("Cannot load namespaces without a loaded world.")
            namespace = SpreadsheetNamespace(self.world).build()
            self.code_globals["namespaces"].spreadsheet = namespace
            self.code_globals["spreadsheet"] = namespace
            self.code_globals.update(namespace.__dict__)
            return namespace

        raise ValueError(f"Unknown namespace: {name}")

    def get_workbook_view(self):
        """Return a read-only view of loaded workbooks."""
        return self.workbooks

    def _build_workbook_view(self):
        """Build a read-only view as a dict of pandas DataFrames keyed by file/sheet."""
        view = {}
        for path, workbook in self.workbooks.items():
            file_key = os.path.basename(path)
            for sheet in workbook.worksheets:
                rows = list(sheet.values)
                if not rows:
                    df = pd.DataFrame()
                else:
                    header = list(rows[0])
                    if not any(h is not None and str(h).strip() != "" for h in header):
                        header = [f"col_{i + 1}" for i in range(len(header))]
                    else:
                        header = [
                            (str(h).strip() if h is not None and str(h).strip() != "" else f"col_{i + 1}")
                            for i, h in enumerate(header)
                        ]
                    data = rows[1:]
                    df = pd.DataFrame(data, columns=header)

                base_key = f"{file_key}::{sheet.title}"
                key = base_key
                suffix = 2
                while key in view:
                    key = f"{base_key}#{suffix}"
                    suffix += 1
                view[key] = df
        return view

    def run(self, code: str, extra_globals: Optional[dict] = None) -> dict:
        """Execute code in the sandbox and return the runner result."""
        if extra_globals:
            original = {}
            for key, value in extra_globals.items():
                original[key] = self.code_globals.get(key, None)
                self.code_globals[key] = value
            try:
                return self.runner.run(code, self.code_globals, self.code_locals)
            finally:
                for key, value in original.items():
                    if value is None and key in self.code_globals:
                        del self.code_globals[key]
                    else:
                        self.code_globals[key] = value
        return self.runner.run(code, self.code_globals, self.code_locals)
