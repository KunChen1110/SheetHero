"""Validation runtime for execution results."""

import ast
import os
import re
from typing import Any, Dict

from openpyxl import load_workbook

from ...log.logger_registry import LoggerRegistry
from ...task_families import detect_task_family, get_task_family_validation_mode
from ..base.runtime import StageRuntime
from ...prompt.prompt_builder import PromptBuilder
from .history import ValidationHistory
from .llm_client import ValidationLLMClient
from .parser import ValidationResponseParser

logger = LoggerRegistry.setup_logger(__name__)


class ValidationRuntime(StageRuntime):
    """Runs validation steps to assess execution results."""

    def __init__(self, client, deployment: str, excel_context_understanding: str,
                 progress_log_file=None, prompt_profile: str = "online_rich"):
        super().__init__(progress_log_file)
        self.excel_context_understanding = excel_context_understanding
        self.history_formatter = ValidationHistory()
        self.llm_client = ValidationLLMClient(client, deployment)
        self.parser = ValidationResponseParser()
        self.prompt_builder = PromptBuilder(profile=prompt_profile)

    @staticmethod
    def _large_workbook_threshold_bytes() -> int:
        return int(os.getenv("SHEETHERO_LARGE_WORKBOOK_THRESHOLD_BYTES", "10000000"))

    @classmethod
    def _is_large_workbook(cls, output_path: str) -> bool:
        if not output_path or not os.path.exists(output_path):
            return False
        try:
            return os.path.getsize(output_path) >= cls._large_workbook_threshold_bytes()
        except OSError:
            return False

    @staticmethod
    def _render_validation_result(validation_result: Dict[str, Any]) -> str:
        status = "PASSED" if validation_result.get("validation_passed") else "FAILED"
        confidence = float(validation_result.get("confidence_score", 0.0) or 0.0)
        issues = validation_result.get("issues_found") or ["None identified."]
        feedback = validation_result.get("improvement_feedback") or ""
        final_assessment = validation_result.get("final_assessment") or ""
        issue_lines = "\n".join(f"- {issue}" for issue in issues)
        return (
            f"VALIDATION_STATUS: {status}\n"
            f"CONFIDENCE_SCORE: {confidence:.2f}\n"
            "ISSUES_FOUND:\n"
            f"{issue_lines}\n"
            "IMPROVEMENT_FEEDBACK:\n"
            f"{feedback}\n"
            "FINAL_ASSESSMENT:\n"
            f"{final_assessment}"
        )

    @staticmethod
    def _looks_like_file_path(s: str) -> bool:
        """True if s looks like a file path (e.g. ends with .xlsx or starts with / or contains \\)."""
        if not s or not isinstance(s, str):
            return False
        s = s.strip()
        if ".xlsx" in s or ".xls" in s or s.startswith("/") or "\\" in s or s.startswith("C:"):
            return True
        return False

    @classmethod
    def _resolve_contract_expectations(
        cls,
        user_question: str,
        need_detail: bool | None,
        need_summary: bool | None,
        need_highlight: bool | None,
    ) -> tuple[bool | None, bool | None, bool | None]:
        family = detect_task_family(user_question)
        if family is None:
            return need_detail, need_summary, need_highlight
        if need_detail is None:
            need_detail = family.requires_detailed_table
        if need_summary is None:
            need_summary = family.requires_summary_metrics
        if need_highlight is None:
            need_highlight = family.requires_highlight
        return need_detail, need_summary, need_highlight

    @staticmethod
    def _load_first_table_rows(output_path: str) -> tuple[list[str], list[list[Any]], str] | None:
        if not output_path or not os.path.exists(output_path):
            return None
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception:
            return None
        for sheet in workbook.worksheets:
            header_row_idx = None
            header: list[str] = []
            for row_idx in range(1, min(sheet.max_row, 20) + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 20) + 1)
                ]
                non_empty = [str(v).strip() for v in row_values if v not in (None, "")]
                if len(non_empty) >= 2:
                    header_row_idx = row_idx
                    header = [str(v).strip() if v not in (None, "") else "" for v in row_values]
                    break
            if header_row_idx is None:
                continue
            rows: list[list[Any]] = []
            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 20) + 1)
                ]
                if not any(v not in (None, "") for v in row_values):
                    if rows:
                        break
                    continue
                rows.append(row_values)
            return header, rows, sheet.title
        return None

    @classmethod
    def _inspect_saved_temporal_aggregation_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Temporal aggregation output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 2:
            issues.append("Temporal aggregation output must contain at least two columns.")
            return issues
        if normalized[0] != "Period":
            issues.append("Temporal aggregation output should start with a `Period` column.")
        metric_label = normalized[1]
        if not any(metric_label.startswith(prefix) for prefix in ("Average ", "Total ", "Count ", "Median ", "Minimum ", "Maximum ")):
            issues.append("Temporal aggregation output metric column label is missing the aggregate prefix.")
        if not rows:
            issues.append("Temporal aggregation output does not contain any data rows.")
            return issues
        first_period = str(rows[0][0]).strip() if rows[0] and rows[0][0] is not None else ""
        if not first_period:
            issues.append("Temporal aggregation output has an empty period label in the first data row.")
        return issues

    @classmethod
    def _inspect_saved_grouped_aggregation_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Grouped aggregation output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 2:
            issues.append("Grouped aggregation output must contain at least two columns.")
            return issues
        metric_label = normalized[-1]
        if not any(metric_label.startswith(prefix) for prefix in ("Average ", "Total ", "Count ", "Median ", "Minimum ", "Maximum ")):
            issues.append("Grouped aggregation output metric column label is missing the aggregate prefix.")
        if not rows:
            issues.append("Grouped aggregation output does not contain any data rows.")
        return issues

    @classmethod
    def _inspect_saved_relational_join_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Relational join output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 2:
            issues.append("Relational join output must contain at least two columns.")
        if not rows:
            issues.append("Relational join output does not contain any data rows.")
            return issues
        non_empty_column_count = sum(1 for value in rows[0] if value not in (None, ""))
        if non_empty_column_count < 2:
            issues.append("Relational join output first data row is not populated across multiple columns.")
        return issues

    @classmethod
    def _inspect_saved_allocation_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Allocation output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 3:
            issues.append("Allocation output must contain entity columns, resource columns, and status.")
            return issues
        if "Allocation Status" not in normalized:
            issues.append("Allocation output should include an `Allocation Status` column.")
        if "Allocated Quantity" not in normalized:
            issues.append("Allocation output should include an `Allocated Quantity` column.")
        if not rows:
            issues.append("Allocation output does not contain any data rows.")
        return issues

    @classmethod
    def _inspect_saved_dependency_schedule_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Dependency-schedule output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if "Start Time" not in normalized or "End Time" not in normalized:
            issues.append("Dependency-schedule output should include `Start Time` and `End Time` columns.")
        if "Task ID" not in normalized:
            issues.append("Dependency-schedule output should include a `Task ID` column.")
        if not rows:
            issues.append("Dependency-schedule output does not contain any scheduled task rows.")
        return issues

    @classmethod
    def _inspect_saved_assignment_schedule_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Assignment-schedule output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 3:
            issues.append("Assignment-schedule output must contain multiple schedule columns.")
            return issues
        scheduling_markers = ("day", "time", "slot", "room", "session")
        entity_markers = ("student", "candidate", "participant", "attendee", "member", "name", "id")
        if not any(any(marker in column.lower() for marker in scheduling_markers) for column in normalized):
            issues.append("Assignment-schedule output should include at least one scheduling/location column.")
        if not any(any(marker in column.lower() for marker in entity_markers) for column in normalized):
            issues.append("Assignment-schedule output should include at least one entity-identifying column.")
        if not rows:
            issues.append("Assignment-schedule output does not contain any assigned rows.")
        return issues

    @classmethod
    def _inspect_saved_region_growth_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Region-growth output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 3:
            issues.append("Region-growth output must contain region, average, and growth columns.")
            return issues
        if normalized[0] != "Region":
            issues.append("Region-growth output should start with a `Region` column.")
        if not any("Avg Penetration" in value for value in normalized[1:]):
            issues.append("Region-growth output is missing an average-penetration column.")
        if not any(value.startswith("Growth (") for value in normalized[1:]):
            issues.append("Region-growth output is missing a growth column.")
        if not rows:
            issues.append("Region-growth output does not contain any data rows.")
        return issues

    @classmethod
    def _inspect_saved_regression_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Regression output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if normalized[:2] != ["Factor", "Weight"]:
            issues.append("Regression output should start with `Factor` and `Weight` columns.")
        if not rows:
            issues.append("Regression output does not contain any coefficient rows.")
            return issues
        first_factor = str(rows[0][0]).strip() if rows[0] and rows[0][0] is not None else ""
        if not first_factor:
            issues.append("Regression output has an empty factor label in the first data row.")
        return issues

    @classmethod
    def _inspect_saved_correlation_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Correlation output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 2:
            issues.append("Correlation output must contain at least two numeric columns.")
        if len(rows) < 2:
            issues.append("Correlation output does not contain enough matrix rows.")
            return issues
        first_row = rows[0]
        non_empty = [value for value in first_row if value not in (None, "")]
        if len(non_empty) < 2:
            issues.append("Correlation output first matrix row is not populated.")
        return issues

    @classmethod
    def _inspect_saved_comparative_multi_sheet_workbook(cls, output_path: str) -> list[str]:
        issues = cls._inspect_saved_generic_workbook(
            output_path,
            need_detail=True,
            need_summary=False,
        )
        if issues:
            return issues
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception as exc:
            return [f"Unable to open saved workbook: {exc}"]
        required_sheets = {"AvgByStoreType", "HolidayVsNonHoliday"}
        missing = sorted(required_sheets.difference(workbook.sheetnames))
        if missing:
            issues.append(
                "Comparative multi-sheet output is missing required sheet(s): "
                + ", ".join(missing)
            )
        return issues

    @classmethod
    def _inspect_saved_generic_workbook(
        cls,
        output_path: str,
        need_detail: bool | None,
        need_summary: bool | None,
    ) -> list[str]:
        issues: list[str] = []
        if not output_path or not os.path.exists(output_path):
            return [f"Saved output file not found: {output_path}"]
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception as exc:
            return [f"Unable to open saved workbook: {exc}"]

        found_header = False
        max_data_rows = 0
        found_metrics = False

        for sheet in workbook.worksheets:
            header_row_idx = None
            for row_idx in range(1, min(sheet.max_row, 20) + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 12) + 1)
                ]
                non_empty = [str(v).strip() for v in row_values if v not in (None, "")]
                if len(non_empty) >= 2:
                    header_row_idx = row_idx
                    found_header = True
                    break
            if header_row_idx is None:
                continue

            data_rows = 0
            metrics_in_sheet = False
            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 12) + 1)
                ]
                if not any(v not in (None, "") for v in row_values):
                    continue

                metric_found = False
                for col_idx in range(0, len(row_values) - 1):
                    first = row_values[col_idx]
                    second = row_values[col_idx + 1]
                    label = str(first).strip() if first not in (None, "") else ""
                    if label and any(
                        marker in label.lower()
                        for marker in (
                            "total",
                            "average",
                            "avg",
                            "count",
                            "max",
                            "min",
                            "growth",
                            "fastest",
                            "region",
                            "sum",
                        )
                    ):
                        metrics_in_sheet = True
                        metric_found = True
                        break
                if metric_found:
                    continue

                first = row_values[0] if row_values else None
                second = row_values[1] if len(row_values) > 1 else None
                if first not in (None, "") and second not in (None, ""):
                    data_rows += 1

            max_data_rows = max(max_data_rows, data_rows)
            found_metrics = found_metrics or metrics_in_sheet
            if (need_detail is not True or data_rows >= 1) and (need_summary is not True or metrics_in_sheet):
                return []

        if not found_header:
            issues.append("Saved workbook does not contain a sheet with a detectable table header.")
        if need_detail is True and max_data_rows < 1:
            issues.append("Saved workbook does not contain a detailed table with at least one data row.")
        if need_summary is True and not found_metrics:
            issues.append("Saved workbook does not contain the required summary metrics block.")
        return issues

    @staticmethod
    def _parse_output_contract_flag(understanding_output: str, key: str) -> bool | None:
        if not understanding_output:
            return None
        pattern = rf"(?:\*\*)?\s*{re.escape(key)}\s*(?:\*\*)?\s*:\s*(YES|NO|TRUE|FALSE)"
        m = re.search(pattern, understanding_output, flags=re.IGNORECASE)
        if not m:
            return None
        return m.group(1).strip().upper() in {"YES", "TRUE"}

    @staticmethod
    def _extract_rows_written(text: str) -> list[int]:
        if not text:
            return []
        out = []
        for m in re.findall(r"Wrote\s+(\d+)\s+rows\s+to", text, flags=re.IGNORECASE):
            try:
                out.append(int(m))
            except Exception:
                    continue
        return out

    @staticmethod
    def _column_letters_to_index(col: str) -> int:
        total = 0
        for ch in (col or "").upper():
            if "A" <= ch <= "Z":
                total = total * 26 + (ord(ch) - ord("A") + 1)
        return total

    @classmethod
    def _extract_written_column_counts(cls, text: str) -> list[int]:
        if not text:
            return []
        counts: list[int] = []
        for start_col, end_col in re.findall(
            r"Wrote\s+\d+\s+rows\s+to\s+[A-Za-z0-9_ ]+!([A-Z]+)\d+:([A-Z]+)\d+",
            text,
            flags=re.IGNORECASE,
        ):
            start_idx = cls._column_letters_to_index(start_col)
            end_idx = cls._column_letters_to_index(end_col)
            if start_idx and end_idx and end_idx >= start_idx:
                counts.append(end_idx - start_idx + 1)
        return counts

    @classmethod
    def _extract_write_ranges(cls, text: str) -> list[tuple[str, int, int, int, int]]:
        if not text:
            return []
        ranges: list[tuple[str, int, int, int, int]] = []
        pattern = r"Wrote\s+\d+\s+rows\s+to\s+([^!\n]+)!([A-Z]+)(\d+):([A-Z]+)(\d+)"
        for sheet_name, start_col, start_row, end_col, end_row in re.findall(
            pattern,
            text,
            flags=re.IGNORECASE,
        ):
            start_idx = cls._column_letters_to_index(start_col)
            end_idx = cls._column_letters_to_index(end_col)
            try:
                start_row_int = int(start_row)
                end_row_int = int(end_row)
            except Exception:
                continue
            if not start_idx or not end_idx:
                continue
            ranges.append((sheet_name.strip(), start_row_int, start_idx, end_row_int, end_idx))
        return ranges

    @staticmethod
    def _ranges_overlap(
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        other_start_row: int,
        other_start_col: int,
        other_end_row: int,
        other_end_col: int,
    ) -> bool:
        rows_overlap = not (end_row < other_start_row or other_end_row < start_row)
        cols_overlap = not (end_col < other_start_col or other_end_col < start_col)
        return rows_overlap and cols_overlap

    @classmethod
    def _find_overlapping_write_ranges(
        cls,
        text: str,
    ) -> list[tuple[tuple[str, int, int, int, int], tuple[str, int, int, int, int]]]:
        ranges = cls._extract_write_ranges(text)
        overlaps: list[tuple[tuple[str, int, int, int, int], tuple[str, int, int, int, int]]] = []
        for idx, current in enumerate(ranges):
            current_sheet, current_start_row, current_start_col, current_end_row, current_end_col = current
            for previous in ranges[:idx]:
                prev_sheet, prev_start_row, prev_start_col, prev_end_row, prev_end_col = previous
                if current_sheet != prev_sheet:
                    continue
                if cls._ranges_overlap(
                    current_start_row,
                    current_start_col,
                    current_end_row,
                    current_end_col,
                    prev_start_row,
                    prev_start_col,
                    prev_end_row,
                    prev_end_col,
                ):
                    overlaps.append((previous, current))
        return overlaps

    @staticmethod
    def _index_to_column_letters(index: int) -> str:
        letters = []
        value = index
        while value > 0:
            value, remainder = divmod(value - 1, 26)
            letters.append(chr(ord("A") + remainder))
        return "".join(reversed(letters)) or "A"

    @classmethod
    def _format_write_range(cls, write_range: tuple[str, int, int, int, int]) -> str:
        sheet_name, start_row, start_col, end_row, end_col = write_range
        start_cell = f"{cls._index_to_column_letters(start_col)}{start_row}"
        end_cell = f"{cls._index_to_column_letters(end_col)}{end_row}"
        return f"{sheet_name}!{start_cell}:{end_cell}"

    @staticmethod
    def _extract_highlight_rows(text: str) -> list[int]:
        if not text:
            return []
        all_rows: list[int] = []
        blocks = re.findall(r"Highlighted row\(s\)\s*\[([^\]]*)\]", text, flags=re.IGNORECASE)
        for raw in blocks:
            for token in re.findall(r"-?\d+", raw):
                try:
                    all_rows.append(int(token))
                except Exception:
                    continue
        return all_rows

    @staticmethod
    def _extract_summary_rows(text: str) -> list[int]:
        if not text:
            return []
        out: list[int] = []
        for raw in re.findall(r"Added summary row at row\s+(\d+)", text, flags=re.IGNORECASE):
            try:
                out.append(int(raw))
            except Exception:
                continue
        return out

    @staticmethod
    def _has_summary_write_signal_from_code(code_text: str) -> bool:
        if not code_text:
            return False
        lower_code = code_text.lower()
        if "add_summary_row(" in lower_code:
            return True
        if '["metric", "value"]' in lower_code or "['metric', 'value']" in lower_code:
            return True
        if "write_dataframe_to_sheet(" in lower_code and (
            "summary" in lower_code or "metric" in lower_code
        ):
            return True
        return False

    @staticmethod
    def _is_regression_request(user_question: str) -> bool:
        q = (user_question or "").lower()
        markers = ("regression", "coefficient", "coefficients", "weight", "weights", "predict")
        return any(m in q for m in markers)

    @staticmethod
    def _extract_reported_columns(text: str) -> list[str]:
        if not text:
            return []
        matches = re.findall(r"Columns:\s*(\[[^\]]*\])", text)
        for raw in reversed(matches):
            try:
                parsed = ast.literal_eval(raw)
                if isinstance(parsed, list):
                    return [str(v).strip() for v in parsed if str(v).strip()]
            except Exception:
                continue
        return []

    @staticmethod
    def _header_is_target_like(header: str) -> bool:
        h = (header or "").strip().lower()
        target_markers = ("sales", "target", "label", "outcome", "revenue", "y")
        return any(m in h for m in target_markers)

    @staticmethod
    def _header_is_non_feature_like(header: str) -> bool:
        h = (header or "").strip().lower()
        non_feature_markers = ("id", "date", "time", "timestamp", "note", "comment")
        return any(m in h for m in non_feature_markers)

    def _expected_regression_predictors(self, columns: list[str]) -> list[str]:
        predictors: list[str] = []
        for h in columns:
            if self._header_is_target_like(h):
                continue
            if self._header_is_non_feature_like(h):
                continue
            predictors.append(h)
        return predictors

    @staticmethod
    def _extract_feature_cols_from_code(code_text: str) -> list[str]:
        if not code_text:
            return []
        m = re.search(r"feature_cols\s*=\s*\[([^\]]*)\]", code_text, flags=re.IGNORECASE | re.DOTALL)
        if not m:
            return []
        body = m.group(1)
        return [s.strip() for s in re.findall(r"['\"]([^'\"]+)['\"]", body)]

    @staticmethod
    def _extract_weight_labels(text: str) -> list[str]:
        if not text:
            return []
        m = re.search(r"Weights:\s*(\[\[.*\]\])", text, flags=re.IGNORECASE)
        if not m:
            return []
        raw = m.group(1).strip()
        try:
            parsed = ast.literal_eval(raw)
            labels = []
            for row in parsed[1:]:
                if isinstance(row, (list, tuple)) and row:
                    labels.append(str(row[0]).strip())
            return labels
        except Exception:
            return []

    @staticmethod
    def _is_dependency_schedule_request(user_question: str) -> bool:
        q = (user_question or "").lower()
        if "dependency" in q or "dependencies" in q or "topological" in q or "dag" in q:
            return True
        if "depends on" in q or "task id" in q:
            return True
        if ("schedule" in q or "scheduling" in q) and (
            "start time" in q
            or "end time" in q
            or "depends on" in q
            or "task id" in q
        ):
            return True
        return False

    @staticmethod
    def _is_region_growth_chart_request(user_question: str) -> bool:
        q = (user_question or "").lower()
        return "line chart" in q and "growth rate" in q and "region" in q and (
            "penetration" in q or "internet" in q
        )

    @staticmethod
    def _is_missing_data_scan_request(user_question: str) -> bool:
        q = (user_question or "").lower()
        return "missing data" in q and ("identify where" in q or "where values are missing" in q or "check the file" in q)

    @staticmethod
    def _is_room_inconsistency_request(user_question: str) -> bool:
        q = (user_question or "").lower()
        return "room identifiers" in q or "room identifier" in q or "c80" in q or "c 80" in q

    @classmethod
    def _inspect_saved_region_growth_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        if not output_path or not os.path.exists(output_path):
            return [f"Saved output file not found: {output_path}"]
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception as exc:
            return [f"Unable to open saved workbook: {exc}"]

        output_sheet = workbook["Output"] if "Output" in workbook.sheetnames else workbook.worksheets[0]
        header = [
            output_sheet.cell(row=1, column=col_idx).value
            for col_idx in range(1, 4)
        ]
        normalized_header = [str(v).strip() if v is not None else "" for v in header]
        if normalized_header != [
            "Region",
            "Avg Penetration (2020-2024)",
            "Growth (2020-2024)",
        ]:
            issues.append("Region-growth output table headers are incorrect or missing.")

        data_rows = 0
        for row_idx in range(2, output_sheet.max_row + 1):
            first = output_sheet.cell(row=row_idx, column=1).value
            second = output_sheet.cell(row=row_idx, column=2).value
            third = output_sheet.cell(row=row_idx, column=3).value
            if all(v not in (None, "") for v in (first, second, third)):
                data_rows += 1
        if data_rows < 2:
            issues.append("Region-growth output does not contain enough complete data rows.")

        summary_labels = set()
        for row_idx in range(1, output_sheet.max_row + 1):
            for col_idx in range(1, min(output_sheet.max_column, 8) + 1, 3):
                label = output_sheet.cell(row=row_idx, column=col_idx).value
                value = output_sheet.cell(row=row_idx, column=col_idx + 1).value
                if label not in (None, "") and value not in (None, ""):
                    summary_labels.add(str(label).strip())
        if "Fastest Growth Region" not in summary_labels:
            issues.append("Region-growth output is missing the `Fastest Growth Region` summary label.")
        if not any("Growth (2020-2024)" in label for label in summary_labels):
            issues.append("Region-growth output is missing the growth summary metric.")

        return issues

    @staticmethod
    def _code_uses_literal_xlsx_name(code_text: str) -> bool:
        if not code_text:
            return False
        return bool(re.search(r"['\"][^'\"]+\.xlsx['\"]", code_text, flags=re.IGNORECASE))

    @staticmethod
    def _parse_time_to_minutes(value: Any) -> int | None:
        if value is None:
            return None
        if hasattr(value, "hour") and hasattr(value, "minute"):
            try:
                return int(value.hour) * 60 + int(value.minute)
            except Exception:
                return None
        text = str(value).strip()
        match = re.match(r"^(\d{1,2}):(\d{2})$", text)
        if not match:
            return None
        hour = int(match.group(1))
        minute = int(match.group(2))
        if hour < 0 or hour > 23 or minute < 0 or minute > 59:
            return None
        return (hour * 60) + minute

    @staticmethod
    def _parse_numeric(value: Any) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return None
        try:
            return float(match.group(0))
        except Exception:
            return None

    @classmethod
    def _inspect_saved_schedule_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        if not output_path or not os.path.exists(output_path):
            return ["Saved output workbook path does not exist on disk."]

        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception as exc:
            return [f"Could not open saved output workbook: {exc}"]

        required_headers = ["Task ID", "Task Name", "Priority", "Start Time", "End Time"]
        schedule_sheet = None
        header_row_index = None
        header_columns: dict[str, int] = {}

        for sheet in workbook.worksheets:
            for row_idx in range(1, min(sheet.max_row, 20) + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 12) + 1)
                ]
                normalized = {
                    str(v).strip(): idx
                    for idx, v in enumerate(row_values, start=1)
                    if v is not None and str(v).strip()
                }
                if all(header in normalized for header in required_headers):
                    schedule_sheet = sheet
                    header_row_index = row_idx
                    header_columns = {header: normalized[header] for header in required_headers}
                    break
            if schedule_sheet is not None:
                break

        if schedule_sheet is None or header_row_index is None:
            return ["Saved workbook does not contain a sheet with required schedule headers."]

        data_rows: list[dict[str, Any]] = []
        summary_label = None
        summary_value = None

        for row_idx in range(header_row_index + 1, schedule_sheet.max_row + 1):
            task_id = schedule_sheet.cell(row=row_idx, column=header_columns["Task ID"]).value
            task_name = schedule_sheet.cell(row=row_idx, column=header_columns["Task Name"]).value
            priority = schedule_sheet.cell(row=row_idx, column=header_columns["Priority"]).value
            start_value = schedule_sheet.cell(row=row_idx, column=header_columns["Start Time"]).value
            end_value = schedule_sheet.cell(row=row_idx, column=header_columns["End Time"]).value

            label_candidate = "" if task_id is None else str(task_id).strip()
            if label_candidate and "total duration" in label_candidate.lower():
                summary_label = label_candidate
                summary_value = task_name
                break

            if not any(v not in (None, "") for v in (task_id, task_name, priority, start_value, end_value)):
                continue

            if not all(v not in (None, "") for v in (task_id, task_name, priority, start_value, end_value)):
                issues.append(f"Schedule row {row_idx} is partially empty.")
                continue

            data_rows.append(
                {
                    "row": row_idx,
                    "task_id": str(task_id).strip(),
                    "task_name": str(task_name).strip(),
                    "priority": str(priority).strip(),
                    "start": start_value,
                    "end": end_value,
                }
            )

        if not data_rows:
            issues.append("Saved schedule sheet contains headers but no complete task rows.")
            return issues

        total_minutes = 0
        previous_end = None
        seen_task_ids: set[str] = set()

        for item in data_rows:
            if not item["task_id"]:
                issues.append(f"Schedule row {item['row']} has empty Task ID.")
                continue
            if item["task_id"] in seen_task_ids:
                issues.append(f"Schedule output duplicates Task ID `{item['task_id']}`.")
            seen_task_ids.add(item["task_id"])

            start_minutes = cls._parse_time_to_minutes(item["start"])
            end_minutes = cls._parse_time_to_minutes(item["end"])
            if start_minutes is None or end_minutes is None:
                issues.append(
                    f"Schedule row {item['row']} has invalid time format; expected `HH:MM` in Start Time and End Time."
                )
                continue
            if end_minutes <= start_minutes:
                issues.append(f"Schedule row {item['row']} has non-positive duration.")
                continue
            if previous_end is not None and start_minutes != previous_end:
                issues.append(
                    f"Schedule timing is not contiguous: row {item['row']} starts at {item['start']} after previous end."
                )
            previous_end = end_minutes
            total_minutes += (end_minutes - start_minutes)

        if summary_label is None:
            issues.append("Saved schedule sheet is missing a total-duration summary row.")
            return issues

        numeric_summary = cls._parse_numeric(summary_value)
        if numeric_summary is None:
            issues.append("Total-duration summary value is missing or non-numeric.")
            return issues

        label_lower = summary_label.lower()
        if "hour" in label_lower:
            expected_value = total_minutes / 60.0
            tolerance = 0.05
        elif "minute" in label_lower:
            expected_value = float(total_minutes)
            tolerance = 1.0
        else:
            expected_value = total_minutes / 60.0
            tolerance = 0.05

        if abs(numeric_summary - expected_value) > tolerance:
            issues.append(
                "Total-duration summary does not match the actual schedule rows "
                f"(expected about {expected_value:.2f}, found {numeric_summary:.2f})."
            )

        return issues

    @staticmethod
    def _issues_only_reference_earlier_failures(issues: list[str]) -> bool:
        if not issues:
            return False
        failure_markers = (
            "syntax error",
            "execution error",
            "nameerror",
            "typeerror",
            "attributeerror",
            "keyerror",
            "traceback",
        )
        normalized = [str(issue or "").strip().lower() for issue in issues if str(issue or "").strip()]
        if not normalized:
            return False
        return all(any(marker in issue for marker in failure_markers) for issue in normalized)

    @classmethod
    def _collect_schedule_code_issues(cls, all_code: str) -> list[str]:
        issues: list[str] = []
        lower_code = (all_code or "").lower()
        uses_schedule_helper = "build_dependency_schedule(" in lower_code
        duration_hour_lines = [
            line.strip().lower()
            for line in (all_code or "").splitlines()
            if "duration (hours)" in line.lower()
        ]
        filename_role_guess = (
            re.search(r"for\s+\w+\s+in\s+all_files\s+if\s+.*['\"]tasks?['\"]\s+in\s+\w+", lower_code)
            or re.search(r"for\s+\w+\s+in\s+all_files\s+if\s+.*['\"]dependenc", lower_code)
            or re.search(r"(task_table|dependency_table)\s*=\s*tables\[\d+\]", lower_code)
            or cls._code_uses_literal_xlsx_name(all_code)
        )
        if "pd.merge(" in lower_code:
            issues.append(
                "Scheduling task incorrectly merges task and dependency tables; these should be processed separately."
            )
        if "same_schema" in lower_code or "schema mismatch between files" in lower_code:
            issues.append(
                "Scheduling task incorrectly assumes task and dependency tables must share the same schema."
            )
        if filename_role_guess:
            issues.append(
                "Scheduling task identifies task/dependency tables from filenames, literal input basenames, or list order instead of verified headers."
            )
        if not uses_schedule_helper:
            issues.append(
                "Scheduling code does not use the runtime helper `build_dependency_schedule(...)`."
            )
        if uses_schedule_helper and "find_table_by_headers(" not in lower_code:
            issues.append(
                "Scheduling code uses the schedule helper but does not use `find_table_by_headers(...)` for role selection."
            )
        if uses_schedule_helper and "load_all_tables(" not in lower_code:
            issues.append(
                "Scheduling code uses the schedule helper but does not use `load_all_tables()` for linear multi-file loading."
            )
        if uses_schedule_helper:
            return issues
        if (
            re.search(r"for\s+\w+(?:\s*,\s*\w+)?\s+in\s+task_df\.iterrows\(\)", lower_code)
            and re.search(r"row\[['\"]depends on['\"]\]", lower_code)
        ):
            issues.append(
                "Scheduling code reads `Depends on` from the task table; root detection and edges must come from the dependency table."
            )
        if (
            re.search(r"for\s+\w+(?:\s*,\s*\w+)?\s+in\s+dep_df\.iterrows\(\)", lower_code)
            and any(
                marker in lower_code
                for marker in (
                    "row['task name']",
                    'row["task name"]',
                    "row['duration (hours)']",
                    'row["duration (hours)"]',
                    "row['priority']",
                    'row["priority"]',
                )
            )
        ):
            issues.append(
                "Scheduling code reads task metadata from the dependency table; keep table roles separate."
            )
        if (
            "if pd.isna(depends_on)" in lower_code
            and "schedule_order.append(task_id)" in lower_code
            and re.search(r"else:\s*.*task_id\s*=\s*row\[['\"]task id['\"]\]", lower_code, flags=re.DOTALL)
        ):
            issues.append(
                "Scheduling code uses `task_id` in the root-task branch before assigning it from the current dependency row."
            )
        if (
            "dropna(subset=['depends on']" in lower_code
            or 'dropna(subset=["depends on"]' in lower_code
            or "['depends on'].notna(" in lower_code
            or '["depends on"].notna(' in lower_code
        ):
            issues.append(
                "Scheduling task drops blank dependencies even though they represent root tasks."
            )
        if re.search(r"adjacency\[\s*row\[['\"]task id['\"]\]\s*\]\.append\(\s*none\s*\)", lower_code):
            issues.append(
                "Scheduling code appends `None` edges for root tasks instead of skipping edge creation."
            )
        reversed_edge_patterns = (
            r"adjacency\[\s*(?:from_task|src_task|task_id|task|current_task)\s*\]\.append\(\s*(?:to_task|depends_on|dep_on|dependency|dep)\s*\)",
            r"in_degree\[\s*(?:to_task|depends_on|dep_on|dependency|dep)\s*\]\s*\+=\s*1",
        )
        if any(re.search(pattern, lower_code) for pattern in reversed_edge_patterns):
            issues.append(
                "Scheduling code reverses dependency direction; edges must be `depends_on -> task_id`."
            )
        if (
            "duration (hours)" in lower_code
            and duration_hour_lines
            and not any(re.search(r"(?:\*\s*60|60\s*\*)", line) for line in duration_hour_lines)
        ):
            issues.append(
                "Scheduling code uses `Duration (hours)` but does not convert hours to minutes before time arithmetic."
            )
        if (
            "add_summary_row(" in lower_code
            and "total duration" in lower_code
            and re.search(r"len\s*\(\s*schedule_order\s*\)\s*\*", lower_code)
        ):
            issues.append(
                "Scheduling summary is derived from task count times a constant instead of real scheduled durations."
            )
        if "start time" not in lower_code or "end time" not in lower_code:
            issues.append(
                "Scheduling code does not clearly build `Start Time` and `End Time` output columns."
            )
        topo_markers = ("in_degree", "adjacency", "queue", "deque(", "topological")
        if not any(marker in lower_code for marker in topo_markers):
            issues.append(
                "Scheduling code does not show explicit dependency-order construction."
            )
        return issues

    def run(self, execution_result: Dict[str, Any], user_question: str,
            understanding_output: str) -> Dict[str, Any]:
        logger.info("Starting validation on execution results")

        run_success = bool(execution_result.get("success", False))
        final_answer = execution_result.get("answer", "") or ""
        steps = execution_result.get("execution_summary", {}).get("execution_steps", [])
        successful_steps = [
            step for step in steps
            if bool(step.get("success"))
        ]
        failed_steps = [
            step for step in steps
            if not bool(step.get("success"))
        ]
        last_successful_step = successful_steps[-1] if successful_steps else {}
        latest_result = str(last_successful_step.get("result") or "")
        latest_code = str(last_successful_step.get("code") or "")
        all_results = " ".join(str(s.get("result") or "") for s in steps)
        all_code = "\n".join(str(s.get("code") or "") for s in steps)
        need_detail = self._parse_output_contract_flag(understanding_output, "requires_detailed_table")
        need_highlight = self._parse_output_contract_flag(understanding_output, "requires_highlight")
        need_summary = self._parse_output_contract_flag(understanding_output, "requires_summary_metrics")
        need_detail, need_summary, need_highlight = self._resolve_contract_expectations(
            user_question,
            need_detail,
            need_summary,
            need_highlight,
        )
        family = detect_task_family(user_question)

        # Fast deterministic failure path: no successful sandbox execution means
        # there is nothing meaningful for the validation LLM to assess.
        if not successful_steps:
            issues = []
            if not run_success:
                issues.append("Execution result indicates failure (`success=false`).")
            if steps:
                issues.append(
                    f"No successful execution step produced runtime output (failed code executions: {len(failed_steps)})."
                )
            else:
                issues.append(
                    "No code execution occurred; all turns were blocked by preflight/forbidden/format checks."
                )
            if self._is_dependency_schedule_request(user_question):
                issues.append(
                    "Scheduling task never produced a saved workbook path, so validation cannot pass."
                )
            feedback = (
                "Get to one successful sandbox execution before calling validation again. "
                "Focus on satisfying preflight constraints first: runtime reads, verified headers, "
                "5-column schedule output, and save via save_workbook_to(output_path)."
            )
            final_assessment = (
                "Validation was short-circuited because execution never produced a successful runnable result."
            )
            validation_result = {
                "validation_passed": False,
                "confidence_score": 0.0,
                "issues_found": issues,
                "improvement_feedback": feedback,
                "final_assessment": final_assessment,
                "verified_answer": "",
                "requires_reexecution": True,
            }
            rendered = (
                "VALIDATION_STATUS: FAILED\n"
                "CONFIDENCE_SCORE: 0.00\n"
                "ISSUES_FOUND:\n"
                + "\n".join(f"- {issue}" for issue in issues)
                + "\nIMPROVEMENT_FEEDBACK:\n"
                + feedback
                + "\nFINAL_ASSESSMENT:\n"
                + final_assessment
            )
            self._log_to_file(f"\n**Validation Analysis:**\n```\n{rendered}\n```\n")
            logger.info("Validation fast-failed: no successful execution steps.")
            logger.info("Validation: FAILED")
            return validation_result

        if (
            run_success
            and final_answer
            and not self._looks_like_file_path(final_answer)
            and not self._is_dependency_schedule_request(user_question)
            and need_detail is False
            and need_highlight is False
            and need_summary is False
        ):
            deterministic_issues = []
            answer_lower = str(final_answer).strip().lower()
            if self._is_missing_data_scan_request(user_question):
                if "missing" not in answer_lower:
                    deterministic_issues.append(
                        "Missing-data task ended with a scalar answer that does not mention missing values."
                    )
            if self._is_room_inconsistency_request(user_question):
                if "room" not in answer_lower:
                    deterministic_issues.append(
                        "Room-format task ended with a scalar answer that does not mention room identifiers."
                    )
            if not deterministic_issues:
                validation_result = {
                    "validation_passed": True,
                    "confidence_score": 1.0,
                    "issues_found": ["None identified."],
                    "improvement_feedback": "No improvement needed - scalar answer passed deterministic validation.",
                    "final_assessment": (
                        "Scalar response matched the output contract; no validation-LLM retry was needed."
                    ),
                    "verified_answer": final_answer,
                    "requires_reexecution": False,
                }
                self._log_to_file(
                    f"\n**Validation Analysis:**\n```\n{self._render_validation_result(validation_result)}\n```\n"
                )
                logger.info("Validation: PASSED (deterministic scalar fast-path)")
                return validation_result

        if (
            run_success
            and final_answer
            and self._looks_like_file_path(final_answer)
            and ("Workbook saved to:" in all_results or "SAVED_FILE:" in all_results)
            and not self._is_dependency_schedule_request(user_question)
        ):
            deterministic_issues = []
            latest_rows = self._extract_rows_written(latest_result)
            latest_column_counts = self._extract_written_column_counts(latest_result)
            if need_detail is True and latest_rows and max(latest_rows) < 2:
                deterministic_issues.append(
                    "Saved workbook write evidence does not show a header plus at least one data row."
                )
            if need_detail is True and latest_column_counts and max(latest_column_counts) < 2:
                deterministic_issues.append(
                    "Saved workbook write evidence does not show a multi-column output table."
                )
            if (
                need_highlight is True
                and "highlighted row(s)" not in latest_result.lower()
                and "no_highlight_rows:" not in latest_result.lower()
            ):
                deterministic_issues.append(
                    "Output contract requires highlight, but no highlight evidence was found."
                )
            if self._is_region_growth_chart_request(user_question):
                if "chart saved to" not in latest_result.lower():
                    deterministic_issues.append(
                        "Region-growth output requires an embedded chart, but no chart-save evidence was found."
                    )
                deterministic_issues.extend(
                    self._inspect_saved_region_growth_workbook(final_answer)
                )
            use_lightweight_validation = (
                self._is_large_workbook(final_answer)
                and need_summary is not True
                and need_highlight is not True
                and not self._is_region_growth_chart_request(user_question)
            )
            if use_lightweight_validation:
                if not os.path.exists(final_answer):
                    deterministic_issues.append(f"Saved output file not found: {final_answer}")
            else:
                deterministic_issues.extend(
                    self._inspect_saved_generic_workbook(
                        final_answer,
                        need_detail=need_detail,
                        need_summary=need_summary,
                    )
                )
            if family is not None:
                validation_mode = get_task_family_validation_mode(family.name)
                if validation_mode == "temporal_aggregation":
                    deterministic_issues.extend(
                        self._inspect_saved_temporal_aggregation_workbook(final_answer)
                    )
                elif validation_mode == "grouped_aggregation":
                    deterministic_issues.extend(
                        self._inspect_saved_grouped_aggregation_workbook(final_answer)
                    )
                elif validation_mode == "relational_join":
                    deterministic_issues.extend(
                        self._inspect_saved_relational_join_workbook(final_answer)
                    )
                elif validation_mode == "allocation":
                    deterministic_issues.extend(
                        self._inspect_saved_allocation_workbook(final_answer)
                    )
                elif validation_mode == "dependency_schedule":
                    deterministic_issues.extend(
                        self._inspect_saved_dependency_schedule_workbook(final_answer)
                    )
                elif validation_mode == "relational_assignment":
                    deterministic_issues.extend(
                        self._inspect_saved_assignment_schedule_workbook(final_answer)
                    )
                elif validation_mode == "temporal_growth":
                    deterministic_issues.extend(
                        self._inspect_saved_region_growth_workbook(final_answer)
                    )
                elif validation_mode == "regression":
                    deterministic_issues.extend(
                        self._inspect_saved_regression_workbook(final_answer)
                    )
                elif validation_mode == "correlation":
                    deterministic_issues.extend(
                        self._inspect_saved_correlation_workbook(final_answer)
                    )
                elif validation_mode == "comparative_multi_sheet":
                    deterministic_issues.extend(
                        self._inspect_saved_comparative_multi_sheet_workbook(final_answer)
                    )
            if not deterministic_issues:
                assessment = (
                    "Saved workbook passed lightweight structural validation based on runtime write evidence; "
                    "the full workbook scan was skipped because the output file exceeded the large-file threshold."
                    if use_lightweight_validation
                    else
                    "Saved workbook passed structural output-contract checks; "
                    "no validation-LLM retry was needed."
                )
                validation_result = {
                    "validation_passed": True,
                    "confidence_score": 1.0,
                    "issues_found": ["None identified."],
                    "improvement_feedback": "No improvement needed - saved workbook passed deterministic validation.",
                    "final_assessment": assessment,
                    "verified_answer": final_answer,
                    "requires_reexecution": False,
                }
                self._log_to_file(
                    f"\n**Validation Analysis:**\n```\n{self._render_validation_result(validation_result)}\n```\n"
                )
                logger.info("Validation: PASSED (deterministic workbook fast-path)")
                return validation_result

        if (
            self._is_dependency_schedule_request(user_question)
            and run_success
            and final_answer
            and self._looks_like_file_path(final_answer)
            and ("Workbook saved to:" in all_results or "SAVED_FILE:" in all_results)
        ):
            family_fast_path = execution_result.get("_family_fast_path")
            deterministic_issues = []
            if family_fast_path != "dependency_constrained_schedule":
                deterministic_issues.extend(self._collect_schedule_code_issues(all_code))
            deterministic_issues.extend(self._inspect_saved_schedule_workbook(final_answer))
            if not deterministic_issues:
                validation_result = {
                    "validation_passed": True,
                    "confidence_score": 1.0,
                    "issues_found": ["None identified."],
                    "improvement_feedback": "No improvement needed - saved schedule passed deterministic validation.",
                    "final_assessment": (
                        "Saved schedule workbook passed structural, timing, and summary checks; "
                        "no validation-LLM retry was needed."
                    ),
                    "verified_answer": final_answer,
                    "requires_reexecution": False,
                }
                self._log_to_file(
                    f"\n**Validation Analysis:**\n```\n{self._render_validation_result(validation_result)}\n```\n"
                )
                logger.info("Validation: PASSED (deterministic schedule fast-path)")
                return validation_result

        conversation_history = execution_result.get("conversation_history", [])
        conversation_history_text = self.history_formatter.format(conversation_history)

        prompt_text = self.prompt_builder.build_validation_prompt(
            user_query=user_question,
            excel_context_understanding=self.excel_context_understanding,
            execution_context=understanding_output,
            execution_success=execution_result.get("success", False),
            total_turns=execution_result.get("total_turns", 0),
            final_answer=execution_result.get("answer", "No answer provided"),
            execution_summary=execution_result.get("execution_summary", {}),
            conversation_history_text=conversation_history_text
        )
        messages = [{"role": "user", "content": prompt_text}]

        try:
            validation_analysis = self.llm_client.get_response(messages)

            validation_result = self.parser.parse(validation_analysis)

            # Rule-based hard checks to reduce offline false positives.
            hard_issues = []
            run_success = bool(execution_result.get("success", False))
            if not run_success:
                hard_issues.append("Execution result indicates failure (`success=false`).")
            # If final answer is file path, execution stdout must contain save evidence.
            if final_answer and self._looks_like_file_path(final_answer):
                if "Workbook saved to:" not in all_results:
                    hard_issues.append(
                        "Final answer is a file path but no 'Workbook saved to:' line found in execution output. "
                        "Execution must use save_workbook_to(output_path) and produce save confirmation."
                    )
                    logger.warning("Validation overridden: path answer without save confirmation in stdout")
            elif self._is_dependency_schedule_request(user_question):
                hard_issues.append(
                    "Scheduling task did not finish with a saved output workbook path."
                )

            rows_written = self._extract_rows_written(latest_result)
            max_rows = max(rows_written) if rows_written else 0
            written_column_counts = self._extract_written_column_counts(latest_result)
            max_written_columns = max(written_column_counts) if written_column_counts else 0
            highlighted_rows = self._extract_highlight_rows(latest_result)
            summary_rows = self._extract_summary_rows(latest_result)
            overlapping_ranges = self._find_overlapping_write_ranges(latest_result)
            lower_results = latest_result.lower()

            if need_detail is True and max_rows < 2:
                hard_issues.append(
                    "Output contract requires detailed table, but write evidence does not show a header plus at least one data row "
                    f"(max rows written = {max_rows})."
                )

            if need_highlight is True:
                if "highlighted row(s)" not in lower_results and "no_highlight_rows:" not in lower_results:
                    hard_issues.append("Output contract requires highlight, but no highlight evidence was found.")
                elif max_rows > 0 and highlighted_rows:
                    invalid = [r for r in highlighted_rows if r < 2 or r > max_rows + 2]
                    if invalid:
                        hard_issues.append(
                            "Highlight rows appear out of expected table range: "
                            + ", ".join(str(v) for v in invalid[:5])
                        )

            if need_summary is True:
                has_summary_signal = ("added summary row" in lower_results) or (len(rows_written) >= 2)
                if not has_summary_signal and self._has_summary_write_signal_from_code(latest_code):
                    has_summary_signal = True
                if not has_summary_signal and need_detail is True and max_rows >= 20:
                    has_summary_signal = True
                if not has_summary_signal:
                    hard_issues.append("Output contract requires summary metrics, but summary write evidence is missing.")
                if need_detail is True and summary_rows and max_rows > 0:
                    overlapping_summary_rows = [r for r in summary_rows if r <= max_rows]
                    if overlapping_summary_rows:
                        hard_issues.append(
                            "Summary row overlaps the detailed table at row(s): "
                            + ", ".join(str(v) for v in overlapping_summary_rows[:5])
                        )

            if overlapping_ranges:
                previous, current = overlapping_ranges[0]
                hard_issues.append(
                    "Output writes overlap in the same sheet: "
                    f"{self._format_write_range(previous)} vs {self._format_write_range(current)}."
                )

            if self._is_regression_request(user_question):
                reported_columns = self._extract_reported_columns(all_results)
                expected_predictors = self._expected_regression_predictors(reported_columns)
                if expected_predictors:
                    feature_cols = self._extract_feature_cols_from_code(all_code)
                    missing_from_feature_cols: list[str] = []
                    if feature_cols:
                        used = {c.lower() for c in feature_cols}
                        missing_from_feature_cols = [
                            p for p in expected_predictors
                            if p.lower() not in used
                        ]
                    weight_labels = self._extract_weight_labels(all_results)
                    missing_from_weights: list[str] = []
                    if weight_labels:
                        present_labels = {w.lower() for w in weight_labels}
                        missing_from_weights = [
                            p for p in expected_predictors
                            if p.lower() not in present_labels
                        ]

                    missing = missing_from_feature_cols or missing_from_weights
                    if missing:
                        hard_issues.append(
                            "Regression feature coverage incomplete. Missing predictor(s): "
                            + ", ".join(missing[:6])
                        )

            if self._is_dependency_schedule_request(user_question):
                lower_code = all_code.lower()
                uses_schedule_helper = "build_dependency_schedule(" in lower_code
                duration_hour_lines = [
                    line.strip().lower()
                    for line in all_code.splitlines()
                    if "duration (hours)" in line.lower()
                ]
                filename_role_guess = (
                    re.search(r"for\s+\w+\s+in\s+all_files\s+if\s+.*['\"]tasks?['\"]\s+in\s+\w+", lower_code)
                    or re.search(r"for\s+\w+\s+in\s+all_files\s+if\s+.*['\"]dependenc", lower_code)
                    or re.search(r"(task_table|dependency_table)\s*=\s*tables\[\d+\]", lower_code)
                    or self._code_uses_literal_xlsx_name(all_code)
                )
                if "pd.merge(" in lower_code:
                    hard_issues.append(
                        "Scheduling task incorrectly merges task and dependency tables; these should be processed separately."
                    )
                if "same_schema" in lower_code or "schema mismatch between files" in lower_code:
                    hard_issues.append(
                        "Scheduling task incorrectly assumes task and dependency tables must share the same schema."
                    )
                if filename_role_guess:
                    hard_issues.append(
                        "Scheduling task identifies task/dependency tables from filenames, literal input basenames, or list order instead of verified headers."
                    )
                if not uses_schedule_helper:
                    hard_issues.append(
                        "Scheduling code does not use the runtime helper `build_dependency_schedule(...)`."
                    )
                if uses_schedule_helper and "find_table_by_headers(" not in lower_code:
                    hard_issues.append(
                        "Scheduling code uses the schedule helper but does not use `find_table_by_headers(...)` for role selection."
                    )
                if uses_schedule_helper and "load_all_tables(" not in lower_code:
                    hard_issues.append(
                        "Scheduling code uses the schedule helper but does not use `load_all_tables()` for linear multi-file loading."
                    )
                if uses_schedule_helper:
                    if max_written_columns and max_written_columns < 5:
                        hard_issues.append(
                            "Scheduling output appears to have fewer than 5 columns; expected "
                            "`Task ID`, `Task Name`, `Priority`, `Start Time`, `End Time`."
                        )
                    if final_answer and self._looks_like_file_path(final_answer):
                        hard_issues.extend(self._inspect_saved_schedule_workbook(final_answer))
                else:
                    if (
                        re.search(r"for\s+\w+(?:\s*,\s*\w+)?\s+in\s+task_df\.iterrows\(\)", lower_code)
                        and re.search(r"row\[['\"]depends on['\"]\]", lower_code)
                    ):
                        hard_issues.append(
                            "Scheduling code reads `Depends on` from the task table; root detection and edges must come from the dependency table."
                        )
                    if (
                        re.search(r"for\s+\w+(?:\s*,\s*\w+)?\s+in\s+dep_df\.iterrows\(\)", lower_code)
                        and any(
                            marker in lower_code
                            for marker in (
                                "row['task name']",
                                'row["task name"]',
                                "row['duration (hours)']",
                                'row["duration (hours)"]',
                                "row['priority']",
                                'row["priority"]',
                            )
                        )
                    ):
                        hard_issues.append(
                            "Scheduling code reads task metadata from the dependency table; keep table roles separate."
                        )
                    if (
                        "if pd.isna(depends_on)" in lower_code
                        and "schedule_order.append(task_id)" in lower_code
                        and re.search(r"else:\s*.*task_id\s*=\s*row\[['\"]task id['\"]\]", lower_code, flags=re.DOTALL)
                    ):
                        hard_issues.append(
                            "Scheduling code uses `task_id` in the root-task branch before assigning it from the current dependency row."
                        )
                    if (
                        "dropna(subset=['depends on']" in lower_code
                        or 'dropna(subset=["depends on"]' in lower_code
                        or "['depends on'].notna(" in lower_code
                        or '["depends on"].notna(' in lower_code
                    ):
                        hard_issues.append(
                            "Scheduling task drops blank dependencies even though they represent root tasks."
                        )
                    if max_written_columns and max_written_columns < 5:
                        hard_issues.append(
                            "Scheduling output appears to have fewer than 5 columns; expected "
                            "`Task ID`, `Task Name`, `Priority`, `Start Time`, `End Time`."
                        )
                    if (
                        "itertuples(" in lower_code
                        and any(
                            marker in lower_code
                            for marker in (
                                ".task_id",
                                ".task_name",
                                ".depends_on",
                                ".duration_hours",
                                ".duration__hours",
                            )
                        )
                    ):
                        hard_issues.append(
                            "Scheduling code uses tuple-attribute access for headers with spaces; use exact column names like `row['Task ID']`."
                        )
                    reversed_edge_patterns = (
                        r"adjacency\[\s*(?:from_task|src_task|task_id|task|current_task)\s*\]\.append\(\s*(?:to_task|depends_on|dep_on|dependency|dep)\s*\)",
                        r"in_degree\[\s*(?:to_task|depends_on|dep_on|dependency|dep)\s*\]\s*\+=\s*1",
                    )
                    if any(re.search(pattern, lower_code) for pattern in reversed_edge_patterns):
                        hard_issues.append(
                            "Scheduling code reverses dependency direction; edges must be `depends_on -> task_id`."
                        )
                    if (
                        "duration (hours)" in lower_code
                        and duration_hour_lines
                        and not any(re.search(r"(?:\*\s*60|60\s*\*)", line) for line in duration_hour_lines)
                    ):
                        hard_issues.append(
                            "Scheduling code uses `Duration (hours)` but does not convert hours to minutes before time arithmetic."
                        )
                    if (
                        "add_summary_row(" in lower_code
                        and "total duration" in lower_code
                        and re.search(r"len\s*\(\s*schedule_order\s*\)\s*\*", lower_code)
                    ):
                        hard_issues.append(
                            "Scheduling summary is derived from task count times a constant instead of real scheduled durations."
                        )
                    if "start time" not in lower_code or "end time" not in lower_code:
                        hard_issues.append(
                            "Scheduling code does not clearly build `Start Time` and `End Time` output columns."
                        )
                    topo_markers = ("in_degree", "adjacency", "queue", "deque(", "topological")
                    if not any(marker in lower_code for marker in topo_markers):
                        hard_issues.append(
                            "Scheduling code does not show explicit dependency-order construction."
                        )
                    if final_answer and self._looks_like_file_path(final_answer):
                        hard_issues.extend(self._inspect_saved_schedule_workbook(final_answer))

            if validation_result.get("validation_passed") and validation_result.get("issues_found"):
                hard_issues.append("Validator marked PASSED but still reported issues_found.")

            if hard_issues:
                validation_result["validation_passed"] = False
                merged_issues = list(validation_result.get("issues_found") or [])
                merged_issues.extend(hard_issues)
                validation_result["issues_found"] = merged_issues
                feedback = (validation_result.get("improvement_feedback") or "").strip()
                if not feedback:
                    validation_result["improvement_feedback"] = (
                        "Fix rule-based validation issues first: contract alignment, output shape, and save evidence."
                    )
            elif (
                run_success
                and final_answer
                and self._looks_like_file_path(final_answer)
                and validation_result.get("validation_passed") is False
                and self._issues_only_reference_earlier_failures(validation_result.get("issues_found") or [])
            ):
                validation_result["validation_passed"] = True
                validation_result["issues_found"] = ["None identified."]
                validation_result["improvement_feedback"] = "No improvement needed - later execution turn corrected earlier failures."
                validation_result["final_assessment"] = (
                    "Earlier execution errors were corrected in later turns; final saved output passed rule-based checks."
                )

            logger.info(
                f"Validation completed. Confidence: {validation_result['confidence_score']:.2f}"
            )
            logger.info(
                f"Validation: {'PASSED' if validation_result['validation_passed'] else 'FAILED'}"
            )

            if validation_result['validation_passed']:
                logger.info("Answer validated - ready for final output")
                validation_result['verified_answer'] = execution_result.get("answer", "")
                validation_result['requires_reexecution'] = False
            else:
                logger.warning("Issues found - recommending re-execution")
                validation_result['requires_reexecution'] = True

            self._log_to_file(
                f"\n**Validation Analysis:**\n```\n{self._render_validation_result(validation_result)}\n```\n"
            )

            return validation_result

        except Exception as e:
            logger.error(f"Error during validation: {str(e)}")
            return {
                "validation_passed": False,
                "confidence_score": 0.0,
                "issues_found": [f"Validation process failed: {str(e)}"],
                "improvement_feedback": (
                    "Unable to provide feedback due to validation error. "
                    "Please review the execution manually."
                ),
                "final_assessment": "Unable to validate due to validation error",
                "verified_answer": "",
                "requires_reexecution": False
            }
