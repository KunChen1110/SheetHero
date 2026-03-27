"""Generate a short user-facing final response."""

from __future__ import annotations

import os
import re
from typing import Any, Dict, Optional

from openpyxl import load_workbook

from ...log.logger_registry import LoggerRegistry
from ...task_families import detect_task_family

logger = LoggerRegistry.setup_logger(__name__)


class FinalResponseStage:
    """Produce a short, contentful final response for the user."""

    def __init__(self, client, deployment: str, progress_logger=None):
        self.client = client
        self.deployment = deployment
        self.progress_logger = progress_logger

    @staticmethod
    def _looks_like_file_path(value: str) -> bool:
        text = (value or "").strip().lower()
        return text.endswith((".xlsx", ".xls", ".csv"))

    @staticmethod
    def _compact_question(user_question: str) -> str:
        text = " ".join((user_question or "").strip().split())
        return text[:240]

    @classmethod
    def _is_task_schedule_question(cls, user_question: str) -> bool:
        lowered = cls._compact_question(user_question).lower()
        if not any(token in lowered for token in ("schedule", "scheduling")):
            return False
        return any(
            token in lowered
            for token in (
                "depends on",
                "dependency",
                "dependencies",
                "task id",
                "start time",
                "end time",
                "root task",
                "prerequisite",
            )
        )

    @classmethod
    def _question_content_label(cls, user_question: str, workbook_summary: Dict[str, Any]) -> str:
        text = cls._compact_question(user_question)
        lowered = text.lower()
        headers = [str(item).strip() for item in (workbook_summary.get("headers") or []) if str(item).strip()]
        family = detect_task_family(user_question)

        if family and family.final_label:
            return family.final_label

        if cls._is_task_schedule_question(user_question):
            return "task scheduling table"

        keyword_labels = (
            (("correlation matrix",), "correlation matrix"),
            (("correlation",), "correlation report"),
            (("dashboard",), "dashboard"),
            (("merge", "merged"), "merged spreadsheet"),
            (("fill missing", "missing data"), "completed data table"),
            (("rank", "screening", "candidates"), "candidate screening report"),
            (("inventory", "eoq"), "inventory report"),
            (("cycle", "graph"), "cycle detection report"),
            (("market share", "shipment"), "market share and shipment report"),
            (("mobile reviews",), "mobile review summary"),
            (("diabetes",), "regional diabetes report"),
            (("financial", "cash flow"), "financial report"),
            (("students and their tutors", "students attending", "tutor meeting"), "tutor meeting schedule"),
        )
        for tokens, label in keyword_labels:
            if any(token in lowered for token in tokens):
                return label

        if headers:
            if len(headers) == 2:
                return f"spreadsheet with {headers[0]} and {headers[1]}"
            if len(headers) >= 3:
                return f"spreadsheet with columns {headers[0]}, {headers[1]}, and {headers[2]}"

        cleaned = re.sub(r"^(please|kindly)\s+", "", text, flags=re.IGNORECASE).strip(" .")
        cleaned = re.sub(r"\b(output|create|generate|produce|return|save)\b.*$", "", cleaned, flags=re.IGNORECASE).strip(" ,.")
        return cleaned or "spreadsheet"

    @classmethod
    def _inspect_output_workbook(cls, output_path: str) -> Dict[str, Any]:
        summary: Dict[str, Any] = {
            "sheet_name": "",
            "headers": [],
            "data_rows": 0,
            "metric_label": "",
            "metric_value": "",
            "metrics": {},
        }
        if not output_path or not os.path.exists(output_path):
            return summary
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception:
            return summary

        for sheet in workbook.worksheets:
            header_row_idx = None
            headers = []
            for row_idx in range(1, min(sheet.max_row, 20) + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 12) + 1)
                ]
                non_empty = [str(v).strip() for v in row_values if v not in (None, "")]
                if len(non_empty) >= 2:
                    header_row_idx = row_idx
                    headers = non_empty
                    break
            if header_row_idx is None:
                continue

            data_rows = 0
            metric_label = ""
            metric_value = ""
            metrics: Dict[str, str] = {}
            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 12) + 1)
                ]
                if not any(v not in (None, "") for v in row_values):
                    continue

                row_metric_found = False
                for col_idx in range(0, len(row_values) - 1):
                    first = row_values[col_idx]
                    second = row_values[col_idx + 1]
                    label = str(first).strip() if first not in (None, "") else ""
                    if label and any(marker in label.lower() for marker in ("total", "average", "count", "max", "min")):
                        metrics[label] = "" if second in (None, "") else str(second).strip()
                        if not metric_label:
                            metric_label = label
                            metric_value = "" if second in (None, "") else str(second).strip()
                        row_metric_found = True
                if row_metric_found:
                    continue

                first = row_values[0] if row_values else None
                second = row_values[1] if len(row_values) > 1 else None
                if first not in (None, "") and second not in (None, ""):
                    data_rows += 1

            summary.update(
                {
                    "sheet_name": sheet.title,
                    "headers": headers,
                    "data_rows": data_rows,
                    "metric_label": metric_label,
                    "metric_value": metric_value,
                    "metrics": metrics,
                }
            )
            return summary
        return summary

    @staticmethod
    def _large_workbook_threshold_bytes() -> int:
        return int(os.getenv("SHEETHERO_LARGE_WORKBOOK_THRESHOLD_BYTES", "10000000"))

    @classmethod
    def _is_large_output_file(cls, output_path: str) -> bool:
        if not output_path or not os.path.exists(output_path):
            return False
        try:
            return os.path.getsize(output_path) >= cls._large_workbook_threshold_bytes()
        except OSError:
            return False

    @classmethod
    def _summary_from_execution_result(cls, execution_result: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        summary: Dict[str, Any] = {
            "sheet_name": "",
            "headers": [],
            "data_rows": 0,
            "metric_label": "",
            "metric_value": "",
            "metrics": {},
        }
        execution_steps = ((execution_result or {}).get("execution_summary") or {}).get("execution_steps") or []
        successful_steps = [step for step in execution_steps if step.get("success")]
        if not successful_steps:
            return summary
        latest_result = str(successful_steps[-1].get("result") or "")
        row_counts = [int(value) for value in re.findall(r"Wrote\s+(\d+)\s+rows\s+to", latest_result, flags=re.IGNORECASE)]
        if row_counts:
            summary["data_rows"] = max(max(row_counts) - 1, 0)
        return summary

    @classmethod
    def _fallback_short_answer(
        cls,
        user_question: str,
        final_answer: str,
        validation_passed: bool,
        workbook_summary: Dict[str, Any],
    ) -> str:
        question = cls._compact_question(user_question)
        content_label = cls._question_content_label(user_question, workbook_summary)
        metric_label = workbook_summary.get("metric_label") or ""
        metric_value = workbook_summary.get("metric_value") or ""
        metrics = workbook_summary.get("metrics") or {}
        data_rows = int(workbook_summary.get("data_rows") or 0)

        def _find_metric(keyword: str) -> tuple[str, str]:
            for label, value in metrics.items():
                if keyword in label.lower():
                    return label, value
            return "", ""

        if cls._looks_like_file_path(final_answer):
            if validation_passed:
                total_label, total_value = _find_metric("total")
                average_label, average_value = _find_metric("average")
                if total_label and average_label and total_value and average_value:
                    return (
                        f"Successfully generated the {content_label}; "
                        f"{total_label} is {total_value}, and {average_label} is {average_value}."
                    )
                if metric_label and metric_value:
                    return (
                        f"Successfully generated the {content_label}; "
                        f"{metric_label} is {metric_value}."
                    )
                if data_rows > 0:
                    unit = "record" if data_rows == 1 else "rows"
                    return f"Successfully generated the {content_label} with {data_rows} {unit}."
                return f"Successfully generated the {content_label}."
            return f"Generated the {content_label}, but validation still found issues."

        answer_text = " ".join((final_answer or "").strip().split())
        if answer_text:
            if len(answer_text.split()) >= 5 and re.search(r"[.!?]$", answer_text):
                return answer_text
            lowered = question.lower()
            if "average spending" in lowered:
                return f"The average spending is {answer_text}."
            if "average" in lowered:
                return f"The average result is {answer_text}."
            if "total" in lowered or "sum" in lowered:
                return f"The total result is {answer_text}."
            if "count" in lowered:
                return f"The count is {answer_text}."
            if "maximum" in lowered or "highest" in lowered or "max " in lowered:
                return f"The maximum result is {answer_text}."
            if "minimum" in lowered or "lowest" in lowered or "min " in lowered:
                return f"The minimum result is {answer_text}."
            return f"The result is {answer_text}."
        if validation_passed:
            return f"Successfully completed the request for {question}."
        return f"The request for {question} completed with validation issues."

    def run(
        self,
        user_question: str,
        final_answer: str,
        validation_result: Optional[Dict[str, Any]],
        execution_result: Optional[Dict[str, Any]] = None,
    ) -> str:
        validation_result = validation_result or {}
        validation_passed = bool(validation_result.get("validation_passed"))
        workbook_summary = {}
        if self._looks_like_file_path(final_answer):
            if self._is_large_output_file(final_answer):
                workbook_summary = self._summary_from_execution_result(execution_result)
            else:
                workbook_summary = self._inspect_output_workbook(final_answer)

        fallback = self._fallback_short_answer(
            user_question=user_question,
            final_answer=final_answer,
            validation_passed=validation_passed,
            workbook_summary=workbook_summary,
        )

        # Default to deterministic final wording for both online and offline.
        # This avoids an extra LLM round-trip after the task has already succeeded.
        use_llm = os.getenv("SHEETHERO_FINAL_RESPONSE_LLM", "0").strip() == "1"
        if not use_llm:
            return fallback

        if not self._looks_like_file_path(final_answer):
            return fallback

        prompt = (
            "Write exactly one short user-facing sentence.\n"
            "- If a spreadsheet file was produced, describe what spreadsheet was generated based on the question.\n"
            "- If a key metric is available, include it naturally.\n"
            "- Do not mention file paths, validation modules, code, or internal stages.\n"
            "- Keep it under 28 words.\n\n"
            f"User question: {self._compact_question(user_question)}\n"
            f"Final answer type: {'file' if self._looks_like_file_path(final_answer) else 'text'}\n"
            f"Validation passed: {validation_passed}\n"
            f"Workbook summary: {workbook_summary}\n"
            f"Fallback wording: {fallback}\n"
        )

        try:
            response = self.client.chat.completions.create(
                model=self.deployment,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=80,
            )
            content = (response.choices[0].message.content or "").strip()
            content = " ".join(content.split())
            if content:
                return content
        except Exception as exc:
            logger.warning(f"Final response generation fallback used: {exc}")

        return fallback
