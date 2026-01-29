"""Excel context builder for token-budgeted summaries."""

import os
from typing import Any, Dict

from openpyxl.utils import get_column_letter

from ...agent.utils.token_cost import calculate_token_cost_line
from ...log.logger_registry import LoggerRegistry

logger = LoggerRegistry.setup_logger(__name__)


class ExcelContextBuilder:
    """Builds markdown summaries of Excel workbooks under a token budget."""

    def __init__(self, excel_paths, workbooks: Dict[str, Any]):
        self.excel_paths = excel_paths
        self.workbooks = workbooks

    def build(self, total_token_budget: int = 50000) -> str:
        """
        Generate a markdown summary of Excel sheets for AI context.
        """
        try:
            workbooks = self.workbooks or {}
            overview_parts = []

            if len(workbooks) > 1:
                overview_parts.append(
                    f"📊 **Multiple Excel Files Overview ({len(workbooks)} files)**\n"
                )
            else:
                first_path = self.excel_paths[0] if self.excel_paths else "unknown"
                overview_parts.append(
                    f"📊 **Excel File Overview: {os.path.basename(first_path)}**\n"
                )

            available_tokens = total_token_budget
            tokens_per_file = available_tokens // len(workbooks) if workbooks else 0

            for excel_path, workbook in workbooks.items():
                file_parts = []
                file_parts.append(f"\n{'=' * 60}")
                file_parts.append(f"📁 **File: {os.path.basename(excel_path)}**")
                file_parts.append(f"**Full Path:** {excel_path}")
                file_parts.append(f"**Total Sheets:** {len(workbook.sheetnames)}\n")

                tokens_per_sheet = (
                    tokens_per_file // len(workbook.sheetnames)
                    if workbook.sheetnames else 0
                )

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_parts = []

                    sheet_parts.append(
                        f"\n**📄 Sheet: '{sheet_name}'** (in {os.path.basename(excel_path)})"
                    )
                    sheet_parts.append(
                        f"- Dimensions: {sheet.max_row} rows × {sheet.max_column} columns"
                    )

                    if tokens_per_sheet > 0:
                        preview_result = self._get_sheet_preview_with_token_limit(
                            sheet,
                            tokens_per_sheet,
                            max_rows=min(sheet.max_row, 10000),
                            max_cols=min(sheet.max_column, 1000)
                        )

                        sheet_parts.append(
                            f"- Data Preview ({preview_result['rows_shown']} of {sheet.max_row} rows, "
                            f"{preview_result['cols_shown']} of {sheet.max_column} columns):"
                        )

                        if preview_result['is_truncated']:
                            sheet_parts.append("  ⚠️ Preview truncated to fit token budget")

                        sheet_parts.append("  Data:")
                        markdown_rows = []
                        for row_data in preview_result['formatted_data']:
                            markdown_rows.append(f"| {' | '.join(row_data)} |")

                        if markdown_rows:
                            sheet_parts.append("  " + "\\n".join(markdown_rows))

                        if preview_result['rows_shown'] < sheet.max_row:
                            sheet_parts.append("\n  📊 Sheet Summary:")
                            sheet_parts.append(f"  - Total rows: {sheet.max_row}")
                            sheet_parts.append(f"  - Total columns: {sheet.max_column}")
                            sheet_parts.append(
                                f"  - Rows shown in preview: {preview_result['rows_shown']}"
                            )

                    file_parts.extend(sheet_parts)

                overview_parts.extend(file_parts)

            return "\n".join(overview_parts)

        except Exception as e:
            logger.error(f"Error generating Excel overview: {str(e)}")
            return f"❌ Error generating Excel overview: {str(e)}"

    def _get_sheet_preview_with_token_limit(self, sheet, token_budget: int,
                                            max_rows: int = 10000,
                                            max_cols: int = 1000) -> Dict[str, Any]:
        """
        Generate a sheet preview that fits within the AI's token budget.
        """
        preview_data = []
        formatted_data = []
        tokens_used = 0
        rows_shown = 0

        start_row = 1
        max_data_rows = min(max_rows, sheet.max_row)
        max_data_cols = min(max_cols, sheet.max_column)

        for row_idx in range(start_row, max_data_rows + 1):
            row_cells = []
            formatted_row_cells = []

            for col_idx in range(1, max_data_cols + 1):
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                cell = sheet[cell_ref]
                cell_value = cell.value

                display_value = str(cell_value) if cell_value is not None else ""
                display_value = (
                    display_value
                    .replace("|", "\\|")
                    .replace("\n", " ")
                    .replace("\r", " ")
                )

                formatted_cell = f"{cell_ref}:{display_value}"
                row_cells.append(cell_value)
                formatted_row_cells.append(formatted_cell)

            row_str = " | ".join(formatted_row_cells)
            row_tokens = calculate_token_cost_line(row_str)

            if tokens_used + row_tokens > token_budget:
                if rows_shown < 5:
                    preview_data.append(row_cells)
                    formatted_data.append(formatted_row_cells)
                    rows_shown += 1
                    tokens_used += row_tokens
                break

            preview_data.append(row_cells)
            formatted_data.append(formatted_row_cells)
            rows_shown += 1
            tokens_used += row_tokens

        return {
            'data': preview_data,
            'formatted_data': formatted_data,
            'rows_shown': rows_shown,
            'cols_shown': max_data_cols,
            'start_row': start_row,
            'is_truncated': rows_shown < max_data_rows,
            'tokens_used': tokens_used
        }
