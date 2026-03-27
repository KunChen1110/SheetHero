"""SheetHero orchestrator class."""
import os
from typing import Any, Dict, List, Optional, Union

from openai import OpenAI
import openpyxl
import pandas as pd
import numpy as np
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from .runner import AgentRunner
from ..io.output_policy import OutputPolicy
from ...config.settings import Config
from ...environment import Sandbox
from ...environment.spreadsheet.world import SpreadsheetWorld
from ...log.progress_logger import ProgressLogger
from ...stages.execution.stage import ExecutionStage
from ...stages.final_response.stage import FinalResponseStage
from ...stages.qa.stage import QualityAssuranceStage
from ...stages.cleaning.stage import DataCleaningStage
from ...stages.diagnose.stage import DiagnoseStage
from ...router.diagnose_router import DiagnoseRouter
from ...stages.interact.stage import InteractStage
from ...stages.understanding.stage import UnderstandingStage
from ...stages.understanding.context_builder import ExcelContextBuilder
from ...stages.validation.stage import ValidationStage
from .session import SheetHeroSession
from ..utils.context_extractor import ContextExtractor
from ..utils.sheethero_helpers import (
    append_ui_thought,
    build_default_output_path,
    build_clarification_message,
    build_output_instruction,
    log_final_report,
    log_session_context,
)


class SheetHero:
    def __init__(self, excel_paths: Union[str, List[str]],
                 config: Config,
                 load_excel: bool = True):
        self.config = config
        raw_paths = excel_paths if isinstance(excel_paths, list) else [excel_paths]
        self.excel_paths = [p for p in raw_paths if p]
        self.has_excel = len(self.excel_paths) > 0
        self.output_preferences = OutputPolicy.build_preferences(
            mode=self.config.output_mode,
            file_path=self.config.output_file
        )

        self.progress_logger = ProgressLogger(self.excel_paths)

        if self.output_preferences.get("mode") == "file":
            self._output_path = self.output_preferences.get("file_path")
        else:
            self._output_path = build_default_output_path(self.excel_paths, __file__)
        self._output_existed_before = os.path.exists(self._output_path)

        self.output_instruction = build_output_instruction(
            self.output_preferences,
            self._output_path
        )

        api_key = (self.config.api_key or "").strip()
        base_url = (self.config.base_url or "").strip()
        if not api_key and not base_url:
            raise ValueError("API key is missing and base_url is not set in Config")

        # OpenAI-compatible local servers still require a non-empty api_key field.
        # Use a harmless placeholder when caller chooses base_url-only mode.
        client_kwargs = {"api_key": api_key or "local-key"}
        if base_url:
            client_kwargs["base_url"] = base_url
        self.client = OpenAI(**client_kwargs)
        self.prompt_profile = self._resolve_prompt_profile()
        
        # initialize the sandbox
        self.sandbox = Sandbox(
            excel_paths=self.excel_paths,
            output_preferences=self.output_preferences,
            output_path=self._output_path,
            enabled_namespaces=["spreadsheet"] if self.has_excel else [],
            progress_logger=self.progress_logger,
            load_excel=load_excel and self.has_excel
        )

        # initialize modules
        self.understanding_module = UnderstandingStage(
            self.client,
            self.config.deployment,
            progress_logger=self.progress_logger,
            prompt_profile=self.prompt_profile,
        )
        self.interact_module = InteractStage(
            self.client,
            self.config.deployment,
            progress_logger=self.progress_logger
        )
        self.diagnose_module = DiagnoseStage(
            self.client,
            self.config.deployment,
            self.excel_paths,
            sandbox=self.sandbox,
            token_budget=self.config.total_token_budget,
            progress_logger=self.progress_logger,
            prompt_profile=self.prompt_profile,
        )
        self.diagnose_router = DiagnoseRouter(
            self.client,
            self.config.deployment,
            progress_logger=self.progress_logger,
            prompt_profile=self.prompt_profile,
        )
        self.execution_module = ExecutionStage(
            self.client,
            self.config.deployment,
            self.sandbox,
            excel_context_execution="",
            output_instruction=self.output_instruction,
            progress_log_file=self.progress_logger.file,
            prompt_profile=self.prompt_profile,
        )
        self.validation_module = ValidationStage(
            self.client,
            self.config.deployment,
            "",
            progress_log_file=self.progress_logger.file,
            prompt_profile=self.prompt_profile,
        )
        self.final_response_module = FinalResponseStage(
            self.client,
            self.config.deployment,
            progress_logger=self.progress_logger,
        )
        self.qa_stage = QualityAssuranceStage(
            self.client,
            self.config.deployment,
            progress_logger=self.progress_logger,
            prompt_profile=self.prompt_profile,
        )
        self.qa_stage.max_qa_rounds = self.config.max_qa_rounds
        self.cleaning_module = DataCleaningStage(
            self.client,
            self.config.deployment,
            token_budget=self.config.total_token_budget,
            progress_logger=self.progress_logger
        )

    def _resolve_prompt_profile(self) -> str:
        """Select prompt profile for current runtime."""
        forced = (os.getenv("SHEETHERO_PROMPT_PROFILE") or "").strip().lower()
        if forced in {"offline", "offline_strict"}:
            return "offline_strict"
        if forced in {"online", "online_rich"}:
            return "online_rich"

        base_url = (self.config.base_url or "").lower()
        if any(marker in base_url for marker in ("localhost", "127.0.0.1", "11434", "ollama")):
            return "offline_strict"
        return "online_rich"

    def _resolve_max_cycles(self) -> int:
        """Max execute-validate cycles per request/session."""
        raw = (os.getenv("SHEETHERO_MAX_EXEC_VALIDATE_CYCLES") or "").strip()
        if raw:
            try:
                return max(1, int(raw))
            except ValueError:
                pass
        return max(1, min(int(self.config.max_turns), 4))

    def _resolve_execution_turn_budget(self) -> int:
        """Max execution turns per iteration."""
        raw = (os.getenv("SHEETHERO_EXECUTION_TURN_BUDGET") or "").strip()
        if raw:
            try:
                return max(1, int(raw))
            except ValueError:
                pass
        return max(1, min(int(self.config.max_turns), 5))
    
    # Run the agent
    def run(self, user_question: str) -> Dict[str, Any]:
        if not self.has_excel:
            response = self.interact_module.run(user_question)
            return {
                "type": "final",
                "result": {"final_answer": response},
                "message": response,
            }
        runner = AgentRunner(
            max_turns=self.config.max_turns,
            max_iterations=self._resolve_max_cycles(),
            execution_turn_budget=self._resolve_execution_turn_budget(),
            progress_logger=self.progress_logger,
            token_budget=self.config.total_token_budget
        )
        return runner.run(
            user_question,
            self.understanding_module,
            self.execution_module,
            self.validation_module,
            self.final_response_module,
        )

    def start_session(self, user_question: str) -> SheetHeroSession:
        self.qa_stage.reset()
        return SheetHeroSession(original_query=user_question)

    def step(self, session: SheetHeroSession, user_input: Optional[str] = None) -> Dict[str, Any]:
        qa_stage = self.qa_stage
        current_request = user_input if user_input is not None else session.original_query
    

        # ========== INIT ==========
        if session.state == "init":
            log_session_context(self.progress_logger, session, current_request)
            init_result = self._handle_init(session, current_request)
            if init_result is not None:
                return init_result

        # ========== UNDERSTANDING ==========
        if session.state == "understanding":
            return self._handle_understanding(session, current_request)

        # ========== DIAGNOSING ==========
        if session.state == "diagnosing":
            return self._handle_diagnosing(session, current_request)

        # ========== QA ==========
        if session.state == "qa":
            if user_input is None:
                return {"type": "clarification", "stage": "qa", "message": "Please clarify your request."}

            return self._handle_qa(session, user_input, qa_stage)

        # ========== CLEANING ==========
        if session.state == "cleaning":
            return self._handle_cleaning(session, qa_stage, current_request)

        # ========== EXECUTING ==========
        if session.state == "executing":
            return self._handle_execution(session, current_request)

        # ========== VALIDATION ==========
        if session.state == "validation":
            return self._handle_validation(session, current_request)

        if session.state == "done":
            return {"type": "final", "result": session.result}

        return {"type": "error", "message": f"Unknown session state: {session.state}"}

    def _log(self, message: str, to_terminal: bool = False) -> None:
        if self.progress_logger:
            self.progress_logger.log(message, to_terminal=to_terminal)

    def _handle_init(self, session: SheetHeroSession, current_request: str) -> Optional[Dict[str, Any]]:
        if session.current_workbooks is None:
            session.current_workbooks = self.sandbox.workbooks
        active_workbooks = session.current_workbooks

        if not active_workbooks:
            needs_spreadsheet = self.interact_module.needs_spreadsheet(current_request)
            if not needs_spreadsheet:
                self._log("[INTERACT] case=1 no_spreadsheet_needed", to_terminal=False)
                response = self.interact_module.run(current_request)
                session.understanding = response
                session.result = {"final_answer": response}
                session.state = "done"
                return {"type": "final", "result": session.result, "message": response}

            # Only fall back to previous files when this turn has no new files.
            # If new files were provided but not loaded, ask user to re-upload/fix input.
            if session.previous_workbooks and not self.has_excel:
                matches_context = self.interact_module.context_matches(
                    current_request,
                    session.context_understanding or ""
                )
                if matches_context:
                    session.current_workbooks = session.previous_workbooks
                    active_workbooks = session.current_workbooks
                else:
                    session.context_understanding = self.interact_module.summarize_context(
                        current_request
                    )
                    self._log(
                        f"[INTERACT] case=3 context_mismatch context=\"{session.context_understanding}\"",
                        to_terminal=False
                    )
                    message = (
                        "Your request seems to switch to a different analysis topic. "
                        "Please confirm whether to switch Excel files, or upload a new Excel file."
                    )
                    session.result = {"final_answer": message}
                    session.state = "done"
                    return {"type": "final", "result": session.result, "message": message}

            if not active_workbooks:
                session.context_understanding = self.interact_module.summarize_context(
                    current_request
                )
                self._log(
                    f"[INTERACT] case=2 needs_spreadsheet_missing_workbook context=\"{session.context_understanding}\"",
                    to_terminal=False
                )
                message = "Excel needed, please upload excel"
                session.result = {"final_answer": message}
                session.state = "done"
                return {"type": "final", "result": session.result, "message": message}
        self._log(
            f"[INTERACT] case=4 proceed_with_workbook context=\"{(session.context_understanding or '').strip()}\"",
            to_terminal=False
        )
        session.current_workbooks = active_workbooks
        self._hydrate_sandbox_from_session(active_workbooks)

        session.state = "understanding"
        return {"type": "progress", "stage": "understanding"}

    def _handle_understanding(self, session: SheetHeroSession, current_request: str) -> Dict[str, Any]:
        active_workbooks = session.current_workbooks
        if not (session.spreadsheet_summary or "").strip():
            session.spreadsheet_summary = ExcelContextBuilder(
                self._get_context_paths(session),
                active_workbooks
            ).build(self.config.total_token_budget)

        append_ui_thought(session, "understanding", "running", "Understanding workbook/task context")
        session.understanding = self.understanding_module.run(
            current_request,
            session.spreadsheet_summary or "",
            session.context_understanding
        )
        append_ui_thought(session, "understanding", "done", session.understanding or "")
        session.state = "diagnosing"
        return {"type": "progress", "stage": "diagnosing"}

    def _handle_diagnosing(self, session: SheetHeroSession, current_request: str) -> Dict[str, Any]:
        if not self.config.enable_diagnose:
            append_ui_thought(
                session,
                "diagnosing",
                "skipped",
                "Diagnose disabled by configuration"
            )
            session.state = "executing"
            return {"type": "progress", "stage": "executing"}

        wb_view = self.sandbox.get_workbook_view()
        decision = self.diagnose_router.decide(
            user_question=current_request,
            understanding_output=session.understanding or "",
            workbook_view=wb_view
        )

        if decision.should_diagnose:
            append_ui_thought(session, "diagnosing", "running", "Diagnosing data quality risks")
            if decision.issues:
                question_list = decision.issues
            else:
                question_list = self.diagnose_module.run_readonly(
                    workbooks=wb_view,
                    user_task=current_request
                )
            append_ui_thought(session, "diagnosing", "done", question_list)

            self.qa_stage.reset()
            self.qa_stage.start(question_list=question_list, original_question=session.original_query)

            question_payload = self.qa_stage.next_question_payload()
            if question_payload:
                session.state = "qa"
                return {
                    "type": "clarification",
                    "stage": "qa",
                    "message": question_payload.get("message"),
                    "details_markdown": question_payload.get("details_markdown", ""),
                }

            self.qa_stage.finalize_decision()
            session.state = "cleaning"
            return {"type": "progress", "stage": "cleaning"}

        append_ui_thought(
            session,
            "diagnosing",
            "skipped",
            "No blocking data quality clarification required"
        )
        session.state = "executing"
        return {"type": "progress", "stage": "executing"}

    def _handle_qa(self, session: SheetHeroSession, user_input: str,
                   qa_stage: Optional[QualityAssuranceStage]) -> Dict[str, Any]:
        qa_stage = qa_stage or self.qa_stage
        if qa_stage is None:
            return {"type": "error", "message": "QA session not initialized."}

        qa_stage.consume_user_reply(user_input)
        if qa_stage.get_last_mismatch():
            followup_payload = qa_stage.next_question_payload()
            hint = qa_stage.get_last_mismatch()
            if followup_payload:
                return {
                    "type": "clarification",
                    "stage": "qa",
                    "message": f"{hint}\n\nPlease answer this question:\n{followup_payload.get('message', '')}",
                    "details_markdown": followup_payload.get("details_markdown", ""),
                }
            qa_stage.clear_last_mismatch()
        question_payload = qa_stage.next_question_payload()
        if question_payload:
            return {
                "type": "clarification",
                "stage": "qa",
                "message": question_payload.get("message"),
                "details_markdown": question_payload.get("details_markdown", ""),
            }

        qa_stage.finalize_decision()
        session.state = "cleaning"
        return {"type": "progress", "stage": "cleaning"}

    def _handle_cleaning(self, session: SheetHeroSession,
                         qa_stage: Optional[QualityAssuranceStage],
                         current_request: str) -> Dict[str, Any]:
        actions = qa_stage.export_cleaning_actions() if qa_stage else []
        session.qa_policy_notes = qa_stage.export_interpretation_policies() if qa_stage else []
        append_ui_thought(session, "cleaning", "running", actions)
        self.cleaning_module.apply(
            sandbox=self.sandbox,
            actions=actions
        )
        append_ui_thought(session, "cleaning", "done", actions)
        if qa_stage:
            qa_stage.clear_cleaning_actions()
        session.spreadsheet_summary = ExcelContextBuilder(
            self._get_context_paths(session),
            session.current_workbooks
        ).build(self.config.total_token_budget)
        session.state = "executing"
        return {"type": "progress", "stage": "executing"}

    def _handle_execution(self, session: SheetHeroSession,
                          current_request: str) -> Dict[str, Any]:
        understanding_output = session.understanding
        if not understanding_output:
            return {"type": "error", "message": "Understanding missing before execution."}
        user_query = current_request
        if not (session.spreadsheet_summary or "").strip():
            session.spreadsheet_summary = ExcelContextBuilder(
                self._get_context_paths(session),
                session.current_workbooks
            ).build(self.config.total_token_budget)
        execution_context = session.spreadsheet_summary or ""
        if session.qa_policy_notes:
            execution_context = (
                execution_context.rstrip()
                + "\n\n[User clarification policies]\n- "
                + "\n- ".join(str(item).strip() for item in session.qa_policy_notes if str(item).strip())
            ).strip()
        self.execution_module.runner.excel_context_execution = execution_context
        append_ui_thought(session, "executing", "running", "Generating and running execution code")
        session.execution_validation_cycles += 1
        execution_turn_budget = self._resolve_execution_turn_budget()
        max_cycles = self._resolve_max_cycles()
        self._log(
            f"Execution budgets: turns_per_iteration={execution_turn_budget}, max_cycles={max_cycles}",
            True
        )
        execution_result = self.execution_module.run(
            understanding_output=understanding_output,
            user_question=user_query,
            max_turns=execution_turn_budget
        )
        append_ui_thought(
            session,
            "executing",
            "done",
            {
                "success": execution_result.get("success"),
                "answer": execution_result.get("answer"),
                "total_turns": execution_result.get("total_turns"),
            },
        )
        session.pending_execution_result = execution_result
        session.pending_execution_context = execution_context
        session.state = "validation"
        return {"type": "progress", "stage": "validation"}

    def _handle_validation(self, session: SheetHeroSession,
                           current_request: str) -> Dict[str, Any]:
        execution_result = session.pending_execution_result
        execution_context = session.pending_execution_context
        if not execution_result or execution_context is None:
            return {"type": "error", "message": "Execution result missing before validation."}
        understanding_output = session.understanding or ""
        user_query = current_request

        append_ui_thought(session, "validation", "running", "Validating execution result")
        validation_result = self.validation_module.run(
            execution_result=execution_result,
            user_question=user_query,
            understanding_output=understanding_output
        )
        successful_execs = int(
            execution_result.get("execution_summary", {}).get("successful_executions", 0) or 0
        )
        if successful_execs <= 0:
            session.consecutive_no_success_execution_cycles += 1
        else:
            session.consecutive_no_success_execution_cycles = 0
        max_cycles = self._resolve_max_cycles()
        if (
            not validation_result.get("validation_passed")
            and validation_result.get("requires_reexecution", True)
            and session.execution_validation_cycles >= max_cycles
        ):
            issues = list(validation_result.get("issues_found") or [])
            issues.append(
                f"Reached max execute-validate cycles ({max_cycles}) in session mode."
            )
            validation_result["issues_found"] = issues
            validation_result["requires_reexecution"] = False
            feedback = (validation_result.get("improvement_feedback") or "").strip()
            if not feedback:
                validation_result["improvement_feedback"] = (
                    "Stopped after max execute-validate cycles. "
                    "Please inspect the latest output/log and refine prompt or data."
                )
        if (
            not validation_result.get("validation_passed")
            and validation_result.get("requires_reexecution", True)
            and session.consecutive_no_success_execution_cycles >= 2
        ):
            issues = list(validation_result.get("issues_found") or [])
            issues.append(
                "Stopped after repeated execute-validate cycles with zero successful code executions."
            )
            validation_result["issues_found"] = issues
            validation_result["requires_reexecution"] = False
            feedback = (validation_result.get("improvement_feedback") or "").strip()
            prefix = (
                "Execution never reached a successful sandbox run in two consecutive cycles. "
                "Refine execution prompt/guards instead of continuing validation loops."
            )
            validation_result["improvement_feedback"] = (
                f"{prefix}\n{feedback}" if feedback else prefix
            )
        append_ui_thought(
            session,
            "validation",
            "done",
            {
                "validation_passed": validation_result.get("validation_passed"),
                "confidence_score": validation_result.get("confidence_score"),
                "requires_reexecution": validation_result.get("requires_reexecution"),
                "issues_found": validation_result.get("issues_found", []),
            },
        )

        final_answer = validation_result.get(
            "verified_answer",
            execution_result.get("answer", "")
        )
        short_answer = self.final_response_module.run(
            user_question=current_request,
            final_answer=final_answer,
            validation_result=validation_result,
            execution_result=execution_result,
        )

        session.result = {
            "execution_result": execution_result,
            "validation_result": validation_result,
            "final_answer": final_answer,
            "short_answer": short_answer,
            "edited_existing_file": self._output_existed_before,
        }

        extracted_context = ContextExtractor.extract_workbook_purpose_domain(
            session.understanding or ""
        ).strip()
        if extracted_context:
            session.context_understanding = extracted_context

        if validation_result.get("validation_passed"):
            session.state = "done"
        else:
            if not validation_result.get("requires_reexecution", True):
                session.state = "done"
            else:
                session.understanding = self.understanding_module.enhance(
                    session.understanding,
                    validation_result,
                    user_question=current_request,
                )
                session.state = "executing"
                return {"type": "progress", "stage": "executing"}
        session.pending_execution_result = None
        session.pending_execution_context = None
        self.qa_stage.reset()
        log_final_report(self.progress_logger, session)
        return {"type": "final", "result": session.result}

    def _get_context_paths(self, session: SheetHeroSession) -> List[str]:
        active_workbooks = session.current_workbooks
        if active_workbooks:
            return list(active_workbooks.keys())
        if self.excel_paths:
            return self.excel_paths
        return []

    def _hydrate_sandbox_from_session(self, workbooks: Dict[str, Any]) -> None:
        if not workbooks:
            return
        primary_path = next(iter(workbooks.keys()))
        self.sandbox.workbooks = workbooks
        self.sandbox.world = SpreadsheetWorld(
            workbooks=workbooks,
            output_path=self.sandbox.output_path,
            primary_path=primary_path
        )
        self.sandbox.code_globals.update({
            "openpyxl": openpyxl,
            "workbooks": workbooks,
            "sheet_names": self.sandbox.world.primary_workbook.sheetnames,
            "range_boundaries": range_boundaries,
            "get_column_letter": get_column_letter,
            "column_index_from_string": column_index_from_string,
            "pandas": pd,
            "pd": pd,
            "numpy": np,
            "np": np,
        })
        self.sandbox.load_namespace("spreadsheet")

    _build_clarification_message = staticmethod(build_clarification_message)
