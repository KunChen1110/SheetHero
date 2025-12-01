import os
import time
import csv
from typing import Dict, Any, Optional, Union, List
from openai import OpenAI
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from config.settings import Config
from modules.understanding import UnderstandingModule
from modules.execution import ExecutionModule
from modules.validation import ValidationModule
from utils.excel_toolkit import ExcelToolkit, calculate_token_cost_line
from utils.logger import setup_logger

logger = setup_logger(__name__)


def build_output_preferences(mode: str = "text",
                             file_path: Optional[str] = None) -> Dict[str, Optional[str]]:
    """
    Normalize user preferences for how results should be delivered.

    Args:
        mode: "text" for inline markdown answers, "file" to save results.
        file_path: Optional explicit path for file output.

    Returns:
        Dict with validated mode and file_path.
    """

    # Normalize input: strip whitespace, lowercase, default to "text"
    normalized_mode = (mode or "text").strip().lower()
    if normalized_mode not in {"text", "file"}:
        raise ValueError("output mode must be 'text' or 'file'")

    #For file mode, generate default path if none provided; text mode needs no path
    if normalized_mode == "file":
        normalized_path = file_path or os.path.join(
            os.getcwd(), "sheethero_output.xlsx"
        )
    else:
        normalized_path = None

    # Return absolute path to prevent issues with relative paths
    return {"mode": normalized_mode, "file_path": os.path.abspath(normalized_path) if normalized_path else None}


# Convenience alias for the same functionality
def output_mode(mode: str = "text", file_path: Optional[str] = None) -> Dict[str, Optional[str]]:
    return build_output_preferences(mode, file_path)


"""                                                                                                                                                                                                 
config (Config)
output_preferences (dict)
output_instruction (str)

excel_paths (list)

excel_conntext_understanding (str)
excel_conntext_execution (str)

self.client (OpenAI)
code_globals (dict)
code_locals (dict)
workbooks

understanding_module
execution_module
validation_module
"""
class SheetHero:
    def __init__(self, excel_paths: Union[str, List[str]],
                 config: Config,
                 load_excel: bool = True):

        # Store configuration and normalize Excel paths to a list
        self.config = config
        self.excel_paths = excel_paths if isinstance(excel_paths, list) else [excel_paths]
        self.output_preferences = build_output_preferences(
            mode=self.config.output_mode,
            file_path=self.config.output_file
        )
        
        # === Progress Log File Setup (always enabled) ===
        # Create a timestamped markdown log file to track each session's progress

        from datetime import datetime
        from pathlib import Path
        
        backend_dir = Path(__file__).parent.parent
        log_dir = backend_dir / "loggers"
        log_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if self.excel_paths:
            first_file = os.path.splitext(os.path.basename(self.excel_paths[0]))[0]
            session_id = f"sheethero_{first_file}"
        else:
            session_id = "sheethero"
        
        self._progress_log_path = log_dir / f"{session_id}_{timestamp}.md"
        self._progress_log_file = open(self._progress_log_path, 'w', encoding='utf-8')
        # Write header
        self._progress_log_file.write(f"# SheetHero Verbose Log\n\n")
        self._progress_log_file.write(f"**Session started:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        self._progress_log_file.write("---\n\n")
        self._progress_log_file.flush()


        # Compute output_path BEFORE calling _build_output_instruction (AI instructions)
        if self.output_preferences.get("mode") == "file":
            self._output_path = self.output_preferences.get("file_path")
        else:
            # Default output path based on first input file
            first_input = self.excel_paths[0] if self.excel_paths else "output"
            dir_path = os.path.dirname(first_input)
            base_name = os.path.splitext(os.path.basename(first_input))[0]
            self._output_path = os.path.join(dir_path, f"{base_name}_output.xlsx")
        
        self.output_instruction = self._build_output_instruction()

        # === OpenAI Client Setup ===
        if not self.config.api_key:
            raise ValueError("OpenAI API key is missing in Config")
        client_kwargs = {"api_key": self.config.api_key}
        if self.config.base_url:
            client_kwargs["base_url"] = self.config.base_url
        self.client = OpenAI(**client_kwargs)

        # === Code Execution Environment Setup ===
        # Prepare isolated environment for AI-generated code with essential libraries
        self.code_globals = {
            'math': __import__('math'),      # Mathematical functions (sin, cos, sqrt, etc.)
            'json': __import__('json'),      # JSON parsing/formatting
            're': __import__('re'),          # Regular expressions for text processing
            'os': __import__('os'),          # Operating system interface
            'sys': __import__('sys'),        # System parameters
            'excel_paths': self.excel_paths, # List of all file paths
            'output_preferences': self.output_preferences,  # Output requirements
            'output_path': self._output_path,  # Target save location
        }
        self.code_locals = {}  # Empty dict for code execution (stores variables created by AI code)

        # === Excel File Loading and Context Generation ===
        # Load workbooks and create AI context summaries respecting token limits
        if load_excel:
            # Load the Excel file(s) and libraries (openpyxl, pandas, etc.)
            self._setup_excel_libraries()
            self.workbooks = self.code_globals.get('workbooks', {})

            # Generate markdown summary of Excel contents for AI context
            # Larger context for understanding phase, smaller for execution
            self.excel_context_understanding = self._generate_sheets_markdown_summary(self.config.total_token_budget * 2)
            self.excel_context_execution = self._generate_sheets_markdown_summary(self.config.total_token_budget)
        else:
            # Headless mode: don't load Excel, work with provided context only
            self.workbooks = {}
            self.excel_context_understanding = "Excel file not loaded. Working with provided context only."
            self.excel_context_execution = "Excel file not loaded. Working with provided context only."

        # === Module Initialization ===
        self.understanding_module = UnderstandingModule(
            self.client,
            self.config.deployment,
            self.excel_context_understanding
        )
        self.execution_module = ExecutionModule(
            self.client,
            self.config.deployment,
            self.code_globals,
            self.code_locals,
            self.excel_context_execution,
            self.output_instruction,
            progress_log_file=None  # Will be set after initialization
        )
        self.validation_module = ValidationModule(
            self.client,
            self.config.deployment,
            self.excel_context_understanding,
            progress_log_file=self._progress_log_file
        )

        # Connect progress logger to execution module
        self.execution_module.progress_log_file = self._progress_log_file

    def _log_to_file(self, message: str):
        """Write message to progress log file."""
        if self._progress_log_file:
            self._progress_log_file.write(message + "\n")
            self._progress_log_file.flush()
    
    def _log_progress(self, message: str, to_terminal: bool = False):
        """ Log progress message to file (always) and optionally to terminal.  """

        # Always write to file with markdown formatting
        formatted = self._format_progress_message(message)
        self._log_to_file(formatted)
        
        # Optionally print to terminal (only for key milestones) - currently disabled
        # if to_terminal:
        #     print(message)
    
    def _format_progress_message(self, message: str) -> str:
        """
        Format progress message for markdown output.
        Converts plain text markers into structured Markdown format.
        """
        cleaned = message.strip()

        # Format ASCII separators as markdown horizontal rules
        if cleaned.startswith("=" * 40) or cleaned.startswith("=" * 60) or cleaned.startswith("=" * 80):
            return "\n---\n"

        # Format ASCII separators as bold headings
        elif cleaned.startswith("-" * 40) or cleaned.startswith("-" * 60):
            return "\n**" + cleaned.replace("-", "").strip() + "**\n"

        # Format special markers as markdown headings
        if "[FINAL SUMMARY]" in cleaned:
            return f"\n## {cleaned}\n"
        elif "[STAGE" in cleaned or "[ITERATION" in cleaned:
            return f"\n### {cleaned}\n"
        elif any(prefix in cleaned for prefix in ["[SheetHero]", "[SUCCESS", "[STOPPING", "[CONTINUE", "[MAX ITERATIONS", "[Excel]"]):
            return f"**{cleaned}**"
        
        # Format emoji messages
        if cleaned.startswith("❌") or cleaned.startswith("⚠️"):
            return f"⚠️ {cleaned.lstrip('❌⚠️').strip()}"
        elif cleaned.startswith("✅"):
            return f"✅ {cleaned.lstrip('✅').strip()}"
        
        return cleaned

    def run(self, user_question: str) -> Dict[str, Any]:
        """
        Execute three-stage iterative analysis: Understanding → Execution → Validation.

        Continues iterating until validation passes or maximum attempts are exhausted.
        """

        # Use config defaults if parameters not provided (allows CLI overrides)
        max_turns = self.config.max_turns

        # Log and print that we're starting analysis
        logger.info("Starting iterative three-stage analysis")
        self._log_progress("🚀 [SheetHero] Starting iterative three-stage analysis...", to_terminal=False)
        self._log_progress("="*80, to_terminal=False)

        # === Execution Tracking ===
        overall_start_time = time.time()
        all_execution_results = []    # Store results from each execution stage
        all_validation_results = []   # Store results from each validation stage

        try:
            # ===== STAGE 1: UNDERSTANDING =====
            # Analyze the question and Excel context to develop analysis strategy
            logger.info("Running understanding module")
            self._log_progress("📖 [STAGE 1] UNDERSTANDING MODULE", to_terminal=False)
            self._log_progress("-" * 40, to_terminal=False)

            understanding_start_time = time.time()
            # Call the UnderstandingModule to analyze the question + Excel context
            understanding_output = self.understanding_module.analyze(user_question)
            understanding_duration = time.time() - understanding_start_time

            self._log_progress(f"✅ [STAGE 1] Understanding completed in {understanding_duration:.2f}s", to_terminal=False)
            # Log full understanding output to file
            self._log_to_file(f"\n**Understanding Analysis:**\n```\n{understanding_output}\n```\n")

            # ===== ITERATIVE EXECUTE-VALIDATE LOOP =====
            for iteration in range(max_turns):
                logger.info(f"Starting iteration {iteration + 1}/{max_turns}")
                self._log_progress(f"\n🔄 [ITERATION {iteration + 1}/{max_turns}] EXECUTE-VALIDATE CYCLE", to_terminal=False)
                self._log_progress("="*60, to_terminal=False)

                # ===== STAGE 2: EXECUTION =====
                self._log_progress(f"💻 [ITERATION {iteration + 1}] EXECUTION MODULE", to_terminal=False)
                self._log_progress("-" * 40, to_terminal=False)
                execution_start_time = time.time()

                # If this isn't the first iteration, add validation feedback from previous attempt
                if iteration > 0 and all_validation_results:
                    last_validation = all_validation_results[-1]
                    if last_validation.get('improvement_feedback'):
                        # Build enhanced prompt with previous feedback
                        enhanced_understanding = f"""{understanding_output}

**IMPROVEMENT FEEDBACK FROM PREVIOUS ITERATION:**
{last_validation['improvement_feedback']}

**ISSUES TO ADDRESS:**
{'; '.join(last_validation.get('issues_found', []))}

Please address these specific points in your new analysis approach."""
                    else:
                        enhanced_understanding = understanding_output
                else:
                    enhanced_understanding = understanding_output

                # Generate and execute analysis code based on understanding
                execution_result = self.execution_module.run(enhanced_understanding, user_question)
                execution_duration = time.time() - execution_start_time
                all_execution_results.append(execution_result)

                # Log execution metrics
                status_emoji = "✅" if execution_result["success"] else "❌"
                self._log_progress(f"{status_emoji} [ITERATION {iteration + 1}] Execution completed in {execution_duration:.2f}s", to_terminal=False)
                self._log_progress(f"🔄 [ITERATION {iteration + 1}] Total turns: {execution_result['total_turns']}", to_terminal=False)
                self._log_progress(f"📊 [ITERATION {iteration + 1}] Code executions: {execution_result.get('execution_summary', {}).get('total_code_executions', 0)}", to_terminal=False)

                # ===== STAGE 3: VALIDATION =====
                # Validate execution results against requirements
                logger.info(f"Running validation module for iteration {iteration + 1}")
                self._log_progress(f"\n🔍 [ITERATION {iteration + 1}] VALIDATION MODULE", to_terminal=False)
                self._log_progress("-" * 40, to_terminal=False)
                validation_start_time = time.time()

                # Run the ValidationModule to check if the answer is correct
                validation_result = self.validation_module.reflect(execution_result, user_question, understanding_output)
                validation_duration = time.time() - validation_start_time
                all_validation_results.append(validation_result)

                # Log validation results
                validation_emoji = "✅" if validation_result["validation_passed"] else "⚠️"
                self._log_progress(f"{validation_emoji} [ITERATION {iteration + 1}] Validation completed in {validation_duration:.2f}s", to_terminal=False)
                self._log_progress(f"🎯 [ITERATION {iteration + 1}] Confidence: {validation_result['confidence_score']:.2f}", to_terminal=False)
                self._log_progress(f"📋 [ITERATION {iteration + 1}] Validation: {'PASSED' if validation_result['validation_passed'] else 'FAILED'}", to_terminal=False)

                # === Loop Termination Logic ===
                # Decide whether to stop iterating or try again
                if validation_result['validation_passed']:
                    # Success! Answer is good enough
                    logger.info(f"Validation passed on iteration {iteration + 1}")
                    self._log_progress(f"🎉 [SUCCESS] Validation passed on iteration {iteration + 1}!", to_terminal=False)
                    final_answer = validation_result.get('verified_answer', execution_result['answer'])
                    overall_success = True
                    confidence_score = validation_result['confidence_score']
                    validation_passed = True
                    break  # Exit the loop

                elif not validation_result.get('requires_reexecution', True):
                    # Validation says retry would not improve results
                    logger.warning("Validation indicates no further improvement possible")
                    self._log_progress("🛑 [STOPPING] Validation indicates no further improvement possible", to_terminal=False)
                    final_answer = execution_result['answer']
                    overall_success = False
                    confidence_score = validation_result['confidence_score']
                    validation_passed = False
                    break  # Exit the loop

                else:
                    # Issues identified - prepare for next iteration
                    logger.info(f"Issues found, preparing for iteration {iteration + 2}")
                    self._log_progress(f"🔄 [CONTINUE] Issues found, preparing for iteration {iteration + 2}", to_terminal=False)
                    
                    # Log issues and feedback for next iteration
                    if validation_result.get('issues_found'):
                        self._log_to_file(f"\n**Issues Found:**\n")
                        for issue in validation_result['issues_found']:
                            self._log_to_file(f"- {issue}\n")
                    if validation_result.get('improvement_feedback'):
                        self._log_to_file(f"\n**Improvement Feedback:**\n```\n{validation_result['improvement_feedback']}\n```\n")

                    # Check maximum iteration limit
                    if iteration == max_turns - 1:
                        logger.warning("Reached maximum iterations without validation")
                        self._log_progress("⚠️ [MAX ITERATIONS] Reached maximum iterations without validation", to_terminal=False)
                        final_answer = execution_result['answer']
                        overall_success = False
                        confidence_score = validation_result['confidence_score']
                        validation_passed = False

            # === Final Report Generation ===
            # Extract issues and feedback from final validation attempt
            if all_validation_results:
                final_validation = all_validation_results[-1]
                issues_found = final_validation.get('issues_found', [])
                improvement_feedback = final_validation.get('improvement_feedback', '')
            else:
                issues_found = []
                improvement_feedback = ''

            # Collect conversation histories for debugging
            all_conversation_histories = []
            for exec_result in all_execution_results:
                conv_history = exec_result.get('conversation_history', [])
                if conv_history:
                    all_conversation_histories.append({
                        'iteration': all_execution_results.index(exec_result) + 1,
                        'conversation_history': conv_history
                    })

            # ===== FINAL SUMMARY =====
            total_duration = time.time() - overall_start_time
            total_iterations = len(all_execution_results)

            logger.info(f"Analysis completed. Success: {overall_success}, Iterations: {total_iterations}")
            self._log_progress("\n" + "="*80, to_terminal=False)
            self._log_progress("🎯 [FINAL SUMMARY]", to_terminal=False)
            self._log_progress("="*80, to_terminal=False)
            self._log_progress(f"Overall Success: {'✅ YES' if overall_success else '❌ NO'}", to_terminal=False)
            self._log_progress(f"Total Iterations: {total_iterations}", to_terminal=False)
            self._log_progress(f"Final Answer: {final_answer}", to_terminal=False)
            self._log_progress(f"Confidence Score: {confidence_score:.2f}/1.0", to_terminal=False)
            self._log_progress(f"Validation Passed: {'✅ YES' if validation_passed else '❌ NO'}", to_terminal=False)
            self._log_progress(f"Total Duration: {total_duration:.2f}s", to_terminal=False)
            self._log_progress("="*80, to_terminal=False)
            
            # Close progress log file
            if self._progress_log_file:
                from datetime import datetime
                self._progress_log_file.write(f"\n---\n\n")
                self._progress_log_file.write(f"**Session ended:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                self._progress_log_file.close()
                progress_log_path = str(self._progress_log_path)
                self._progress_log_file = None
            else:
                progress_log_path = None

            # Return comprehensive results
            result = {
                "success": overall_success,
                "answer": final_answer,
                "confidence_score": confidence_score,
                "validation_passed": validation_passed,
                "total_iterations": total_iterations,
                "all_execution_results": all_execution_results,      # Full history
                "all_validation_results": all_validation_results,    # Full history
                "conversation_history": all_conversation_histories,  # For debugging
                "issues_found": issues_found,
                "improvement_feedback": improvement_feedback,
                "total_duration": total_duration,
                "user_question": user_question,
                "understanding_output": understanding_output
            }
            
            # Add log path to result
            if progress_log_path:
                result["verbose_log_path"] = progress_log_path
            
            return result

        except Exception as e:
            # === Error Handling ===
            # If anything goes wrong, log it and return error info
            error_duration = time.time() - overall_start_time
            logger.error(f"Critical error: {str(e)}")
            self._log_progress(f"❌ [SheetHero] Critical error: {str(e)}", to_terminal=False)
            self._log_progress(f"⏱️ [SheetHero] Failed after {error_duration:.2f}s", to_terminal=False)
            
            # Close log file on error
            progress_log_path = None
            if self._progress_log_file:
                from datetime import datetime
                self._progress_log_file.write(f"\n---\n\n")
                self._progress_log_file.write(f"**Session ended:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                self._progress_log_file.close()
                progress_log_path = str(self._progress_log_path)
                self._progress_log_file = None

            # Collect what conversation history we have
            all_conversation_histories = []
            for exec_result in all_execution_results:
                conv_history = exec_result.get('conversation_history', [])
                if conv_history:
                    all_conversation_histories.append({
                        'iteration': all_execution_results.index(exec_result) + 1,
                        'conversation_history': conv_history
                    })

            # Return error results (success=False, answer=error message)
            result = {
                "success": False,
                "answer": f"Analysis failed due to error: {str(e)}",
                "confidence_score": 0.0,
                "validation_passed": False,
                "total_iterations": len(all_execution_results),
                "all_execution_results": all_execution_results,
                "all_validation_results": all_validation_results,
                "conversation_history": all_conversation_histories,
                "issues_found": [f"Critical error: {str(e)}"],
                "improvement_feedback": "Review the error and try again",
                "total_duration": error_duration,
                "user_question": user_question
            }
            
            # Add log path to result
            if progress_log_path:
                result["verbose_log_path"] = progress_log_path
            
            return result
    def _generate_sheets_markdown_summary(self, total_token_budget: int = 50000) -> str:
        """
        Generate a markdown summary of Excel sheets for AI context, with token budgeting.

        Creates a structured description of workbooks including file structure,
        sheet dimensions, and data previews that respect the model's context limits.
        """
        try:
            # Get all workbooks from code_globals (set up in _setup_excel_libraries(execution environment))
            if hasattr(self, 'code_globals') and 'workbooks' in self.code_globals:
                workbooks = self.code_globals['workbooks']
            else:
                # Fallback: use self.workbooks if available
                workbooks = self.workbooks if hasattr(self, 'workbooks') else {}
            
            overview_parts = []
            
            # Build file overview header
            if len(workbooks) > 1:
                overview_parts.append(f"📊 **Multiple Excel Files Overview ({len(workbooks)} files)**\n")
            else:
                first_path = self.excel_paths[0] if self.excel_paths else "unknown"
                overview_parts.append(f"📊 **Excel File Overview: {os.path.basename(first_path)}**\n")
            
            # Allocate token budget equally across files
            available_tokens = total_token_budget
            tokens_per_file = available_tokens // len(workbooks) if workbooks else 0
            
            # Process each workbook
            for excel_path, workbook in workbooks.items():
                file_parts = []
                
                # File header
                file_parts.append(f"\n{'='*60}")
                file_parts.append(f"📁 **File: {os.path.basename(excel_path)}**")
                file_parts.append(f"**Full Path:** {excel_path}")
                file_parts.append(f"**Total Sheets:** {len(workbook.sheetnames)}\n")
                
                # Further divide budget among sheets in this file
                tokens_per_sheet = tokens_per_file // len(workbook.sheetnames) if workbook.sheetnames else 0
                
                # Process each sheet in this workbook
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_parts = []
                    
                    # Sheet header
                    sheet_parts.append(f"\n**📄 Sheet: '{sheet_name}'** (in {os.path.basename(excel_path)})")
                    sheet_parts.append(f"- Dimensions: {sheet.max_row} rows × {sheet.max_column} columns")
                    
                    if tokens_per_sheet > 0:
                        # Generate data preview within token budget
                        preview_result = self._get_sheet_preview_with_token_limit(
                            sheet,
                            tokens_per_sheet,
                            max_rows=min(sheet.max_row, 10000),  # Safety cap for performance
                            max_cols=min(sheet.max_column, 1000)   # Safety cap for performance
                        )
                        
                        sheet_parts.append(f"- Data Preview ({preview_result['rows_shown']} of {sheet.max_row} rows, "
                                           f"{preview_result['cols_shown']} of {sheet.max_column} columns):")
                        
                        # Warn if data was truncated due to token limits
                        if preview_result['is_truncated']:
                            sheet_parts.append("  ⚠️ Preview truncated to fit token budget")
                        
                        # Add actual data in Markdown table format
                        sheet_parts.append("  Data:")
                        markdown_rows = []
                        for row_data in preview_result['formatted_data']:
                            # Join cells with "|" and add borders for Markdown table format
                            markdown_rows.append(f"| {' | '.join(row_data)} |")
                        
                        # Join all rows with newlines (using \n for compactness)
                        if markdown_rows:
                            sheet_parts.append("  " + "\\n".join(markdown_rows))
                        
                        # Add summary stats if not all rows shown
                        if preview_result['rows_shown'] < sheet.max_row:
                            sheet_parts.append(f"\n  📊 Sheet Summary:")
                            sheet_parts.append(f"  - Total rows: {sheet.max_row}")
                            sheet_parts.append(f"  - Total columns: {sheet.max_column}")
                            sheet_parts.append(f"  - Rows shown in preview: {preview_result['rows_shown']}")
                    
                    file_parts.extend(sheet_parts)
                
                overview_parts.extend(file_parts)
            
            # Combine all parts into complete overview
            final_overview = "\n".join(overview_parts)
            return final_overview

        except Exception as e:
            # If context generation fails, return error but don't crash whole analysis
            logger.error(f"Error generating Excel overview: {str(e)}")
            return f"❌ Error generating Excel overview: {str(e)}"

    def _build_output_instruction(self) -> str:
        """ Generate AI instructions for output format based on user preferences  """

        mode = self.output_preferences.get("mode", "text")
        if mode == "file":
            return f"""**OUTPUT REQUIREMENTS:**
1. Save final results to: `output_path` (variable available in code: "{self._output_path}")
2. Use the UNIFIED OUTPUT WORKFLOW:
   - Convert DataFrame to 2D list: `[df.columns.tolist()] + df.values.tolist()`
   - Create output sheet: `create_output_sheet("Output")`
   - Write data: `write_dataframe_to_sheet(data_2d, "Output", "A1")`
   - Add summary if needed: `add_summary_row("Output", row_num, {{"Total": val, "Average": avg}})`
   - Highlight important rows: `highlight_rows("Output", [row_nums], {{"fill_color": "red"}})`
   - Save: `save_workbook_to(output_path)`
3. Return the saved file path in Final Answer
4. DO NOT use DataFrame.to_excel() or pd.ExcelWriter()"""

        return (
            "Final results must be presented directly in the final answer as a clean markdown "
            "table or list. Do not save any files unless the user explicitly asks."
        )

    def _load_workbook_from_path(self, excel_path: str):
        """ Load a spreadsheet file (Excel or CSV) and return an openpyxl Workbook. """
        _, ext = os.path.splitext(excel_path)
        ext = ext.lower()
        excel_extensions = {".xlsx", ".xlsm", ".xltx", ".xltm"}

        # Load Excel files directly
        if ext in excel_extensions:
            return load_workbook(excel_path, data_only=True)

        # Convert CSV files to Workbook format
        if ext == ".csv":
            return self._create_workbook_from_csv(excel_path)

        raise ValueError(
            f"Unsupported file extension '{ext}' for {excel_path}. "
            "Supported formats: .xlsx, .xlsm, .xltx, .xltm, .csv"
        )

    def _create_workbook_from_csv(self, csv_path: str) -> Workbook:
        """ Convert a CSV file into an openpyxl Workbook for unified processing. """

        workbook = Workbook()
        sheet = workbook.active

        # Sheet titles in Excel are limited to 31 characters
        sheet_name = os.path.splitext(os.path.basename(csv_path))[0][:31] or "Sheet1"
        sheet.title = sheet_name

        # Read CSV with UTF-8 BOM handling for Excel-generated files
        with open(csv_path, newline="", encoding="utf-8-sig") as csv_file:
            reader = csv.reader(csv_file)
            for row in reader:
                # Convert each cell value to appropriate type (int/float/str)
                processed_row = [self._infer_cell_value(value) for value in row]
                sheet.append(processed_row)

        return workbook

    @staticmethod
    def _infer_cell_value(value: str):
        """
        Convert CSV string values to numeric types when possible.
        Tries int first (no decimal point), then float, keeps as string if both fail.
        """
        if value is None:
            return value

        stripped = value.strip()
        if stripped == "":
            return ""

        # Attempt numeric conversion for proper Excel handling
        try:
            if '.' not in stripped:
                return int(stripped)
            return float(stripped)
        except ValueError:
            return value

    def _get_sheet_preview_with_token_limit(self, sheet, token_budget: int,
                                            max_rows: int = 10000, max_cols: int = 1000) -> Dict[str, Any]:
        """
        Generate a sheet preview that fits within the AI's token budget.

        Iterates through rows, counting tokens until budget is reached. Always includes
        at least 5 rows for minimal useful data. Escapes markdown-breaking characters.

        Token counting is approximate but sufficient for budget management.
        """
        preview_data = []          # Raw cell values (for modules to use)
        formatted_data = []        # Formatted strings with A1 references for AI
        tokens_used = 0
        rows_shown = 0

        start_row = 1  # Start from first row

        # Calculate actual limits (respect sheet boundaries and safety caps)
        max_data_rows = min(max_rows, sheet.max_row)
        max_data_cols = min(max_cols, sheet.max_column)

        # Iterate through rows until we hit token budget or row limit
        for row_idx in range(start_row, max_data_rows + 1):
            row_cells = []           # Raw values for this row
            formatted_row_cells = []  # Formatted strings for this row

            # Process each cell in the row
            for col_idx in range(1, max_data_cols + 1):
                # Create cell reference like "A1", "B1", etc.
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                cell = sheet[cell_ref]
                cell_value = cell.value

                # Convert value to string for display (handle None/empty cells)
                display_value = str(cell_value) if cell_value is not None else ""

                # Escape characters that break Markdown tables
                # | is the column separator in Markdown
                # \n and \r would create new lines
                display_value = display_value.replace("|", "\\|").replace("\n", " ").replace("\r", " ")

                # Create formatted version: "A1:value"
                formatted_cell = f"{cell_ref}:{display_value}"

                row_cells.append(cell_value)          # Store raw value
                formatted_row_cells.append(formatted_cell)  # Store formatted string

            # Estimate tokens for this row
            row_str = " | ".join(formatted_row_cells)
            row_tokens = calculate_token_cost_line(row_str)

            # Check if adding this row would exceed budget
            if tokens_used + row_tokens > token_budget:
                # Always include at least 5 rows for minimal useful data
                if rows_shown < 5:
                    preview_data.append(row_cells)
                    formatted_data.append(formatted_row_cells)
                    rows_shown += 1
                    tokens_used += row_tokens
                # Stop after hitting budget
                break

            # Add the row to our preview
            preview_data.append(row_cells)
            formatted_data.append(formatted_row_cells)
            rows_shown += 1
            tokens_used += row_tokens

        return {
            'data': preview_data,
            'formatted_data': formatted_data,
            'rows_shown': rows_shown,
            'cols_shown': max_data_cols,
            'start_row': start_row,
            'is_truncated': rows_shown < max_data_rows,  # True if we didn't show all rows
            'tokens_used': tokens_used
        }

    def _setup_excel_libraries(self):
        """
        Load Excel libraries and prepare the code execution environment.

        Imports required libraries, loads workbooks, and creates helper functions
        that AI-generated code can use. Uses non-GUI matplotlib backend for server
        environments. Sets up multi-workbook support for complex analyses.
        """
        try:
            # Import required libraries
            import openpyxl
            from openpyxl.utils import range_boundaries, get_column_letter, column_index_from_string
            import pandas as pd
            import numpy as np
            import matplotlib
            matplotlib.use('Agg')  # Use non-interactive backend (no GUI needed)

            # Load all Excel files
            start_time = time.time()
            workbooks = {}  # Map file paths to workbook objects
            all_sheet_names = []  # List of all sheet names across all files
            
            for excel_path in self.excel_paths:
                logger.info(f"Loading spreadsheet file: {excel_path}")
                workbook = self._load_workbook_from_path(excel_path)
                workbooks[excel_path] = workbook
                all_sheet_names.extend([(excel_path, sheet_name) for sheet_name in workbook.sheetnames])
            
            load_time = time.time() - start_time
            logger.info(f"All spreadsheet files loaded in {load_time:.2f}s")
            self._log_progress(f"📊 [Excel] Loaded {len(self.excel_paths)} file(s) in {load_time:.2f}s")
            
            # Use the first workbook as the primary one for ExcelToolkit
            primary_workbook = workbooks[self.excel_paths[0]]
            primary_path = self.excel_paths[0]
            
            # Add libraries and workbook references to execution environment
            self.code_globals.update({
                'openpyxl': openpyxl,                    # Excel library
                'workbooks': workbooks,                  # Dictionary of all workbooks
                'sheet_names': primary_workbook.sheetnames,  # List of sheet names from primary workbook
                'range_boundaries': range_boundaries,    # Convert "A1:B2" to coordinates
                'get_column_letter': get_column_letter,  # Convert 1 -> "A"
                'column_index_from_string': column_index_from_string,  # Convert "A" -> 1
                'pandas': pd,                            # Data analysis library
                'pd': pd,                                # Short alias
                'numpy': np,                             # Numerical computing library
                'np': np,                                # Short alias
            })

            # Create ExcelToolkit with helper functions for common operations
            # For multi-file support, we'll use the primary workbook but provide access to all
            self.mcp_toolkit = ExcelToolkit(primary_workbook, primary_path, self._output_path)
            excel_helpers = self.mcp_toolkit.get_helper_functions_dict()
            
            # Add multi-workbook helper functions for cross-file analysis
            def get_workbook(file_path: str):
                """ Retrieve workbook by full path or filename."""
                if file_path in workbooks:
                    return workbooks[file_path]
                # Try to find by filename
                for path, wb in workbooks.items():
                    if os.path.basename(path) == os.path.basename(file_path):
                        return wb
                raise ValueError(f"Workbook not found: {file_path}. Available: {list(workbooks.keys())}")
            
            def list_all_workbooks():
                """List all loaded workbook file paths."""
                return list(workbooks.keys())
            
            def get_sheet_from_workbook(file_path: str, sheet_name: str):
                """Get a sheet from a specific workbook."""
                wb = get_workbook(file_path)
                if sheet_name in wb.sheetnames:
                    return wb[sheet_name]
                raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}. Available: {wb.sheetnames}")
            
            def inspector_multi(file_path: str, range_ref: str, sheet_name: Optional[str] = None):
                """ Read a range from a specific workbook by file path. """
                wb = get_workbook(file_path)
                if sheet_name is None:
                    sheet = wb.active
                else:
                    if sheet_name not in wb.sheetnames:
                        raise ValueError(f"Sheet '{sheet_name}' not found in {os.path.basename(file_path)}. Available sheets: {wb.sheetnames}")
                    sheet = wb[sheet_name]
                cell_range = sheet[range_ref]
                if hasattr(cell_range, 'value'):
                    return [[cell_range.value]]
                result = []
                for row in cell_range:
                    row_values = [cell.value for cell in row]
                    result.append(row_values)
                return result
            
            # Add multi-workbook helpers to the execution environment
            excel_helpers.update({
                'get_workbook': get_workbook,
                'list_all_workbooks': list_all_workbooks,
                'get_sheet_from_workbook': get_sheet_from_workbook,
                'inspector_multi': inspector_multi,
            })
            
            self.code_globals.update(excel_helpers)  # Add helpers to sandbox

            logger.info("Excel libraries loaded successfully")
            logger.info(f"Loaded {len(workbooks)} workbook(s)")
            for path, wb in workbooks.items():
                logger.info(f"  - {os.path.basename(path)}: {len(wb.sheetnames)} sheet(s) - {wb.sheetnames}")
            self._log_progress("📦 [SheetHero] Excel libraries loaded successfully", to_terminal=False)
            self._log_progress(f"📊 [SheetHero] Loaded {len(workbooks)} workbook(s):", to_terminal=False)
            for path, wb in workbooks.items():
                self._log_to_file(f"  📄 {os.path.basename(path)}: {len(wb.sheetnames)} sheet(s) - {wb.sheetnames}\n")

        except ImportError as e:
            # Handle missing libraries (e.g., user didn't install openpyxl)
            logger.error(f"Failed to import required libraries: {e}")
            self._log_progress(f"❌ [SheetHero] Failed to import required libraries: {e}", to_terminal=False)
            raise  # Re-raise to stop execution

        except Exception as e:
            # Problem loading the Excel file (corrupted, wrong format, etc.)
            logger.error(f"Failed to load Excel file: {e}")
            self._log_progress(f"❌ [SheetHero] Failed to load Excel file: {e}", to_terminal=False)
            raise  # Re-raise to stop execution