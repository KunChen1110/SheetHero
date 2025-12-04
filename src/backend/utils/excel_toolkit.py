"""
Excel utility library and toolkit for SheetHero AI agent.

Provides a safe, sandboxed interface for AI-generated code to:
- Read, search, and analyze Excel data
- Create visualizations and charts
- Modify workbooks (edit, format, insert formulas)
- Manage output files separately from inputs

Design Philosophy:
- Safe: Validates inputs, clear error messages
- Verbose: Logs actions for debugging AI-generated code
- Convenient: Common patterns in single function calls
- Sandbox-Ready: Controlled API surface for AI execution
"""

# Import standard library modules
import os              # File path operations and cleanup
import re              # Regular expressions for cell reference validation
import tempfile        # Create temporary files for images
import io              # In-memory byte buffers for image handling
from typing import List, Optional, Dict, Union, Any  # Type hints

# Import third-party libraries
import matplotlib.pyplot as plt  # Plotting library
from openpyxl.utils import get_column_letter, column_index_from_string  # Excel coordinate conversion
from openpyxl.utils.cell import coordinate_to_tuple  # Convert "A1" to (1, 1)
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment  # Excel formatting
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart  # Excel charts
from openpyxl.chart.reference import Reference  # Chart data references
from openpyxl.drawing.image import Image  # Excel image handling
from PIL import Image as PILImage  # Python Imaging Library (for image manipulation)
import tiktoken  # OpenAI's token counting library


def calculate_token_cost_line(text: str, model: str = "gpt-4") -> int:
    """
    Calculate token cost of text for AI context management.

    Prevents exceeding model context limits and controls costs.
    Uses tiktoken for accurate counting based on model-specific encodings.

    """
    try:
        # Map models to their token encodings
        model_encodings = {
            "gpt-4": "cl100k_base",
            "gpt-4-turbo": "cl100k_base",
            "gpt-4o": "o200k_base",
            "gpt-3.5-turbo": "cl100k_base",
            "gpt-5-nano-2025-08-07": "o200k_base",
            "text-embedding-ada-002": "cl100k_base",
        }

        # Get the appropriate tokenizer encoding
        encoding_name = model_encodings.get(model, "cl100k_base")  # Default fallback
        encoding = tiktoken.get_encoding(encoding_name)

        # Encode the text and count tokens
        tokens = encoding.encode(text)
        return len(tokens)

    except Exception:
        # Fallback: rough estimate if tiktoken fails
        # Assumes ~3.5 characters per token (conservative estimate)
        char_count = len(text)
        token_count = max(1, int(char_count / 3.5))
        return token_count


class ExcelToolkit:
    """
    Comprehensive Excel manipulation toolkit for AI-generated code.

    Wraps openpyxl to provide a simpler, safer API for reading, analyzing,
    modifying, and visualizing Excel data. Tracks temporary files for cleanup.

   """

    def __init__(self, workbook, excel_path: str, output_path: Optional[str] = None):
        """  Initialize toolkit with workbook and file paths. """

        self.workbook = workbook
        self.excel_path = excel_path
        self.output_path = output_path  # Explicit output path (if provided)
        self._temp_files = []  # Track temporary image files for cleanup
        self._output_workbook = None  # Separate workbook for output (created on demand)

    def get_sheet(self, sheet_name: Optional[str] = None):
        """ Get worksheet by name, or active sheet if no name provided.  """

        if sheet_name is None:
            return self.workbook.active
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {self.workbook.sheetnames}")

    def inspector(self, range_ref: str, sheet_name: Optional[str] = None) -> List[List]:
        """
        Read cell values from a range and return as 2D list.

        Handles single cells ("A1") and ranges ("A1:B5").
        """
        sheet = self.get_sheet(sheet_name)
        cell_range = sheet[range_ref]

        # Handle single cell case (not a range)
        if hasattr(cell_range, 'value'):
            return [[cell_range.value]]

        # Handle multi-cell range
        result = []
        for row in cell_range:
            row_values = [cell.value for cell in row]
            result.append(row_values)
        return result

    def inspector_attribute(self, range_ref: str, attributes: List[str],
                            sheet_name: Optional[str] = None) -> Dict:
        """
        Read formatting attributes from a range of cells.

        Extracts visual properties like color, font, or formulas.
        Useful for analysis based on formatting (e.g., "sum all red cells").
        """
        print(f"🔎 [read_range_attribute] Reading attributes {attributes} for range {range_ref} in sheet '{sheet_name}'")

        # Validate attributes
        if not attributes:
            return {"error": "No attributes specified"}

        valid_attributes = ["color", "font", "formula"]
        invalid_attrs = [attr for attr in attributes if attr not in valid_attributes]
        if invalid_attrs:
            return {"error": f"Invalid attributes: {invalid_attrs}. Valid options: {valid_attributes}"}

        try:
            sheet = self.get_sheet(sheet_name)
            cell_range = sheet[range_ref]
        except (ValueError, KeyError) as e:
            return {"error": str(e)}

        # Normalize to list of cells (handles both single cell and range)
        if hasattr(cell_range, 'coordinate'):
            cells_to_process = [cell_range]
        else:
            cells_to_process = []
            for row in cell_range:
                if hasattr(row, '__iter__'):
                    cells_to_process.extend(row)
                else:
                    cells_to_process.append(row)

        result_attributes = {}

        # Extract each requested attribute
        for attr in attributes:
            result_attributes[attr] = {}

            for cell in cells_to_process:
                cell_coord = cell.coordinate
                attr_value = None

                if attr == "color":
                    # Get background fill color
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb != '00000000':
                        attr_value = f"#{cell.fill.fgColor.rgb}"

                elif attr == "font":
                    # Compile font properties into a string
                    font_details = []
                    if cell.font:
                        if cell.font.color and cell.font.color.rgb != '00000000':
                            font_details.append(f"color:#{cell.font.color.rgb}")
                        if cell.font.name:
                            font_details.append(f"name:{cell.font.name}")
                        if cell.font.size:
                            font_details.append(f"size:{cell.font.size}")
                        if cell.font.bold:
                            font_details.append("bold:True")
                        if cell.font.italic:
                            font_details.append("italic:True")
                        if cell.font.underline and cell.font.underline != 'none':
                            font_details.append(f"underline:{cell.font.underline}")

                    attr_value = "; ".join(font_details) if font_details else None

                elif attr == "formula":
                    # Get formula if cell contains one
                    if cell.data_type == 'f' and cell.value:
                        attr_value = str(cell.value)

                if attr_value is not None:
                    result_attributes[attr][cell_coord] = attr_value

        return {
            "range": range_ref,
            "sheet": sheet_name or sheet.title,
            "attributes": result_attributes,
            "total_cells_processed": len(cells_to_process)
        }

    def search(self, value: Any, sheet_name: Optional[str] = None,
               case_sensitive: bool = False, search_type: str = "partial") -> List[Dict]:
        """
        Search for cells containing a specific value across the sheet.
        Scans all cells to find matches. Supports partial, whole, or strip matching.
        """

        sheet = self.get_sheet(sheet_name)
        matches = []

        # Validate search type
        valid_search_types = ["partial", "whole", "strip"]
        if search_type not in valid_search_types:
            raise ValueError(f"Invalid search_type '{search_type}'. Valid options: {valid_search_types}")

        # Prepare search value
        search_value = str(value) if case_sensitive else str(value).lower()

        # Scan all cells in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell_str = str(cell.value)

                    if not case_sensitive:
                        cell_str = cell_str.lower()

                    is_match = False

                    # Determine match type
                    if search_type == "partial":
                        is_match = search_value in cell_str
                    elif search_type == "whole":
                        is_match = search_value == cell_str
                    elif search_type == "strip":
                        stripped_cell_str = cell_str.strip()
                        is_match = search_value == stripped_cell_str

                    if is_match:
                        matches.append({
                            'coordinate': cell.coordinate,
                            'value': cell.value,
                            'row': cell.row,
                            'column': cell.column
                        })

        return matches

    def get_sheet_as_dataframe(self, sheet_name: Optional[str] = None,
                               header_row: int = 1, max_rows: Optional[int] = None):
        """
        Convert sheet to pandas DataFrame for data analysis.

        Transforms Excel data into pandas DataFrame format, enabling
        powerful analysis (filtering, grouping, statistics).
        """
        import pandas as pd  # Local import (only when needed)
        sheet = self.get_sheet(sheet_name)

        # Extract all cell values
        data = []
        for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if max_rows and i > max_rows:
                break
            data.append(row)

        if not data:
            return pd.DataFrame()  # Return empty DataFrame for empty sheet

        # Use specified header row for column names
        if header_row > 0 and len(data) >= header_row:
            headers = data[header_row - 1]
            data_rows = data[header_row:]
            df = pd.DataFrame(data_rows, columns=headers)
        else:
            df = pd.DataFrame(data)

        return df


#############################################
    def list_sheets(self) -> List[str]:
        """ Return list of all sheet names in the workbook. """

        return self.workbook.sheetnames

    def get_sheet_info(self, sheet_name: Optional[str] = None) -> Dict[str, Any]:
        """ Get information about a specific sheet (name, dimensions)."""

        sheet = self.get_sheet(sheet_name)
        return {
            'name': sheet.title,
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'dimensions': f"{sheet.max_row} rows × {sheet.max_column} columns"
        }

    def get_all_sheets_info(self) -> Dict[str, Dict[str, Any]]:
        """ Get information about all sheets in the workbook. """

        result = {}
        for sheet_name in self.workbook.sheetnames:
            result[sheet_name] = self.get_sheet_info(sheet_name)
        return result

    def read_multiple_sheets(self, sheet_names: List[str], range_ref: Optional[str] = None) -> Dict[str, List[List]]:
        """ Read data from multiple sheets at once. """

        result = {}
        for sheet_name in sheet_names:
            if sheet_name not in self.workbook.sheetnames:
                result[sheet_name] = None
                continue
            
            if range_ref:
                result[sheet_name] = self.inspector(range_ref, sheet_name)
            else:
                # Read entire sheet
                sheet = self.get_sheet(sheet_name)
                max_range = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
                result[sheet_name] = self.inspector(max_range, sheet_name)
        
        return result
#####################################################



    def save_plot_to_excel(self, sheet_name: str, cell_position: str = "A1",
                           figsize: tuple = (10, 6), dpi: int = 100) -> str:
        """
        Save current matplotlib plot as image in Excel sheet.

        Creates sheet if it doesn't exist, saves plot to temporary file,
        inserts image at specified cell, and tracks file for cleanup.
        """
        # Create sheet if it doesn't exist
        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        sheet = self.workbook[sheet_name]

        # Get current figure
        fig = plt.gcf()
        if fig.get_axes():  # Check if figure has content
            fig.set_size_inches(figsize)
            plt.tight_layout()  # Adjust spacing

            # Save to memory buffer first
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=dpi, bbox_inches='tight')
            img_buffer.seek(0)

            # Convert to PIL Image
            pil_img = PILImage.open(img_buffer)

            # Save to temporary file (Excel needs a file path, not just bytes)
            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as tmp_file:
                pil_img.save(tmp_file.name, 'PNG')
                tmp_filename = tmp_file.name

            # Add to Excel
            img = Image(tmp_filename)
            sheet.add_image(img, cell_position)

            # Track temp file for cleanup later
            self._temp_files.append(tmp_filename)

            print(f"✅ Chart saved to sheet '{sheet_name}' at position {cell_position}")
            plt.close(fig)  # Close plot to free memory
            return f"Chart saved to {sheet_name}!{cell_position}"
        else:
            print("⚠️ No plot found to save. Create a plot first.")
            return "No plot to save"

    def save_workbook(self) -> str:
        """
        Save OUTPUT workbook to new file (input files remain unchanged).

        Saves output workbook (sheets created via create_output_sheet()),
        cleans up temporary image files, and clears tracking list.
        """

        # Use explicit output_path if set, otherwise generate from input filename
        if self.output_path:
            filename = self.output_path
        else:
            dir_path = os.path.dirname(self.excel_path)
            base_name = os.path.splitext(os.path.basename(self.excel_path))[0]
            filename = os.path.join(dir_path, f"{base_name}_output.xlsx")

        # Get output workbook (create if needed)
        output_wb = self._get_output_workbook()
        
        # If no sheets were created, add a default one to avoid empty workbook error
        if len(output_wb.sheetnames) == 0:
            output_wb.create_sheet("Output")
        
        # Save the OUTPUT workbook
        output_wb.save(filename)

        # Clean up temporary image files
        for temp_file in self._temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)  # Delete file
            except Exception as e:
                print(f"⚠️ Warning: Could not delete temporary file {temp_file}: {e}")

        # Clear tracking list
        self._temp_files = []

        print(f"💾 Workbook saved to: {filename}")
        return filename

    # ==================== Excel Editing Functions ====================
    # These methods modify the Excel file structure and content

    def insert_rows(self, sheet_name: str, row_index: int, count: int = 1) -> str:
        """
        Insert empty rows at specific position.
        Existing rows at and below insertion point shift down.
        """

        try:
            sheet = self.get_sheet(sheet_name)

            # Validate inputs
            if row_index < 1 or count < 1:
                raise ValueError("Row index and count must be >= 1")

            # Perform insertion
            sheet.insert_rows(row_index, count)

            message = f"✅ Inserted {count} row(s) at row {row_index} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error inserting rows: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def insert_columns(self, sheet_name: str, col_index: Union[int, str], count: int = 1) -> str:
        """
        Insert empty columns at specific position.
        Existing columns at and to the right shift right.
        """
        try:
            sheet = self.get_sheet(sheet_name)

            # Convert column letter to number if needed (e.g., "B" -> 2)
            if isinstance(col_index, str):
                col_index = column_index_from_string(col_index)

            # Validate inputs
            if col_index < 1 or count < 1:
                raise ValueError("Column index and count must be >= 1")

            # Perform insertion
            sheet.insert_cols(col_index, count)

            col_letter = get_column_letter(col_index)
            message = f"✅ Inserted {count} column(s) at column {col_letter} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error inserting columns: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def delete_rows(self, sheet_name: str, start_row: int, count: int = 1) -> str:
        """ Delete rows from the sheet. """

        try:
            sheet = self.get_sheet(sheet_name)

            # Validate inputs
            if start_row < 1 or count < 1:
                raise ValueError("Start row and count must be >= 1")
            if start_row > sheet.max_row:
                raise ValueError(f"Start row {start_row} exceeds sheet max row {sheet.max_row}")

            # Perform deletion
            sheet.delete_rows(start_row, count)

            message = f"✅ Deleted {count} row(s) starting from row {start_row} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error deleting rows: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def delete_columns(self, sheet_name: str, start_col: Union[int, str], count: int = 1) -> str:
        """ Delete columns from the sheet. """

        try:
            sheet = self.get_sheet(sheet_name)

            # Convert column letter to number if needed
            if isinstance(start_col, str):
                start_col = column_index_from_string(start_col)

            # Validate inputs
            if start_col < 1 or count < 1:
                raise ValueError("Start column and count must be >= 1")
            if start_col > sheet.max_column:
                raise ValueError(f"Start column {start_col} exceeds sheet max column {sheet.max_column}")

            # Perform deletion
            sheet.delete_cols(start_col, count)

            col_letter = get_column_letter(start_col)
            message = f"✅ Deleted {count} column(s) starting from column {col_letter} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error deleting columns: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def set_cell_value(self, sheet_name: str, cell_ref: str, value: Any) -> str:
        """ Set the value of a single cell. """

        try:
            sheet = self.get_sheet(sheet_name)

            # Validate cell reference format (must be like "A1", "B5", "AA100")
            if not re.match(r'^[A-Z]+[0-9]+$', cell_ref.upper()):
                raise ValueError(f"Invalid cell reference: {cell_ref}")

            # Set the value
            sheet[cell_ref] = value

            message = f"✅ Set cell {cell_ref} to '{value}' in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error setting cell value: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def set_range_values(self, sheet_name: str, start_cell: str,
                         values_2d_array: List[List[Any]]) -> str:
        """ Set values for a range of cells using a 2D array. """

        try:
            sheet = self.get_sheet(sheet_name)

            # Validate inputs
            if not re.match(r'^[A-Z]+[0-9]+$', start_cell.upper()):
                raise ValueError(f"Invalid cell reference: {start_cell}")

            if not values_2d_array or not isinstance(values_2d_array, list):
                raise ValueError("values_2d_array must be a non-empty list")

            # Convert start cell to row/column numbers
            start_row, start_col = coordinate_to_tuple(start_cell)

            # Write the 2D array to the sheet
            for row_idx, row_values in enumerate(values_2d_array):
                if not isinstance(row_values, list):
                    raise ValueError(f"Row {row_idx} must be a list")

                for col_idx, value in enumerate(row_values):
                    current_row = start_row + row_idx
                    current_col = start_col + col_idx
                    sheet.cell(row=current_row, column=current_col, value=value)

            # Calculate the written range for the success message
            rows_count = len(values_2d_array)
            cols_count = max(len(row) for row in values_2d_array) if values_2d_array else 0
            end_cell = sheet.cell(row=start_row + rows_count - 1,
                                  column=start_col + cols_count - 1).coordinate

            message = f"✅ Set range {start_cell}:{end_cell} ({rows_count}x{cols_count}) in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error setting range values: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def copy_range(self, src_sheet: str, src_range: str, dest_sheet: str, dest_cell: str) -> str:
        """ Copy data from one range to another (within or across sheets). """

        try:
            # Get source and destination sheets
            src_ws = self.get_sheet(src_sheet)
            dest_ws = self.get_sheet(dest_sheet)

            # Validate source range format
            if ':' not in src_range:
                raise ValueError("Source range must be in format 'A1:B2'")

            # Read source data
            source_data = []
            for row in src_ws[src_range]:
                row_data = [cell.value for cell in row]
                source_data.append(row_data)

            # Write to destination if data exists
            if source_data:
                dest_start_row, dest_start_col = coordinate_to_tuple(dest_cell)

                for row_idx, row_values in enumerate(source_data):
                    for col_idx, value in enumerate(row_values):
                        dest_row = dest_start_row + row_idx
                        dest_col = dest_start_col + col_idx
                        dest_ws.cell(row=dest_row, column=dest_col, value=value)

                rows_count = len(source_data)
                cols_count = len(source_data[0]) if source_data else 0
                dest_end_cell = dest_ws.cell(row=dest_start_row + rows_count - 1,
                                             column=dest_start_col + cols_count - 1).coordinate

                message = f"✅ Copied {src_sheet}!{src_range} to {dest_sheet}!{dest_cell}:{dest_end_cell}"
                print(message)
                return message
            else:
                message = "⚠️ No data found in source range"
                print(message)
                return message

        except Exception as e:
            error_msg = f"❌ Error copying range: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def apply_formatting(self, sheet_name: str, range_ref: str, format_dict: Dict[str, Any]) -> str:
        """
        Apply visual formatting to range of cells.
        Supports background color, font properties, borders, and alignment.
        """

        try:
            sheet = self.get_sheet(sheet_name)

            # Handle both single cell and range
            if ':' in range_ref:
                cell_range = sheet[range_ref]
                cells = []
                for row in cell_range:
                    if hasattr(row, '__iter__'):
                        cells.extend(row)
                    else:
                        cells.append(row)
            else:
                cells = [sheet[range_ref]]

            # Apply each formatting property
            for cell in cells:
                # Background fill color
                if 'fill_color' in format_dict:
                    color = self._parse_color(format_dict['fill_color'])
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

                # Font properties
                font_kwargs = {}
                if 'font_color' in format_dict:
                    font_kwargs['color'] = self._parse_color(format_dict['font_color'])
                if 'font_size' in format_dict:
                    font_kwargs['size'] = format_dict['font_size']
                if 'font_name' in format_dict:
                    font_kwargs['name'] = format_dict['font_name']
                if 'bold' in format_dict:
                    font_kwargs['bold'] = format_dict['bold']
                if 'italic' in format_dict:
                    font_kwargs['italic'] = format_dict['italic']
                if 'underline' in format_dict:
                    font_kwargs['underline'] = 'single' if format_dict['underline'] else None

                if font_kwargs:
                    cell.font = Font(**font_kwargs)

                # Border
                if 'border' in format_dict:
                    border_style = format_dict['border']
                    side = Side(style=border_style)
                    cell.border = Border(left=side, right=side, top=side, bottom=side)

                # Alignment
                if 'alignment' in format_dict:
                    horizontal = format_dict['alignment']
                    cell.alignment = Alignment(horizontal=horizontal)

            message = f"✅ Applied formatting to {range_ref} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error applying formatting: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def create_chart(self, sheet_name: str, chart_type: str, data_range: str,
                     position: str = "A1", title: str = "",
                     x_axis_title: str = "", y_axis_title: str = "") -> str:
        """
        Create chart in Excel sheet.
        Supports bar, line, pie, scatter, and area charts.
        """

        try:
            sheet = self.get_sheet(sheet_name)

            # Map chart type names to openpyxl classes
            chart_classes = {
                'bar': BarChart,
                'line': LineChart,
                'pie': PieChart,
                'scatter': ScatterChart,
                'area': AreaChart
            }

            # Validate chart type
            if chart_type.lower() not in chart_classes:
                raise ValueError(f"Unsupported chart type: {chart_type}. Available: {list(chart_classes.keys())}")

            # Create chart instance
            chart_class = chart_classes[chart_type.lower()]
            chart = chart_class()

            # Set chart properties
            if title:
                chart.title = title
            if x_axis_title and hasattr(chart, 'x_axis'):
                chart.x_axis.title = x_axis_title
            if y_axis_title and hasattr(chart, 'y_axis'):
                chart.y_axis.title = y_axis_title

            # Add data to chart
            data = Reference(sheet, range_string=data_range)
            chart.add_data(data, titles_from_data=True)

            # Place chart on sheet
            sheet.add_chart(chart, position)

            message = f"✅ Created {chart_type} chart from {data_range} at {position} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error creating chart: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def add_formula(self, sheet_name: str, cell_ref: str, formula: str) -> str:
        """
        Add Excel formula to cell.
        Automatically prepends "=" if missing. Formula calculates when file opens.
        """

        try:
            sheet = self.get_sheet(sheet_name)

            # Validate cell reference
            if not re.match(r'^[A-Z]+[0-9]+$', cell_ref.upper()):
                raise ValueError(f"Invalid cell reference: {cell_ref}")

            # Ensure formula starts with "="
            if not formula.startswith('='):
                formula = '=' + formula

            # Set the formula
            sheet[cell_ref] = formula

            message = f"✅ Added formula '{formula}' to cell {cell_ref} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error adding formula: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def _parse_color(self, color: str) -> str:
        """
        Convert color names to hex format.
        Supports common color names and hex codes. Used internally by apply_formatting().
        """

        # Color name to hex mapping
        color_names = {
            'red': 'FF0000', 'green': '00FF00', 'blue': '0000FF',
            'yellow': 'FFFF00', 'orange': 'FFA500', 'purple': '800080',
            'pink': 'FFC0CB', 'brown': 'A52A2A', 'black': '000000',
            'white': 'FFFFFF', 'gray': '808080', 'grey': '808080'
        }

        if color.startswith('#'):
            return color[1:]  # Remove "#" prefix
        elif color.lower() in color_names:
            return color_names[color.lower()]
        else:
            # Assume it's already a hex code
            return color

    def _get_output_workbook(self):
        """Get or create the output workbook (separate from input files)."""

        from openpyxl import Workbook as OpenpyxlWorkbook
        if self._output_workbook is None:
            self._output_workbook = OpenpyxlWorkbook()
            # Remove the default sheet created by openpyxl
            if 'Sheet' in self._output_workbook.sheetnames:
                del self._output_workbook['Sheet']
        return self._output_workbook

    def create_output_sheet(self, sheet_name: str = "Output") -> str:
        """ Create a new sheet in the OUTPUT workbook (separate from input files). """

        try:
            output_wb = self._get_output_workbook()
            
            if sheet_name in output_wb.sheetnames:
                # Remove existing sheet and recreate
                del output_wb[sheet_name]
            
            output_wb.create_sheet(sheet_name)
            message = f"✅ Created output sheet '{sheet_name}' (in new output file)"
            print(message)
            return message
        except Exception as e:
            error_msg = f"❌ Error creating sheet: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def write_dataframe_to_sheet(self, data: List[List], sheet_name: str, 
                                  start_cell: str = "A1", include_header: bool = True) -> str:
        """
        Write 2D list (DataFrame-like data) to output workbook.

        Writes to new output file, keeping input files unchanged.
        Use instead of DataFrame.to_excel() to avoid file path confusion.
        """

        try:
            output_wb = self._get_output_workbook()
            
            # Create sheet if not exists in output workbook
            if sheet_name not in output_wb.sheetnames:
                output_wb.create_sheet(sheet_name)
            
            sheet = output_wb[sheet_name]
            
            # Convert start cell to row/column numbers
            start_row, start_col = coordinate_to_tuple(start_cell)
            
            # Write data
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    current_row = start_row + row_idx
                    current_col = start_col + col_idx
                    sheet.cell(row=current_row, column=current_col, value=value)
            
            rows_written = len(data)
            cols_written = max(len(row) for row in data) if data else 0
            end_row = start_row + rows_written - 1
            end_col = start_col + cols_written - 1
            end_cell = f"{get_column_letter(end_col)}{end_row}"
            
            message = f"✅ Wrote {rows_written} rows to {sheet_name}!{start_cell}:{end_cell}"
            print(message)
            return message
            
        except Exception as e:
            error_msg = f"❌ Error writing data to sheet: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def highlight_rows(self, sheet_name: str, row_numbers: List[int], 
                       format_dict: Dict[str, Any] = None) -> str:
        """
        Highlight entire rows in output workbook (default: red background).

        Use for emphasizing important records (max spending days, etc.).
        Applies formatting to all cells in specified rows.
        """
        try:
            if format_dict is None:
                format_dict = {"fill_color": "red"}
            
            output_wb = self._get_output_workbook()
            if sheet_name not in output_wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in output workbook. Create it first with create_output_sheet().")
            sheet = output_wb[sheet_name]
            max_col = sheet.max_column
            
            for row_num in row_numbers:
                # Apply formatting to all cells in the row
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_num, column=col_idx)
                    
                    # Apply fill color
                    if 'fill_color' in format_dict:
                        color = self._parse_color(format_dict['fill_color'])
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
                    # Apply font properties
                    font_kwargs = {}
                    if 'font_color' in format_dict:
                        font_kwargs['color'] = self._parse_color(format_dict['font_color'])
                    if 'bold' in format_dict:
                        font_kwargs['bold'] = format_dict['bold']
                    if font_kwargs:
                        cell.font = Font(**font_kwargs)
            
            message = f"✅ Highlighted row(s) {row_numbers} in sheet '{sheet_name}'"
            print(message)
            return message
            
        except Exception as e:
            error_msg = f"❌ Error highlighting rows: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def save_workbook_to(self, output_path: str) -> str:
        """ Save the output workbook to a specific path (keeps input files unchanged). """

        try:
            output_wb = self._get_output_workbook()
            
            # If no sheets were created, add a default one
            if len(output_wb.sheetnames) == 0:
                output_wb.create_sheet("Output")
            
            output_wb.save(output_path)
            
            # Clean up temporary files
            for temp_file in self._temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                except Exception:
                    pass
            self._temp_files = []
            
            message = f"💾 Workbook saved to: {output_path}"
            print(message)
            return output_path
            
        except Exception as e:
            error_msg = f"❌ Error saving workbook: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def add_summary_row(self, sheet_name: str, row_number: int, 
                        summary_data: Dict[str, Any]) -> str:
        """ Add a summary row with labeled statistics to the output workbook.  """

        try:
            output_wb = self._get_output_workbook()
            if sheet_name not in output_wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in output workbook.")
            sheet = output_wb[sheet_name]
            col = 1
            
            for label, value in summary_data.items():
                sheet.cell(row=row_number, column=col, value=label)
                sheet.cell(row=row_number, column=col + 1, value=value)
                col += 3  # Space between each stat
            
            message = f"✅ Added summary row at row {row_number} in sheet '{sheet_name}'"
            print(message)
            return message
            
        except Exception as e:
            error_msg = f"❌ Error adding summary row: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def get_helper_functions_dict(self) -> Dict:
        """
        Return dictionary of helper functions for AI code sandbox.

        Exposes whitelisted methods to AI-generated code for security and
        sandboxing. Prevents accidental access to internal methods.
        """

        return {
            # Reading functions
            'get_sheet': self.get_sheet,
            'inspector': self.inspector,
            'inspector_attribute': self.inspector_attribute,
            'search': self.search,
            'get_sheet_as_dataframe': self.get_sheet_as_dataframe,

            # Visualization
            'save_plot_to_excel': self.save_plot_to_excel,

            # File operations
            'save_workbook': self.save_workbook,

            # Multi-sheet operations
            'list_sheets': self.list_sheets,
            'get_sheet_info': self.get_sheet_info,
            'get_all_sheets_info': self.get_all_sheets_info,
            'read_multiple_sheets': self.read_multiple_sheets,

            #additional editing tools
            'insert_rows': self.insert_rows,
            'insert_columns': self.insert_columns,
            'delete_rows': self.delete_rows,
            'delete_columns': self.delete_columns,
            'set_cell_value': self.set_cell_value,
            'set_range_values': self.set_range_values,
            'copy_range': self.copy_range,
            'apply_formatting': self.apply_formatting,
            'create_chart': self.create_chart,
            'add_formula': self.add_formula,

            # New unified output functions
            'create_output_sheet': self.create_output_sheet,
            'write_dataframe_to_sheet': self.write_dataframe_to_sheet,
            'highlight_rows': self.highlight_rows,
            'save_workbook_to': self.save_workbook_to,
            'add_summary_row': self.add_summary_row
        }