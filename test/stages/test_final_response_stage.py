import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))

from backend.stages.final_response.stage import FinalResponseStage
from openpyxl import Workbook


def test_question_content_label_uses_helper_metadata_for_multi_key_join():
    label = FinalResponseStage._question_content_label(
        "Join the spreadsheet tables by the shared student id and semester keys.",
        {},
    )

    assert label == "multi-key join report"


def test_question_content_label_uses_helper_metadata_for_missing_data_scan():
    label = FinalResponseStage._question_content_label(
        "Identify which sheets have missing values and report the affected columns.",
        {},
    )

    assert label == "missing data report"


def test_question_content_label_uses_helper_metadata_for_dependency_schedule():
    label = FinalResponseStage._question_content_label(
        "Schedule tasks based on dependencies and include start time and end time columns.",
        {},
    )

    assert label == "task schedule"


def test_fallback_short_answer_generates_generic_average_subject():
    answer = FinalResponseStage._fallback_short_answer(
        user_question="What is the average revenue per region?",
        final_answer="42",
        validation_passed=True,
        workbook_summary={},
    )

    assert answer == "The average revenue per region is 42."


def test_file_short_answer_includes_sample_rows_from_workbook(tmp_path):
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Region", "Average_Rate_2020_2024", "Growth_2020_to_2024", "Rank_by_Growth"])
    sheet.append(["Europe", 87.04, 4.7, 1])
    sheet.append(["North America", 82.22, 3.4, 2])
    workbook.save(output_path)

    stage = FinalResponseStage(client=None, deployment="test")
    answer = stage.run(
        user_question="Calculate the average rate, growth, and ranking by region.",
        final_answer=str(output_path),
        validation_result={"validation_passed": True},
        execution_result=None,
    )

    assert "Generated" in answer
    assert "Europe" in answer
    assert "Average_Rate_2020_2024=87.04" in answer
    assert "Rank_by_Growth=1" in answer


def test_file_short_answer_prefers_execution_result_summary(tmp_path):
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    workbook.active.append(["A", "B"])
    workbook.active.append([1, 2])
    workbook.save(output_path)

    stage = FinalResponseStage(client=None, deployment="test")
    answer = stage.run(
        user_question="Create a dashboard.",
        final_answer=str(output_path),
        validation_result={"validation_passed": True},
        execution_result={
            "execution_summary": {
                "execution_steps": [
                    {
                        "success": True,
                        "result": "Workbook saved.\nRESULT_SUMMARY: Total profit is 100 and top region is Europe.",
                    }
                ]
            }
        },
    )

    assert answer == "Total profit is 100 and top region is Europe."


def test_file_short_answer_ignores_path_only_execution_result_summary(tmp_path):
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    workbook.active.append(["EmpID", "Name", "Department"])
    workbook.active.append(["E002", "Sarah Jones", "Finance"])
    workbook.save(output_path)

    stage = FinalResponseStage(client=None, deployment="test")
    answer = stage.run(
        user_question="Fill any missing employee department values.",
        final_answer=str(output_path),
        validation_result={"validation_passed": True},
        execution_result={
            "execution_summary": {
                "execution_steps": [
                    {
                        "success": True,
                        "result": f"RESULT_SUMMARY: Final Output saved at: {output_path}",
                    }
                ]
            }
        },
    )

    assert "Sarah Jones" in answer
    assert "Finance" in answer
    assert "Final Output saved at" not in answer
