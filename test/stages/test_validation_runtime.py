import os
import sys
from types import SimpleNamespace

from openpyxl.styles import PatternFill
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))

from backend.stages.validation.runtime import ValidationRuntime


class _ClientStub:
    pass


def test_validation_runtime_rule_passes_saved_workbook_without_llm(tmp_path):
    output_path = tmp_path / "out.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Output"
    sheet.append(["Date", "Category", "Daily Spending (£)", "Notes"])
    sheet.append(["2025-11-01", "Food", 20, "Lunch"])
    sheet.append(["2025-11-02", "Travel", 15, "Bus"])
    sheet.append(["Total Spending in November", 35])
    for cell in sheet[2]:
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    workbook.save(output_path)

    runtime = ValidationRuntime(
        client=_ClientStub(),
        deployment="offline-test",
        excel_context_understanding="",
        prompt_profile="offline_strict",
    )
    runtime.llm_client = SimpleNamespace(
        get_response=lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("validation LLM should not be called")
        )
    )

    execution_result = {
        "success": True,
        "answer": str(output_path),
        "conversation_history": [],
        "total_turns": 3,
        "execution_summary": {
            "total_code_executions": 1,
            "successful_executions": 1,
            "failed_executions": 0,
            "execution_steps": [
                {
                    "turn": 3,
                    "success": True,
                    "code": "save_workbook_to(output_path)",
                    "result": (
                        "Wrote 3 rows to Output!A1:D3\n"
                        "Added summary row at row 4 in sheet 'Output'\n"
                        "Highlighted row(s) [2] in sheet 'Output'\n"
                        f"Workbook saved to: {output_path}\n"
                    ),
                }
            ],
        },
    }

    understanding_output = "\n".join(
        [
            "requires_detailed_table: YES",
            "requires_highlight: YES",
            "requires_summary_metrics: YES",
        ]
    )

    result = runtime.run(
        execution_result=execution_result,
        user_question=(
            "Merge the two spending tables, calculate the average daily spending and total spending "
            "in November, highlight the maximum spending days, and output a new spreadsheet."
        ),
        understanding_output=understanding_output,
    )

    assert result["validation_passed"] is True
    assert result["verified_answer"] == str(output_path)
    assert result["requires_reexecution"] is False


def test_validation_runtime_rejects_highlighted_row_that_is_not_numeric_max(tmp_path):
    output_path = tmp_path / "wrong_highlight.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Output"
    sheet.append(["Date", "Category", "Daily Spending (£)", "Notes"])
    for day, value in enumerate(
        [106.11, 43.62, 89.53, 100, 61.14, 92.15, 76.3, 116.58, 32.59, 58.96,
         56.22, 93.13, 67.38, 73.28, 67.03, 51.77, 101.05, 42.52, 107.96,
         22.64, 1000, 38.45, 115.04, 71.52, 111.18, 102.98, 57.13, 32.5,
         99.25, 35.74],
        start=1,
    ):
        sheet.append([f"2025-11-{day:02d}", "Food", value, ""])
    sheet.append(["Total", 3123.75, "", "Average", 104.12, "", "Max", 1000])
    for cell in sheet[20]:
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    workbook.save(output_path)

    runtime = ValidationRuntime(
        client=_ClientStub(),
        deployment="offline-test",
        excel_context_understanding="",
        prompt_profile="offline_strict",
    )
    runtime.llm_client = SimpleNamespace(
        get_response=lambda *_args, **_kwargs: (
            "VALIDATION_STATUS: PASSED\n"
            "CONFIDENCE_SCORE: 1.00\n"
            "ISSUES_FOUND:\n"
            "- None identified.\n"
            "IMPROVEMENT_FEEDBACK:\n"
            "No improvement needed.\n"
            "FINAL_ASSESSMENT:\n"
            "Looks correct."
        )
    )

    execution_result = {
        "success": True,
        "answer": str(output_path),
        "conversation_history": [],
        "total_turns": 1,
        "execution_summary": {
            "total_code_executions": 1,
            "successful_executions": 1,
            "failed_executions": 0,
            "execution_steps": [
                {
                    "turn": 1,
                    "success": True,
                    "code": "highlight_rows('Output', [20], {'fill_color': 'red'})",
                    "result": (
                        "Wrote 31 rows to Output!A1:D31\n"
                        "Added summary row at row 32 in sheet 'Output'\n"
                        "Highlighted row(s) [20] in sheet 'Output'\n"
                        f"Workbook saved to: {output_path}\n"
                    ),
                }
            ],
        },
    }

    result = runtime.run(
        execution_result=execution_result,
        user_question=(
            "Merge the two spending tables, calculate average and total spending in November, "
            "highlight the maximum spending day, and output a new spreadsheet."
        ),
        understanding_output="\n".join(
            [
                "requires_detailed_table: YES",
                "requires_highlight: YES",
                "requires_summary_metrics: YES",
            ]
        ),
    )

    assert result["validation_passed"] is False
    assert any("highlighted row" in issue.lower() and "maximum" in issue.lower() for issue in result["issues_found"])
