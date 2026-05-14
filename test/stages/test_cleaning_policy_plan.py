import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))

from openpyxl import Workbook

from backend.stages.cleaning.stage import DataCleaningStage


class _SandboxStub:
    def __init__(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Cabin"
        ws["A2"] = None
        ws["A3"] = None
        self.workbooks = {"titanic.xlsx": wb}


class _DropRowSandboxStub:
    def __init__(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Date", "Daily Spending (£)"])
        ws.append(["2025-11-03", 89.53])
        ws.append(["2025-11-04", None])
        ws.append(["2025-11-05", 12.0])
        self.workbooks = {"tc01_input01.xlsx": wb}


def test_apply_policy_plan_leaves_missing_values_blank_without_llm():
    stage = DataCleaningStage(client=None, deployment="offline-test", prompt_profile="offline_strict")
    sandbox = _SandboxStub()

    report = stage.apply_policy_plans(
        sandbox,
        [
            {
                "policy_kind": "missing_value",
                "sheet_key": "titanic.xlsx::Sheet1",
                "column": "Cabin",
                "affected_rows": [2, 3],
                "resolution": "leave_blank",
            }
        ],
    )

    assert report["applied_actions"] == ["Leave missing `Cabin` values unchanged in `titanic.xlsx::Sheet1`."]
    ws = sandbox.workbooks["titanic.xlsx"]["Sheet1"]
    assert ws["A2"].value is None
    assert ws["A3"].value is None


def test_deterministic_cleaning_can_drop_exact_excel_row():
    stage = DataCleaningStage(client=None, deployment="offline-test", prompt_profile="offline_strict")
    sandbox = _DropRowSandboxStub()

    report = stage.apply(
        sandbox,
        ["Drop Excel row 3 in `tc01_input01.xlsx::Sheet1`."],
    )

    assert report["applied_actions"] == ["Drop Excel row 3 in `tc01_input01.xlsx::Sheet1`."]
    ws = sandbox.workbooks["tc01_input01.xlsx"]["Sheet1"]
    assert ws.max_row == 3
    assert ws["A3"].value == "2025-11-05"
