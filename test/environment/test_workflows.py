from pathlib import Path
import os
import sys

import pandas as pd
from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))

from backend.environment.spreadsheet.loader import load_world
from backend.environment.namespace.namespace import SpreadsheetNamespace
from backend.environment.spreadsheet.tools.output import ExcelOutputWriter
from backend.environment.spreadsheet.tools.cross_workbook import extract_sheet_table
from backend.environment.spreadsheet.tools.workflows import (
    _build_weighted_period_output,
    _build_grouped_assignment_join,
    _extract_period_records,
    _find_first_period_cell,
    _merge_on_shared_period,
    _resolve_column_name,
    _select_contiguous_labeled_columns,
    build_relational_join_enrichment_report,
    build_cycle_detection_report,
    build_region_share_cost_report,
    build_financial_dashboard_report,
    build_multi_source_utilisation_summary_report,
    build_two_dimension_mean_count_summary_report,
    build_room_format_report,
    build_multi_source_group_comparison_report,
    build_cash_flow_efficiency_report,
    build_dependency_schedule,
    build_inventory_eoq_report,
    build_time_series_aggregation_report,
    load_all_tables,
)


def _write_workbook(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_identifier_format_report_is_not_room_specific(tmp_path: Path):
    workbook_path = tmp_path / "venues.xlsx"
    _write_workbook(
        workbook_path,
        [
            ["Venue Code", "Status"],
            ["A10", "Open"],
            ["A 10", "Closed"],
            ["B20", "Open"],
        ],
    )
    world = load_world([str(workbook_path)], output_path=str(tmp_path / "output.xlsx"))

    report = build_room_format_report(world)

    assert "Venue Code" in report["answer"]
    assert "A10" in report["answer"]
    assert "A 10" in report["answer"]


def test_identifier_format_report_groups_punctuation_case_and_spacing_variants(tmp_path: Path):
    workbook_path = tmp_path / "products.xlsx"
    _write_workbook(
        workbook_path,
        [
            ["ProductCode", "Name"],
            ["A-101", "Alpha"],
            ["a101", "Alpha"],
            ["A 101", "Alpha"],
            ["B-202", "Beta"],
            ["b202", "Beta"],
        ],
    )
    world = load_world([str(workbook_path)], output_path=str(tmp_path / "output.xlsx"))

    report = build_room_format_report(world)

    assert "A101" in report["answer"]
    assert "B202" in report["answer"]
    assert "A-101" in report["answer"]
    assert "a101" in report["answer"]
    assert "B-202" in report["answer"]


def test_select_contiguous_labeled_columns_stops_at_blank_or_stop_marker():
    header_row = ["", "Year", "North", "South", "West", "In %", "Ignore"]

    columns = _select_contiguous_labeled_columns(
        header_row,
        start_col_idx=2,
        stop_markers=("in %",),
    )

    assert columns == [(2, "North"), (3, "South"), (4, "West")]


def test_extract_period_records_reads_contiguous_period_block():
    rows = [
        ["Region", "North", "South"],
        ["2020", 10, 20],
        ["2021", 15, 25],
        ["", "", ""],
        ["2022", 30, 40],
    ]

    records = _extract_period_records(
        rows,
        period_col_idx=0,
        labeled_columns=[(1, "North"), (2, "South")],
        is_period_value=lambda value: str(value).isdigit(),
        period_transform=int,
    )

    assert records == [
        {"Year": 2020, "North": 10, "South": 20},
        {"Year": 2021, "North": 15, "South": 25},
    ]


def test_merge_on_shared_period_keeps_only_overlap():
    left = [
        {"Time": "Q1 2024", "North": 12.5},
        {"Time": "Q2 2024", "North": 13.0},
    ]
    right = [
        {"Time": "Q2 2024", "Shipment": 100.0},
        {"Time": "Q3 2024", "Shipment": 110.0},
    ]

    overlap_df = _merge_on_shared_period(left, right, period_col="Time")

    assert overlap_df.to_dict(orient="records") == [
        {"Time": "Q2 2024", "North": 13.0, "Shipment": 100.0},
    ]


def test_build_grouped_assignment_join_preserves_schedule_rows():
    assignment_df = pd.DataFrame(
        {
            "Assigned Tutor": ["Alice", "Alice", "Bob"],
            "Student": ["S1", "S2", "S3"],
        }
    )
    schedule_df = pd.DataFrame(
        {
            "Tutor Name": ["Alice", "Bob", "Cara"],
            "Time Slot": ["Mon 9", "Tue 10", "Wed 11"],
            "Room": ["R1", "R2", "R3"],
        }
    )

    output_df = _build_grouped_assignment_join(
        assignment_df=assignment_df,
        assignment_col="Assigned Tutor",
        entity_col="Student",
        schedule_df=schedule_df,
        resource_col="Tutor Name",
        schedule_cols=["Time Slot", "Room"],
    )

    assert output_df.to_dict(orient="records") == [
        {"Tutor Name": "Alice", "Time Slot": "Mon 9", "Room": "R1", "Student": "S1, S2"},
        {"Tutor Name": "Bob", "Time Slot": "Tue 10", "Room": "R2", "Student": "S3"},
        {"Tutor Name": "Cara", "Time Slot": "Wed 11", "Room": "R3", "Student": ""},
    ]


def test_find_first_period_cell_returns_row_and_column():
    rows = [
        ["", "", ""],
        ["Overview", "North", "South"],
        ["2021", 10, 20],
    ]

    assert _find_first_period_cell(rows, is_period_value=lambda value: str(value).isdigit()) == (2, 0)


def test_build_weighted_period_output_scales_value_columns():
    overlap_df = pd.DataFrame(
        {
            "Time": ["Q1 2024", "Q2 2024"],
            "Vivo": [10.0, 20.0],
            "Samsung": [30.0, 40.0],
            "Shipment": [100.0, 50.0],
        }
    )

    output_df = _build_weighted_period_output(
        overlap_df,
        period_col="Time",
        value_columns=["Vivo", "Samsung"],
        weight_col="Shipment",
        output_period_col="Year",
        output_label_template="{name} (Unit shipment)",
    )

    assert output_df.to_dict(orient="records") == [
        {"Year": "Q1 2024", "Vivo (Unit shipment)": 10.0, "Samsung (Unit shipment)": 30.0},
        {"Year": "Q2 2024", "Vivo (Unit shipment)": 10.0, "Samsung (Unit shipment)": 20.0},
    ]


def test_load_all_tables_prefers_structured_data_sheet_over_overview(tmp_path: Path):
    workbook_path = tmp_path / "market_share.xlsx"
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Statistic as Excel data file"])
    data_sheet = workbook.create_sheet("Data")
    data_sheet.append(["Year", "Vivo", "Samsung"])
    data_sheet.append(["2024", 10, 20])
    workbook.save(workbook_path)

    world = load_world([str(workbook_path)], output_path=str(tmp_path / "output.xlsx"))
    tables = load_all_tables(world, require_primary_key=False, stop_at_note_row=False)

    assert len(tables) == 1
    assert tables[0]["sheet_name"] == "Data"
    assert tables[0]["header"] == ["Year", "Vivo", "Samsung"]


def test_load_all_tables_promotes_period_header_row_and_infers_time_column(tmp_path: Path):
    workbook_path = tmp_path / "market_share.xlsx"
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Statistic as Excel data file"])
    data_sheet = workbook.create_sheet("Data")
    data_sheet.append([None])
    data_sheet.append([None])
    data_sheet.append([None, "Smartphone market share held by vendors in India Q1 2017-Q3 2025"])
    data_sheet.append([None, "Smartphone market share held by vendors in India from 1st quarter 2017 to 3rd quarter 2025"])
    data_sheet.append([None, None, "Vivo", "Samsung", "Xiaomi", "Others", None])
    data_sheet.append([None, "Q1 2017", 11, 28, 14, 47, "in %"])
    data_sheet.append([None, "Q2 2017", 13, 24, 17, 46, "in %"])
    workbook.save(workbook_path)

    world = load_world([str(workbook_path)], output_path=str(tmp_path / "output.xlsx"))
    tables = load_all_tables(world, require_primary_key=False, stop_at_note_row=False)

    assert len(tables) == 1
    assert tables[0]["header"] == ["Time", "Vivo", "Samsung", "Xiaomi", "Others"]
    assert tables[0]["rows"] == [
        ["Q1 2017", 11, 28, 14, 47],
        ["Q2 2017", 13, 24, 17, 46],
    ]


def test_namespace_exposes_financial_helpers(tmp_path: Path):
    workbook_path = tmp_path / "params.xlsx"
    _write_workbook(
        workbook_path,
        [
            ["Parameter", "Value"],
            ["Annual_Demand_units", 2400],
        ],
    )
    world = load_world([str(workbook_path)], output_path=str(tmp_path / "output.xlsx"))
    namespace = SpreadsheetNamespace(world).build()

    assert callable(namespace.build_inventory_eoq_report)
    assert callable(namespace.build_financial_dashboard_report)
    assert callable(namespace.build_region_share_cost_report)
    assert callable(namespace.build_two_dimension_mean_count_summary_report)
    assert callable(namespace.build_multi_source_group_comparison_report)
    assert callable(namespace.build_multi_source_utilisation_summary_report)


def test_build_inventory_eoq_report_accepts_benchmark_parameter_aliases(tmp_path: Path):
    workbook_path = tmp_path / "inventory.xlsx"
    _write_workbook(
        workbook_path,
        [
            ["Parameter", "Value"],
            ["Product", "A4 Copy Paper (Reams)"],
            ["Annual_Demand_units", 2400],
            ["Order_Cost_USD", 25.00],
            ["Holding_Cost_per_unit_USD", 1.50],
            ["Unit_Cost_USD", 8.00],
            ["Lead_Time_days", 7],
            ["Working_Days_per_year", 250],
        ],
    )
    world = load_world([str(workbook_path)], output_path=str(tmp_path / "output.xlsx"))

    report = build_inventory_eoq_report(world)

    assert report["output_df"].iloc[0, 0] == "EOQ"
    assert report["base_metrics"]["EOQ"] > 0


def test_build_region_share_cost_report_accepts_generic_benchmark_labels(tmp_path: Path):
    prevalence_path = tmp_path / "prevalence.xlsx"
    population_path = tmp_path / "population.xlsx"
    expenditure_path = tmp_path / "expenditure.xlsx"

    _write_workbook(
        prevalence_path,
        [
            ["Country", "Region", "ObesityPrevalence_pct"],
            ["A", "North", 12.0],
        ],
    )
    _write_workbook(
        population_path,
        [
            ["Region", "Obese_Population_millions"],
            ["North", 20.0],
            ["South", 30.0],
        ],
    )
    _write_workbook(
        expenditure_path,
        [
            ["Region", "Healthcare_Expenditure_billion_USD"],
            ["North", 5.0],
            ["South", 9.0],
        ],
    )

    world = load_world(
        [str(prevalence_path), str(population_path), str(expenditure_path)],
        output_path=str(tmp_path / "output.xlsx"),
    )

    report = build_region_share_cost_report(world)

    assert report["output_df"].columns.tolist() == [
        "Region",
        "Obese_Pop_millions",
        "Global_Share_pct",
        "Expenditure_BillionUSD",
        "Avg_Exp_per_Person_USD",
    ]
    assert report["output_df"]["Region"].tolist() == ["South", "North"]
    assert round(report["output_df"]["Global_Share_pct"].sum(), 6) == 100.0
    assert report["detail_data"][-1] == ["Total", 50.0, "100.00", None, None]


def test_build_two_dimension_mean_count_summary_report_accepts_compound_type_alias(tmp_path: Path):
    workbook_path = tmp_path / "reviews.xlsx"
    _write_workbook(
        workbook_path,
        [
            ["ReviewID", "Country", "HotelType", "Rating"],
            ["R1", "UK", "Budget", 4.0],
            ["R2", "UK", "Budget", 5.0],
            ["R3", "France", "Luxury", 3.5],
        ],
    )
    world = load_world([str(workbook_path)], output_path=str(tmp_path / "output.xlsx"))

    report = build_two_dimension_mean_count_summary_report(world)

    assert report["output_df"].columns.tolist() == ["Country", "HotelType", "avg_rating", "num_reviews"]
    assert report["output_df"]["num_reviews"].tolist() == [1, 2]


def test_build_multi_source_group_comparison_report_accepts_storeid_and_weeklysales_aliases(tmp_path: Path):
    sales_path = tmp_path / "sales.xlsx"
    store_path = tmp_path / "stores.xlsx"
    _write_workbook(
        sales_path,
        [
            ["StoreID", "Week", "IsHoliday", "WeeklySales_USD", "Temperature_F", "FuelPrice_USD"],
            [1, "2024-03-01", True, 1000, 70, 3.5],
            [1, "2024-03-08", False, 800, 68, 3.4],
            [2, "2024-03-01", True, 1200, 72, 3.6],
        ],
    )
    _write_workbook(
        store_path,
        [
            ["StoreID", "StoreType", "Size_sqft", "Region"],
            [1, "A", 10000, "North"],
            [2, "B", 15000, "South"],
        ],
    )
    world = load_world([str(sales_path), str(store_path)], output_path=str(tmp_path / "output.xlsx"))

    report = build_multi_source_group_comparison_report(world)

    assert report["avg_by_type_df"].columns.tolist() == ["StoreType", "WeeklySales_USD", "Temperature_F", "FuelPrice_USD"]
    assert report["holiday_df"].columns.tolist() == ["Feature", "Holiday Average", "Non-Holiday Average", "Difference"]


def test_build_multi_source_utilisation_summary_report_accepts_department_based_school_tables(tmp_path: Path):
    staff_path = tmp_path / "staff.xlsx"
    schedule_path = tmp_path / "schedule.xlsx"
    service_path = tmp_path / "service.xlsx"
    booking_path = tmp_path / "booking.xlsx"
    _write_workbook(
        staff_path,
        [
            ["StaffID", "Name", "Department", "ContractHours_per_week"],
            ["S1", "A", "Arts", 40],
            ["S2", "B", "Sports", 40],
        ],
    )
    _write_workbook(
        schedule_path,
        [
            ["StaffID", "Week", "ActualHours"],
            ["S1", "2024-W10", 38],
            ["S2", "2024-W10", 39],
        ],
    )
    _write_workbook(
        service_path,
        [
            ["Department", "VisitsPerMonth", "MaxCapacity_per_month"],
            ["Arts", 98, 100],
            ["Sports", 95, 100],
        ],
    )
    _write_workbook(
        booking_path,
        [
            ["BookingID", "StudentID", "Department", "BooksOrEquipment"],
            ["B1", "U1", "Arts", "Book"],
        ],
    )
    world = load_world(
        [str(staff_path), str(schedule_path), str(service_path), str(booking_path)],
        output_path=str(tmp_path / "output.xlsx"),
    )

    report = build_multi_source_utilisation_summary_report(world)

    assert "Department" in report["output_df"].columns
    assert "Service Utilisation (%)" in report["output_df"].columns
    assert "Staff Utilisation (%)" in report["output_df"].columns
    assert report["highlight_rows"]


def test_build_multi_source_utilisation_summary_report_accepts_university_section_tables(tmp_path: Path):
    students_path = tmp_path / "students.csv"
    courses_path = tmp_path / "courses.csv"
    sections_path = tmp_path / "sections.csv"
    enrollments_path = tmp_path / "enrollments.csv"
    rooms_path = tmp_path / "rooms.csv"

    pd.DataFrame(
        {
            "StudentID": ["S1", "S2", "S3"],
            "StudentName": ["A", "B", "C"],
            "School": ["ENG", "ENG", "SCI"],
        }
    ).to_csv(students_path, index=False)
    pd.DataFrame(
        {
            "CourseID": ["C1", "C2"],
            "CourseName": ["Math", "Physics"],
        }
    ).to_csv(courses_path, index=False)
    pd.DataFrame(
        {
            "SectionID": ["SEC1", "SEC2"],
            "CourseID": ["C1", "C2"],
            "Instructor": ["Ivy", "Noah"],
            "RoomID": ["R1", "R2"],
            "Term": ["T1", "T1"],
            "Capacity": [20, 10],
            "ScheduledHours": [3, 2],
        }
    ).to_csv(sections_path, index=False)
    pd.DataFrame(
        {
            "EnrollID": ["E1", "E2", "E3", "E4"],
            "StudentID": ["S1", "S2", "S3", "S1"],
            "SectionID": ["SEC1", "SEC1", "SEC2", "SEC2"],
            "EnrollStatus": ["Registered", "Waitlisted", "Registered", "Registered"],
        }
    ).to_csv(enrollments_path, index=False)
    pd.DataFrame(
        {
            "RoomID": ["R1", "R2"],
            "Building": ["North", "South"],
            "Capacity": [20, 10],
        }
    ).to_csv(rooms_path, index=False)

    world = load_world(
        [
            str(students_path),
            str(courses_path),
            str(sections_path),
            str(enrollments_path),
            str(rooms_path),
        ],
        output_path=str(tmp_path / "output.xlsx"),
    )

    report = build_multi_source_utilisation_summary_report(world)

    assert report["sheet_outputs"].keys() == {"Section_Utilisation", "Instructor_Load", "Room_Utilisation"}
    assert report["sheet_outputs"]["Section_Utilisation"]["SectionID"].tolist() == ["SEC1", "SEC2"]
    assert report["sheet_outputs"]["Instructor_Load"]["Scheduled_Hours"].tolist() == [3, 2]
    assert report["sheet_outputs"]["Room_Utilisation"]["Building"].tolist() == ["North", "South"]


def test_build_financial_dashboard_report_accepts_metric_target_aliases(tmp_path: Path):
    pnl_path = tmp_path / "pnl.csv"
    sales_path = tmp_path / "sales.csv"
    target_path = tmp_path / "targets.csv"

    pd.DataFrame(
        {
            "Month": ["April", "May", "June"],
            "Revenue_USD": [75000, 82000, 79000],
            "COGS_USD": [31500, 34440, 33180],
            "OperatingExpenses_USD": [28000, 30000, 29000],
            "InterestPaid_USD": [1500, 1500, 1500],
        }
    ).to_csv(pnl_path, index=False)
    pd.DataFrame(
        {
            "Month": ["April", "May", "June"],
            "NewCustomers": [150, 180, 165],
            "MarketingSpend_USD": [8000, 9500, 8800],
        }
    ).to_csv(sales_path, index=False)
    pd.DataFrame(
        {
            "Metric": ["Gross_Profit", "Net_Profit", "Gross_Profit_Margin", "Net_Profit_Margin", "CAC", "MER"],
            "Target": [130000.0, 40000.0, 0.55, 0.18, 55.0, 9.0],
        }
    ).to_csv(target_path, index=False)

    world = load_world(
        [str(pnl_path), str(sales_path), str(target_path)],
        output_path=str(tmp_path / "output.xlsx"),
    )

    report = build_financial_dashboard_report(world)

    assert list(report["output_df"].columns) == [
        "Metric",
        "Q2_Actual",
        "Target",
        "Variance",
        "Assessment",
    ]
    assert len(report["output_df"]) == 6
    assert report["output_df"].to_dict(orient="records") == [
        {
            "Metric": "Gross_Profit",
            "Q2_Actual": 136880.0,
            "Target": 130000.0,
            "Variance": 6880.0,
            "Assessment": "Exceeding Target",
        },
        {
            "Metric": "Net_Profit",
            "Q2_Actual": 45380.0,
            "Target": 40000.0,
            "Variance": 5380.0,
            "Assessment": "Exceeding Target",
        },
        {
            "Metric": "Gross_Profit_Margin",
            "Q2_Actual": 0.58,
            "Target": 0.55,
            "Variance": 0.03,
            "Assessment": "Exceeding Target",
        },
        {
            "Metric": "Net_Profit_Margin",
            "Q2_Actual": 0.1923,
            "Target": 0.18,
            "Variance": 0.0123,
            "Assessment": "Exceeding Target",
        },
        {
            "Metric": "CAC",
            "Q2_Actual": 53.13,
            "Target": 55.0,
            "Variance": -1.87,
            "Assessment": "Exceeding Target (lower is better)",
        },
        {
            "Metric": "Marketing_Efficiency_Ratio",
            "Q2_Actual": 8.97,
            "Target": 9.0,
            "Variance": -0.03,
            "Assessment": "Below Target",
        },
    ]
    assert report["detail_data"][0] == ["Metric", "Q2_Actual", "Target", "Variance", "Assessment"]


def test_build_cash_flow_efficiency_report_accepts_structured_year_table(tmp_path: Path):
    csv_path = tmp_path / "cashflow.csv"
    pd.DataFrame(
        {
            "Year": [2017, 2018, 2019],
            "NetIncome_M_USD": [-1961, -976, -862],
            "OperatingCashFlow_M_USD": [477, 2098, 2405],
            "CapEx_M_USD": [1086, 2101, 1437],
        }
    ).to_csv(csv_path, index=False)

    world = load_world([str(csv_path)], output_path=str(tmp_path / "output.xlsx"))

    report = build_cash_flow_efficiency_report(world)

    assert list(report["output_df"].columns) == [
        "Year",
        "Net Income",
        "Operating Cash Flow",
        "Capital Expenditures",
        "OCF/Net Income",
        "Free Cash Flow",
        "Anomaly Note",
    ]
    assert report["output_df"].iloc[-1]["Year"] == "Total"
    assert report["output_df"].iloc[0]["Anomaly Note"] == "Negative net income year"


def test_load_all_tables_synthesizes_time_value_headers_for_period_series_sheet(tmp_path: Path):
    workbook_path = tmp_path / "shipments.xlsx"
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Statistic as Excel data file"])
    data_sheet = workbook.create_sheet("Data")
    data_sheet.append([None])
    data_sheet.append([None])
    data_sheet.append([None, "India smartphone unit shipments Q2 2012-Q2 2025"])
    data_sheet.append([None, "Total number of smartphone unit shipments in India from 2nd quarter of 2012 to 2nd quarter of 2025"])
    data_sheet.append([None])
    data_sheet.append([None, "Q2 2012", 4])
    data_sheet.append([None, "Q3 2012", 4])
    data_sheet.append([None, "Q4 2012", 5])
    workbook.save(workbook_path)

    world = load_world([str(workbook_path)], output_path=str(tmp_path / "output.xlsx"))
    tables = load_all_tables(world, require_primary_key=False, stop_at_note_row=False)

    assert len(tables) == 1
    assert tables[0]["header"] == ["Time", "Value"]
    assert tables[0]["rows"] == [
        ["Q2 2012", 4],
        ["Q3 2012", 4],
        ["Q4 2012", 5],
    ]


def test_build_time_series_aggregation_report_infers_year_column(tmp_path: Path):
    workbook_path = tmp_path / "admissions.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["Year", "TotalAdmissions"])
    sheet.append([2010, 100])
    sheet.append([2011, 130])
    sheet.append([2012, 125])
    workbook.save(workbook_path)

    world = load_world([str(workbook_path)], output_path=str(tmp_path / "output.xlsx"))
    report = build_time_series_aggregation_report(
        world,
        date_col="Date",
        value_col=None,
        period="year",
        aggregate="sum",
        sort_desc=False,
    )

    assert report["output_df"].to_dict(orient="records") == [
        {"Period": "2010", "Total TotalAdmissions": 100},
        {"Period": "2011", "Total TotalAdmissions": 130},
        {"Period": "2012", "Total TotalAdmissions": 125},
    ]


def test_extract_sheet_table_flattens_two_row_matrix_headers():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Region", "Electronics", None, None, None, "Clothing", None, None, None])
    sheet.append([None, "Q1", "Q2", "Q3", "Q4", "Q1", "Q2", "Q3", "Q4"])
    sheet.append(["North", 10, 20, 30, 40, 1, 2, 3, 4])
    sheet.append(["South", 11, 21, 31, 41, 5, 6, 7, 8])

    extracted = extract_sheet_table(
        sheet,
        "A1:I10",
        drop_blank_rows=True,
        drop_empty_primary_key=False,
        stop_at_note_row=True,
    )

    assert extracted["header"] == [
        "Region",
        "Electronics_Q1",
        "Electronics_Q2",
        "Electronics_Q3",
        "Electronics_Q4",
        "Clothing_Q1",
        "Clothing_Q2",
        "Clothing_Q3",
        "Clothing_Q4",
    ]
    assert extracted["rows"] == [
        ["North", 10, 20, 30, 40, 1, 2, 3, 4],
        ["South", 11, 21, 31, 41, 5, 6, 7, 8],
    ]


def test_load_all_tables_flattens_real_sales_matrix_benchmark_headers():
    dataset_path = (
        Path(__file__).resolve().parents[2]
        / "dataset"
        / "SystemEvaluationBenchmark"
        / "median_clean_cases"
        / "Case01"
        / "tc01_input01.csv"
    )
    world = load_world([str(dataset_path)], output_path=str(dataset_path.parent / "output.xlsx"))

    tables = load_all_tables(world, require_primary_key=False, stop_at_note_row=True)

    assert len(tables) == 1
    assert tables[0]["header"][:5] == [
        "Region",
        "Electronics_Q1",
        "Electronics_Q2",
        "Electronics_Q3",
        "Electronics_Q4",
    ]
    assert "Home_Appliances_Q4" in tables[0]["header"]


def test_load_all_tables_exposes_basename_friendly_file_alias(tmp_path: Path):
    workbook_path = tmp_path / "students.csv.xlsx"
    _write_workbook(
        workbook_path,
        [
            ["StudentID", "StudentName"],
            ["S1", "Alice"],
        ],
    )

    world = load_world([str(workbook_path)], output_path=str(tmp_path / "output.xlsx"))
    tables = load_all_tables(world, require_primary_key=False, stop_at_note_row=False)

    assert len(tables) == 1
    assert tables[0]["file"] == "students.csv.xlsx"
    assert tables[0]["file_name"] == "students.csv.xlsx"
    assert tables[0]["file_path"] == str(workbook_path)


def test_relational_join_enrichment_handles_bridge_join_large_revenue_case():
    dataset_root = (
        Path(__file__).resolve().parents[2]
        / "dataset"
        / "SystemEvaluationBenchmark"
        / "large_clean_cases"
        / "Case01"
    )
    input_paths = [
        str((dataset_root / f"tc01_input0{i}.csv").resolve())
        for i in range(1, 5)
    ]
    world = load_world(input_paths, output_path=str((dataset_root / "output.xlsx").resolve()))

    report = build_relational_join_enrichment_report(world, how="left")

    assert len(report["output_df"]) == 120
    assert {"StoreID", "ProductID", "Month", "Region", "Category", "UnitPrice", "RevenueTarget"} <= set(
        report["output_df"].columns
    )
    assert report["metadata"]["join_path"] == [
        "tc01_input01.csv on StoreID",
        "tc01_input02.csv on ProductID",
        "tc01_input04.csv on StoreID, Month, Category",
    ]


def test_resolve_column_name_accepts_camel_case_and_spaceless_variants():
    columns = ["TaskID", "TaskName", "Duration_hours", "DependsOn"]

    assert _resolve_column_name(columns, "Task ID") == "TaskID"
    assert _resolve_column_name(columns, "Task Name") == "TaskName"
    assert _resolve_column_name(columns, "Duration (hours)") == "Duration_hours"
    assert _resolve_column_name(columns, "Depends on") == "DependsOn"


def test_build_dependency_schedule_accepts_compact_task_headers():
    task_df = pd.DataFrame(
        {
            "TaskID": ["T1", "T2"],
            "TaskName": ["Design", "Build"],
            "Priority": ["High", "Medium"],
            "Duration_hours": [2, 3],
        }
    )
    dependency_df = pd.DataFrame(
        {
            "TaskID": ["T1", "T2"],
            "DependsOn": ["", "T1"],
        }
    )

    report = build_dependency_schedule(task_df, dependency_df, start_time="09:00")

    assert report["scheduled_task_ids"] == ["T1", "T2"]
    assert report["detail_data"][1] == ["T1", "Design", "High", "09:00", "11:00"]
    assert report["detail_data"][2] == ["T2", "Build", "Medium", "11:00", "14:00"]


def test_build_cycle_detection_report_accepts_single_dataframe():
    edges_df = pd.DataFrame(
        {
            "Source": ["A", "B", "C"],
            "Target": ["B", "C", "A"],
        }
    )

    report = build_cycle_detection_report(edges_df, from_col="Source", to_col="Target")

    assert report["contains_cycle"] is True
    assert report["detail_data"][0] == ["GraphID", "Contains_Cycle (True / False)"]
    assert report["detail_data"][1] == ["graph_1", True]


def test_output_writer_normalizes_helper_result_dict_to_scalar():
    writer = ExcelOutputWriter(workbook=None, excel_path="input.xlsx", output_path=None, temp_files=[])

    normalized = writer._normalize_output_value(
        {
            "contains_cycle": True,
            "detail_data": [["GraphID", "Contains_Cycle"], ["graph_1", True]],
        }
    )

    assert normalized is True
