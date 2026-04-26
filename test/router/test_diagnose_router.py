import os
import sys
from pathlib import Path

import pandas as pd

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "../.."))

from openpyxl import Workbook, load_workbook

from backend.router.diagnose_router import DiagnoseRouter


def test_self_reporting_scan_task_skips_router_issue_generation_for_generic_identifier_formats():
    question = (
        "Check whether the venue codes are inconsistent, for example A10 versus A 10, "
        "and report the issue."
    )

    assert DiagnoseRouter._is_self_reporting_issue_task(question) is True


def test_real_task3_messy_header_only_triggers_period_clarification():
    dataset_path = (
        Path(__file__).resolve().parents[2]
        / "dataset"
        / "DevelopmentBenchmark"
        / "Task03"
        / "tc03_input01.xlsx"
    )
    workbook = load_workbook(dataset_path)
    router = DiagnoseRouter(client=None, deployment="offline-test", prompt_profile="offline_strict")
    question = (
        "Here is a table showing internet penetration rates from 2009 to 2024. "
        "Please calculate the average internet penetration rate for each region over the years 2020–2024. "
        "Identify the region with the fastest growth rate and sort the regions by growth rate. "
        "Also provide a line chart where each region is represented by a different color."
    )

    decision = router.decide(question, "", {str(dataset_path): workbook})

    issue_types = [issue.get("issue_type") for issue in decision.issues]
    assert decision.should_diagnose is True
    assert issue_types == ["missing_period_endpoint"]


def test_extract_requested_years_prefers_last_explicit_range():
    question = (
        "The source table covers 2009 to 2024. "
        "Please calculate the average for the years 2020-2024 and rank the regions."
    )

    assert DiagnoseRouter._extract_requested_years(question) == [2020, 2021, 2022, 2023, 2024]


def test_multi_header_matrix_task_does_not_trigger_missing_value_qa():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Region", "Electronics", None, None, None, "Clothing", None, None, None])
    sheet.append([None, "Q1", "Q2", "Q3", "Q4", "Q1", "Q2", "Q3", "Q4"])
    sheet.append(["North", 10, 20, 30, 40, 1, 2, 3, 4])
    sheet.append(["South", 11, 21, 31, 41, 5, 6, 7, 8])

    router = DiagnoseRouter(client=None, deployment="offline-test", prompt_profile="offline_strict")
    question = (
        "Here is a sales matrix with two header rows. "
        "Forward-fill the blank category headers so every quarterly column is associated with the correct category, "
        "then compute annual totals and a Q4 leaderboard."
    )

    decision = router.decide(question, "", {"sales.csv": workbook})

    issue_types = [issue.get("issue_type") for issue in decision.issues]
    assert "missing_value" not in issue_types


def test_clean_schedule_task_skips_diagnose_without_data_evidence():
    workbook = Workbook()
    task_sheet = workbook.active
    task_sheet.title = "tc04_input01"
    task_sheet.append(["TaskID", "TaskName", "Duration_hours", "Priority"])
    task_sheet.append(["T1", "Setup", 2, "High"])
    task_sheet.append(["T2", "Collect", 4, "High"])
    dep_sheet = workbook.create_sheet("tc04_input02")
    dep_sheet.append(["TaskID", "DependsOn"])
    dep_sheet.append(["T2", "T1"])

    router = DiagnoseRouter(client=None, deployment="offline-test", prompt_profile="offline_strict")
    question = (
        "Here is a table of sprint tasks with durations and priorities, and another table of task dependencies. "
        "All work starts at 09:00. Schedule the tasks respecting dependencies and output a new Excel sheet "
        "with columns TaskID, TaskName, Priority, StartTime, EndTime."
    )

    decision = router.decide(question, "", {"schedule.csv": workbook})

    assert decision.should_diagnose is False
    assert decision.issues == []


def test_clean_market_share_task_skips_diagnose_without_data_evidence():
    workbook = Workbook()
    units_sheet = workbook.active
    units_sheet.title = "tc05_input01"
    units_sheet.append(["Quarter", "Total_EV_Units_thousand"])
    units_sheet.append(["2021Q1", 12.0])
    units_sheet.append(["2021Q2", 15.3])
    share_sheet = workbook.create_sheet("tc05_input02")
    share_sheet.append(["Quarter", "Tata", "MG", "Hyundai", "Mahindra", "Kia", "Others"])
    share_sheet.append(["2021Q1", 72.5, 8.3, 6.2, 3.1, 4.8, 5.1])
    share_sheet.append(["2021Q2", 71.2, 8.9, 6.8, 3.2, 4.5, 5.4])

    router = DiagnoseRouter(client=None, deployment="offline-test", prompt_profile="offline_strict")
    question = (
        "Here are two tables. One gives the total EV units sold in India by quarter, and the other gives market "
        "share by brand. Find the overlapping time period, then estimate the number of EVs sold for each brand."
    )

    decision = router.decide(question, "", {"market.csv": workbook})

    assert decision.should_diagnose is False
    assert decision.issues == []


def test_enrollment_relationship_table_does_not_trigger_duplicate_conflict_qa():
    workbook = Workbook()
    student_sheet = workbook.active
    student_sheet.title = "students"
    student_sheet.append(["StudentID", "StudentName"])
    student_sheet.append(["ST001", "Alice"])
    student_sheet.append(["ST002", "Bob"])

    enrollment_sheet = workbook.create_sheet("enrollments")
    enrollment_sheet.append(["StudentID", "CourseCode"])
    enrollment_sheet.append(["ST001", "CS101"])
    enrollment_sheet.append(["ST001", "MA102"])
    enrollment_sheet.append(["ST002", "CS101"])

    router = DiagnoseRouter(client=None, deployment="offline-test", prompt_profile="offline_strict")
    question = (
        "Merge the student table with the enrollment table and list, for each course, the enrolled students."
    )

    decision = router.decide(question, "", {"enrollment.csv": workbook})

    issue_types = [issue.get("issue_type") for issue in decision.issues]
    assert "duplicate_conflicting_rows" not in issue_types


def test_movie_genre_membership_table_does_not_trigger_duplicate_conflict_qa():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "genres"
    sheet.append(["MovieID", "Genre"])
    sheet.append([1, "Drama"])
    sheet.append([1, "Sci-Fi"])
    sheet.append([2, "Romance"])
    sheet.append([2, "Drama"])

    router = DiagnoseRouter(client=None, deployment="offline-test", prompt_profile="offline_strict")
    question = "Merge the movie files on MovieID and output one combined table."

    decision = router.decide(question, "", {"movies.csv": workbook})

    issue_types = [issue.get("issue_type") for issue in decision.issues]
    assert "duplicate_conflicting_rows" not in issue_types


def test_large_integer_score_column_does_not_trigger_semantic_anomaly():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "scores"
    sheet.append(["MovieID", "AudienceScore", "CriticScore"])
    sheet.append([1, 9.1, 56019])
    sheet.append([2, 8.3, 110058])
    sheet.append([3, 8.3, 331527])
    sheet.append([4, 8.5, 452967])
    sheet.append([5, 8.2, 108485])

    router = DiagnoseRouter(client=None, deployment="offline-test", prompt_profile="offline_strict")
    question = "Merge the movie files on MovieID and output one combined table."

    decision = router.decide(question, "", {"scores.csv": workbook})

    issue_types = [issue.get("issue_type") for issue in decision.issues]
    assert "semantic_anomaly" not in issue_types


def test_large_clean_revenue_case_does_not_trigger_missing_period_endpoint():
    dataset_root = (
        Path(__file__).resolve().parents[2]
        / "dataset"
        / "SystemEvaluationBenchmark"
        / "large_clean_cases"
        / "Case01"
    )
    workbook_view = {
        path.name: pd.read_csv(path)
        for path in sorted(dataset_root.glob("tc01_input*.csv"))
    }

    router = DiagnoseRouter(client=None, deployment="offline-test", prompt_profile="offline_strict")
    question = (
        "You have four clean tables: stores, products, monthly sales, and monthly targets. "
        "Join the tables using StoreID and ProductID, keep only the H2 2024 months (2024-07 to 2024-12), "
        "compute net revenue as UnitsSold × UnitPrice × (1 - DiscountPct), aggregate the results by Region "
        "and Category, and compare each aggregate to the summed H2 revenue target. Then create a Top 10 store "
        "leaderboard by H2 revenue and a category summary with revenue, gross profit, and discounted-unit share. "
        "Output everything into a new Excel workbook with three sheets: Region_Category_H2, Store_Leaderboard, "
        "and Category_Summary."
    )

    decision = router.decide(question, "", workbook_view)

    issue_types = [issue.get("issue_type") for issue in decision.issues]
    assert "missing_period_endpoint" not in issue_types
    assert decision.should_diagnose is False
