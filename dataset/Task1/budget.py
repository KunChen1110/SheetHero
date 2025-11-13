import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === 1. 读取两张表 ===
# 请根据你的文件名修改路径
df1 = pd.read_excel("input1.xlsx")
df2 = pd.read_excel("input2.xlsx")

# === 2. 合并数据 ===
df = pd.concat([df1, df2], ignore_index=True)

# === 3. 转换日期格式 ===
df["Date"] = pd.to_datetime(df["Date"])

# === 4. 只保留 2025 年 11 月的数据 ===
df_nov = df[df["Date"].dt.month == 11]

# === 5. 计算每日总支出 ===
daily_total = df_nov.groupby(df_nov["Date"].dt.date)["Daily Spending (£)"].sum().reset_index()
daily_total.columns = ["Date", "Total Spending (£)"]

# === 6. 计算总支出与平均支出 ===
total_spending = daily_total["Total Spending (£)"].sum()
average_spending = daily_total["Total Spending (£)"].mean()

print(f"✅ 总支出（11月）: £{total_spending:.2f}")
print(f"✅ 平均每日支出: £{average_spending:.2f}")

# === 7. 合并回原始数据（可选）===
df_nov["Date"] = pd.to_datetime(df_nov["Date"], errors="coerce")
daily_total["Date"] = pd.to_datetime(daily_total["Date"], errors="coerce")
merged = df_nov.merge(daily_total, on="Date", how="left")

# === 8. 输出到 Excel ===
output_file = "merged_spending_november.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    merged.to_excel(writer, index=False, sheet_name="Merged Data")
    daily_total.to_excel(writer, index=False, sheet_name="Daily Summary")

# === 9. 标红最高支出日 ===
wb = load_workbook(output_file)
ws = wb["Daily Summary"]

max_spending = daily_total["Total Spending (£)"].max()
fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for row in range(2, ws.max_row + 1):
    if ws.cell(row, 2).value == max_spending:
        for col in range(1, 3):  # 标红整行
            ws.cell(row, col).fill = fill

wb.save(output_file)
print("🎯 结果已保存到 merged_spending_november.xlsx，最高支出日已标红。")
