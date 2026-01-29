import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Read 2 files
df1 = pd.read_excel("tc01_input01.xlsx")
df2 = pd.read_excel("tc01_input02.xlsx")

# === concat the data ===
df = pd.concat([df1, df2], ignore_index=True)

# === change the date ===
df["Date"] = pd.to_datetime(df["Date"])

# === find out november spendings   ==
df_nov = df[df["Date"].dt.month == 11]

# === calculate daily spendings ===
daily_total = df_nov.groupby(df_nov["Date"].dt.date)["Daily Spending (£)"].sum().reset_index()
daily_total.columns = ["Date", "Total Spending (£)"]

# === calculate the total and average  ===
total_spending = daily_total["Total Spending (£)"].sum()
average_spending = daily_total["Total Spending (£)"].sum() / 28

print(f"✅ Total spendings: £{total_spending:.2f}")
print(f"✅ Average daily spendings: £{average_spending:.2f}")

# === ）===
df_nov["Date"] = pd.to_datetime(df_nov["Date"], errors="coerce")
daily_total["Date"] = pd.to_datetime(daily_total["Date"], errors="coerce")
merged = df_nov.merge(daily_total, on="Date", how="left")

# === output to Excel ===
output_file = "output1.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    merged.to_excel(writer, index=False, sheet_name="Merged Data")
    daily_total.to_excel(writer, index=False, sheet_name="Daily Summary")

# === mark the highest sending day with ===
wb = load_workbook(output_file)
ws = wb["Daily Summary"]

max_spending = daily_total["Total Spending (£)"].max()
fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for row in range(2, ws.max_row + 1):
    if ws.cell(row, 2).value == max_spending:
        for col in range(1, 3):  
            ws.cell(row, col).fill = fill

wb.save(output_file)
print("🎯 Result successfully written to merged_spending_november.xlsx ")
