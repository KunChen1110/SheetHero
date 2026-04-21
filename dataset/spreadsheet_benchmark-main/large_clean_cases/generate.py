"""Generate a large clean spreadsheet benchmark dataset.

Each case uses 3-5 clean input tables. Every input table is designed to be in
the 500-1000 cell range and every task is composed of multiple atomic
operations such as join/filter/aggregate/rank/export.
"""

import csv
import json
import os
from collections import Counter, defaultdict
from openpyxl import Workbook


BASE = os.path.dirname(os.path.abspath(__file__))


def ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def write_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


def save_wb(path, wb):
    wb.save(path)


def pct(numerator, denominator):
    if not denominator:
        return 0.0
    return round((numerator / denominator) * 100, 2)


def weighted_avg(pairs):
    total_weight = sum(weight for _, weight in pairs)
    if not total_weight:
        return 0.0
    return round(sum(value * weight for value, weight in pairs) / total_weight, 2)


def top_n(items, key_idx, n=10, reverse=True):
    return sorted(items, key=lambda row: row[key_idx], reverse=reverse)[:n]


def build_case01():
    case_dir = os.path.join(BASE, "Case01")
    ensure_dir(case_dir)

    regions = ["North", "South", "East", "West", "Central", "Northeast", "Northwest"]
    cities = ["Leeds", "Bristol", "Manchester", "Liverpool", "York", "Oxford", "Bath", "Derby", "Exeter", "Norwich", "Reading", "Luton"]
    managers = ["Avery", "Blake", "Casey", "Dakota", "Emerson", "Finley", "Harper", "Jordan", "Kendall", "Logan", "Morgan", "Parker"]
    formats = ["Flagship", "Urban", "Suburban", "Mall"]
    categories = ["Electronics", "Home", "Sports", "Beauty", "Office"]
    launch = ["2023Q1", "2023Q2", "2023Q3", "2023Q4", "2024Q1"]

    stores = [["StoreID", "StoreName", "Region", "City", "Manager", "Format", "OpenYear"]]
    for i in range(84):
        stores.append([
            f"S{i+1:03d}",
            f"Store_{i+1:03d}",
            regions[i % len(regions)],
            cities[i % len(cities)],
            managers[i % len(managers)],
            formats[i % len(formats)],
            2014 + (i % 10),
        ])

    products = [["ProductID", "ProductName", "Category", "UnitPrice", "UnitCost", "LaunchQuarter"]]
    for i in range(90):
        category = categories[i % len(categories)]
        unit_cost = 8 + ((i * 5) % 28)
        unit_price = unit_cost + 10 + (i % 8)
        products.append([
            f"P{i+1:03d}",
            f"{category}_Item_{i+1:03d}",
            category,
            unit_price,
            unit_cost,
            launch[i % len(launch)],
        ])

    months = [f"2024-{m:02d}" for m in range(1, 13)]
    sales = [["SaleID", "StoreID", "ProductID", "Month", "UnitsSold", "DiscountPct", "Channel"]]
    targets = [["TargetID", "StoreID", "Category", "Month", "RevenueTarget", "UnitsTarget"]]
    for i in range(120):
        store_id = f"S{((i * 3) % 84) + 1:03d}"
        product_id = f"P{((i * 7) % 90) + 1:03d}"
        product_category = products[((i * 7) % 90) + 1][2]
        month = months[i % len(months)]
        units = 28 + ((i * 9) % 64)
        discount = [0, 5, 10, 15][i % 4]
        channel = ["Online", "InStore", "Marketplace"][i % 3]
        sales.append([f"SL{i+1:03d}", store_id, product_id, month, units, discount, channel])

        target_revenue = 1100 + ((i * 37) % 3600)
        target_units = 25 + ((i * 11) % 70)
        targets.append([f"T{i+1:03d}", store_id, product_category, month, target_revenue, target_units])

    write_csv(os.path.join(case_dir, "tc01_input01.csv"), stores)
    write_csv(os.path.join(case_dir, "tc01_input02.csv"), products)
    write_csv(os.path.join(case_dir, "tc01_input03.csv"), sales)
    write_csv(os.path.join(case_dir, "tc01_input04.csv"), targets)

    store_map = {row[0]: row for row in stores[1:]}
    product_map = {row[0]: row for row in products[1:]}

    region_cat = defaultdict(lambda: {"revenue": 0.0, "units": 0, "target": 0})
    store_perf = defaultdict(lambda: {"region": "", "revenue": 0.0, "units": 0})
    category_summary = defaultdict(lambda: {"revenue": 0.0, "profit": 0.0, "discounted_units": 0, "total_units": 0})

    target_lookup = defaultdict(int)
    for _, store_id, category, month, revenue_target, _ in targets[1:]:
        if month >= "2024-07":
            region = store_map[store_id][2]
            region_cat[(region, category)]["target"] += int(revenue_target)
            target_lookup[(store_id, category)] += int(revenue_target)

    for _, store_id, product_id, month, units_sold, discount_pct, _ in sales[1:]:
        if month < "2024-07":
            continue
        units = int(units_sold)
        discount = int(discount_pct) / 100
        _, _, category, unit_price, unit_cost, _ = product_map[product_id]
        region = store_map[store_id][2]
        revenue = round(units * float(unit_price) * (1 - discount), 2)
        profit = round(units * (float(unit_price) * (1 - discount) - float(unit_cost)), 2)

        region_cat[(region, category)]["revenue"] += revenue
        region_cat[(region, category)]["units"] += units
        store_perf[store_id]["region"] = region
        store_perf[store_id]["revenue"] += revenue
        store_perf[store_id]["units"] += units
        category_summary[category]["revenue"] += revenue
        category_summary[category]["profit"] += profit
        category_summary[category]["total_units"] += units
        if discount > 0:
            category_summary[category]["discounted_units"] += units

    region_rows = []
    for (region, category), metrics in sorted(region_cat.items()):
        region_rows.append([
            region,
            category,
            round(metrics["revenue"], 2),
            metrics["units"],
            metrics["target"],
            pct(metrics["revenue"], metrics["target"]),
        ])

    leaderboard = []
    for store_id, metrics in store_perf.items():
        leaderboard.append([
            store_id,
            store_map[store_id][1],
            metrics["region"],
            round(metrics["revenue"], 2),
            metrics["units"],
        ])
    leaderboard = top_n(leaderboard, 3, n=10)

    category_rows = []
    for category, metrics in sorted(category_summary.items()):
        category_rows.append([
            category,
            round(metrics["revenue"], 2),
            round(metrics["profit"], 2),
            metrics["total_units"],
            pct(metrics["discounted_units"], metrics["total_units"]),
        ])

    wb = Workbook()
    ws = wb.active
    ws.title = "Region_Category_H2"
    ws.append(["Region", "Category", "H2_Revenue", "H2_Units", "H2_Target", "Achievement_Pct"])
    for row in region_rows:
        ws.append(row)

    ws2 = wb.create_sheet("Store_Leaderboard")
    ws2.append(["StoreID", "StoreName", "Region", "H2_Revenue", "H2_Units"])
    for row in leaderboard:
        ws2.append(row)

    ws3 = wb.create_sheet("Category_Summary")
    ws3.append(["Category", "H2_Revenue", "Gross_Profit", "H2_Units", "Discounted_Unit_Share_Pct"])
    for row in category_rows:
        ws3.append(row)

    save_wb(os.path.join(case_dir, "tc01_output01.xlsx"), wb)

    best_region_row = max(region_rows, key=lambda row: row[5])
    best_store = leaderboard[0]
    best_category = max(category_rows, key=lambda row: row[2])
    return {
        "task_id": "Test 1",
        "title": "Large Retail H2 Revenue and Target Analysis",
        "scenario": "Retail operations / sales performance analysis",
        "category": "join four clean tables; filter H2 months; aggregate revenue by region and category; compare against targets; rank stores; summarise category profitability",
        "atomic_operations": [
            "join stores with sales on StoreID",
            "join products with sales on ProductID",
            "filter rows to H2 2024 months",
            "compute net revenue after discount",
            "aggregate revenue and units by Region and Category",
            "aggregate H2 targets by Region and Category",
            "compute target achievement percentage",
            "rank stores by H2 revenue",
            "summarise category revenue and gross profit",
            "export results to a multi-sheet workbook"
        ],
        "number_of_atomic_operations": 10,
        "spreadsheets": [
            "Case01/tc01_input01.csv",
            "Case01/tc01_input02.csv",
            "Case01/tc01_input03.csv",
            "Case01/tc01_input04.csv",
        ],
        "prompt": "You have four clean tables: stores, products, monthly sales, and monthly targets. Join the tables using StoreID and ProductID, keep only the H2 2024 months (2024-07 to 2024-12), compute net revenue as UnitsSold × UnitPrice × (1 - DiscountPct), aggregate the results by Region and Category, and compare each aggregate to the summed H2 revenue target. Then create a Top 10 store leaderboard by H2 revenue and a category summary with revenue, gross profit, and discounted-unit share. Output everything into a new Excel workbook with three sheets: Region_Category_H2, Store_Leaderboard, and Category_Summary.",
        "answer": f"Top H2 target achievement: {best_region_row[0]} / {best_region_row[1]} at {best_region_row[5]}%. Highest-revenue store: {best_store[1]} ({best_store[2]}) with {best_store[3]:.2f}. Most profitable category in H2: {best_category[0]} with gross profit {best_category[2]:.2f}.",
        "expected_output_file": ["Case01/tc01_output01.xlsx"],
        "feedback": "The four input tables are clean and fully keyed. I plan to compute net revenue after discount, restrict the analysis to 2024-07 through 2024-12, sum targets over the same scope, and then produce the three requested sheets.\n\n---",
        "conversations": [
            {
                "role": "user",
                "content": "Join the four clean retail tables, keep only H2 2024, compare actual revenue against targets by region and category, rank stores by H2 revenue, and summarise category profitability."
            },
            {
                "role": "assistant",
                "content": "I will treat DiscountPct as a percentage reduction applied before aggregation, limit the scope to 2024-07 through 2024-12, and aggregate targets over the same region/category scope. Then I'll write the three result sheets."
            },
            {
                "role": "user",
                "content": "Proceed."
            },
            {
                "role": "assistant",
                "content": f"Completed. Best H2 achievement is {best_region_row[0]} / {best_region_row[1]} at {best_region_row[5]}%. The top store is {best_store[1]} with H2 revenue {best_store[3]:.2f}, and {best_category[0]} leads gross profit."
            },
        ],
    }


def build_case02():
    case_dir = os.path.join(BASE, "Case02")
    ensure_dir(case_dir)

    categories = ["Raw_Materials", "Packaging", "Components", "Services", "Tools"]
    regions = ["UK", "EU", "NA", "APAC"]
    buyers = [["BuyerID", "BuyerName", "Team", "Region", "ExperienceYears", "Manager", "ActiveFlag"]]
    for i in range(80):
        buyers.append([
            f"B{i+1:03d}",
            f"Buyer_{i+1:03d}",
            ["Strategic", "Direct", "Indirect", "Capex"][i % 4],
            regions[i % len(regions)],
            2 + (i % 11),
            f"Lead_{(i % 10) + 1:02d}",
            "Yes",
        ])

    suppliers = [["SupplierID", "SupplierName", "Category", "Country", "Region", "PreferredFlag", "ContractTier"]]
    countries = ["UK", "Germany", "France", "Poland", "India", "Japan", "Canada", "Spain"]
    for i in range(84):
        suppliers.append([
            f"SUP{i+1:03d}",
            f"Supplier_{i+1:03d}",
            categories[i % len(categories)],
            countries[i % len(countries)],
            regions[i % len(regions)],
            "Yes" if i % 3 == 0 else "No",
            ["Gold", "Silver", "Bronze"][i % 3],
        ])

    po_rows = [["POID", "SupplierID", "BuyerID", "OrderMonth", "Category", "OrderValue", "PromisedDays"]]
    delivery_rows = [["POID", "SupplierID", "ActualDays", "ReceivedQty", "AcceptedQty", "FreightCost", "DeliveryRegion"]]
    quality_rows = [["POID", "SupplierID", "InspectionScore", "DefectUnits", "ReturnedUnits", "AuditPass"]]
    months = [f"2024-{m:02d}" for m in range(1, 13)]
    for i in range(120):
        supplier_id = f"SUP{((i * 5) % 84) + 1:03d}"
        buyer_id = f"B{((i * 7) % 80) + 1:03d}"
        category = suppliers[((i * 5) % 84) + 1][2]
        promised = 4 + (i % 10)
        actual = promised - 1 + ((i * 3) % 5)
        received = 420 + ((i * 13) % 360)
        defects = 8 + ((i * 7) % 35)
        returns = 2 + ((i * 5) % 12)
        accepted = received - defects
        po_rows.append([
            f"PO{i+1:03d}",
            supplier_id,
            buyer_id,
            months[i % len(months)],
            category,
            9000 + ((i * 421) % 26000),
            promised,
        ])
        delivery_rows.append([
            f"PO{i+1:03d}",
            supplier_id,
            actual,
            received,
            accepted,
            380 + ((i * 17) % 760),
            regions[(i + 1) % len(regions)],
        ])
        quality_rows.append([
            f"PO{i+1:03d}",
            supplier_id,
            71 + ((i * 4) % 28),
            defects,
            returns,
            "Pass" if i % 6 != 0 else "Review",
        ])

    write_csv(os.path.join(case_dir, "tc02_input01.csv"), suppliers)
    write_csv(os.path.join(case_dir, "tc02_input02.csv"), buyers)
    write_csv(os.path.join(case_dir, "tc02_input03.csv"), po_rows)
    write_csv(os.path.join(case_dir, "tc02_input04.csv"), delivery_rows)
    write_csv(os.path.join(case_dir, "tc02_input05.csv"), quality_rows)

    supplier_map = {row[0]: row for row in suppliers[1:]}
    buyer_map = {row[0]: row for row in buyers[1:]}
    po_map = {row[0]: row for row in po_rows[1:]}
    delivery_map = {row[0]: row for row in delivery_rows[1:]}
    quality_map = {row[0]: row for row in quality_rows[1:]}

    supplier_stats = defaultdict(lambda: {"orders": 0, "ontime": 0, "inspection": [], "accepted": 0, "received": 0, "defects": 0, "value": 0})
    buyer_stats = defaultdict(lambda: {"orders": 0, "value": 0, "score_pairs": []})
    risk_rows = []

    for poid, supplier_id, buyer_id, _, category, order_value, promised_days in po_rows[1:]:
        actual_days = int(delivery_map[poid][2])
        received_qty = int(delivery_map[poid][3])
        accepted_qty = int(delivery_map[poid][4])
        inspection = int(quality_map[poid][2])
        defect_units = int(quality_map[poid][3]) + int(quality_map[poid][4])
        ontime = 1 if actual_days <= int(promised_days) else 0
        supplier_stats[supplier_id]["orders"] += 1
        supplier_stats[supplier_id]["ontime"] += ontime
        supplier_stats[supplier_id]["inspection"].append(inspection)
        supplier_stats[supplier_id]["accepted"] += accepted_qty
        supplier_stats[supplier_id]["received"] += received_qty
        supplier_stats[supplier_id]["defects"] += defect_units
        supplier_stats[supplier_id]["value"] += int(order_value)

    supplier_rows = []
    for supplier_id, stats in supplier_stats.items():
        ontime_pct = pct(stats["ontime"], stats["orders"])
        acceptance_pct = pct(stats["accepted"], stats["received"])
        inspection_avg = round(sum(stats["inspection"]) / len(stats["inspection"]), 2)
        score = round(0.5 * ontime_pct + 0.3 * inspection_avg + 0.2 * acceptance_pct, 2)
        defect_rate = pct(stats["defects"], stats["received"])
        supplier_rows.append([
            supplier_id,
            supplier_map[supplier_id][1],
            supplier_map[supplier_id][2],
            stats["orders"],
            ontime_pct,
            inspection_avg,
            acceptance_pct,
            defect_rate,
            score,
            stats["value"],
        ])
        if score < 87 or defect_rate > 8:
            risk_rows.append([
                supplier_id,
                supplier_map[supplier_id][1],
                round(score, 2),
                round(defect_rate, 2),
                supplier_map[supplier_id][6],
            ])

    supplier_rows.sort(key=lambda row: row[8], reverse=True)
    risk_rows.sort(key=lambda row: (row[2], -row[3]))

    supplier_score_lookup = {row[0]: row[8] for row in supplier_rows}
    for poid, supplier_id, buyer_id, _, _, order_value, _ in po_rows[1:]:
        buyer_stats[buyer_id]["orders"] += 1
        buyer_stats[buyer_id]["value"] += int(order_value)
        buyer_stats[buyer_id]["score_pairs"].append((supplier_score_lookup[supplier_id], int(order_value)))

    buyer_rows = []
    for buyer_id, stats in buyer_stats.items():
        buyer_rows.append([
            buyer_id,
            buyer_map[buyer_id][1],
            buyer_map[buyer_id][2],
            buyer_map[buyer_id][3],
            stats["orders"],
            stats["value"],
            weighted_avg(stats["score_pairs"]),
        ])
    buyer_rows.sort(key=lambda row: row[6], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Supplier_Scorecard"
    ws.append(["SupplierID", "SupplierName", "Category", "PO_Count", "OnTime_Pct", "Inspection_Avg", "Acceptance_Pct", "Defect_Rate_Pct", "Weighted_Score", "OrderValue"])
    for row in supplier_rows:
        ws.append(row)

    ws2 = wb.create_sheet("Buyer_Summary")
    ws2.append(["BuyerID", "BuyerName", "Team", "Region", "PO_Count", "Managed_OrderValue", "Weighted_Supplier_Score"])
    for row in buyer_rows:
        ws2.append(row)

    ws3 = wb.create_sheet("Risk_Watchlist")
    ws3.append(["SupplierID", "SupplierName", "Weighted_Score", "Defect_Rate_Pct", "ContractTier"])
    for row in risk_rows:
        ws3.append(row)

    save_wb(os.path.join(case_dir, "tc02_output01.xlsx"), wb)

    best_supplier = supplier_rows[0]
    best_buyer = buyer_rows[0]
    riskiest = risk_rows[0]
    return {
        "task_id": "Test 2",
        "title": "Large Procurement Supplier Scorecard",
        "scenario": "Procurement / supplier performance analytics",
        "category": "join five clean tables; calculate on-time rate, inspection average, acceptance rate, and weighted supplier score; summarise buyers; build supplier risk watchlist",
        "atomic_operations": [
            "join purchase orders with suppliers on SupplierID",
            "join purchase orders with buyers on BuyerID",
            "join deliveries to purchase orders on POID",
            "join quality results to purchase orders on POID",
            "compute supplier on-time delivery rate",
            "compute supplier inspection average",
            "compute supplier acceptance rate",
            "compute supplier defect rate",
            "compute weighted supplier score",
            "aggregate supplier scores to buyer-level summary",
            "filter risky suppliers into a watchlist",
            "export results to a multi-sheet workbook"
        ],
        "number_of_atomic_operations": 12,
        "spreadsheets": [
            "Case02/tc02_input01.csv",
            "Case02/tc02_input02.csv",
            "Case02/tc02_input03.csv",
            "Case02/tc02_input04.csv",
            "Case02/tc02_input05.csv",
        ],
        "prompt": "You have five clean procurement tables: suppliers, buyers, purchase orders, deliveries, and quality results. Join the records by SupplierID, BuyerID, and POID. For each supplier, calculate PO count, on-time delivery rate, average inspection score, acceptance rate, defect rate, and a weighted score defined as 50% on-time rate + 30% inspection average + 20% acceptance rate. Then create a buyer summary showing managed order value and weighted supplier score, and a risk watchlist of suppliers with low scores or high defect rates. Output the results into a new Excel workbook with sheets Supplier_Scorecard, Buyer_Summary, and Risk_Watchlist.",
        "answer": f"Top supplier: {best_supplier[1]} with weighted score {best_supplier[8]:.2f}. Best buyer portfolio: {best_buyer[1]} with weighted supplier score {best_buyer[6]:.2f}. Highest-risk supplier on the watchlist: {riskiest[1]} (score {riskiest[2]:.2f}, defect rate {riskiest[3]:.2f}%).",
        "expected_output_file": ["Case02/tc02_output01.xlsx"],
        "feedback": "The tables are already clean and keyed, so I will score suppliers at the supplier level after joining each PO with its delivery and quality record, and then roll those supplier scores up to the buyer summary.\n\n---",
        "conversations": [
            {
                "role": "user",
                "content": "Join the five procurement tables, score suppliers, summarise buyer portfolios, and build a risk watchlist."
            },
            {
                "role": "assistant",
                "content": "I will compute supplier metrics from PO-level joins first, apply the weighted-score formula exactly as given, then aggregate those scores into the buyer summary and create the risk watchlist."
            },
            {
                "role": "user",
                "content": "Proceed."
            },
            {
                "role": "assistant",
                "content": f"Completed. {best_supplier[1]} has the top weighted score at {best_supplier[8]:.2f}. {best_buyer[1]} leads the buyer summary, and {riskiest[1]} appears first on the risk watchlist."
            },
        ],
    }


def build_case03():
    case_dir = os.path.join(BASE, "Case03")
    ensure_dir(case_dir)

    schools = ["School_A", "School_B", "School_C", "School_D", "School_E"]
    levels = ["UG", "PG"]
    instructors = [f"Instructor_{i+1:03d}" for i in range(20)]
    buildings = ["North", "South", "East", "West", "Central"]

    students = [["StudentID", "StudentName", "School", "ProgramLevel", "EntryYear", "Status", "Advisor"]]
    for i in range(84):
        students.append([
            f"STU{i+1:03d}",
            f"Student_{i+1:03d}",
            schools[i % len(schools)],
            levels[i % len(levels)],
            2021 + (i % 4),
            "Active",
            f"Advisor_{(i % 12) + 1:02d}",
        ])

    courses = [["CourseID", "CourseName", "School", "Subject", "Credits", "CoreFlag", "Level"]]
    subjects = ["Analytics", "Finance", "Biology", "History", "Design", "Economics"]
    for i in range(80):
        courses.append([
            f"C{i+1:03d}",
            f"Course_{i+1:03d}",
            schools[i % len(schools)],
            subjects[i % len(subjects)],
            2 + (i % 3),
            "Core" if i % 2 == 0 else "Elective",
            levels[(i + 1) % len(levels)],
        ])

    rooms = [["RoomID", "Building", "RoomType", "Capacity", "Floor", "AVReady", "Campus"]]
    for i in range(80):
        rooms.append([
            f"R{i+1:03d}",
            buildings[i % len(buildings)],
            ["Lecture", "Lab", "Seminar", "Studio"][i % 4],
            24 + ((i * 3) % 46),
            1 + (i % 6),
            "Yes",
            "Main",
        ])

    sections = [["SectionID", "CourseID", "Instructor", "RoomID", "Term", "Capacity", "ScheduledHours", "DeliveryMode"]]
    for i in range(90):
        sections.append([
            f"SEC{i+1:03d}",
            f"C{((i * 5) % 80) + 1:03d}",
            instructors[i % len(instructors)],
            f"R{((i * 7) % 80) + 1:03d}",
            "2024-Fall",
            26 + ((i * 4) % 40),
            2 + (i % 4),
            ["Onsite", "Hybrid", "Online"][i % 3],
        ])

    enrollments = [["EnrollID", "StudentID", "SectionID", "EnrollStatus", "UnitsTaken", "GradeBasis", "WaitlistPosition"]]
    for i in range(120):
        section_id = f"SEC{((i * 3) % 90) + 1:03d}"
        status = ["Enrolled", "Enrolled", "Enrolled", "Waitlist"][i % 4]
        wait_pos = 0 if status == "Enrolled" else 1 + (i % 8)
        course_id = sections[((i * 3) % 90) + 1][1]
        units = courses[int(course_id[1:])][4]
        enrollments.append([
            f"EN{i+1:03d}",
            f"STU{((i * 7) % 84) + 1:03d}",
            section_id,
            status,
            units,
            "Letter",
            wait_pos,
        ])

    write_csv(os.path.join(case_dir, "tc03_input01.csv"), students)
    write_csv(os.path.join(case_dir, "tc03_input02.csv"), courses)
    write_csv(os.path.join(case_dir, "tc03_input03.csv"), sections)
    write_csv(os.path.join(case_dir, "tc03_input04.csv"), enrollments)
    write_csv(os.path.join(case_dir, "tc03_input05.csv"), rooms)

    course_map = {row[0]: row for row in courses[1:]}
    room_map = {row[0]: row for row in rooms[1:]}
    section_map = {row[0]: row for row in sections[1:]}

    enroll_counts = Counter()
    wait_counts = Counter()
    for _, _, section_id, status, _, _, _ in enrollments[1:]:
        if status == "Enrolled":
            enroll_counts[section_id] += 1
        else:
            wait_counts[section_id] += 1

    section_rows = []
    instructor_stats = defaultdict(lambda: {"sections": 0, "enrolled": 0, "hours": 0})
    room_rows = defaultdict(lambda: {"sections": 0, "fill_pairs": [], "hours": 0})
    for section_id, course_id, instructor, room_id, term, capacity, scheduled_hours, delivery_mode in sections[1:]:
        enrolled = enroll_counts[section_id]
        wait = wait_counts[section_id]
        cap = int(capacity)
        fill_rate = pct(enrolled, cap)
        wait_rate = pct(wait, cap)
        course = course_map[course_id]
        room = room_map[room_id]
        section_rows.append([
            section_id,
            course_id,
            course[1],
            course[3],
            instructor,
            room[1],
            cap,
            enrolled,
            wait,
            fill_rate,
            wait_rate,
            delivery_mode,
        ])
        instructor_stats[instructor]["sections"] += 1
        instructor_stats[instructor]["enrolled"] += enrolled
        instructor_stats[instructor]["hours"] += int(scheduled_hours)
        room_rows[room[1]]["sections"] += 1
        room_rows[room[1]]["fill_pairs"].append((fill_rate, 1))
        room_rows[room[1]]["hours"] += int(scheduled_hours)

    section_rows.sort(key=lambda row: row[9], reverse=True)

    instructor_rows = []
    for instructor, stats in instructor_stats.items():
        instructor_rows.append([
            instructor,
            stats["sections"],
            stats["enrolled"],
            stats["hours"],
            round(stats["enrolled"] / stats["sections"], 2),
        ])
    instructor_rows.sort(key=lambda row: row[2], reverse=True)

    room_summary_rows = []
    for building, stats in room_rows.items():
        room_summary_rows.append([
            building,
            stats["sections"],
            weighted_avg(stats["fill_pairs"]),
            stats["hours"],
        ])
    room_summary_rows.sort(key=lambda row: row[2], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Section_Utilisation"
    ws.append(["SectionID", "CourseID", "CourseName", "Subject", "Instructor", "Building", "Capacity", "Enrolled", "Waitlisted", "FillRate_Pct", "WaitRate_Pct", "DeliveryMode"])
    for row in section_rows:
        ws.append(row)

    ws2 = wb.create_sheet("Instructor_Load")
    ws2.append(["Instructor", "SectionCount", "TotalEnrolled", "ScheduledHours", "AvgEnrolledPerSection"])
    for row in instructor_rows:
        ws2.append(row)

    ws3 = wb.create_sheet("Room_Utilisation")
    ws3.append(["Building", "SectionCount", "AvgFillRate_Pct", "ScheduledHours"])
    for row in room_summary_rows:
        ws3.append(row)

    save_wb(os.path.join(case_dir, "tc03_output01.xlsx"), wb)

    fullest = section_rows[0]
    busiest_instructor = instructor_rows[0]
    top_building = room_summary_rows[0]
    return {
        "task_id": "Test 3",
        "title": "Large University Section Utilisation Analysis",
        "scenario": "Education operations / enrollment analytics",
        "category": "join five clean tables; compute section fill and waitlist rates; summarise instructor teaching load; measure room utilisation by building",
        "atomic_operations": [
            "join sections with courses on CourseID",
            "join sections with rooms on RoomID",
            "join enrollments with sections on SectionID",
            "count enrolled students per section",
            "count waitlisted students per section",
            "compute section fill rate",
            "compute section waitlist rate",
            "aggregate section load by instructor",
            "aggregate room utilisation by building",
            "rank sections by fill rate",
            "export results to a multi-sheet workbook"
        ],
        "number_of_atomic_operations": 11,
        "spreadsheets": [
            "Case03/tc03_input01.csv",
            "Case03/tc03_input02.csv",
            "Case03/tc03_input03.csv",
            "Case03/tc03_input04.csv",
            "Case03/tc03_input05.csv",
        ],
        "prompt": "You have five clean university tables: students, courses, sections, enrollments, and rooms. Join the tables using StudentID, CourseID, SectionID, and RoomID. For each section, calculate enrolled count, waitlisted count, fill rate, and waitlist rate. Then summarise instructor load with section count, total enrolled, and scheduled hours, and create a room-utilisation summary by building using average section fill rate. Output the results into a new Excel workbook with sheets Section_Utilisation, Instructor_Load, and Room_Utilisation.",
        "answer": f"Highest-fill section: {fullest[0]} ({fullest[2]}) at {fullest[9]}% fill. Busiest instructor: {busiest_instructor[0]} with {busiest_instructor[2]} enrolled students across {busiest_instructor[1]} sections. Highest average building utilisation: {top_building[0]} at {top_building[2]}%.",
        "expected_output_file": ["Case03/tc03_output01.xlsx"],
        "feedback": "The tables are already rectangular and clean. I will count enrolled and waitlisted students at the section level first, then derive instructor and building summaries from those section-level results.\n\n---",
        "conversations": [
            {
                "role": "user",
                "content": "Join the five university tables, calculate section fill and waitlist rates, summarise instructor load, and add a building-level room utilisation sheet."
            },
            {
                "role": "assistant",
                "content": "I will compute section-level metrics first, then roll them up into instructor and building summaries so the final workbook stays consistent across all three sheets."
            },
            {
                "role": "user",
                "content": "Proceed."
            },
            {
                "role": "assistant",
                "content": f"Completed. {fullest[0]} is the fullest section at {fullest[9]}%. {busiest_instructor[0]} has the largest teaching load by enrollment, and {top_building[0]} has the highest average building utilisation."
            },
        ],
    }


def build_case04():
    case_dir = os.path.join(BASE, "Case04")
    ensure_dir(case_dir)

    specialties = ["Cardiology", "Dermatology", "Neurology", "Orthopedics", "General", "Pediatrics"]
    clinics = [["ClinicID", "ClinicName", "City", "Region", "SpecialtyHub", "OpenDays", "CapacityBand"]]
    for i in range(80):
        clinics.append([
            f"CL{i+1:03d}",
            f"Clinic_{i+1:03d}",
            ["London", "Bristol", "Leeds", "Oxford", "York"][i % 5],
            ["North", "South", "East", "West"][i % 4],
            specialties[i % len(specialties)],
            ["Mon-Fri", "Mon-Sat"][i % 2],
            ["Small", "Medium", "Large"][i % 3],
        ])

    doctors = [["DoctorID", "DoctorName", "ClinicID", "Specialty", "Grade", "SessionsPerWeek", "ActiveFlag"]]
    for i in range(84):
        doctors.append([
            f"D{i+1:03d}",
            f"Doctor_{i+1:03d}",
            f"CL{((i * 3) % 80) + 1:03d}",
            specialties[i % len(specialties)],
            ["Consultant", "Senior", "Associate"][i % 3],
            4 + (i % 6),
            "Yes",
        ])

    patients = [["PatientID", "PatientName", "AgeBand", "Gender", "InsuranceType", "ChronicFlag", "HomeRegion"]]
    for i in range(90):
        patients.append([
            f"PT{i+1:03d}",
            f"Patient_{i+1:03d}",
            ["18-29", "30-44", "45-59", "60+"][i % 4],
            ["F", "M"][i % 2],
            ["Private", "Public", "Mixed"][i % 3],
            "Yes" if i % 4 == 0 else "No",
            ["North", "South", "East", "West"][i % 4],
        ])

    appointments = [["ApptID", "ClinicID", "DoctorID", "PatientID", "VisitMonth", "Status", "BookedMinutes"]]
    billing = [["ApptID", "ChargeAmount", "CollectedAmount", "PayerType", "ProcedureCode", "ClaimStatus"]]
    for i in range(120):
        clinic_id = f"CL{((i * 5) % 80) + 1:03d}"
        doctor_id = f"D{((i * 7) % 84) + 1:03d}"
        patient_id = f"PT{((i * 9) % 90) + 1:03d}"
        status = ["Completed", "Completed", "Completed", "NoShow", "Cancelled"][i % 5]
        charge = 95 + ((i * 13) % 260)
        collected = 0 if status != "Completed" else charge - ((i * 7) % 26)
        appointments.append([
            f"A{i+1:03d}",
            clinic_id,
            doctor_id,
            patient_id,
            f"2024-{(i % 12) + 1:02d}",
            status,
            15 + ((i * 5) % 45),
        ])
        billing.append([
            f"A{i+1:03d}",
            charge,
            collected,
            ["Private", "Public", "Mixed"][i % 3],
            f"PROC{(i % 18) + 1:02d}",
            "Submitted" if status == "Completed" else "NotFiled",
        ])

    write_csv(os.path.join(case_dir, "tc04_input01.csv"), clinics)
    write_csv(os.path.join(case_dir, "tc04_input02.csv"), doctors)
    write_csv(os.path.join(case_dir, "tc04_input03.csv"), patients)
    write_csv(os.path.join(case_dir, "tc04_input04.csv"), appointments)
    write_csv(os.path.join(case_dir, "tc04_input05.csv"), billing)

    clinic_map = {row[0]: row for row in clinics[1:]}
    doctor_map = {row[0]: row for row in doctors[1:]}
    billing_map = {row[0]: row for row in billing[1:]}

    clinic_rows = defaultdict(lambda: {"appointments": 0, "noshow": 0, "charges": 0, "collected": 0, "completed": 0})
    doctor_rows = defaultdict(lambda: {"completed": 0, "charges": 0, "collected": 0, "minutes": 0})
    payer_rows = defaultdict(lambda: {"charges": 0, "collected": 0, "completed": 0})

    for appt_id, clinic_id, doctor_id, _, visit_month, status, booked_minutes in appointments[1:]:
        bill = billing_map[appt_id]
        charge = int(bill[1])
        collected = int(bill[2])
        clinic_key = (clinic_id, clinic_map[clinic_id][4])
        clinic_rows[clinic_key]["appointments"] += 1
        clinic_rows[clinic_key]["charges"] += charge
        clinic_rows[clinic_key]["collected"] += collected
        if status == "NoShow":
            clinic_rows[clinic_key]["noshow"] += 1
        if status == "Completed":
            clinic_rows[clinic_key]["completed"] += 1
            doctor_rows[doctor_id]["completed"] += 1
            doctor_rows[doctor_id]["charges"] += charge
            doctor_rows[doctor_id]["collected"] += collected
            doctor_rows[doctor_id]["minutes"] += int(booked_minutes)
            payer_rows[bill[3]]["charges"] += charge
            payer_rows[bill[3]]["collected"] += collected
            payer_rows[bill[3]]["completed"] += 1

    clinic_summary = []
    for (clinic_id, specialty), stats in clinic_rows.items():
        clinic_summary.append([
            clinic_id,
            clinic_map[clinic_id][1],
            specialty,
            stats["appointments"],
            pct(stats["noshow"], stats["appointments"]),
            round(stats["charges"], 2),
            round(stats["collected"], 2),
            pct(stats["collected"], stats["charges"]),
        ])
    clinic_summary.sort(key=lambda row: row[4], reverse=True)

    doctor_summary = []
    for doctor_id, stats in doctor_rows.items():
        doctor_summary.append([
            doctor_id,
            doctor_map[doctor_id][1],
            doctor_map[doctor_id][3],
            doctor_map[doctor_id][2],
            stats["completed"],
            round(stats["charges"], 2),
            round(stats["collected"], 2),
            pct(stats["collected"], stats["charges"]),
            round(stats["minutes"] / stats["completed"], 2) if stats["completed"] else 0,
        ])
    doctor_summary.sort(key=lambda row: row[6], reverse=True)

    payer_summary = []
    for payer, stats in payer_rows.items():
        payer_summary.append([
            payer,
            stats["completed"],
            round(stats["charges"], 2),
            round(stats["collected"], 2),
            pct(stats["collected"], stats["charges"]),
        ])
    payer_summary.sort(key=lambda row: row[4], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Clinic_Summary"
    ws.append(["ClinicID", "ClinicName", "Specialty", "Appointments", "NoShowRate_Pct", "Charges", "Collected", "CollectionRate_Pct"])
    for row in clinic_summary:
        ws.append(row)

    ws2 = wb.create_sheet("Doctor_Productivity")
    ws2.append(["DoctorID", "DoctorName", "Specialty", "ClinicID", "CompletedVisits", "Charges", "Collected", "CollectionRate_Pct", "AvgMinutesPerVisit"])
    for row in doctor_summary:
        ws2.append(row)

    ws3 = wb.create_sheet("Payer_Collection")
    ws3.append(["PayerType", "CompletedVisits", "Charges", "Collected", "CollectionRate_Pct"])
    for row in payer_summary:
        ws3.append(row)

    save_wb(os.path.join(case_dir, "tc04_output01.xlsx"), wb)

    top_noshow = clinic_summary[0]
    top_doctor = doctor_summary[0]
    best_payer = payer_summary[0]
    return {
        "task_id": "Test 4",
        "title": "Large Clinic Appointment and Billing Analysis",
        "scenario": "Healthcare operations / appointment and revenue analysis",
        "category": "join five clean tables; compute no-show and collection rates by clinic and specialty; summarise doctor productivity; compare payer collection performance",
        "atomic_operations": [
            "join appointments with clinics on ClinicID",
            "join appointments with doctors on DoctorID",
            "join appointments with patients on PatientID",
            "join billing with appointments on ApptID",
            "count appointments by clinic",
            "compute clinic no-show rate",
            "sum clinic charges and collections",
            "compute clinic collection rate",
            "filter completed visits for doctor productivity",
            "aggregate doctor productivity metrics",
            "aggregate payer-level collection metrics",
            "export results to a multi-sheet workbook"
        ],
        "number_of_atomic_operations": 12,
        "spreadsheets": [
            "Case04/tc04_input01.csv",
            "Case04/tc04_input02.csv",
            "Case04/tc04_input03.csv",
            "Case04/tc04_input04.csv",
            "Case04/tc04_input05.csv",
        ],
        "prompt": "You have five clean clinic tables: clinics, doctors, patients, appointments, and billing. Join the data using ClinicID, DoctorID, PatientID, and ApptID. For each clinic, calculate appointment count, no-show rate, total charges, total collected, and collection rate, keeping the clinic's specialty in the result. Then summarise doctor productivity using completed visits, charges, collections, and average booked minutes per completed visit, and add a payer-level collection summary. Output a new Excel workbook with sheets Clinic_Summary, Doctor_Productivity, and Payer_Collection.",
        "answer": f"Highest no-show clinic: {top_noshow[1]} ({top_noshow[2]}) at {top_noshow[4]}%. Top doctor by collected revenue: {top_doctor[1]} with {top_doctor[6]:.2f} collected. Best payer collection rate: {best_payer[0]} at {best_payer[4]}%.",
        "expected_output_file": ["Case04/tc04_output01.xlsx"],
        "feedback": "The input tables are clean and keyed one-to-many, so I will calculate clinic metrics from all appointments, but doctor and payer productivity only from completed visits where collection values are meaningful.\n\n---",
        "conversations": [
            {
                "role": "user",
                "content": "Join the clinic, doctor, patient, appointment, and billing tables. Compute clinic no-show and collection metrics, doctor productivity, and payer collection performance."
            },
            {
                "role": "assistant",
                "content": "I will compute clinic metrics across all appointments, while doctor productivity and payer collection will be based on completed visits so the revenue measures stay comparable."
            },
            {
                "role": "user",
                "content": "Proceed."
            },
            {
                "role": "assistant",
                "content": f"Completed. {top_noshow[1]} has the highest no-show rate at {top_noshow[4]}%. {top_doctor[1]} leads doctor collections, and {best_payer[0]} has the strongest payer collection rate."
            },
        ],
    }


def build_case05():
    case_dir = os.path.join(BASE, "Case05")
    ensure_dir(case_dir)

    zones = ["A", "B", "C", "D", "E"]
    categories = ["Electronics", "Apparel", "Home", "Beauty", "Auto"]
    warehouses = [["WarehouseID", "WarehouseName", "Region", "Manager", "AreaSqm", "DockCount", "ShiftPattern"]]
    for i in range(80):
        warehouses.append([
            f"W{i+1:03d}",
            f"Warehouse_{i+1:03d}",
            ["North", "South", "East", "West"][i % 4],
            f"Manager_{(i % 14) + 1:02d}",
            1800 + ((i * 25) % 2200),
            4 + (i % 7),
            ["Day", "Night", "Split"][i % 3],
        ])

    skus = [["SKU", "SKUName", "Category", "UnitWeightKg", "UnitValue", "ShelfLifeDays", "HazmatFlag"]]
    for i in range(90):
        skus.append([
            f"SKU{i+1:03d}",
            f"SKU_{i+1:03d}",
            categories[i % len(categories)],
            round(0.5 + ((i * 2) % 18) / 2, 1),
            12 + ((i * 9) % 70),
            60 + ((i * 5) % 240),
            "Yes" if i % 11 == 0 else "No",
        ])

    inventory = [["WarehouseID", "SKU", "BeginningStock", "ReceiptsQty", "EndingStock", "DamagedQty", "CycleCountAccuracy", "StorageZone"]]
    for i in range(90):
        inventory.append([
            f"W{((i * 3) % 80) + 1:03d}",
            f"SKU{i+1:03d}",
            300 + ((i * 11) % 420),
            90 + ((i * 7) % 180),
            140 + ((i * 13) % 260),
            2 + (i % 15),
            93 + (i % 7),
            zones[i % len(zones)],
        ])

    orders = [["OrderID", "WarehouseID", "SKU", "OrderMonth", "OrderedQty", "Priority", "Channel"]]
    shipments = [["OrderID", "ShipMonth", "ShippedQty", "TransitDays", "OnTimeFlag", "Carrier", "FreightCost"]]
    for i in range(120):
        warehouse_id = f"W{((i * 5) % 80) + 1:03d}"
        sku = f"SKU{((i * 7) % 90) + 1:03d}"
        ordered = 18 + ((i * 4) % 75)
        shipped = max(8, ordered - (i % 9))
        ontime = "Yes" if i % 5 != 0 else "No"
        orders.append([
            f"O{i+1:03d}",
            warehouse_id,
            sku,
            f"2024-{(i % 12) + 1:02d}",
            ordered,
            ["High", "Medium", "Low"][i % 3],
            ["B2B", "D2C", "Marketplace"][i % 3],
        ])
        shipments.append([
            f"O{i+1:03d}",
            f"2024-{(i % 12) + 1:02d}",
            shipped,
            1 + (i % 6),
            ontime,
            ["CarrierA", "CarrierB", "CarrierC", "CarrierD"][i % 4],
            120 + ((i * 9) % 260),
        ])

    write_csv(os.path.join(case_dir, "tc05_input01.csv"), warehouses)
    write_csv(os.path.join(case_dir, "tc05_input02.csv"), skus)
    write_csv(os.path.join(case_dir, "tc05_input03.csv"), inventory)
    write_csv(os.path.join(case_dir, "tc05_input04.csv"), orders)
    write_csv(os.path.join(case_dir, "tc05_input05.csv"), shipments)

    warehouse_map = {row[0]: row for row in warehouses[1:]}
    sku_map = {row[0]: row for row in skus[1:]}
    shipment_map = {row[0]: row for row in shipments[1:]}
    inventory_map = {(row[0], row[1]): row for row in inventory[1:]}

    warehouse_summary = defaultdict(lambda: {"orders": 0, "ordered": 0, "shipped": 0, "ontime": 0, "freight": 0})
    category_turns = defaultdict(lambda: {"shipped": 0, "avg_inventory": 0, "records": 0})
    backlog_rows = []

    for order_id, warehouse_id, sku, _, ordered_qty, _, _ in orders[1:]:
        shipment = shipment_map[order_id]
        shipped_qty = int(shipment[2])
        ordered_qty = int(ordered_qty)
        ontime = 1 if shipment[4] == "Yes" else 0
        freight = int(shipment[6])
        category = sku_map[sku][2]

        warehouse_summary[warehouse_id]["orders"] += 1
        warehouse_summary[warehouse_id]["ordered"] += ordered_qty
        warehouse_summary[warehouse_id]["shipped"] += shipped_qty
        warehouse_summary[warehouse_id]["ontime"] += ontime
        warehouse_summary[warehouse_id]["freight"] += freight

        inv = inventory_map.get((warehouse_id, sku))
        if inv:
            avg_inv = (int(inv[2]) + int(inv[4])) / 2
            category_turns[category]["shipped"] += shipped_qty
            category_turns[category]["avg_inventory"] += avg_inv
            category_turns[category]["records"] += 1
            backlog_rows.append([
                order_id,
                warehouse_id,
                warehouse_map[warehouse_id][2],
                sku,
                category,
                ordered_qty,
                shipped_qty,
                ordered_qty - shipped_qty,
                shipment[4],
            ])

    warehouse_rows = []
    for warehouse_id, stats in warehouse_summary.items():
        warehouse_rows.append([
            warehouse_id,
            warehouse_map[warehouse_id][1],
            warehouse_map[warehouse_id][2],
            stats["orders"],
            pct(stats["shipped"], stats["ordered"]),
            pct(stats["ontime"], stats["orders"]),
            round(stats["freight"], 2),
        ])
    warehouse_rows.sort(key=lambda row: row[4], reverse=True)

    category_rows = []
    for category, stats in category_turns.items():
        turns = round(stats["shipped"] / stats["avg_inventory"], 2) if stats["avg_inventory"] else 0
        category_rows.append([
            category,
            stats["shipped"],
            round(stats["avg_inventory"], 2),
            turns,
        ])
    category_rows.sort(key=lambda row: row[3], reverse=True)

    backlog_rows.sort(key=lambda row: row[7], reverse=True)
    backlog_rows = backlog_rows[:20]

    wb = Workbook()
    ws = wb.active
    ws.title = "Warehouse_Service"
    ws.append(["WarehouseID", "WarehouseName", "Region", "OrderCount", "FillRate_Pct", "OnTimeRate_Pct", "FreightCost"])
    for row in warehouse_rows:
        ws.append(row)

    ws2 = wb.create_sheet("Category_Turns")
    ws2.append(["Category", "ShippedQty", "Summed_AvgInventory", "InventoryTurnsProxy"])
    for row in category_rows:
        ws2.append(row)

    ws3 = wb.create_sheet("Backlog_Alert")
    ws3.append(["OrderID", "WarehouseID", "Region", "SKU", "Category", "OrderedQty", "ShippedQty", "BacklogQty", "OnTimeFlag"])
    for row in backlog_rows:
        ws3.append(row)

    save_wb(os.path.join(case_dir, "tc05_output01.xlsx"), wb)

    best_wh = warehouse_rows[0]
    top_turn = category_rows[0]
    backlog_top = backlog_rows[0]
    return {
        "task_id": "Test 5",
        "title": "Large Warehouse Fulfilment and Inventory Analysis",
        "scenario": "Logistics / warehouse performance analytics",
        "category": "join five clean tables; compute warehouse fill rate and on-time rate; derive category inventory-turns proxy; identify largest order backlogs",
        "atomic_operations": [
            "join orders with warehouses on WarehouseID",
            "join orders with SKUs on SKU",
            "join shipments with orders on OrderID",
            "join inventory snapshots on WarehouseID and SKU",
            "compute warehouse fill rate",
            "compute warehouse on-time rate",
            "sum warehouse freight cost",
            "aggregate shipped quantity by category",
            "compute average inventory proxy by category",
            "compute category inventory-turns proxy",
            "compute backlog quantity per order",
            "rank backlog orders descending",
            "export results to a multi-sheet workbook"
        ],
        "number_of_atomic_operations": 13,
        "spreadsheets": [
            "Case05/tc05_input01.csv",
            "Case05/tc05_input02.csv",
            "Case05/tc05_input03.csv",
            "Case05/tc05_input04.csv",
            "Case05/tc05_input05.csv",
        ],
        "prompt": "You have five clean logistics tables: warehouses, SKUs, inventory snapshots, orders, and shipments. Join the tables using WarehouseID, SKU, and OrderID. For each warehouse, calculate order count, fill rate (ShippedQty / OrderedQty), on-time rate, and total freight cost. Then summarise category performance with total shipped quantity and an inventory-turns proxy defined as shipped quantity divided by summed average inventory across matched warehouse-SKU records. Finally, create a backlog alert sheet showing the largest order shortfalls sorted by backlog quantity descending. Output a new Excel workbook with sheets Warehouse_Service, Category_Turns, and Backlog_Alert.",
        "answer": f"Best warehouse fill rate: {best_wh[1]} ({best_wh[2]}) at {best_wh[4]}%. Fastest-turning category: {top_turn[0]} with turns proxy {top_turn[3]:.2f}. Largest backlog order: {backlog_top[0]} for {backlog_top[3]} with backlog {backlog_top[7]}.",
        "expected_output_file": ["Case05/tc05_output01.xlsx"],
        "feedback": "The inputs are clean and aligned by WarehouseID, SKU, and OrderID. I will compute service metrics at warehouse level, derive category turns from matched inventory rows, and sort the backlog sheet by absolute shortfall.\n\n---",
        "conversations": [
            {
                "role": "user",
                "content": "Join the five warehouse tables, measure service levels, estimate category inventory turns, and build a backlog alert sheet."
            },
            {
                "role": "assistant",
                "content": "I will calculate fill rate and on-time rate at warehouse level first, then compute category turns only where order and inventory records align on the same warehouse and SKU."
            },
            {
                "role": "user",
                "content": "Proceed."
            },
            {
                "role": "assistant",
                "content": f"Completed. {best_wh[1]} has the highest fill rate at {best_wh[4]}%. {top_turn[0]} leads the category turns proxy, and {backlog_top[0]} is the largest backlog order."
            },
        ],
    }


def build_case06():
    case_dir = os.path.join(BASE, "Case06")
    ensure_dir(case_dir)

    departments = [["DepartmentID", "DepartmentName", "Function", "Region", "BudgetOwner", "TargetUtilisation", "HiringPlan"]]
    dept_names = ["Consulting", "Data", "Finance", "HR", "Operations", "Sales", "Support", "Legal"]
    for i in range(80):
        departments.append([
            f"DP{i+1:03d}",
            dept_names[i % len(dept_names)],
            ["Client", "Corporate"][i % 2],
            ["UK", "EU", "NA", "APAC"][i % 4],
            f"Owner_{(i % 12) + 1:02d}",
            72 + (i % 12),
            1 + (i % 6),
        ])

    employees = [["EmployeeID", "EmployeeName", "DepartmentID", "RoleFamily", "Grade", "Location", "HireYear", "ManagerID"]]
    role_families = ["Analyst", "Engineer", "Consultant", "Specialist", "Manager"]
    for i in range(84):
        employees.append([
            f"E{i+1:03d}",
            f"Employee_{i+1:03d}",
            f"DP{((i * 3) % 80) + 1:03d}",
            role_families[i % len(role_families)],
            ["G1", "G2", "G3", "G4"][i % 4],
            ["London", "Leeds", "Bristol", "Manchester"][i % 4],
            2018 + (i % 7),
            f"M{(i % 18) + 1:03d}",
        ])

    timesheets = [["EntryID", "EmployeeID", "Week", "ProjectType", "BillableHours", "NonBillableHours", "LeaveHours"]]
    for i in range(120):
        timesheets.append([
            f"TS{i+1:03d}",
            f"E{((i * 5) % 84) + 1:03d}",
            f"2024-W{(i % 12) + 1:02d}",
            ["Client", "Internal", "Support"][i % 3],
            20 + ((i * 3) % 21),
            4 + ((i * 5) % 10),
            i % 4,
        ])

    training = [["TrainID", "EmployeeID", "Program", "RequiredFlag", "CompletedFlag", "Score"]]
    programs = ["Security", "Compliance", "Data_Privacy", "Leadership", "AI_Foundations"]
    for i in range(110):
        required = "Yes" if i % 3 != 0 else "No"
        completed = "Yes" if i % 5 != 0 else "No"
        training.append([
            f"TR{i+1:03d}",
            f"E{((i * 7) % 84) + 1:03d}",
            programs[i % len(programs)],
            required,
            completed,
            65 + ((i * 4) % 31),
        ])

    performance = [["EmployeeID", "Q1", "Q2", "Q3", "Q4", "ManagerRating", "Potential"]]
    for i in range(84):
        performance.append([
            f"E{i+1:03d}",
            68 + ((i * 2) % 28),
            70 + ((i * 3) % 25),
            69 + ((i * 5) % 27),
            71 + ((i * 7) % 23),
            3 + (i % 3),
            ["Low", "Medium", "High"][i % 3],
        ])

    write_csv(os.path.join(case_dir, "tc06_input01.csv"), departments)
    write_csv(os.path.join(case_dir, "tc06_input02.csv"), employees)
    write_csv(os.path.join(case_dir, "tc06_input03.csv"), timesheets)
    write_csv(os.path.join(case_dir, "tc06_input04.csv"), training)
    write_csv(os.path.join(case_dir, "tc06_input05.csv"), performance)

    dept_map = {row[0]: row for row in departments[1:]}
    emp_map = {row[0]: row for row in employees[1:]}
    perf_map = {row[0]: row for row in performance[1:]}

    hours_by_emp = defaultdict(lambda: {"billable": 0, "nonbillable": 0, "leave": 0})
    for _, employee_id, _, _, billable, nonbillable, leave in timesheets[1:]:
        hours_by_emp[employee_id]["billable"] += int(billable)
        hours_by_emp[employee_id]["nonbillable"] += int(nonbillable)
        hours_by_emp[employee_id]["leave"] += int(leave)

    training_by_emp = defaultdict(lambda: {"required": 0, "completed_required": 0, "score_pairs": []})
    for _, employee_id, _, required_flag, completed_flag, score in training[1:]:
        if required_flag == "Yes":
            training_by_emp[employee_id]["required"] += 1
            if completed_flag == "Yes":
                training_by_emp[employee_id]["completed_required"] += 1
        training_by_emp[employee_id]["score_pairs"].append((int(score), 1))

    employee_rows = []
    dept_stats = defaultdict(lambda: {"headcount": 0, "util_pairs": [], "perf_pairs": [], "req": 0, "completed_req": 0, "high_potential": 0})
    risk_rows = []
    for employee_id, employee_name, department_id, role_family, grade, location, hire_year, manager_id in employees[1:]:
        hrs = hours_by_emp[employee_id]
        total_work = hrs["billable"] + hrs["nonbillable"]
        util = pct(hrs["billable"], total_work)
        train = training_by_emp[employee_id]
        completion = pct(train["completed_required"], train["required"])
        perf = perf_map[employee_id]
        avg_perf = round((int(perf[1]) + int(perf[2]) + int(perf[3]) + int(perf[4])) / 4, 2)
        if avg_perf >= 90:
            tier = "Top"
        elif avg_perf >= 80:
            tier = "Strong"
        elif avg_perf >= 72:
            tier = "Core"
        else:
            tier = "At_Risk"
        employee_rows.append([
            employee_id,
            employee_name,
            dept_map[department_id][1],
            role_family,
            util,
            completion,
            avg_perf,
            perf[6],
            tier,
        ])
        dept_stats[department_id]["headcount"] += 1
        dept_stats[department_id]["util_pairs"].append((util, 1))
        dept_stats[department_id]["perf_pairs"].append((avg_perf, 1))
        dept_stats[department_id]["req"] += train["required"]
        dept_stats[department_id]["completed_req"] += train["completed_required"]
        if perf[6] == "High":
            dept_stats[department_id]["high_potential"] += 1
        if util < 65 or completion < 100 or tier == "At_Risk":
            risk_rows.append([
                employee_id,
                employee_name,
                dept_map[department_id][1],
                util,
                completion,
                avg_perf,
                tier,
            ])

    employee_rows.sort(key=lambda row: (row[6], row[4]), reverse=True)
    risk_rows.sort(key=lambda row: (row[6], row[3], row[4]))

    dept_rows = []
    for department_id, stats in dept_stats.items():
        dept_rows.append([
            department_id,
            dept_map[department_id][1],
            dept_map[department_id][3],
            stats["headcount"],
            weighted_avg(stats["util_pairs"]),
            pct(stats["completed_req"], stats["req"]),
            weighted_avg(stats["perf_pairs"]),
            stats["high_potential"],
        ])
    dept_rows.sort(key=lambda row: row[6], reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Employee_Summary"
    ws.append(["EmployeeID", "EmployeeName", "Department", "RoleFamily", "Utilisation_Pct", "RequiredTrainingCompletion_Pct", "AvgPerformance", "Potential", "PerformanceTier"])
    for row in employee_rows:
        ws.append(row)

    ws2 = wb.create_sheet("Department_Summary")
    ws2.append(["DepartmentID", "DepartmentName", "Region", "Headcount", "AvgUtilisation_Pct", "RequiredTrainingCompletion_Pct", "AvgPerformance", "HighPotentialCount"])
    for row in dept_rows:
        ws2.append(row)

    ws3 = wb.create_sheet("Risk_List")
    ws3.append(["EmployeeID", "EmployeeName", "Department", "Utilisation_Pct", "RequiredTrainingCompletion_Pct", "AvgPerformance", "PerformanceTier"])
    for row in risk_rows[:25]:
        ws3.append(row)

    save_wb(os.path.join(case_dir, "tc06_output01.xlsx"), wb)

    top_employee = employee_rows[0]
    top_department = dept_rows[0]
    risk_employee = risk_rows[0]
    return {
        "task_id": "Test 6",
        "title": "Large Workforce Utilisation and Performance Analysis",
        "scenario": "HR analytics / workforce planning",
        "category": "join five clean tables; calculate employee utilisation, required-training completion, and average performance; summarise departments; create employee risk list",
        "atomic_operations": [
            "join employees with departments on DepartmentID",
            "join timesheets with employees on EmployeeID",
            "join training records with employees on EmployeeID",
            "join performance records with employees on EmployeeID",
            "compute employee utilisation",
            "compute required-training completion rate",
            "compute average Q1-Q4 performance",
            "assign performance tier",
            "aggregate employee metrics to department summaries",
            "count high-potential employees by department",
            "filter employees into a risk list",
            "export results to a multi-sheet workbook"
        ],
        "number_of_atomic_operations": 12,
        "spreadsheets": [
            "Case06/tc06_input01.csv",
            "Case06/tc06_input02.csv",
            "Case06/tc06_input03.csv",
            "Case06/tc06_input04.csv",
            "Case06/tc06_input05.csv",
        ],
        "prompt": "You have five clean HR tables: departments, employees, timesheets, training records, and performance scores. Join the data using DepartmentID and EmployeeID. For each employee, calculate utilisation as BillableHours / (BillableHours + NonBillableHours), required-training completion rate, average performance across Q1-Q4, and assign a performance tier (Top, Strong, Core, At_Risk). Then produce a department summary with headcount, average utilisation, required-training completion, average performance, and high-potential count. Finally, create a risk list for employees with low utilisation, incomplete required training, or At_Risk performance. Output the results into a new Excel workbook with sheets Employee_Summary, Department_Summary, and Risk_List.",
        "answer": f"Top employee: {top_employee[1]} with average performance {top_employee[6]:.2f} and utilisation {top_employee[4]}%. Highest-performing department summary row: {top_department[1]} ({top_department[2]}) with average performance {top_department[6]:.2f}. First employee on the risk list: {risk_employee[1]} ({risk_employee[2]}), tier {risk_employee[6]}.",
        "expected_output_file": ["Case06/tc06_output01.xlsx"],
        "feedback": "All input tables are clean and keyed. I will compute employee-level utilisation and training completion first, then aggregate those employee metrics to the department level and build the risk list from the same employee summary.\n\n---",
        "conversations": [
            {
                "role": "user",
                "content": "Join the five HR tables, compute employee utilisation and performance, summarise departments, and generate a risk list."
            },
            {
                "role": "assistant",
                "content": "I will derive a single employee-level summary first so the department summary and risk list are both based on the same utilisation, training, and performance calculations."
            },
            {
                "role": "user",
                "content": "Proceed."
            },
            {
                "role": "assistant",
                "content": f"Completed. {top_employee[1]} leads the employee summary, {top_department[1]} is the strongest department by average performance, and {risk_employee[1]} appears first on the risk list."
            },
        ],
    }


def main():
    ensure_dir(BASE)
    cases = [
        build_case01(),
        build_case02(),
        build_case03(),
        build_case04(),
        build_case05(),
        build_case06(),
    ]

    with open(os.path.join(BASE, "dataset.json"), "w", encoding="utf-8") as f:
        json.dump(cases, f, indent=4, ensure_ascii=False)

    print(f"Generated {len(cases)} large clean cases in {BASE}")


if __name__ == "__main__":
    main()
