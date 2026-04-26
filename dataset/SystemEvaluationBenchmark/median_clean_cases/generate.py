"""Generate all input and output files for median_clean_cases Cases 01-10."""
import os, csv, random, itertools
from openpyxl import Workbook
from collections import defaultdict, Counter

BASE = os.path.dirname(os.path.abspath(__file__))

def wb_save(wb, path):
    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# CASE 01  Product Sales Matrix  (multi-header CSV, 12 regions)
# ─────────────────────────────────────────────────────────────────────────────
def case01():
    d = os.path.join(BASE, "Case01")
    regions = ["North","South","East","West","Central","Northwest","Northeast",
               "Southeast","Southwest","Midwest","Pacific","Mountain"]
    random.seed(101)
    # Build raw data: (region, [e_q1..q4], [c_q1..q4], [h_q1..q4])
    raw = []
    for r in regions:
        e = [random.randint(80, 260) for _ in range(4)]
        c = [random.randint(38, 115) for _ in range(4)]
        h = [random.randint(60, 175) for _ in range(4)]
        raw.append((r, e, c, h))

    rows = [
        ["Region","Electronics","","","","Clothing","","","","Home_Appliances","","",""],
        ["","Q1","Q2","Q3","Q4","Q1","Q2","Q3","Q4","Q1","Q2","Q3","Q4"],
    ]
    for r, e, c, h in raw:
        rows.append([r] + e + c + h)
    with open(os.path.join(d, "tc01_input01.csv"), "w", newline="") as f:
        csv.writer(f).writerows(rows)

    wb = Workbook()
    ws = wb.active; ws.title = "Annual_Sales"
    ws.append(["Region","Electronics_Annual","Clothing_Annual","Home_Appliances_Annual","Grand_Total"])
    for r, e, c, h in raw:
        ea, ca, ha = sum(e), sum(c), sum(h)
        ws.append([r, ea, ca, ha, ea+ca+ha])

    ws2 = wb.create_sheet("Q4_Leaders")
    ws2.append(["Region","Product","Q4_Sales"])
    q4 = []
    for r, e, c, h in raw:
        q4 += [(r,"Electronics",e[3]),(r,"Clothing",c[3]),(r,"Home_Appliances",h[3])]
    for row in sorted(q4, key=lambda x: -x[2]):
        ws2.append(list(row))

    ws3 = wb.create_sheet("Category_Total")
    ws3.append(["Category","Q1_Total","Q2_Total","Q3_Total","Q4_Total","Annual_Total"])
    for cat_idx, cat_name in enumerate(["Electronics","Clothing","Home_Appliances"]):
        data_col = [raw[i][cat_idx+1] for i in range(len(raw))]
        qtots = [sum(row[q] for row in data_col) for q in range(4)]
        ws3.append([cat_name] + qtots + [sum(qtots)])

    wb_save(wb, os.path.join(d, "tc01_output01.xlsx"))
    print("Case01 done")


# ─────────────────────────────────────────────────────────────────────────────
# CASE 02  Employee Performance & Bonus  (50 employees, two CSVs)
# ─────────────────────────────────────────────────────────────────────────────
def case02():
    d = os.path.join(BASE, "Case02")
    random.seed(202)
    first = ["Alice","Bob","Carol","David","Eva","Frank","Grace","Henry","Iris","Jack",
             "Karen","Leo","Maria","Nathan","Olivia","Paul","Quinn","Rachel","Sam","Tina",
             "Uma","Victor","Wendy","Xavier","Yara","Zach","Ava","Blake","Chloe","Dylan",
             "Emma","Finn","Gina","Hugo","Isla","Jake","Kira","Liam","Maya","Nora",
             "Owen","Petra","Ravi","Sofia","Tom","Ursula","Vince","Wren","Yusuf","Zoe"]
    last  = ["Chen","Martinez","White","Kim","Brown","Liu","Park","Nguyen","Scott","Adams",
             "Hill","Turner","Santos","Kowalski","Ahmad","Patel","Osei","Rossi","Müller","Jones",
             "Garcia","Tanaka","Singh","Ferreira","Hassan","Novak","Weber","Andersen","Dupont","Ito",
             "Nkosi","Cruz","Petrov","Yamamoto","Bakker","Svensson","Ali","Markov","Costa","Diallo",
             "Fischer","Herrera","Ivanova","Johansson","Kaur","Larsson","Moreau","Nielsen","Ozturk","Pereira"]
    depts = ["Engineering","Engineering","Engineering","Engineering","Engineering",
             "Engineering","Engineering","Engineering","Engineering","Engineering",
             "Marketing","Marketing","Marketing","Marketing","Marketing",
             "Marketing","Marketing","Marketing","Marketing","Marketing",
             "Sales","Sales","Sales","Sales","Sales",
             "Sales","Sales","Sales","Sales","Sales",
             "HR","HR","HR","HR","HR",
             "Finance","Finance","Finance","Finance","Finance",
             "Operations","Operations","Operations","Operations","Operations",
             "Legal","Legal","Legal","Legal","Legal"]
    base_ranges = {
        "Engineering":(85000,120000),"Marketing":(65000,90000),"Sales":(58000,78000),
        "HR":(55000,72000),"Finance":(75000,100000),"Operations":(60000,82000),"Legal":(90000,115000),
    }
    employees = [["EmpID","Name","Department","BaseSalary_USD"]]
    scores_data = [["EmpID","Q1_Score","Q2_Score","Q3_Score","Q4_Score"]]
    for i in range(50):
        eid = f"E{i+1:02d}"
        name = f"{first[i]} {last[i]}"
        dept = depts[i]
        lo, hi = base_ranges[dept]
        salary = random.randint(lo//1000, hi//1000) * 1000
        employees.append([eid, name, dept, salary])
        qs = [random.randint(55, 99) for _ in range(4)]
        scores_data.append([eid] + qs)

    with open(os.path.join(d, "tc02_input01.csv"), "w", newline="") as f:
        csv.writer(f).writerows(employees)
    with open(os.path.join(d, "tc02_input02.csv"), "w", newline="") as f:
        csv.writer(f).writerows(scores_data)

    emp_map   = {r[0]: r for r in employees[1:]}
    score_map = {r[0]: r for r in scores_data[1:]}
    wb = Workbook()
    ws = wb.active; ws.title = "Bonus_Report"
    ws.append(["EmpID","Name","Department","BaseSalary","Avg_Score","Bonus_USD","Qualifies_High_Performer"])
    results = []
    for eid, sr in score_map.items():
        er = emp_map[eid]
        avg = sum(int(x) for x in sr[1:]) / 4
        salary = int(er[3])
        bonus = round(salary * (avg / 100) * 0.15, 2)
        results.append([eid, er[1], er[2], salary, round(avg,2), bonus, "Yes" if avg>=85 else "No"])
    results.sort(key=lambda x: -x[5])
    for r in results:
        ws.append(r)

    ws2 = wb.create_sheet("Dept_Summary")
    ws2.append(["Department","Headcount","Avg_Score","Total_Bonus_USD","High_Performers"])
    dept_agg = defaultdict(lambda: {"scores":[],"bonuses":[],"hp":0})
    for r in results:
        dept_agg[r[2]]["scores"].append(r[4])
        dept_agg[r[2]]["bonuses"].append(r[5])
        if r[6]=="Yes": dept_agg[r[2]]["hp"] += 1
    for dept in sorted(dept_agg):
        ag = dept_agg[dept]
        ws2.append([dept, len(ag["scores"]), round(sum(ag["scores"])/len(ag["scores"]),2),
                    round(sum(ag["bonuses"]),2), ag["hp"]])

    ws3 = wb.create_sheet("Top10_Earners")
    ws3.append(["EmpID","Name","Department","Bonus_USD"])
    for r in results[:10]:
        ws3.append([r[0],r[1],r[2],r[5]])

    wb_save(wb, os.path.join(d, "tc02_output01.xlsx"))
    print("Case02 done")


# ─────────────────────────────────────────────────────────────────────────────
# CASE 03  Monthly Climate Data  (10 cities, multi-header CSV)
# ─────────────────────────────────────────────────────────────────────────────
def case03():
    d = os.path.join(BASE, "Case03")
    cities = ["London","Paris","Berlin","Madrid","Rome","Amsterdam","Vienna","Barcelona","Prague","Lisbon"]
    # (temp_c, rainfall_mm, humidity_%) per month
    monthly = {
        "London":  [(5,55,82),(6,40,78),(8,45,73),(11,42,68),(15,48,66),(18,45,65),(20,40,63),(20,50,65),(17,52,70),(13,60,76),(9,62,80),(6,58,83)],
        "Paris":   [(4,50,85),(5,38,80),(9,42,72),(13,40,67),(17,45,65),(21,42,62),(24,38,60),(23,47,62),(19,48,68),(14,55,74),(8,55,80),(5,52,85)],
        "Berlin":  [(1,42,84),(2,36,80),(6,38,72),(12,38,65),(17,50,62),(21,62,63),(23,55,61),(23,50,62),(18,42,68),(12,35,72),(6,45,80),(2,50,85)],
        "Madrid":  [(8,35,70),(10,28,65),(14,32,55),(17,36,50),(21,40,45),(28,18,38),(32,8,32),(32,10,33),(25,22,45),(18,42,58),(12,48,65),(8,40,70)],
        "Rome":    [(8,80,72),(9,70,68),(12,65,60),(16,60,55),(21,45,52),(26,25,45),(29,18,42),(29,22,43),(25,60,55),(18,95,63),(13,110,68),(9,90,72)],
        "Amsterdam":[(4,62,84),(4,48,80),(7,52,74),(11,44,68),(15,52,67),(18,58,67),(20,68,70),(20,72,72),(17,68,76),(13,70,80),(8,72,83),(5,68,85)],
        "Vienna":  [(0,38,80),(1,32,76),(6,40,67),(11,42,60),(16,58,60),(20,68,60),(22,72,58),(22,65,58),(17,48,64),(11,40,70),(5,50,78),(1,44,82)],
        "Barcelona":[(10,42,65),(11,35,62),(13,38,58),(16,42,55),(19,42,60),(24,26,58),(27,15,55),(27,20,57),(24,65,63),(19,80,66),(14,68,68),(11,52,67)],
        "Prague":  [(-1,28,80),(0,24,76),(5,30,68),(10,32,60),(15,48,58),(18,62,60),(20,68,60),(20,60,60),(16,38,66),(10,34,74),(4,44,80),(0,38,82)],
        "Lisbon":  [(11,82,72),(12,68,68),(14,52,64),(16,44,60),(18,34,58),(22,10,52),(25,3,48),(25,5,48),(23,28,56),(18,62,64),(14,82,68),(12,90,72)],
    }
    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    header0 = ["Month"] + [c for c in cities for _ in range(3)]
    header1 = [""]      + ["Temp_C","Rainfall_mm","Humidity_%"] * len(cities)
    rows = [header0, header1]
    for i, m in enumerate(months):
        row = [m]
        for c in cities:
            row += list(monthly[c][i])
        rows.append(row)
    with open(os.path.join(d, "tc03_input01.csv"), "w", newline="") as f:
        csv.writer(f).writerows(rows)

    wb = Workbook()
    ws = wb.active; ws.title = "City_Summary"
    ws.append(["City","Annual_Avg_Temp_C","Hottest_Month","Coldest_Month","Total_Rainfall_mm","Avg_Humidity_%"])
    for c in cities:
        temps = [monthly[c][i][0] for i in range(12)]
        rain  = [monthly[c][i][1] for i in range(12)]
        hum   = [monthly[c][i][2] for i in range(12)]
        ws.append([c, round(sum(temps)/12,1), months[temps.index(max(temps))],
                   months[temps.index(min(temps))], sum(rain), round(sum(hum)/12,1)])

    ws2 = wb.create_sheet("Summer_Avg_Temp")
    ws2.append(["City","Summer_Avg_Temp_C (Jun-Aug)"])
    for c in cities:
        summer = [monthly[c][i][0] for i in [5,6,7]]
        ws2.append([c, round(sum(summer)/3,1)])

    ws3 = wb.create_sheet("Driest_Month_Per_City")
    ws3.append(["City","Driest_Month","Rainfall_mm"])
    for c in cities:
        rain_monthly = [monthly[c][i][1] for i in range(12)]
        driest_idx = rain_monthly.index(min(rain_monthly))
        ws3.append([c, months[driest_idx], min(rain_monthly)])

    wb_save(wb, os.path.join(d, "tc03_output01.xlsx"))
    print("Case03 done")


# ─────────────────────────────────────────────────────────────────────────────
# CASE 04  Hospital Records  (60 patients, 100 admissions, multi-sheet XLSX)
# ─────────────────────────────────────────────────────────────────────────────
def case04():
    d = os.path.join(BASE, "Case04")
    random.seed(404)
    first_m = ["James","Robert","Michael","William","David","Richard","Joseph","Thomas",
               "Charles","Christopher","Daniel","Matthew","Anthony","Mark","Donald","Steven",
               "Paul","Andrew","Kenneth","George","Joshua","Kevin","Brian","Timothy","Ronald",
               "Edward","Jason","Jeffrey","Ryan","Jacob"]
    first_f = ["Mary","Patricia","Jennifer","Linda","Barbara","Elizabeth","Susan","Jessica",
               "Sarah","Karen","Lisa","Nancy","Betty","Margaret","Sandra","Ashley","Dorothy",
               "Kimberly","Emily","Donna","Michelle","Carol","Amanda","Melissa","Deborah",
               "Stephanie","Rebecca","Laura","Sharon","Cynthia"]
    last_n  = ["Walker","Connor","Johnson","Davis","Lee","Brown","Harris","Clark","Lewis","Young",
               "King","Wright","Scott","Green","Hall","Baker","Nelson","Carter","Mitchell","Perez",
               "Roberts","Turner","Phillips","Campbell","Parker","Evans","Edwards","Collins","Stewart","Morris"]
    blood_types = ["A+","A-","B+","B-","O+","O-","AB+","AB-"]
    patients = []
    for i in range(60):
        gender = "M" if i < 30 else "F"
        fname = first_m[i % 30] if gender == "M" else first_f[i % 30]
        lname = last_n[i % 30]
        age = random.randint(18, 82)
        bt = random.choice(blood_types)
        patients.append((f"P{i+1:03d}", f"{fname} {lname}", age, gender, bt))

    wards = ["Cardiology","General","Orthopedics","Oncology","Neurology","Pediatrics"]
    diagnoses = {
        "Cardiology":  ["Hypertension","Arrhythmia","Heart Failure","Angina","Myocarditis"],
        "General":     ["Appendicitis","Gastroenteritis","Bronchitis","Influenza","Pneumonia","Minor Injury"],
        "Orthopedics": ["Hip Fracture","Knee Replacement","Spinal Stenosis","Post-op Review","Fracture"],
        "Oncology":    ["Breast Cancer","Colon Cancer","Cervical Cancer","Chemotherapy","Lymphoma"],
        "Neurology":   ["Stroke","Migraine","Epilepsy","Parkinson","Multiple Sclerosis"],
        "Pediatrics":  ["Asthma","RSV","Ear Infection","Tonsillitis","Fever"],
    }
    admissions = []
    # 100 admissions spread across 60 patients (some repeat)
    adm_id = 1
    year_months = [(2024, m) for m in range(1, 13)]
    for step in range(100):
        # pick patient - first 40 steps use random patient, next 60 include repeats for first 20
        if step < 60:
            pid_idx = random.randint(0, 59)
        else:
            pid_idx = random.randint(0, 19)  # repeats for first 20 patients
        pid = patients[pid_idx][0]
        ward = random.choice(wards)
        diag = random.choice(diagnoses[ward])
        yr, mo = random.choice(year_months)
        day = random.randint(1, 25)
        los = random.randint(1, 10)
        admit = f"{yr}-{mo:02d}-{day:02d}"
        # discharge = admit + los days (simple string, not parsed)
        day2 = min(day + los, 28)
        discharge = f"{yr}-{mo:02d}-{day2:02d}"
        admissions.append((f"A{adm_id:03d}", pid, ward, admit, discharge, diag, los))
        adm_id += 1

    wb_in = Workbook()
    ws_p = wb_in.active; ws_p.title = "Patients"
    ws_p.append(["PatientID","Name","Age","Gender","BloodType"])
    for p in patients:
        ws_p.append(list(p))

    ws_a = wb_in.create_sheet("Admissions")
    ws_a.append(["AdmissionID","PatientID","Ward","AdmitDate","DischargeDate","Diagnosis","LOS_Days"])
    for a in admissions:
        ws_a.append(list(a))
    wb_in.save(os.path.join(d, "tc04_input01.xlsx"))

    # Output
    wb = Workbook()
    ws = wb.active; ws.title = "Ward_Summary"
    ws.append(["Ward","Num_Admissions","Avg_LOS_Days","Most_Common_Diagnosis"])
    ward_data = defaultdict(lambda: {"count":0,"los":[],"diags":[]})
    for a in admissions:
        ward_data[a[2]]["count"] += 1
        ward_data[a[2]]["los"].append(a[6])
        ward_data[a[2]]["diags"].append(a[5])
    for ward in sorted(ward_data):
        info = ward_data[ward]
        common = Counter(info["diags"]).most_common(1)[0][0]
        ws.append([ward, info["count"], round(sum(info["los"])/len(info["los"]),1), common])

    ws2 = wb.create_sheet("Repeat_Patients")
    ws2.append(["PatientID","Name","Num_Admissions","Wards_Visited"])
    pat_adm = defaultdict(lambda: {"count":0,"wards":set()})
    for a in admissions:
        pat_adm[a[1]]["count"] += 1
        pat_adm[a[1]]["wards"].add(a[2])
    pat_map = {p[0]: p[1] for p in patients}
    for pid, info in sorted(pat_adm.items()):
        if info["count"] > 1:
            ws2.append([pid, pat_map[pid], info["count"], ", ".join(sorted(info["wards"]))])

    ws3 = wb.create_sheet("Monthly_Admissions")
    ws3.append(["Month","Num_Admissions","Avg_LOS_Days"])
    month_data = defaultdict(lambda: {"count":0,"los":[]})
    for a in admissions:
        mo = a[3][:7]  # "YYYY-MM"
        month_data[mo]["count"] += 1
        month_data[mo]["los"].append(a[6])
    for mo in sorted(month_data):
        info = month_data[mo]
        ws3.append([mo, info["count"], round(sum(info["los"])/len(info["los"]),1)])

    wb_save(wb, os.path.join(d, "tc04_output01.xlsx"))
    print("Case04 done")


# ─────────────────────────────────────────────────────────────────────────────
# CASE 05  Retail Inventory + Monthly Sales  (30 products, 6 months)
# ─────────────────────────────────────────────────────────────────────────────
def case05():
    d = os.path.join(BASE, "Case05")
    random.seed(505)
    products = [
        ["ProductID","ProductName","Category","StockQty","UnitCost_USD","ReorderLevel"],
        ["PR01","Laptop 14\"","Electronics",30,420.00,8],
        ["PR02","Laptop 15\"","Electronics",20,495.00,6],
        ["PR03","Wireless Mouse","Electronics",100,18.50,25],
        ["PR04","Mechanical Keyboard","Electronics",55,72.00,15],
        ["PR05","USB-C Hub","Electronics",70,32.00,18],
        ["PR06","Monitor 24\"","Electronics",25,185.00,8],
        ["PR07","Webcam HD","Electronics",45,48.00,12],
        ["PR08","Office Chair","Furniture",22,185.00,6],
        ["PR09","Standing Desk","Furniture",10,340.00,4],
        ["PR10","Desk Lamp","Furniture",55,28.00,12],
        ["PR11","Bookshelf","Furniture",18,95.00,5],
        ["PR12","Filing Cabinet","Furniture",14,115.00,4],
        ["PR13","A4 Notebooks (10pk)","Stationery",180,8.50,50],
        ["PR14","Ballpoint Pens (20pk)","Stationery",220,4.20,60],
        ["PR15","Sticky Notes (5pk)","Stationery",200,3.80,60],
        ["PR16","Stapler","Stationery",60,12.00,15],
        ["PR17","Scissors Set","Stationery",40,7.50,10],
        ["PR18","Coffee Beans 1kg","Pantry",65,14.00,22],
        ["PR19","Instant Coffee 200g","Pantry",95,6.50,35],
        ["PR20","Green Tea 50 bags","Pantry",80,5.20,28],
        ["PR21","Protein Bars (12pk)","Pantry",50,18.00,15],
        ["PR22","Mineral Water 12pk","Pantry",120,8.00,40],
        ["PR23","Hand Sanitiser 500ml","Health",120,3.90,35],
        ["PR24","Paracetamol 24pk","Health",90,2.80,28],
        ["PR25","Face Masks 50pk","Health",65,9.50,22],
        ["PR26","First Aid Kit","Health",30,22.00,8],
        ["PR27","Vitamin C 60 tabs","Health",75,11.00,20],
        ["PR28","Whiteboard A1","Office",12,45.00,3],
        ["PR29","Marker Set (8pc)","Office",55,8.00,15],
        ["PR30","Printer Paper (500sh)","Office",90,6.50,25],
    ]
    sale_prices = {
        "PR01":649.99,"PR02":769.99,"PR03":28.99,"PR04":109.99,"PR05":49.99,
        "PR06":279.99,"PR07":74.99,"PR08":279.99,"PR09":519.99,"PR10":39.99,
        "PR11":149.99,"PR12":179.99,"PR13":12.99,"PR14":6.49,"PR15":5.99,
        "PR16":18.99,"PR17":11.99,"PR18":21.99,"PR19":9.99,"PR20":7.99,
        "PR21":26.99,"PR22":11.99,"PR23":5.99,"PR24":4.29,"PR25":14.99,
        "PR26":34.99,"PR27":16.99,"PR28":64.99,"PR29":12.99,"PR30":9.99,
    }
    months = ["Jan","Feb","Mar","Apr","May","Jun"]
    # Generate sales qty: roughly realistic, using seed
    sales_rows = [["ProductID","Month","UnitsSold","SalePrice_USD"]]
    base_qty = {
        "PR01":11,"PR02":8,"PR03":38,"PR04":22,"PR05":28,"PR06":12,"PR07":18,
        "PR08":7,"PR09":3,"PR10":20,"PR11":5,"PR12":4,"PR13":65,"PR14":85,
        "PR15":75,"PR16":22,"PR17":15,"PR18":24,"PR19":38,"PR20":30,
        "PR21":18,"PR22":55,"PR23":45,"PR24":32,"PR25":24,"PR26":10,
        "PR27":28,"PR28":4,"PR29":20,"PR30":38,
    }
    sales_qty = {}
    for pid, base in base_qty.items():
        for m in months:
            vary = random.randint(-3, 6)
            sales_qty[(pid, m)] = max(1, base + vary)
    for pid in [f"PR{i:02d}" for i in range(1, 31)]:
        for m in months:
            sales_rows.append([pid, m, sales_qty[(pid, m)], sale_prices[pid]])

    with open(os.path.join(d, "tc05_input01.csv"), "w", newline="") as f:
        csv.writer(f).writerows(products)
    with open(os.path.join(d, "tc05_input02.csv"), "w", newline="") as f:
        csv.writer(f).writerows(sales_rows)

    prod_map = {r[0]: r for r in products[1:]}
    cat_revenue = defaultdict(float)
    cat_cost    = defaultdict(float)
    for row in sales_rows[1:]:
        pid, month, qty, sp = row[0], row[1], int(row[2]), float(row[3])
        pr = prod_map[pid]; cat = pr[2]; cost = float(pr[4])
        cat_revenue[cat] += qty * sp
        cat_cost[cat]    += qty * cost

    wb = Workbook()
    ws = wb.active; ws.title = "Category_Revenue"
    ws.append(["Category","Total_Revenue_USD","Total_COGS_USD","Gross_Profit_USD","Margin_%"])
    for cat in sorted(cat_revenue):
        rev = round(cat_revenue[cat], 2); cogs = round(cat_cost[cat], 2)
        gp = round(rev - cogs, 2)
        margin = round((gp/rev)*100, 1) if rev else 0
        ws.append([cat, rev, cogs, gp, margin])

    ws2 = wb.create_sheet("Low_Stock_Alert")
    ws2.append(["ProductID","ProductName","Remaining_Stock","Reorder_Level"])
    low_stock = []
    for pid, pr in prod_map.items():
        total_sold = sum(sales_qty[(pid, m)] for m in months)
        rem = int(pr[3]) - total_sold
        if rem < int(pr[5]):
            low_stock.append((pid, pr[1], rem, int(pr[5])))
    for row in sorted(low_stock, key=lambda x: x[2]):
        ws2.append(list(row))

    ws3 = wb.create_sheet("Monthly_Revenue")
    ws3.append(["Month","Total_Revenue_USD","Total_Units_Sold"])
    for m in months:
        rev = sum(sales_qty[(pid,m)] * sale_prices[pid] for pid in [f"PR{i:02d}" for i in range(1,31)])
        units = sum(sales_qty[(pid,m)] for pid in [f"PR{i:02d}" for i in range(1,31)])
        ws3.append([m, round(rev,2), units])

    wb_save(wb, os.path.join(d, "tc05_output01.xlsx"))
    print("Case05 done")


# ─────────────────────────────────────────────────────────────────────────────
# CASE 06  Student Exam Results  (60 students, 4 classes)
# ─────────────────────────────────────────────────────────────────────────────
def case06():
    d = os.path.join(BASE, "Case06")
    random.seed(606)
    first_names = [
        "Aisha","Ben","Clara","Daniel","Elena","Finn","Grace","Hassan","Isabella","Jamal",
        "Kira","Luca","Maya","Nate","Olivia","Pedro","Quinn","Ravi","Sofia","Tom",
        "Uma","Victor","Wendy","Xian","Yara","Zach","Amara","Boris","Celia","Diego",
        "Eva","Felix","Gia","Hugo","Iris","Jonas","Kemi","Lars","Mia","Noah",
        "Opal","Pita","Rosa","Sven","Tara","Umar","Vera","Will","Xena","Yuki",
        "Zara","Abel","Bea","Carlos","Dana","Eli","Fern","Gus","Hana","Ivan",
    ]
    last_names = [
        "Patel","Foster","Zhang","Osei","Rossi","McCarthy","Liu","Ali","Müller","Brown",
        "Johansson","Ferrari","Singh","Williams","Chen","Andrade","Park","Sharma","Kim","Carter",
        "Hassan","Novak","Weber","Tanaka","Ferreira","Fischer","Diallo","Petrov","Cruz","Herrera",
        "Svensson","Bakker","Costa","Yamamoto","Kowalski","Ahmad","Nkosi","Larsson","Moreau","Nielsen",
        "Ozturk","Pereira","Markov","Ivanova","Ito","Santos","Dupont","Andersen","Garcia","Jones",
        "Adams","Hill","Turner","Scott","Green","Hall","Baker","Nelson","Mitchell","Perez",
    ]
    classes = ["10A","10A","10A","10A","10A","10A","10A","10A","10A","10A",
               "10A","10A","10A","10A","10A","10B","10B","10B","10B","10B",
               "10B","10B","10B","10B","10B","10B","10B","10B","10B","10B",
               "10C","10C","10C","10C","10C","10C","10C","10C","10C","10C",
               "10C","10C","10C","10C","10C","10D","10D","10D","10D","10D",
               "10D","10D","10D","10D","10D","10D","10D","10D","10D","10D"]
    # Generate scores: mix of profiles
    students = [["StudentID","Name","Class","Math","Physics","Chemistry","English","History"]]
    for i in range(60):
        sid = f"S{i+1:02d}"
        name = f"{first_names[i]} {last_names[i]}"
        cls = classes[i]
        # Varied performance levels
        profile = random.choice(["high","mid","low","mixed"])
        if profile == "high":
            scores = [random.randint(78, 99) for _ in range(5)]
        elif profile == "mid":
            scores = [random.randint(62, 85) for _ in range(5)]
        elif profile == "low":
            scores = [random.randint(38, 68) for _ in range(5)]
        else:
            scores = [random.randint(45, 99) for _ in range(5)]
        students.append([sid, name, cls] + scores)

    with open(os.path.join(d, "tc06_input01.csv"), "w", newline="") as f:
        csv.writer(f).writerows(students)

    subjects = ["Math","Physics","Chemistry","English","History"]
    wb = Workbook()
    ws = wb.active; ws.title = "GPA_Report"
    ws.append(["StudentID","Name","Class","Math","Physics","Chemistry","English","History",
               "Avg_Score","Grade","Pass"])
    for row in students[1:]:
        sid, name, cls = row[0], row[1], row[2]
        scores = [int(x) for x in row[3:]]
        avg = round(sum(scores)/len(scores), 1)
        grade = "A" if avg>=90 else "B" if avg>=80 else "C" if avg>=70 else "D" if avg>=60 else "F"
        passed = "Yes" if all(s>=60 for s in scores) else "No"
        ws.append(row + [avg, grade, passed])

    ws2 = wb.create_sheet("Top_Per_Subject")
    ws2.append(["Subject","Top_Student","Score"])
    for i, subj in enumerate(subjects):
        col_idx = i + 3
        best = max(students[1:], key=lambda r: int(r[col_idx]))
        ws2.append([subj, best[1], int(best[col_idx])])

    ws3 = wb.create_sheet("Fail_Analysis")
    ws3.append(["StudentID","Name","Failed_Subjects","Num_Fails"])
    for row in students[1:]:
        fails = [subjects[i] for i, s in enumerate(row[3:]) if int(s) < 60]
        if fails:
            ws3.append([row[0], row[1], ", ".join(fails), len(fails)])

    ws4 = wb.create_sheet("Subject_Stats")
    ws4.append(["Subject","Min","Max","Class_Average","Students_Below_60"])
    for i, subj in enumerate(subjects):
        scores = [int(r[i+3]) for r in students[1:]]
        ws4.append([subj, min(scores), max(scores), round(sum(scores)/len(scores),1),
                    sum(1 for s in scores if s<60)])

    ws5 = wb.create_sheet("Class_Summary")
    ws5.append(["Class","Headcount","Avg_Score","Pass_Rate_%","Top_Student"])
    class_data = defaultdict(lambda: {"rows":[]})
    for row in students[1:]:
        class_data[row[2]]["rows"].append(row)
    for cls in sorted(class_data):
        rows_c = class_data[cls]["rows"]
        avgs = [round(sum(int(x) for x in r[3:])/5, 1) for r in rows_c]
        passes = sum(1 for r in rows_c if all(int(x)>=60 for x in r[3:]))
        top_row = max(rows_c, key=lambda r: sum(int(x) for x in r[3:]))
        ws5.append([cls, len(rows_c), round(sum(avgs)/len(avgs),1),
                    round(passes/len(rows_c)*100,1), top_row[1]])

    wb_save(wb, os.path.join(d, "tc06_output01.xlsx"))
    print("Case06 done")


# ─────────────────────────────────────────────────────────────────────────────
# CASE 07  Logistics Operations  (40 orders, 12 drivers, 3-sheet XLSX)
# ─────────────────────────────────────────────────────────────────────────────
def case07():
    d = os.path.join(BASE, "Case07")
    random.seed(707)

    city_list = ["London","Manchester","Birmingham","Leeds","Bristol","Sheffield",
                 "Edinburgh","Cardiff","Liverpool","Newcastle","Nottingham","Leicester"]
    # Full route matrix between all pairs (both directions)
    dist_base = {
        ("London","Manchester"):333,("London","Birmingham"):189,("London","Bristol"):189,
        ("London","Leeds"):340,("London","Sheffield"):266,("London","Liverpool"):354,
        ("London","Nottingham"):209,("London","Leicester"):160,
        ("Manchester","Leeds"):69,("Manchester","Birmingham"):145,("Manchester","Sheffield"):60,
        ("Manchester","Liverpool"):56,("Manchester","Nottingham"):148,("Manchester","Newcastle"):230,
        ("Birmingham","Bristol"):145,("Birmingham","Leeds"):210,("Birmingham","Sheffield"):168,
        ("Birmingham","Nottingham"):80,("Birmingham","Leicester"):61,
        ("Leeds","Sheffield"):58,("Leeds","Newcastle"):160,("Leeds","Nottingham"):120,
        ("Bristol","Cardiff"):68,("Sheffield","Nottingham"):68,("Sheffield","Leeds"):58,
        ("Edinburgh","Newcastle"):175,("Liverpool","Manchester"):56,
        ("Nottingham","Leicester"):48,("Newcastle","Leeds"):160,("Cardiff","Bristol"):68,
    }
    routes = {}
    for (a, b), dist in dist_base.items():
        speed = random.randint(75, 95)
        routes[(a, b)] = (dist, speed)
        routes[(b, a)] = (dist, speed)

    drivers = [
        ("D01","Tom Ellis",     600, 1.20, "2024-03-01"),
        ("D02","Sarah Miles",   400, 0.95, "2024-03-01"),
        ("D03","James Ford",    700, 1.35, "2024-03-02"),
        ("D04","Lucy Holt",     350, 0.88, "2024-03-02"),
        ("D05","Mark Stone",    800, 1.50, "2024-03-03"),
        ("D06","Nina Ross",     500, 1.10, "2024-03-01"),
        ("D07","Craig Bell",    650, 1.28, "2024-03-03"),
        ("D08","Amy Walsh",     450, 1.02, "2024-03-04"),
        ("D09","Dean Price",    750, 1.42, "2024-03-04"),
        ("D10","Emma Sharp",    550, 1.15, "2024-03-05"),
        ("D11","Gary Owen",     900, 1.60, "2024-03-05"),
        ("D12","Helen Nash",    300, 0.82, "2024-03-06"),
    ]

    priorities = ["High","Medium","Low"]
    orders = []
    # Generate 40 orders using routes that exist
    route_keys = [k for k in routes if k[0] in city_list and k[1] in city_list]
    random.shuffle(route_keys)
    for i in range(40):
        orig, dest = route_keys[i % len(route_keys)]
        wt = random.randint(80, 880)
        prio = random.choice(priorities)
        mo = random.randint(3, 5); day = random.randint(1, 28)
        orders.append((f"ORD{i+1:03d}", orig, dest, wt, prio, f"2024-{mo:02d}-{day:02d}"))

    wb_in = Workbook()
    ws_o = wb_in.active; ws_o.title = "Orders"
    ws_o.append(["OrderID","OriginCity","DestCity","WeightKG","Priority","RequestedDate"])
    for o in orders: ws_o.append(list(o))

    ws_d = wb_in.create_sheet("Drivers")
    ws_d.append(["DriverID","Name","TruckCapacityKG","CostPerKM_GBP","AvailableFrom"])
    for dr in drivers: ws_d.append(list(dr))

    ws_r = wb_in.create_sheet("Routes")
    ws_r.append(["Origin","Destination","DistanceKM","AvgSpeedKMH"])
    for (orig, dest), (dist, speed) in sorted(routes.items()):
        ws_r.append([orig, dest, dist, speed])
    wb_in.save(os.path.join(d, "tc07_input01.xlsx"))

    # Output: assign cheapest capable driver per order
    wb = Workbook()
    ws = wb.active; ws.title = "Order_Assignments"
    ws.append(["OrderID","Origin","Dest","WeightKG","Priority",
               "AssignedDriver","DriverCostPerKM","DistanceKM","EstTimeHrs","TotalCost_GBP"])
    assignments = []
    for o in orders:
        oid, orig, dest, wt, prio, _ = o
        wt = int(wt)
        dist, speed = routes.get((orig, dest), (200, 80))
        capable = [dr for dr in drivers if int(dr[2]) >= wt]
        if capable:
            cheapest = min(capable, key=lambda dr: float(dr[3]))
            cost = round(dist * float(cheapest[3]), 2)
            time_h = round(dist / speed, 2)
            assignments.append([oid,orig,dest,wt,prio,cheapest[1],float(cheapest[3]),dist,time_h,cost])
        else:
            assignments.append([oid,orig,dest,wt,prio,"UNASSIGNED",0,dist,0,0])
    for a in assignments: ws.append(a)

    ws2 = wb.create_sheet("High_Priority")
    ws2.append(["OrderID","Origin","Dest","WeightKG","AssignedDriver","TotalCost_GBP"])
    for a in assignments:
        if a[4]=="High":
            ws2.append([a[0],a[1],a[2],a[3],a[5],a[9]])

    ws3 = wb.create_sheet("Driver_Utilisation")
    ws3.append(["DriverName","Orders_Assigned","Total_Distance_KM","Total_Revenue_GBP"])
    driver_util = defaultdict(lambda: {"orders":0,"dist":0,"rev":0.0})
    for a in assignments:
        if a[5]!="UNASSIGNED":
            driver_util[a[5]]["orders"] += 1
            driver_util[a[5]]["dist"]   += a[7]
            driver_util[a[5]]["rev"]    += a[9]
    for name in sorted(driver_util):
        u = driver_util[name]
        ws3.append([name, u["orders"], u["dist"], round(u["rev"],2)])

    wb_save(wb, os.path.join(d, "tc07_output01.xlsx"))
    print("Case07 done")


# ─────────────────────────────────────────────────────────────────────────────
# CASE 08  Budget vs Actual Variance  (60 cost centers, 6 departments)
# ─────────────────────────────────────────────────────────────────────────────
def case08():
    d = os.path.join(BASE, "Case08")
    random.seed(808)

    dept_categories = {
        "Engineering": ["Salaries","Software_Licenses","Hardware","Training","Cloud_Infrastructure","R&D"],
        "Marketing":   ["Salaries","Advertising","Events","Research","Social_Media","PR_Agency"],
        "Operations":  ["Salaries","Facilities","Logistics","Maintenance","Security","Utilities"],
        "Finance":     ["Salaries","Audit_Fees","Compliance","Tax_Advisory","ERP_Licenses","Insurance"],
        "HR":          ["Salaries","Recruitment","Benefits_Admin","L&D_Programs","Wellness","HR_Software"],
        "Legal":       ["Salaries","External_Counsel","IP_Management","Contract_Review","Regulatory","Litigation"],
        "Sales":       ["Salaries","Commissions","CRM_Software","Travel","Conferences","Demo_Equipment"],
        "IT":          ["Salaries","Network_Infra","Cybersecurity","Helpdesk","SaaS_Tools","Data_Storage"],
        "Procurement": ["Salaries","Supplier_Audits","ERP_Modules","Logistics_Software","Vendor_Mgmt","Contracts"],
        "Facilities":  ["Salaries","Rent","Cleaning","Catering","Parking","Fit_Out"],
    }
    budget_rows = [["Department","CostCenter","Category","Budget_USD"]]
    actual_rows = [["Department","CostCenter","Category","Actual_USD"]]

    for dept, cats in dept_categories.items():
        for j, cat in enumerate(cats):
            cc = f"CC-{dept[:3].upper()}-{j+1:02d}"
            if cat == "Salaries":
                bud = random.randint(120, 600) * 1000
            else:
                bud = random.randint(20, 200) * 1000
            # actual: +/- 20%
            variance_pct = random.uniform(-0.15, 0.20)
            act = int(bud * (1 + variance_pct) // 1000) * 1000
            budget_rows.append([dept, cc, cat, bud])
            actual_rows.append([dept, cc, cat, act])

    with open(os.path.join(d, "tc08_input01.csv"), "w", newline="") as f:
        csv.writer(f).writerows(budget_rows)
    with open(os.path.join(d, "tc08_input02.csv"), "w", newline="") as f:
        csv.writer(f).writerows(actual_rows)

    bud_map = {r[1]: (r[0], r[2], int(r[3])) for r in budget_rows[1:]}
    act_map = {r[1]: int(r[3]) for r in actual_rows[1:]}

    wb = Workbook()
    ws = wb.active; ws.title = "Variance_Detail"
    ws.append(["Department","CostCenter","Category","Budget_USD","Actual_USD",
               "Variance_USD","Variance_%","Status"])
    rows_out = []
    for cc, (dept, cat, bud) in bud_map.items():
        act = act_map.get(cc, 0)
        var = act - bud
        var_pct = round((var / bud) * 100, 1) if bud else 0
        status = ("Over (>10%)" if var_pct > 10 else
                  "Over" if var_pct > 0 else
                  "Under" if var_pct < 0 else "On Budget")
        rows_out.append([dept, cc, cat, bud, act, var, var_pct, status])
    rows_out.sort(key=lambda x: -x[5])
    for r in rows_out: ws.append(r)

    ws2 = wb.create_sheet("Dept_Summary")
    ws2.append(["Department","Total_Budget","Total_Actual","Total_Variance","Variance_%"])
    dept_agg = defaultdict(lambda: {"b":0,"a":0})
    for r in rows_out:
        dept_agg[r[0]]["b"] += r[3]; dept_agg[r[0]]["a"] += r[4]
    for dept in sorted(dept_agg):
        b = dept_agg[dept]["b"]; a = dept_agg[dept]["a"]
        v = a - b; vp = round((v/b)*100,1)
        ws2.append([dept, b, a, v, vp])

    ws3 = wb.create_sheet("Over_Budget_Items")
    ws3.append(["Department","CostCenter","Category","Budget_USD","Actual_USD","Overspend_USD"])
    for r in rows_out:
        if r[5] > 0:
            ws3.append([r[0],r[1],r[2],r[3],r[4],r[5]])

    wb_save(wb, os.path.join(d, "tc08_output01.xlsx"))
    print("Case08 done")


# ─────────────────────────────────────────────────────────────────────────────
# CASE 09  Supplier Price Comparison Matrix  (25 products, 8 suppliers)
# ─────────────────────────────────────────────────────────────────────────────
def case09():
    d = os.path.join(BASE, "Case09")
    random.seed(909)

    product_names = [
        "Steel Bolts M8 (100pk)","Aluminium Sheet 1m²","Copper Wire 10m","PVC Pipe 3m",
        "Rubber Gasket Set","Stainless Screws (200pk)","Electrical Cable 50m","Paint Primer 5L",
        "Safety Gloves (pair)","Drill Bits Set (10pc)","Hex Nuts M10 (50pk)","Wire Mesh 1m²",
        "Silicone Sealant 300ml","Industrial Adhesive 1L","Welding Rods (50pk)",
        "Hydraulic Hose 5m","Carbon Fibre Sheet A4","Titanium Bolts M6 (50pk)",
        "Anti-Corrosion Spray 500ml","Circuit Breaker 16A","LED Strip 5m",
        "Pneumatic Fitting 8mm","Neoprene Sheet 1m²","Bearing 6205 (each)",
        "Grease Cartridge 400g",
    ]
    suppliers = ["SupplierA","SupplierB","SupplierC","SupplierD",
                 "SupplierE","SupplierF","SupplierG","SupplierH"]
    base_prices = [18.5,45.0,12.8,8.2,6.5,22.0,38.5,15.0,4.8,28.0,
                   5.5,32.0,9.8,14.5,42.0,28.5,85.0,35.0,11.0,22.5,
                   18.0,7.5,55.0,12.0,8.5]

    rows = [["Product"] + suppliers]
    product_data = []
    for i, prod in enumerate(product_names):
        base = base_prices[i]
        prices = []
        for _ in suppliers:
            # vary ±20% around base
            p = round(base * random.uniform(0.82, 1.20), 2)
            prices.append(p)
        rows.append([prod] + prices)
        product_data.append((prod, prices))

    with open(os.path.join(d, "tc09_input01.csv"), "w", newline="") as f:
        csv.writer(f).writerows(rows)

    wb = Workbook()
    ws = wb.active; ws.title = "Price_Analysis"
    ws.append(["Product","Min_Price","Best_Supplier","Max_Price","Worst_Supplier",
               "Avg_Price","Price_Range","Savings_vs_Max_%"])
    for prod, prices in product_data:
        mn = min(prices); mx = max(prices)
        best  = suppliers[prices.index(mn)]
        worst = suppliers[prices.index(mx)]
        avg   = round(sum(prices)/len(prices), 2)
        rng   = round(mx - mn, 2)
        savings = round((mx - mn) / mx * 100, 1)
        ws.append([prod, mn, best, mx, worst, avg, rng, savings])

    ws2 = wb.create_sheet("Supplier_Ranking")
    ws2.append(["Supplier","Times_Cheapest","Times_Most_Expensive","Avg_Rank_Score"])
    cheapest_count   = {s: 0 for s in suppliers}
    expensive_count  = {s: 0 for s in suppliers}
    for prod, prices in product_data:
        mn = min(prices); mx = max(prices)
        cheapest_count[suppliers[prices.index(mn)]]  += 1
        expensive_count[suppliers[prices.index(mx)]] += 1
    for s in suppliers:
        ws2.append([s, cheapest_count[s], expensive_count[s],
                    cheapest_count[s] - expensive_count[s]])

    ws3 = wb.create_sheet("Top5_Savings")
    ws3.append(["Product","Best_Supplier","Min_Price","Max_Price","Savings_USD","Savings_%"])
    savings_list = []
    for prod, prices in product_data:
        mn = min(prices); mx = max(prices)
        best = suppliers[prices.index(mn)]
        savings_list.append((prod, best, mn, mx, round(mx-mn,2), round((mx-mn)/mx*100,1)))
    for row in sorted(savings_list, key=lambda x: -x[4])[:5]:
        ws3.append(list(row))

    wb_save(wb, os.path.join(d, "tc09_output01.xlsx"))
    print("Case09 done")


# ─────────────────────────────────────────────────────────────────────────────
# CASE 10  Football League  (12 teams, 66 matches, 72 players, 3-sheet XLSX)
# ─────────────────────────────────────────────────────────────────────────────
def case10():
    d = os.path.join(BASE, "Case10")
    random.seed(1010)

    teams = [
        ("T01","FC United",      "Old Stadium",       55000),
        ("T02","City Athletic",  "Blue Arena",        48000),
        ("T03","Rovers FC",      "Riverside Ground",  32000),
        ("T04","Athletic Club",  "North Park",        42000),
        ("T05","Town FC",        "The Meadow",        28000),
        ("T06","Wanderers",      "East End Park",     38000),
        ("T07","Rangers FC",     "Highfield",         45000),
        ("T08","Dynamo",         "Central Ground",    36000),
        ("T09","United Stars",   "South Arena",       40000),
        ("T10","Phoenix FC",     "Westfield",         34000),
        ("T11","Olympic Club",   "Riverside Arena",   29000),
        ("T12","Sparta FC",      "Highland Park",     31000),
    ]
    team_ids = [t[0] for t in teams]
    pairs = list(itertools.combinations(team_ids, 2))  # 66 matches
    matches = []
    for i, (home, away) in enumerate(pairs, 1):
        hg = random.randint(0, 4)
        ag = random.randint(0, 3)
        mo = (i % 9) + 1; day = (i % 25) + 1
        matches.append((f"M{i:02d}", home, away, hg, ag, f"2024-{mo:02d}-{day:02d}"))

    positions = ["FW","FW","MF","MF","MF","DF","DF","GK"]
    players = []
    pid = 1
    for t in teams:
        tid = t[0]
        for j in range(6):  # 6 players per team = 72 total
            pos = positions[j % len(positions)]
            if pos == "FW":
                goals = random.randint(2, 14)
            elif pos == "MF":
                goals = random.randint(1, 7)
            elif pos == "DF":
                goals = random.randint(0, 3)
            else:
                goals = 0
            assists = max(0, goals - random.randint(0, 3))
            players.append((f"PL{pid:02d}", tid, f"Player_{pid}", pos, goals, assists))
            pid += 1

    wb_in = Workbook()
    ws_t = wb_in.active; ws_t.title = "Teams"
    ws_t.append(["TeamID","TeamName","Stadium","Capacity"])
    for t in teams: ws_t.append(list(t))

    ws_m = wb_in.create_sheet("Matches")
    ws_m.append(["MatchID","HomeTeamID","AwayTeamID","HomeGoals","AwayGoals","MatchDate"])
    for m in matches: ws_m.append(list(m))

    ws_p = wb_in.create_sheet("Players")
    ws_p.append(["PlayerID","TeamID","PlayerName","Position","GoalsScored","Assists"])
    for p in players: ws_p.append(list(p))
    wb_in.save(os.path.join(d, "tc10_input01.xlsx"))

    # Output
    stats = {t[0]: {"name":t[1],"W":0,"D":0,"L":0,"GF":0,"GA":0} for t in teams}
    for m in matches:
        _, home, away, hg, ag, _ = m
        stats[home]["GF"] += hg; stats[home]["GA"] += ag
        stats[away]["GF"] += ag; stats[away]["GA"] += hg
        if hg > ag:   stats[home]["W"] += 1; stats[away]["L"] += 1
        elif hg < ag: stats[away]["W"] += 1; stats[home]["L"] += 1
        else:         stats[home]["D"] += 1; stats[away]["D"] += 1

    wb = Workbook()
    ws = wb.active; ws.title = "League_Table"
    ws.append(["Pos","TeamName","P","W","D","L","GF","GA","GD","Pts"])
    table = []
    for tid, s in stats.items():
        p = s["W"]+s["D"]+s["L"]; pts = s["W"]*3+s["D"]
        gd = s["GF"]-s["GA"]
        table.append([s["name"],p,s["W"],s["D"],s["L"],s["GF"],s["GA"],gd,pts])
    table.sort(key=lambda x: (-x[8],-x[7],x[0]))
    for i, r in enumerate(table, 1):
        ws.append([i]+r)

    ws2 = wb.create_sheet("Top_Scorers")
    ws2.append(["Rank","PlayerName","TeamName","Position","Goals","Assists"])
    team_name_map = {t[0]: t[1] for t in teams}
    top = sorted(players, key=lambda p: (-p[4], -p[5]))[:15]
    for i, p in enumerate(top, 1):
        ws2.append([i, p[2], team_name_map[p[1]], p[3], p[4], p[5]])

    ws3 = wb.create_sheet("Top4_Teams")
    ws3.append(["Pos","TeamName","Pts","GD","Status"])
    for i, r in enumerate(table[:4], 1):
        ws3.append([i, r[0], r[8], r[7],
                    "Champions League" if i<=2 else "Europa League"])

    ws4 = wb.create_sheet("Team_Goal_Stats")
    ws4.append(["TeamName","Goals_Scored","Goals_Conceded","GD","Top_Scorer","Top_Scorer_Goals"])
    team_players = defaultdict(list)
    for p in players:
        team_players[p[1]].append(p)
    for tid, s in stats.items():
        tp = max(team_players[tid], key=lambda p: p[4])
        ws4.append([s["name"], s["GF"], s["GA"], s["GF"]-s["GA"], tp[2], tp[4]])

    wb_save(wb, os.path.join(d, "tc10_output01.xlsx"))
    print("Case10 done")


if __name__ == "__main__":
    case01(); case02(); case03(); case04(); case05()
    case06(); case07(); case08(); case09(); case10()
    print("All cases generated.")
