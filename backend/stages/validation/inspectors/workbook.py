"""Saved-workbook validation inspectors."""

from datetime import date, datetime, time
import os
import re
from typing import Any

from openpyxl import load_workbook


class WorkbookValidationInspectorMixin:
    _NUMERIC_VALUE_HEADER_MARKERS = (
        "amount",
        "average",
        "avg",
        "balance",
        "cost",
        "count",
        "fee",
        "growth",
        "income",
        "margin",
        "measure",
        "payment",
        "price",
        "profit",
        "quantity",
        "qty",
        "rate",
        "revenue",
        "score",
        "spending",
        "sum",
        "total",
        "value",
        "weight",
    )
    _NON_MEASURE_HEADER_MARKERS = (
        "date",
        "day",
        "id",
        "identifier",
        "index",
        "name",
        "note",
        "rank",
        "time",
        "timestamp",
        "year",
    )
    _SUMMARY_LABEL_MARKERS = (
        "average",
        "avg",
        "count",
        "max",
        "maximum",
        "min",
        "minimum",
        "sum",
        "total",
    )

    @staticmethod
    def _load_first_table_rows(output_path: str) -> tuple[list[str], list[list[Any]], str] | None:
        if not output_path or not os.path.exists(output_path):
            return None
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception:
            return None
        for sheet in workbook.worksheets:
            header_row_idx = None
            header: list[str] = []
            for row_idx in range(1, min(sheet.max_row, 20) + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 20) + 1)
                ]
                non_empty = [str(v).strip() for v in row_values if v not in (None, "")]
                if len(non_empty) >= 2:
                    header_row_idx = row_idx
                    header = [str(v).strip() if v not in (None, "") else "" for v in row_values]
                    break
            if header_row_idx is None:
                continue
            rows: list[list[Any]] = []
            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 20) + 1)
                ]
                if not any(v not in (None, "") for v in row_values):
                    if rows:
                        break
                    continue
                rows.append(row_values)
            return header, rows, sheet.title
        return None

    @classmethod
    def _inspect_saved_temporal_aggregation_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Temporal aggregation output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 2:
            issues.append("Temporal aggregation output must contain at least two columns.")
            return issues
        first_col = normalized[0]
        second_col = normalized[1]
        derived_temporal_output = (
            first_col == "Year"
            and "YoY_Growth_pct" in normalized
            and "High_Growth" in normalized
        )
        if first_col != "Period" and not derived_temporal_output:
            issues.append("Temporal aggregation output should start with a `Period` column.")
        metric_label = normalized[1]
        if (
            not derived_temporal_output
            and not any(metric_label.startswith(prefix) for prefix in ("Average ", "Total ", "Count ", "Median ", "Minimum ", "Maximum "))
        ):
            issues.append("Temporal aggregation output metric column label is missing the aggregate prefix.")
        if not rows:
            issues.append("Temporal aggregation output does not contain any data rows.")
            return issues
        first_period = str(rows[0][0]).strip() if rows[0] and rows[0][0] is not None else ""
        if not first_period:
            issues.append("Temporal aggregation output has an empty period label in the first data row.")
        return issues

    @classmethod
    def _inspect_saved_grouped_aggregation_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Grouped aggregation output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 2:
            issues.append("Grouped aggregation output must contain at least two columns.")
            return issues
        metric_label = normalized[-1]
        if not any(metric_label.startswith(prefix) for prefix in ("Average ", "Total ", "Count ", "Median ", "Minimum ", "Maximum ")):
            issues.append("Grouped aggregation output metric column label is missing the aggregate prefix.")
        if not rows:
            issues.append("Grouped aggregation output does not contain any data rows.")
        return issues

    @classmethod
    def _inspect_saved_relational_join_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Relational join output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 2:
            issues.append("Relational join output must contain at least two columns.")
        if not rows:
            issues.append("Relational join output does not contain any data rows.")
            return issues
        non_empty_column_count = sum(1 for value in rows[0] if value not in (None, ""))
        if non_empty_column_count < 2:
            issues.append("Relational join output first data row is not populated across multiple columns.")
        return issues

    @classmethod
    def _inspect_saved_allocation_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Allocation output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 3:
            issues.append("Allocation output must contain entity columns, resource columns, and status.")
            return issues
        if "Allocation Status" not in normalized:
            issues.append("Allocation output should include an `Allocation Status` column.")
        if "Allocated Quantity" not in normalized:
            issues.append("Allocation output should include an `Allocated Quantity` column.")
        if not rows:
            issues.append("Allocation output does not contain any data rows.")
        return issues

    @classmethod
    def _inspect_saved_dependency_schedule_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Dependency-schedule output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if "Start Time" not in normalized or "End Time" not in normalized:
            issues.append("Dependency-schedule output should include `Start Time` and `End Time` columns.")
        if "Task ID" not in normalized:
            issues.append("Dependency-schedule output should include a `Task ID` column.")
        if not rows:
            issues.append("Dependency-schedule output does not contain any scheduled task rows.")
        return issues

    @classmethod
    def _inspect_saved_assignment_schedule_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Assignment-schedule output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 3:
            issues.append("Assignment-schedule output must contain multiple schedule columns.")
            return issues
        scheduling_markers = ("day", "time", "slot", "room", "session")
        entity_markers = ("student", "candidate", "participant", "attendee", "member", "name", "id")
        if not any(any(marker in column.lower() for marker in scheduling_markers) for column in normalized):
            issues.append("Assignment-schedule output should include at least one scheduling/location column.")
        if not any(any(marker in column.lower() for marker in entity_markers) for column in normalized):
            issues.append("Assignment-schedule output should include at least one entity-identifying column.")
        if not rows:
            issues.append("Assignment-schedule output does not contain any assigned rows.")
        return issues

    @classmethod
    def _inspect_saved_regression_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Regression output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if normalized[:2] != ["Factor", "Weight"]:
            issues.append("Regression output should start with `Factor` and `Weight` columns.")
        if not rows:
            issues.append("Regression output does not contain any coefficient rows.")
            return issues
        first_factor = str(rows[0][0]).strip() if rows[0] and rows[0][0] is not None else ""
        if not first_factor:
            issues.append("Regression output has an empty factor label in the first data row.")
        return issues

    @classmethod
    def _inspect_saved_correlation_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        loaded = cls._load_first_table_rows(output_path)
        if loaded is None:
            return ["Correlation output workbook does not contain a readable result table."]
        header, rows, _sheet_name = loaded
        normalized = [str(value).strip() for value in header if str(value).strip()]
        if len(normalized) < 2:
            issues.append("Correlation output must contain at least two numeric columns.")
        if len(rows) < 2:
            issues.append("Correlation output does not contain enough matrix rows.")
            return issues
        first_row = rows[0]
        non_empty = [value for value in first_row if value not in (None, "")]
        if len(non_empty) < 2:
            issues.append("Correlation output first matrix row is not populated.")
        return issues

    @classmethod
    def _inspect_saved_comparative_multi_sheet_workbook(cls, output_path: str) -> list[str]:
        issues = cls._inspect_saved_generic_workbook(output_path, need_detail=True, need_summary=False)
        if issues:
            return issues
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception as exc:
            return [f"Unable to open saved workbook: {exc}"]
        required_sheets = {"AvgByStoreType", "HolidayVsNonHoliday"}
        missing = sorted(required_sheets.difference(workbook.sheetnames))
        if missing:
            issues.append(
                "Comparative multi-sheet output is missing required sheet(s): " + ", ".join(missing)
            )
        return issues

    @classmethod
    def _inspect_saved_generic_workbook(
        cls,
        output_path: str,
        need_detail: bool | None,
        need_summary: bool | None,
    ) -> list[str]:
        issues: list[str] = []
        if not output_path or not os.path.exists(output_path):
            return [f"Saved output file not found: {output_path}"]
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception as exc:
            return [f"Unable to open saved workbook: {exc}"]

        found_header = False
        max_data_rows = 0
        found_metrics = False

        for sheet in workbook.worksheets:
            header_row_idx = None
            for row_idx in range(1, min(sheet.max_row, 20) + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 12) + 1)
                ]
                non_empty = [str(v).strip() for v in row_values if v not in (None, "")]
                if len(non_empty) >= 2:
                    header_row_idx = row_idx
                    found_header = True
                    break
            if header_row_idx is None:
                continue

            data_rows = 0
            metrics_in_sheet = False
            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 12) + 1)
                ]
                if not any(v not in (None, "") for v in row_values):
                    continue

                metric_found = False
                for col_idx in range(0, len(row_values) - 1):
                    first = row_values[col_idx]
                    second = row_values[col_idx + 1]
                    label = str(first).strip() if first not in (None, "") else ""
                    if label and any(
                        marker in label.lower()
                        for marker in (
                            "total",
                            "average",
                            "avg",
                            "count",
                            "max",
                            "min",
                            "growth",
                            "fastest",
                            "region",
                            "sum",
                            "rows used",
                            "latest year",
                        )
                    ):
                        metrics_in_sheet = True
                        metric_found = True
                        break
                if metric_found:
                    continue

                first = row_values[0] if row_values else None
                second = row_values[1] if len(row_values) > 1 else None
                if first not in (None, "") and second not in (None, ""):
                    data_rows += 1

            max_data_rows = max(max_data_rows, data_rows)
            if need_summary is True and need_detail is not True and data_rows >= 1:
                metrics_in_sheet = True
            found_metrics = found_metrics or metrics_in_sheet
            if (need_detail is not True or data_rows >= 1) and (need_summary is not True or metrics_in_sheet):
                return []

        if not found_header:
            issues.append("Saved workbook does not contain a sheet with a detectable table header.")
        if need_detail is True and max_data_rows < 1:
            issues.append("Saved workbook does not contain a detailed table with at least one data row.")
        if need_summary is True and not found_metrics:
            issues.append("Saved workbook does not contain the required summary metrics block.")
        return issues

    @classmethod
    def _question_requests_max_highlight(cls, user_question: str) -> bool:
        lower = (user_question or "").lower()
        if not any(marker in lower for marker in ("highlight", "mark", "colour", "color")):
            return False
        return any(
            marker in lower
            for marker in (
                "max",
                "maximum",
                "highest",
                "largest",
                "greatest",
                "peak",
                "top",
            )
        )

    @staticmethod
    def _cell_has_non_default_fill(cell) -> bool:
        fill = getattr(cell, "fill", None)
        if fill is None or fill.fill_type in (None, "none"):
            return False
        color = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
        color_type = getattr(color, "type", None)
        color_value = (
            getattr(color, "rgb", None)
            or getattr(color, "indexed", None)
            or getattr(color, "theme", None)
        )
        if color_type == "indexed" and color_value in (None, 64):
            return False
        if str(color_value).upper() in {"00000000", "00FFFFFF", "FFFFFFFF", "64"}:
            return False
        return True

    @classmethod
    def _row_is_highlighted(cls, sheet, row_idx: int, max_col: int) -> bool:
        filled_cells = 0
        populated_cells = 0
        for col_idx in range(1, max_col + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value not in (None, ""):
                populated_cells += 1
            if cls._cell_has_non_default_fill(cell):
                filled_cells += 1
        if filled_cells == 0:
            return False
        if populated_cells <= 1:
            return filled_cells >= 1
        return filled_cells >= min(2, populated_cells)

    @classmethod
    def _row_is_summary_like(cls, row_values: list[Any]) -> bool:
        first_non_empty = ""
        for value in row_values:
            if value not in (None, ""):
                first_non_empty = str(value).strip().lower()
                break
        if not first_non_empty:
            return False
        return any(marker in first_non_empty for marker in cls._SUMMARY_LABEL_MARKERS)

    @staticmethod
    def _numeric_cell_value(value: Any) -> float | None:
        if value is None or value == "" or isinstance(value, bool):
            return None
        if isinstance(value, (datetime, date, time)):
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(",", "")
        if not text:
            return None
        if not re.fullmatch(r"[-+]?\d+(?:\.\d+)?%?", text):
            return None
        try:
            parsed = float(text.rstrip("%"))
        except ValueError:
            return None
        if text.endswith("%"):
            parsed /= 100.0
        return parsed

    @classmethod
    def _measure_column_priority(cls, header: str) -> int:
        lower = (header or "").strip().lower()
        if any(marker in lower for marker in cls._NON_MEASURE_HEADER_MARKERS):
            return -1
        if any(marker in lower for marker in cls._NUMERIC_VALUE_HEADER_MARKERS):
            return 2
        return 1

    @classmethod
    def _inspect_sheet_highlighted_max_rows(cls, sheet) -> list[str]:
        max_col = min(sheet.max_column, 50)
        header_row_idx = None
        headers: list[str] = []
        for row_idx in range(1, min(sheet.max_row, 20) + 1):
            row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
            non_empty = [str(value).strip() for value in row_values if value not in (None, "")]
            if len(non_empty) >= 2:
                header_row_idx = row_idx
                headers = [str(value).strip() if value not in (None, "") else "" for value in row_values]
                break
        if header_row_idx is None:
            return []

        data_row_numbers: list[int] = []
        highlighted_rows: set[int] = set()
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row_values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]
            if not any(value not in (None, "") for value in row_values):
                if data_row_numbers:
                    break
                continue
            if cls._row_is_summary_like(row_values):
                continue
            data_row_numbers.append(row_idx)
            if cls._row_is_highlighted(sheet, row_idx, max_col):
                highlighted_rows.add(row_idx)

        if not data_row_numbers or not highlighted_rows:
            return []

        candidate_columns: list[tuple[int, int, list[tuple[int, float]]]] = []
        for col_idx in range(1, max_col + 1):
            header = headers[col_idx - 1] if col_idx <= len(headers) else ""
            priority = cls._measure_column_priority(header)
            if priority < 0:
                continue
            values: list[tuple[int, float]] = []
            for row_idx in data_row_numbers:
                parsed = cls._numeric_cell_value(sheet.cell(row=row_idx, column=col_idx).value)
                if parsed is not None:
                    values.append((row_idx, parsed))
            if len(values) < 2:
                continue
            candidate_columns.append((priority, col_idx, values))

        if not candidate_columns:
            return []

        candidate_columns.sort(key=lambda item: item[0], reverse=True)
        for _priority, _col_idx, values in candidate_columns:
            max_value = max(value for _row_idx, value in values)
            max_rows = {row_idx for row_idx, value in values if abs(value - max_value) <= 1e-9}
            if highlighted_rows.issubset(max_rows):
                return []

        best_priority, best_col_idx, best_values = candidate_columns[0]
        best_max_value = max(value for _row_idx, value in best_values)
        best_max_rows = sorted(row_idx for row_idx, value in best_values if abs(value - best_max_value) <= 1e-9)
        header = headers[best_col_idx - 1] if best_col_idx <= len(headers) else f"column {best_col_idx}"
        highlighted_preview = ", ".join(str(row) for row in sorted(highlighted_rows)[:5])
        expected_preview = ", ".join(str(row) for row in best_max_rows[:5])
        return [
            "Saved workbook highlighted row(s) do not match the maximum value row(s): "
            f"highlighted {highlighted_preview}, expected {expected_preview} for `{header}`."
        ]

    @classmethod
    def _inspect_saved_highlighted_max_rows(cls, output_path: str, user_question: str) -> list[str]:
        if not cls._question_requests_max_highlight(user_question):
            return []
        if not output_path or not os.path.exists(output_path):
            return []
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception:
            return []

        for sheet in workbook.worksheets:
            issues = cls._inspect_sheet_highlighted_max_rows(sheet)
            if issues:
                return issues
        return []

    @classmethod
    def _inspect_saved_region_growth_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        if not output_path or not os.path.exists(output_path):
            return [f"Saved output file not found: {output_path}"]
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception as exc:
            return [f"Unable to open saved workbook: {exc}"]

        output_sheet = workbook["Output"] if "Output" in workbook.sheetnames else workbook.worksheets[0]
        header = [output_sheet.cell(row=1, column=col_idx).value for col_idx in range(1, 4)]
        normalized_header = [str(v).strip() if v is not None else "" for v in header]
        if normalized_header[0] != "Region":
            issues.append("Region-growth output table headers are incorrect or missing.")
        avg_match = re.fullmatch(r"Avg Penetration \((\d{4})-(\d{4})\)", normalized_header[1])
        growth_match = re.fullmatch(r"Growth(?: Rate)? \((\d{4})-(\d{4})\)", normalized_header[2])
        if avg_match is None or growth_match is None:
            issues.append("Region-growth output table headers are incorrect or missing.")
        elif avg_match.groups() != growth_match.groups():
            issues.append("Region-growth output uses inconsistent year windows across headers.")

        data_rows = 0
        for row_idx in range(2, output_sheet.max_row + 1):
            first = output_sheet.cell(row=row_idx, column=1).value
            second = output_sheet.cell(row=row_idx, column=2).value
            third = output_sheet.cell(row=row_idx, column=3).value
            if all(v not in (None, "") for v in (first, second, third)):
                data_rows += 1
        if data_rows < 2:
            issues.append("Region-growth output does not contain enough complete data rows.")

        summary_labels = set()
        for row_idx in range(1, output_sheet.max_row + 1):
            for col_idx in range(1, min(output_sheet.max_column, 8) + 1, 3):
                label = output_sheet.cell(row=row_idx, column=col_idx).value
                value = output_sheet.cell(row=row_idx, column=col_idx + 1).value
                if label not in (None, "") and value not in (None, ""):
                    summary_labels.add(str(label).strip())
        if "Fastest Growth Region" not in summary_labels:
            issues.append("Region-growth output is missing the `Fastest Growth Region` summary label.")
        if not any(label.startswith("Growth") for label in summary_labels):
            issues.append("Region-growth output is missing the growth summary metric.")

        return issues

    @staticmethod
    def _parse_time_to_minutes(value: Any) -> int | None:
        if value is None:
            return None
        if hasattr(value, "hour") and hasattr(value, "minute"):
            try:
                return int(value.hour) * 60 + int(value.minute)
            except Exception:
                return None
        text = str(value).strip()
        match = re.match(r"^(\d{1,2}):(\d{2})$", text)
        if not match:
            return None
        hour = int(match.group(1))
        minute = int(match.group(2))
        if hour < 0 or minute < 0 or minute > 59:
            return None
        return (hour * 60) + minute

    @staticmethod
    def _parse_numeric(value: Any) -> float | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return None
        try:
            return float(match.group(0))
        except Exception:
            return None

    @classmethod
    def _inspect_saved_schedule_workbook(cls, output_path: str) -> list[str]:
        issues: list[str] = []
        if not output_path or not os.path.exists(output_path):
            return ["Saved output workbook path does not exist on disk."]

        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception as exc:
            return [f"Could not open saved output workbook: {exc}"]

        required_headers = ["Task ID", "Task Name", "Priority", "Start Time", "End Time"]
        schedule_sheet = None
        header_row_index = None
        header_columns: dict[str, int] = {}

        for sheet in workbook.worksheets:
            for row_idx in range(1, min(sheet.max_row, 20) + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 12) + 1)
                ]
                normalized = {
                    str(v).strip(): idx
                    for idx, v in enumerate(row_values, start=1)
                    if v is not None and str(v).strip()
                }
                if all(header in normalized for header in required_headers):
                    schedule_sheet = sheet
                    header_row_index = row_idx
                    header_columns = {header: normalized[header] for header in required_headers}
                    break
            if schedule_sheet is not None:
                break

        if schedule_sheet is None or header_row_index is None:
            return ["Saved workbook does not contain a sheet with required schedule headers."]

        data_rows: list[dict[str, Any]] = []
        summary_label = None
        summary_value = None

        for row_idx in range(header_row_index + 1, schedule_sheet.max_row + 1):
            task_id = schedule_sheet.cell(row=row_idx, column=header_columns["Task ID"]).value
            task_name = schedule_sheet.cell(row=row_idx, column=header_columns["Task Name"]).value
            priority = schedule_sheet.cell(row=row_idx, column=header_columns["Priority"]).value
            start_value = schedule_sheet.cell(row=row_idx, column=header_columns["Start Time"]).value
            end_value = schedule_sheet.cell(row=row_idx, column=header_columns["End Time"]).value

            label_candidate = "" if task_id is None else str(task_id).strip()
            if label_candidate and "total duration" in label_candidate.lower():
                summary_label = label_candidate
                summary_value = task_name
                break

            if not any(v not in (None, "") for v in (task_id, task_name, priority, start_value, end_value)):
                continue

            if not all(v not in (None, "") for v in (task_id, task_name, priority, start_value, end_value)):
                issues.append(f"Schedule row {row_idx} is partially empty.")
                continue

            data_rows.append(
                {
                    "row": row_idx,
                    "task_id": str(task_id).strip(),
                    "task_name": str(task_name).strip(),
                    "priority": str(priority).strip(),
                    "start": start_value,
                    "end": end_value,
                }
            )

        if not data_rows:
            issues.append("Saved schedule sheet contains headers but no complete task rows.")
            return issues

        total_minutes = 0
        previous_end = None
        seen_task_ids: set[str] = set()

        for item in data_rows:
            if not item["task_id"]:
                issues.append(f"Schedule row {item['row']} has empty Task ID.")
                continue
            if item["task_id"] in seen_task_ids:
                issues.append(f"Schedule output duplicates Task ID `{item['task_id']}`.")
            seen_task_ids.add(item["task_id"])

            start_minutes = cls._parse_time_to_minutes(item["start"])
            end_minutes = cls._parse_time_to_minutes(item["end"])
            if start_minutes is None or end_minutes is None:
                issues.append(
                    f"Schedule row {item['row']} has invalid time format; expected `HH:MM` in Start Time and End Time."
                )
                continue
            if end_minutes <= start_minutes:
                issues.append(f"Schedule row {item['row']} has non-positive duration.")
                continue
            if previous_end is not None and start_minutes != previous_end:
                issues.append(
                    f"Schedule timing is not contiguous: row {item['row']} starts at {item['start']} after previous end."
                )
            previous_end = end_minutes
            total_minutes += (end_minutes - start_minutes)

        if summary_label is None:
            issues.append("Saved schedule sheet is missing a total-duration summary row.")
            return issues

        numeric_summary = cls._parse_numeric(summary_value)
        if numeric_summary is None:
            issues.append("Total-duration summary value is missing or non-numeric.")
            return issues

        label_lower = summary_label.lower()
        if "hour" in label_lower:
            expected_value = total_minutes / 60.0
            tolerance = 0.05
        elif "minute" in label_lower:
            expected_value = float(total_minutes)
            tolerance = 1.0
        else:
            expected_value = total_minutes / 60.0
            tolerance = 0.05

        if abs(numeric_summary - expected_value) > tolerance:
            issues.append(
                "Total-duration summary does not match the actual schedule rows "
                f"(expected about {expected_value:.2f}, found {numeric_summary:.2f})."
            )

        return issues
