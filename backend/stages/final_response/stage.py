"""Generate a short user-facing final response."""

from __future__ import annotations

import os
import re
from typing import Any, Dict, Optional

from openpyxl import load_workbook

from ...log.logger_registry import LoggerRegistry
from ...skills import detect_skill, get_helper_final_response_label, select_helper
from ..base.stage import Stage

logger = LoggerRegistry.setup_logger(__name__)


class FinalResponseStage(Stage):
    """Produce a short, contentful final response for the user."""

    def __init__(self, client, deployment: str, progress_logger=None):
        self.client = client
        self.deployment = deployment
        self.progress_logger = progress_logger

    @staticmethod
    def _looks_like_file_path(value: str) -> bool:
        text = (value or "").strip().lower()
        return text.endswith((".xlsx", ".xls", ".csv"))

    @staticmethod
    def _compact_question(user_question: str) -> str:
        text = " ".join((user_question or "").strip().split())
        return text[:240]

    @classmethod
    def _question_content_label(cls, user_question: str, workbook_summary: Dict[str, Any]) -> str:
        text = cls._compact_question(user_question)
        headers = [str(item).strip() for item in (workbook_summary.get("headers") or []) if str(item).strip()]

        routed_label = cls._routed_content_label(user_question)
        if routed_label:
            return routed_label

        if headers:
            if len(headers) == 2:
                return f"spreadsheet with {headers[0]} and {headers[1]}"
            if len(headers) >= 3:
                return f"spreadsheet with columns {headers[0]}, {headers[1]}, and {headers[2]}"

        cleaned = re.sub(r"^(please|kindly)\s+", "", text, flags=re.IGNORECASE).strip(" .")
        cleaned = re.sub(r"\b(output|create|generate|produce|return|save)\b.*$", "", cleaned, flags=re.IGNORECASE).strip(" ,.")
        return cleaned or "spreadsheet"

    @staticmethod
    def _routed_content_label(user_question: str) -> Optional[str]:
        skill = detect_skill(user_question)
        if skill is None:
            return None
        helper = select_helper(skill, user_question)
        if helper is None:
            return None
        return get_helper_final_response_label(helper.name)

    @classmethod
    def _inspect_output_workbook(cls, output_path: str) -> Dict[str, Any]:
        summary: Dict[str, Any] = {
            "sheet_name": "",
            "headers": [],
            "data_rows": 0,
            "metric_label": "",
            "metric_value": "",
            "metrics": {},
            "sample_rows": [],
        }
        if not output_path or not os.path.exists(output_path):
            return summary
        try:
            workbook = load_workbook(output_path, data_only=True)
        except Exception:
            return summary

        for sheet in workbook.worksheets:
            header_row_idx = None
            headers = []
            for row_idx in range(1, min(sheet.max_row, 20) + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 12) + 1)
                ]
                non_empty = [str(v).strip() for v in row_values if v not in (None, "")]
                if len(non_empty) >= 2:
                    header_row_idx = row_idx
                    headers = non_empty
                    break
            if header_row_idx is None:
                continue

            data_rows = 0
            metric_label = ""
            metric_value = ""
            metrics: Dict[str, str] = {}
            sample_rows: list[dict[str, str]] = []
            for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
                row_values = [
                    sheet.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, min(sheet.max_column, 12) + 1)
                ]
                if not any(v not in (None, "") for v in row_values):
                    continue

                row_metric_found = False
                for col_idx in range(0, len(row_values) - 1):
                    first = row_values[col_idx]
                    second = row_values[col_idx + 1]
                    label = str(first).strip() if first not in (None, "") else ""
                    if label and any(marker in label.lower() for marker in ("total", "average", "count", "max", "min")):
                        metrics[label] = "" if second in (None, "") else str(second).strip()
                        if not metric_label:
                            metric_label = label
                            metric_value = "" if second in (None, "") else str(second).strip()
                        row_metric_found = True
                if row_metric_found:
                    continue

                first = row_values[0] if row_values else None
                second = row_values[1] if len(row_values) > 1 else None
                if first not in (None, "") and second not in (None, ""):
                    data_rows += 1
                    if len(sample_rows) < 3:
                        row_summary: dict[str, str] = {}
                        for header, value in zip(headers, row_values):
                            if value in (None, ""):
                                continue
                            row_summary[str(header)] = cls._format_summary_value(value)
                            if len(row_summary) >= 5:
                                break
                        if row_summary:
                            sample_rows.append(row_summary)

            summary.update(
                {
                    "sheet_name": sheet.title,
                    "headers": headers,
                    "data_rows": data_rows,
                    "metric_label": metric_label,
                    "metric_value": metric_value,
                    "metrics": metrics,
                    "sample_rows": sample_rows,
                }
            )
            return summary
        return summary

    @staticmethod
    def _format_summary_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float):
            rounded = round(value, 4)
            if rounded.is_integer():
                return str(int(rounded))
            return f"{rounded:.4f}".rstrip("0").rstrip(".")
        return str(value).strip()

    @classmethod
    def _format_metric_facts(cls, metrics: Dict[str, Any], limit: int = 4) -> list[str]:
        facts: list[str] = []
        for label, value in metrics.items():
            clean_label = str(label).strip()
            clean_value = cls._format_summary_value(value)
            if clean_label and clean_value:
                facts.append(f"{clean_label}: {clean_value}")
            if len(facts) >= limit:
                break
        return facts

    @classmethod
    def _format_sample_row_facts(cls, sample_rows: list[dict[str, str]], limit: int = 3) -> list[str]:
        facts: list[str] = []
        for row in sample_rows[:limit]:
            if not row:
                continue
            items = [(str(k).strip(), str(v).strip()) for k, v in row.items() if str(k).strip() and str(v).strip()]
            if not items:
                continue
            key_label, key_value = items[0]
            details = ", ".join(f"{label}={value}" for label, value in items[1:4])
            if details:
                facts.append(f"{key_label} {key_value}: {details}")
            else:
                facts.append(f"{key_label}: {key_value}")
        return facts

    @staticmethod
    def _large_workbook_threshold_bytes() -> int:
        return int(os.getenv("SHEETHERO_LARGE_WORKBOOK_THRESHOLD_BYTES", "10000000"))

    @classmethod
    def _is_large_output_file(cls, output_path: str) -> bool:
        if not output_path or not os.path.exists(output_path):
            return False
        try:
            return os.path.getsize(output_path) >= cls._large_workbook_threshold_bytes()
        except OSError:
            return False

    @classmethod
    def _summary_from_execution_result(cls, execution_result: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        summary: Dict[str, Any] = {
            "sheet_name": "",
            "headers": [],
            "data_rows": 0,
            "metric_label": "",
            "metric_value": "",
            "metrics": {},
            "sample_rows": [],
        }
        execution_steps = ((execution_result or {}).get("execution_summary") or {}).get("execution_steps") or []
        successful_steps = [step for step in execution_steps if step.get("success")]
        if not successful_steps:
            return summary
        latest_result = str(successful_steps[-1].get("result") or "")
        row_counts = [int(value) for value in re.findall(r"Wrote\s+(\d+)\s+rows\s+to", latest_result, flags=re.IGNORECASE)]
        if row_counts:
            summary["data_rows"] = max(max(row_counts) - 1, 0)
        return summary

    @classmethod
    def _summary_text_from_execution_result(cls, execution_result: Optional[Dict[str, Any]]) -> str:
        execution_steps = ((execution_result or {}).get("execution_summary") or {}).get("execution_steps") or []
        for step in reversed(execution_steps):
            if not step.get("success"):
                continue
            result_text = str(step.get("result") or "")
            match = re.search(r"RESULT_SUMMARY:\s*(.+)", result_text, flags=re.IGNORECASE)
            if match:
                summary = " ".join(match.group(1).strip().split())
                lowered = summary.lower()
                if cls._looks_like_file_path(summary) or (
                    ".xlsx" in lowered
                    and any(prefix in lowered for prefix in ("saved at", "saved to", "output saved", "workbook saved"))
                ):
                    continue
                if summary:
                    return summary
        return ""

    @classmethod
    def _infer_scalar_subject(cls, user_question: str, metric_kind: str) -> str:
        question = cls._compact_question(user_question)
        lowered = question.lower().strip(" ?.")
        metric_patterns = {
            "average": (r"\baverage\s+(.+)", r"\bmean\s+(.+)"),
            "total": (r"\btotal\s+(.+)", r"\bsum(?:\s+of)?\s+(.+)"),
            "count": (r"\bcount(?:\s+of)?\s+(.+)",),
            "maximum": (r"\b(?:maximum|max(?:imum)?|highest)\s+(.+)",),
            "minimum": (r"\b(?:minimum|min(?:imum)?|lowest)\s+(.+)",),
        }
        for pattern in metric_patterns.get(metric_kind, ()):
            match = re.search(pattern, lowered, flags=re.IGNORECASE)
            if not match:
                continue
            subject = (match.group(1) or "").strip(" ?.,")
            subject = re.sub(r"\b(for|in|within)\s+the\s+.*$", "", subject, flags=re.IGNORECASE).strip(" ,.")
            if subject:
                return subject
        return "result"

    @classmethod
    def _fallback_short_answer(
        cls,
        user_question: str,
        final_answer: str,
        validation_passed: bool,
        workbook_summary: Dict[str, Any],
    ) -> str:
        question = cls._compact_question(user_question)
        content_label = cls._question_content_label(user_question, workbook_summary)
        metric_label = workbook_summary.get("metric_label") or ""
        metric_value = workbook_summary.get("metric_value") or ""
        metrics = workbook_summary.get("metrics") or {}
        sample_rows = workbook_summary.get("sample_rows") or []
        data_rows = int(workbook_summary.get("data_rows") or 0)

        def _find_metric(keyword: str) -> tuple[str, str]:
            for label, value in metrics.items():
                if keyword in label.lower():
                    return label, value
            return "", ""

        if cls._looks_like_file_path(final_answer):
            if validation_passed:
                row_phrase = ""
                if data_rows > 0:
                    unit = "record" if data_rows == 1 else "rows"
                    row_phrase = f" with {data_rows} {unit}"
                lead = f"Generated the {content_label}{row_phrase}."

                facts = cls._format_metric_facts(metrics)
                if not facts:
                    facts = cls._format_sample_row_facts(sample_rows)
                if facts:
                    return f"{lead} Key results: {'; '.join(facts)}."

                total_label, total_value = _find_metric("total")
                average_label, average_value = _find_metric("average")
                if total_label and average_label and total_value and average_value:
                    return (
                        f"Successfully generated the {content_label}; "
                        f"{total_label} is {total_value}, and {average_label} is {average_value}."
                    )
                if metric_label and metric_value:
                    return (
                        f"Successfully generated the {content_label}; "
                        f"{metric_label} is {metric_value}."
                    )
                if data_rows > 0:
                    unit = "record" if data_rows == 1 else "rows"
                    return f"Successfully generated the {content_label} with {data_rows} {unit}."
                return f"Successfully generated the {content_label}."
            return f"Generated the {content_label}, but validation still found issues."

        answer_text = " ".join((final_answer or "").strip().split())
        if answer_text:
            if len(answer_text.split()) >= 5 and re.search(r"[.!?]$", answer_text):
                return answer_text
            lowered = question.lower()
            if "average" in lowered:
                return f"The average {cls._infer_scalar_subject(user_question, 'average')} is {answer_text}."
            if "total" in lowered or "sum" in lowered:
                return f"The total {cls._infer_scalar_subject(user_question, 'total')} is {answer_text}."
            if "count" in lowered:
                return f"The count of {cls._infer_scalar_subject(user_question, 'count')} is {answer_text}."
            if "maximum" in lowered or "highest" in lowered or "max " in lowered:
                return f"The maximum {cls._infer_scalar_subject(user_question, 'maximum')} is {answer_text}."
            if "minimum" in lowered or "lowest" in lowered or "min " in lowered:
                return f"The minimum {cls._infer_scalar_subject(user_question, 'minimum')} is {answer_text}."
            return f"The result is {answer_text}."
        if validation_passed:
            return f"Successfully completed the request for {question}."
        return f"The request for {question} completed with validation issues."

    def run(
        self,
        user_question: str,
        final_answer: str,
        validation_result: Optional[Dict[str, Any]],
        execution_result: Optional[Dict[str, Any]] = None,
    ) -> str:
        validation_result = validation_result or {}
        validation_passed = bool(validation_result.get("validation_passed"))
        workbook_summary = {}
        if self._looks_like_file_path(final_answer):
            execution_summary_text = self._summary_text_from_execution_result(execution_result)
            if execution_summary_text:
                return execution_summary_text
            if self._is_large_output_file(final_answer):
                workbook_summary = self._summary_from_execution_result(execution_result)
            else:
                workbook_summary = self._inspect_output_workbook(final_answer)

        fallback = self._fallback_short_answer(
            user_question=user_question,
            final_answer=final_answer,
            validation_passed=validation_passed,
            workbook_summary=workbook_summary,
        )
        return fallback
