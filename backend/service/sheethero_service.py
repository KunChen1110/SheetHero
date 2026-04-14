"""Service layer for orchestrating SheetHero single-task sessions."""

from __future__ import annotations

import csv
import os
from dataclasses import dataclass, field
from typing import Any, Dict, Optional

from openpyxl import load_workbook

from ..agent import SheetHero
from ..config.settings import Config
from ..agent.core.session import SheetHeroSession


@dataclass
class DialogueMemory:
    """Cross-session memory for dialogue continuity."""

    last_context_understanding: str = ""
    last_workbooks: Optional[Dict[str, Any]] = None


class SheetHeroService:
    """Dialogue-level service: each new user task creates a new session."""

    _TEXT_PREVIEW_MAX_ROWS = 20
    _TEXT_PREVIEW_MAX_COLS = 8
    _TEXT_PREVIEW_MAX_CELL_LEN = 40

    def __init__(self, config: Config, load_excel: bool = True) -> None:
        self.config = config
        self.load_excel = load_excel
        self._agent: Optional[SheetHero] = None
        self._session: Optional[SheetHeroSession] = None
        self._memory = DialogueMemory()

    def submit_turn(self, prompt: str, excel_paths: list[str]) -> Dict[str, object]:
        # Clarification must continue via submit_clarification().
        if self._session is not None and self._session.state == "qa":
            return {
                "type": "clarification",
                "stage": "qa",
                "message": "Session is awaiting clarification. Please call submit_clarification().",
            }
        response = self._prepare_turn(prompt, excel_paths)
        return self._finalize_response(self._auto_step_until_blocked(response))

    def submit_clarification(self, user_reply: str) -> Dict[str, object]:
        response = self._prepare_clarification(user_reply)
        return self._finalize_response(self._auto_step_until_blocked(response))

    def _prepare_clarification(self, user_reply: str) -> Dict[str, object]:
        if self._agent is None or self._session is None:
            return {
                "type": "error",
                "message": "No active session for clarification.",
            }
        if self._session.state != "qa":
            return {
                "type": "error",
                "message": "Session is not awaiting clarification.",
            }
        return self._agent.step(self._session, user_reply)

    def _prepare_turn(self, prompt: str, excel_paths: list[str]) -> Dict[str, object]:
        incoming_paths = self._resolve_excel_paths(excel_paths)

        # Cache prior session memory before switching task/session.
        self._cache_session_memory()

        # Build a new agent for each new task turn.
        self._agent = SheetHero(
            excel_paths=incoming_paths,
            config=self.config,
            load_excel=self.load_excel,
        )

        # New task -> always create a new session.
        self._session = self._agent.start_session(prompt)
        self._restore_session_memory(session=self._session)
        return self._agent.step(self._session)

    def _resolve_excel_paths(self, excel_paths: list[str]) -> list[str]:
        """Normalize current-turn workbook paths.

        Note: no_excel means no files provided in this turn.
        We intentionally do not back-fill from history here.
        """
        return [p for p in (excel_paths or []) if p]

    def _auto_step_until_blocked(
        self,
        response: Dict[str, object],
        max_auto_steps: int = 50,
    ) -> Dict[str, object]:
        if self._agent is None or self._session is None:
            return {"type": "error", "message": "Agent session not initialized."}

        for _ in range(max_auto_steps):
            response_type = response.get("type")
            if response_type == "progress":
                response = self._agent.step(self._session)
                continue
            return response

        return {
            "type": "error",
            "message": "Exceeded internal turn limit.",
        }

    def _finalize_response(self, response: Dict[str, object]) -> Dict[str, object]:
        if self._session:
            new_thoughts = self._session.ui_thoughts[self._session.ui_thought_cursor:]
            self._session.ui_thought_cursor = len(self._session.ui_thoughts)
            if new_thoughts:
                response["ui_thoughts"] = new_thoughts
        if response.get("type") == "final":
            self._apply_text_output_mode(response)
            result_meta = self._build_result_meta(response)
            if result_meta:
                response.update(result_meta)
                result = response.get("result")
                if isinstance(result, dict):
                    result["result_meta"] = result_meta
        if response.get("type") == "final" and not response.get("message"):
            response["message"] = self._extract_message(response)
        self._cache_session_memory()
        return response

    def _auto_step_stream(
        self,
        initial_response: Dict[str, object],
        max_auto_steps: int = 50,
    ):
        response = initial_response
        for _ in range(max_auto_steps):
            yield self._finalize_response(response)
            if response.get("type") != "progress":
                return
            if self._agent is None or self._session is None:
                yield self._finalize_response(
                    {"type": "error", "message": "Agent session not initialized."}
                )
                return
            response = self._agent.step(self._session)
        yield self._finalize_response(
            {"type": "error", "message": "Exceeded internal turn limit."}
        )

    def stream_turn(self, prompt: str, excel_paths: list[str]):
        initial = self._prepare_turn(prompt, excel_paths)
        yield from self._auto_step_stream(initial)

    def stream_clarification(self, user_reply: str):
        initial = self._prepare_clarification(user_reply)
        yield from self._auto_step_stream(initial)

    def _cache_session_memory(self) -> None:
        if self._session is None:
            return
        context = (self._session.context_understanding or "").strip()
        if context:
            self._memory.last_context_understanding = context
        active = self._session.current_workbooks
        if active:
            self._memory.last_workbooks = active

    def _restore_session_memory(self, session: SheetHeroSession) -> None:
        if self._memory.last_context_understanding:
            session.context_understanding = self._memory.last_context_understanding
        session.previous_workbooks = self._memory.last_workbooks

    @staticmethod
    def _extract_message(response: Dict[str, object]) -> str:
        result = response.get("result")
        if isinstance(result, dict):
            rendered_text = result.get("rendered_text")
            if rendered_text is not None:
                return str(rendered_text)
            short_answer = result.get("short_answer")
            if short_answer is not None:
                return str(short_answer)
            final_answer = result.get("final_answer")
            if final_answer is not None:
                return str(final_answer)
        return ""

    @staticmethod
    def _looks_like_file_path(value: object) -> bool:
        text = str(value or "").strip().lower()
        return text.endswith((".xlsx", ".xls", ".csv"))

    @classmethod
    def _compact_cell(cls, value: object) -> str:
        text = " ".join(str(value if value is not None else "").split())
        if len(text) > cls._TEXT_PREVIEW_MAX_CELL_LEN:
            return text[: cls._TEXT_PREVIEW_MAX_CELL_LEN - 1] + "…"
        return text

    @classmethod
    def _render_tabular_preview(cls, rows: list[list[object]]) -> tuple[str, bool, int, int]:
        non_empty_rows = [row for row in rows if any(cell not in (None, "") for cell in row)]
        if not non_empty_rows:
            return "", False, 0, 0

        body_rows = non_empty_rows[1:]
        truncated = len(body_rows) > cls._TEXT_PREVIEW_MAX_ROWS
        preview_body = body_rows[: cls._TEXT_PREVIEW_MAX_ROWS]

        def _effective_width(row: list[object]) -> int:
            width = 0
            for idx, cell in enumerate(row, start=1):
                if cell not in (None, ""):
                    width = idx
            return width

        observed_rows = [non_empty_rows[0]] + preview_body
        effective_width = max((_effective_width(row) for row in observed_rows), default=0)
        visible_cols = min(max(effective_width, 1), cls._TEXT_PREVIEW_MAX_COLS)

        header = [cls._compact_cell(cell) for cell in non_empty_rows[0][:visible_cols]]
        col_truncated = effective_width > cls._TEXT_PREVIEW_MAX_COLS

        lines = [" | ".join(header + (["..."] if col_truncated else []))]
        for row in preview_body:
            compact_row = [cls._compact_cell(cell) for cell in row[:visible_cols]]
            if col_truncated:
                compact_row.append("...")
            lines.append(" | ".join(compact_row))

        return "\n".join(lines), truncated or col_truncated, len(preview_body), len(body_rows)

    @classmethod
    def _extract_output_rows(cls, output_path: str) -> tuple[list[list[object]], str]:
        lowered = output_path.lower()
        if lowered.endswith(".csv"):
            with open(output_path, "r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                rows = [row for row in reader]
            return rows, "Output"

        workbook = load_workbook(output_path, data_only=True)
        for sheet in workbook.worksheets:
            rows: list[list[object]] = []
            for row in sheet.iter_rows(values_only=True):
                row_values = list(row)
                if any(cell not in (None, "") for cell in row_values):
                    rows.append(row_values)
            if rows:
                return rows, sheet.title
        return [], ""

    @classmethod
    def _build_text_rendering_from_file(cls, output_path: str) -> Optional[Dict[str, object]]:
        if not output_path or not os.path.exists(output_path):
            return None
        try:
            rows, sheet_name = cls._extract_output_rows(output_path)
        except Exception:
            return None

        preview_text, truncated, preview_rows, total_rows = cls._render_tabular_preview(rows)
        if not preview_text:
            return None

        intro = f"Text preview of generated output ({sheet_name or 'Output'}):"
        note = ""
        if truncated:
            note = (
                f"\n\nPreview truncated to the first {cls._TEXT_PREVIEW_MAX_ROWS} rows"
                f" and {cls._TEXT_PREVIEW_MAX_COLS} columns where needed. "
                "Please enable file mode for the full output."
            )
        return {
            "rendered_text": f"{intro}\n{preview_text}{note}",
            "truncated": truncated,
            "preview_rows": preview_rows,
            "total_rows": total_rows,
        }

    def _apply_text_output_mode(self, response: Dict[str, object]) -> None:
        if (self.config.output_mode or "text").strip().lower() != "text":
            return
        result = response.get("result")
        if not isinstance(result, dict):
            return

        final_answer = result.get("final_answer")
        if not self._looks_like_file_path(final_answer):
            existing_rendered = result.get("rendered_text")
            if existing_rendered is not None:
                result["rendered_text"] = existing_rendered
                result["truncated"] = bool(result.get("truncated", False))
                result["preview_rows"] = result.get("preview_rows")
                result["total_rows"] = result.get("total_rows")
            else:
                result["rendered_text"] = result.get("short_answer") or final_answer or ""
                result["truncated"] = False
                result["preview_rows"] = None
                result["total_rows"] = None
            return

        output_path = str(final_answer).strip()
        rendered = self._build_text_rendering_from_file(output_path)
        fallback = (
            (result.get("short_answer") or "").strip()
            or "Structured output was generated. Please enable file mode for the full result."
        )

        if rendered is None:
            rendered_text = fallback
            truncated = False
            preview_rows = None
            total_rows = None
        else:
            rendered_text = str(rendered["rendered_text"])
            truncated = bool(rendered["truncated"])
            preview_rows = rendered["preview_rows"]
            total_rows = rendered["total_rows"]

        result["raw_final_answer"] = output_path
        result["final_answer"] = rendered_text
        result["short_answer"] = None
        result["rendered_text"] = rendered_text
        result["truncated"] = truncated
        result["preview_rows"] = preview_rows
        result["total_rows"] = total_rows

        if not bool(result.get("edited_existing_file")) and os.path.exists(output_path):
            try:
                os.remove(output_path)
            except OSError:
                pass

    @classmethod
    def _build_result_meta(cls, response: Dict[str, object]) -> Dict[str, object]:
        if response.get("type") != "final":
            return {}

        result = response.get("result")
        if not isinstance(result, dict):
            return {}

        final_answer = result.get("final_answer")
        raw_final_answer = result.get("raw_final_answer")
        source_answer = raw_final_answer if raw_final_answer is not None else final_answer
        text_mode = result.get("rendered_text") is not None or (not cls._looks_like_file_path(source_answer))

        if not text_mode and cls._looks_like_file_path(source_answer):
            output_path = str(source_answer).strip()
            has_output_file = bool(output_path) and os.path.exists(output_path)
            edited_existing_file = bool(result.get("edited_existing_file"))
            return {
                "result_kind": "file",
                "has_output_file": has_output_file,
                "file_created": has_output_file and not edited_existing_file,
                "output_path": output_path if has_output_file else None,
                "output_dir": os.path.dirname(output_path) if has_output_file else None,
                "truncated": False,
                "preview_rows": None,
                "total_rows": None,
            }

        return {
            "result_kind": "text",
            "has_output_file": False,
            "file_created": False,
            "output_path": None,
            "output_dir": None,
            "truncated": bool(result.get("truncated", False)),
            "preview_rows": result.get("preview_rows"),
            "total_rows": result.get("total_rows"),
        }
