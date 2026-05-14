"""Excel editing helpers."""

import re
from typing import Any, List, Optional, Union

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_to_tuple


class ExcelEditor:
    """Editing helpers for worksheets."""

    def __init__(self, workbook, reader):
        self.workbook = workbook
        self.reader = reader

    def insert_rows(self, sheet_name: str, row_index: int, count: int = 1) -> str:
        """Insert empty rows at specific position."""
        try:
            sheet = self.reader.get_sheet(sheet_name)

            if row_index < 1 or count < 1:
                raise ValueError("Row index and count must be >= 1")

            sheet.insert_rows(row_index, count)

            message = (
                f"✅ Inserted {count} row(s) at row {row_index} in sheet '{sheet_name}'"
            )
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error inserting rows: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def insert_columns(self, sheet_name: str, col_index: Union[int, str],
                       count: int = 1) -> str:
        """Insert empty columns at specific position."""
        try:
            sheet = self.reader.get_sheet(sheet_name)

            if isinstance(col_index, str):
                col_index = column_index_from_string(col_index)

            if col_index < 1 or count < 1:
                raise ValueError("Column index and count must be >= 1")

            sheet.insert_cols(col_index, count)

            col_letter = get_column_letter(col_index)
            message = (
                f"✅ Inserted {count} column(s) at column {col_letter} in sheet '{sheet_name}'"
            )
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error inserting columns: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def delete_rows(self, sheet_name: str, start_row: int, count: int = 1) -> str:
        """Delete rows from the sheet."""
        try:
            sheet = self.reader.get_sheet(sheet_name)

            if start_row < 1 or count < 1:
                raise ValueError("Start row and count must be >= 1")
            if start_row > sheet.max_row:
                raise ValueError(
                    f"Start row {start_row} exceeds sheet max row {sheet.max_row}"
                )

            sheet.delete_rows(start_row, count)

            message = (
                f"✅ Deleted {count} row(s) starting from row {start_row} in sheet '{sheet_name}'"
            )
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error deleting rows: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def delete_columns(self, sheet_name: str, start_col: Union[int, str],
                       count: int = 1) -> str:
        """Delete columns from the sheet."""
        try:
            sheet = self.reader.get_sheet(sheet_name)

            if isinstance(start_col, str):
                start_col = column_index_from_string(start_col)

            if start_col < 1 or count < 1:
                raise ValueError("Start column and count must be >= 1")
            if start_col > sheet.max_column:
                raise ValueError(
                    f"Start column {start_col} exceeds sheet max column {sheet.max_column}"
                )

            sheet.delete_cols(start_col, count)

            col_letter = get_column_letter(start_col)
            message = (
                f"✅ Deleted {count} column(s) starting from column {col_letter} in sheet '{sheet_name}'"
            )
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error deleting columns: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def set_cell_value(self, sheet_name: str, cell_ref: str, value: Any) -> str:
        """Set the value of a single cell."""
        try:
            sheet = self.reader.get_sheet(sheet_name)

            if not re.match(r'^[A-Z]+[0-9]+$', cell_ref.upper()):
                raise ValueError(f"Invalid cell reference: {cell_ref}")

            sheet[cell_ref] = value

            message = (
                f"✅ Set cell {cell_ref} to '{value}' in sheet '{sheet_name}'"
            )
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error setting cell value: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def set_range_values(self, sheet_name: str, start_cell: str,
                         values_2d_array: List[List[Any]]) -> str:
        """Set values for a range of cells using a 2D array."""
        try:
            sheet = self.reader.get_sheet(sheet_name)

            if not re.match(r'^[A-Z]+[0-9]+$', start_cell.upper()):
                raise ValueError(f"Invalid cell reference: {start_cell}")

            if not values_2d_array or not isinstance(values_2d_array, list):
                raise ValueError("values_2d_array must be a non-empty list")

            start_row, start_col = coordinate_to_tuple(start_cell)

            for row_idx, row_values in enumerate(values_2d_array):
                if not isinstance(row_values, list):
                    raise ValueError(f"Row {row_idx} must be a list")

                for col_idx, value in enumerate(row_values):
                    current_row = start_row + row_idx
                    current_col = start_col + col_idx
                    sheet.cell(row=current_row, column=current_col, value=value)

            rows_count = len(values_2d_array)
            cols_count = max(len(row) for row in values_2d_array) if values_2d_array else 0
            end_cell = sheet.cell(
                row=start_row + rows_count - 1,
                column=start_col + cols_count - 1
            ).coordinate

            message = (
                f"✅ Set range {start_cell}:{end_cell} ({rows_count}x{cols_count}) in sheet '{sheet_name}'"
            )
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error setting range values: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def copy_range(self, src_sheet: str, src_range: str,
                   dest_sheet: str, dest_cell: str) -> str:
        """Copy data from one range to another (within or across sheets)."""
        try:
            src_ws = self.reader.get_sheet(src_sheet)
            dest_ws = self.reader.get_sheet(dest_sheet)

            if ':' not in src_range:
                raise ValueError("Source range must be in format 'A1:B2'")

            source_data = []
            for row in src_ws[src_range]:
                row_data = [cell.value for cell in row]
                source_data.append(row_data)

            if source_data:
                dest_start_row, dest_start_col = coordinate_to_tuple(dest_cell)

                for row_idx, row_values in enumerate(source_data):
                    for col_idx, value in enumerate(row_values):
                        dest_row = dest_start_row + row_idx
                        dest_col = dest_start_col + col_idx
                        dest_ws.cell(row=dest_row, column=dest_col, value=value)

                rows_count = len(source_data)
                cols_count = len(source_data[0]) if source_data else 0
                dest_end_cell = dest_ws.cell(
                    row=dest_start_row + rows_count - 1,
                    column=dest_start_col + cols_count - 1
                ).coordinate

                message = (
                    f"✅ Copied {src_sheet}!{src_range} to {dest_sheet}!"
                    f"{dest_cell}:{dest_end_cell}"
                )
                print(message)
                return message

            message = "⚠️ No data found in source range"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error copying range: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def add_formula(self, sheet_name: str, cell_ref: str, formula: str) -> str:
        """
        Add Excel formula to cell.
        Automatically prepends "=" if missing.
        """
        try:
            sheet = self.reader.get_sheet(sheet_name)

            if not re.match(r'^[A-Z]+[0-9]+$', cell_ref.upper()):
                raise ValueError(f"Invalid cell reference: {cell_ref}")

            if not formula.startswith('='):
                formula = '=' + formula

            sheet[cell_ref] = formula

            message = (
                f"✅ Added formula '{formula}' to cell {cell_ref} in sheet '{sheet_name}'"
            )
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error adding formula: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)
