"""Output workbook writer helpers."""

import os
import re
from typing import Any, Dict, List, Optional

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple


class ExcelOutputWriter:
    """Write to a separate output workbook and manage artifacts."""

    def __init__(self, workbook, excel_path: str,
                 output_path: Optional[str], temp_files: list):
        self.workbook = workbook
        self.excel_path = excel_path
        self.output_path = output_path
        self._temp_files = temp_files
        self._output_workbook = None
        self._write_regions: Dict[str, List[tuple[int, int, int, int, str]]] = {}

    def save_workbook(self) -> str:
        """
        Save OUTPUT workbook to new file (input files remain unchanged).
        """
        if self.output_path:
            filename = self._normalize_output_path(self.output_path)
        else:
            dir_path = os.path.dirname(self.excel_path)
            base_name = os.path.splitext(os.path.basename(self.excel_path))[0]
            filename = os.path.join(dir_path, f"{base_name}_output.xlsx")
            filename = self._normalize_output_path(filename)

        output_wb = self._get_output_workbook()

        if len(output_wb.sheetnames) == 0:
            output_wb.create_sheet("Output")

        output_wb.save(filename)

        for temp_file in self._temp_files:
            try:
                if os.path.exists(temp_file):
                    os.unlink(temp_file)
            except Exception as e:
                print(f"⚠️ Warning: Could not delete temporary file {temp_file}: {e}")

        self._temp_files = []

        print(f"💾 Workbook saved to: {filename}")
        return filename

    def _get_output_workbook(self):
        """Get or create the output workbook (separate from input files)."""
        from openpyxl import Workbook as OpenpyxlWorkbook
        if self._output_workbook is None:
            self._output_workbook = OpenpyxlWorkbook()
            if 'Sheet' in self._output_workbook.sheetnames:
                del self._output_workbook['Sheet']
        return self._output_workbook

    def create_output_sheet(self, sheet_name: str = "Output") -> str:
        """Create a new sheet in the OUTPUT workbook."""
        try:
            output_wb = self._get_output_workbook()

            if sheet_name in output_wb.sheetnames:
                del output_wb[sheet_name]

            output_wb.create_sheet(sheet_name)
            self._clear_sheet_write_regions(sheet_name)
            message = f"✅ Created output sheet '{sheet_name}' (in new output file)"
            print(message)
            return message
        except Exception as e:
            error_msg = f"❌ Error creating sheet: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def write_dataframe_to_sheet(self, data: Any, sheet_name: str,
                                 start_cell: Any = "A1",
                                 include_header: bool = True) -> str:
        """
        Write a pandas DataFrame or 2D row data to output workbook.
        """
        try:
            output_wb = self._get_output_workbook()

            if sheet_name not in output_wb.sheetnames:
                output_wb.create_sheet(sheet_name)

            sheet = output_wb[sheet_name]

            normalized_start_cell = self._normalize_start_cell(start_cell)
            start_row, start_col = coordinate_to_tuple(normalized_start_cell)
            rows = self._coerce_tabular_rows(data, include_header=include_header)
            rows_written = len(rows)
            cols_written = max(len(row) for row in rows) if rows else 0
            end_row = start_row + rows_written - 1
            end_col = start_col + cols_written - 1
            end_cell = f"{get_column_letter(end_col)}{end_row}"

            self._ensure_non_overlapping_range(
                sheet_name,
                start_row,
                start_col,
                end_row,
                end_col,
                "write_dataframe_to_sheet",
            )

            for row_idx, row_data in enumerate(rows):
                for col_idx, value in enumerate(row_data):
                    current_row = start_row + row_idx
                    current_col = start_col + col_idx
                    normalized_value = self._normalize_output_value(value)
                    sheet.cell(row=current_row, column=current_col, value=normalized_value)

            self._register_write_range(
                sheet_name,
                start_row,
                start_col,
                end_row,
                end_col,
                "write_dataframe_to_sheet",
            )

            message = f"✅ Wrote {rows_written} rows to {sheet_name}!{normalized_start_cell}:{end_cell}"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error writing data to sheet: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def _coerce_tabular_rows(self, data: Any, include_header: bool = True) -> List[List[Any]]:
        """Accept DataFrame-like objects or row lists and normalize into a 2D list."""
        if data is None:
            return []

        if hasattr(data, "columns") and hasattr(data, "values"):
            header = [list(data.columns)] if include_header else []
            try:
                values = data.values.tolist()
            except Exception:
                values = []
            return header + [list(row) if isinstance(row, (list, tuple)) else [row] for row in values]

        if isinstance(data, tuple):
            return [list(data)]

        if isinstance(data, list):
            if not data:
                return []
            if all(not isinstance(row, (list, tuple)) for row in data):
                return [list(data)]
            normalized_rows: List[List[Any]] = []
            for row in data:
                if isinstance(row, list):
                    normalized_rows.append(row)
                elif isinstance(row, tuple):
                    normalized_rows.append(list(row))
                else:
                    normalized_rows.append([row])
            return normalized_rows

        return [[data]]

    @staticmethod
    def _normalize_start_cell(start_cell: Any) -> str:
        if isinstance(start_cell, str):
            return start_cell
        if isinstance(start_cell, int):
            if start_cell < 1:
                raise ValueError("start_cell row number must be >= 1.")
            return f"A{start_cell}"
        raise ValueError("start_cell must be an Excel cell string like 'A1' or a 1-based row number.")

    _TIME_DECIMAL_RE = re.compile(r"^\s*(\d+)\.0:(\d+)\.0\s*$")

    def _normalize_output_value(self, value: Any) -> Any:
        """Normalize common model formatting artifacts before writing."""
        if isinstance(value, dict):
            for scalar_key in ("contains_cycle", "Contains_Cycle", "success", "value", "answer"):
                scalar_value = value.get(scalar_key)
                if isinstance(scalar_value, (str, int, float, bool)):
                    return scalar_value
        if not isinstance(value, str):
            return value

        matched = self._TIME_DECIMAL_RE.match(value)
        if matched:
            hour = int(matched.group(1))
            minute = int(matched.group(2))
            return f"{hour:02d}:{minute:02d}"

        return value

    def highlight_rows(self, sheet_name: str, row_numbers: List[int],
                       format_dict: Dict[str, Any] = None) -> str:
        """
        Highlight entire rows in output workbook.
        """
        try:
            if format_dict is None:
                format_dict = {"fill_color": "red"}

            output_wb = self._get_output_workbook()
            if sheet_name not in output_wb.sheetnames:
                raise ValueError(
                    f"Sheet '{sheet_name}' not found in output workbook. Create it first with create_output_sheet()."
                )
            sheet = output_wb[sheet_name]
            max_col = sheet.max_column
            normalized_rows = self._normalize_row_numbers(row_numbers)
            if not normalized_rows:
                raise ValueError("No valid row numbers provided for highlighting.")

            for row_num in normalized_rows:
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_num, column=col_idx)

                    if 'fill_color' in format_dict:
                        color = self._parse_color(format_dict['fill_color'])
                        cell.fill = PatternFill(
                            start_color=color, end_color=color, fill_type='solid'
                        )

                    font_kwargs = {}
                    if 'font_color' in format_dict:
                        font_kwargs['color'] = self._parse_color(format_dict['font_color'])
                    if 'bold' in format_dict:
                        font_kwargs['bold'] = format_dict['bold']
                    if font_kwargs:
                        cell.font = Font(**font_kwargs)

            message = f"✅ Highlighted row(s) {normalized_rows} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error highlighting rows: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def _normalize_row_numbers(self, row_numbers: Any) -> List[int]:
        """Normalize nested/index-like row inputs into a flat positive-int list."""
        if row_numbers is None:
            return []

        def _iter_items(value: Any):
            if value is None:
                return
            if isinstance(value, (list, tuple, set)):
                for item in value:
                    yield from _iter_items(item)
                return
            if hasattr(value, "tolist"):
                try:
                    converted = value.tolist()
                    yield from _iter_items(converted)
                    return
                except Exception:
                    pass
            yield value

        normalized: List[int] = []
        for item in _iter_items(row_numbers):
            try:
                row_num = int(item)
            except Exception:
                continue
            if row_num >= 1:
                normalized.append(row_num)

        # Deduplicate while preserving order.
        seen = set()
        deduped: List[int] = []
        for row in normalized:
            if row in seen:
                continue
            seen.add(row)
            deduped.append(row)
        return deduped

    def save_workbook_to(self, output_path: str) -> str:
        """Save the output workbook to a specific path."""
        try:
            output_wb = self._get_output_workbook()

            if len(output_wb.sheetnames) == 0:
                output_wb.create_sheet("Output")

            output_path = self._normalize_output_path(output_path)
            output_wb.save(output_path)

            for temp_file in self._temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                except Exception:
                    pass
            self._temp_files = []

            message = f"💾 Workbook saved to: {output_path}"
            print(message)
            return output_path

        except Exception as e:
            error_msg = f"❌ Error saving workbook: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    def add_summary_row(self, sheet_name: str, row_number: int,
                        summary_data: Any) -> str:
        """Add a summary row with labeled statistics to the output workbook."""
        try:
            output_wb = self._get_output_workbook()
            if sheet_name not in output_wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found in output workbook.")
            sheet = output_wb[sheet_name]
            summary_items = self._coerce_summary_items(summary_data)
            if not summary_items:
                raise ValueError("summary_data must contain at least one label/value pair.")
            start_col = 1
            end_col = start_col + ((len(summary_items) - 1) * 3) + 1
            self._ensure_non_overlapping_range(
                sheet_name,
                row_number,
                start_col,
                row_number,
                end_col,
                "add_summary_row",
            )
            col = 1

            for label, value in summary_items:
                sheet.cell(row=row_number, column=col, value=label)
                sheet.cell(row=row_number, column=col + 1, value=value)
                col += 3

            self._register_write_range(
                sheet_name,
                row_number,
                start_col,
                row_number,
                end_col,
                "add_summary_row",
            )

            message = f"✅ Added summary row at row {row_number} in sheet '{sheet_name}'"
            print(message)
            return message

        except Exception as e:
            error_msg = f"❌ Error adding summary row: {str(e)}"
            print(error_msg)
            raise Exception(error_msg)

    @staticmethod
    def _coerce_summary_items(summary_data: Any) -> List[tuple[Any, Any]]:
        if summary_data is None:
            return []
        if isinstance(summary_data, dict):
            normalized_keys = {str(key).strip().lower() for key in summary_data.keys()}
            if normalized_keys == {"task", "value"}:
                return [(summary_data.get("Task"), summary_data.get("Value"))]
            if normalized_keys == {"metric", "value"}:
                return [(summary_data.get("Metric"), summary_data.get("Value"))]
            if normalized_keys == {"label", "value"}:
                return [(summary_data.get("Label"), summary_data.get("Value"))]
            return list(summary_data.items())
        if isinstance(summary_data, (list, tuple)):
            items: List[tuple[Any, Any]] = []
            for item in summary_data:
                if isinstance(item, (list, tuple)) and len(item) >= 2:
                    items.append((item[0], item[1]))
                else:
                    raise ValueError(
                        "summary_data list items must be `[label, value]` pairs."
                    )
            if items and str(items[0][0]).strip().lower() in {"metric", "task", "label"} and str(items[0][1]).strip().lower() == "value":
                items = items[1:]
            return items
        raise ValueError("summary_data must be a dict or list of [label, value] pairs.")

    def _parse_color(self, color: str) -> str:
        color_names = {
            'red': 'FF0000', 'green': '00FF00', 'blue': '0000FF',
            'yellow': 'FFFF00', 'orange': 'FFA500', 'purple': '800080',
            'pink': 'FFC0CB', 'brown': 'A52A2A', 'black': '000000',
            'white': 'FFFFFF', 'gray': '808080', 'grey': '808080'
        }

        if color.startswith('#'):
            return color[1:]
        if color.lower() in color_names:
            return color_names[color.lower()]
        return color

    def _normalize_output_path(self, output_path: str) -> str:
        if not output_path:
            return output_path
        expanded = os.path.expanduser(output_path)
        if os.path.isdir(expanded):
            expanded = os.path.join(expanded, "output.xlsx")
        os.makedirs(os.path.dirname(expanded), exist_ok=True)
        return expanded

    def _clear_sheet_write_regions(self, sheet_name: str) -> None:
        self._write_regions.pop(sheet_name, None)

    def _register_write_range(
        self,
        sheet_name: str,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        source: str,
    ) -> None:
        self._write_regions.setdefault(sheet_name, []).append(
            (start_row, start_col, end_row, end_col, source)
        )

    def _ensure_non_overlapping_range(
        self,
        sheet_name: str,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        source: str,
    ) -> None:
        if end_row < start_row or end_col < start_col:
            raise ValueError("Output write range is empty or invalid.")

        for existing_start_row, existing_start_col, existing_end_row, existing_end_col, existing_source in self._write_regions.get(sheet_name, []):
            if not self._ranges_overlap(
                start_row,
                start_col,
                end_row,
                end_col,
                existing_start_row,
                existing_start_col,
                existing_end_row,
                existing_end_col,
            ):
                continue
            new_range = self._format_range(sheet_name, start_row, start_col, end_row, end_col)
            existing_range = self._format_range(
                sheet_name,
                existing_start_row,
                existing_start_col,
                existing_end_row,
                existing_end_col,
            )
            raise ValueError(
                "WRITE_RANGE_OVERLAP: attempted to write "
                f"{new_range} via `{source}`, but it overlaps existing range "
                f"{existing_range} from `{existing_source}`. "
                "Write later blocks strictly below or to the right of earlier output."
            )

    @staticmethod
    def _ranges_overlap(
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        other_start_row: int,
        other_start_col: int,
        other_end_row: int,
        other_end_col: int,
    ) -> bool:
        rows_overlap = not (end_row < other_start_row or other_end_row < start_row)
        cols_overlap = not (end_col < other_start_col or other_end_col < start_col)
        return rows_overlap and cols_overlap

    @staticmethod
    def _format_range(sheet_name: str, start_row: int, start_col: int, end_row: int, end_col: int) -> str:
        start_cell = f"{get_column_letter(start_col)}{start_row}"
        end_cell = f"{get_column_letter(end_col)}{end_row}"
        return f"{sheet_name}!{start_cell}:{end_cell}"
